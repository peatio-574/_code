# -*- coding: utf-8 -*-
import sys
import time

import random
from pathlib import Path

import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

sys.path.append(str(Path(__file__).parent.parent))

from PlayWright import Playwright_, logger

import pandas

def extract_jd_data(input_file, output_file):
    """
    读取xlsx所有sheet页，提取商品标准名称、京东零售价金额、链接三个字段
    并保存至新的xlsx，保持sheet名称一致
    """
    try:
        # 读取Excel文件的所有sheet名称
        excel_file = pd.ExcelFile(input_file)
        sheet_names = excel_file.sheet_names

        logger.info(f'共找到 {len(sheet_names)} 个sheet页: {sheet_names}')

        # 创建ExcelWriter用于写入新文件
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:

            for sheet_name in sheet_names:
                logger.info(f'正在处理sheet: {sheet_name}')

                # 读取当前sheet
                df = pd.read_excel(input_file, sheet_name=sheet_name)

                # 查找目标列（支持模糊匹配）
                target_columns = []
                for col in df.columns:
                    col_str = str(col).strip()
                    if '商品标准名称' in col_str or '商品名称' in col_str:
                        target_columns.append(('商品标准名称', col))
                    elif '京东零售价金额' in col_str:
                        target_columns.append(('京东零售价金额', col))
                    elif '链接' in col_str or 'url' in col_str.lower():
                        target_columns.append(('链接', col))

                # 检查是否包含所有必需字段
                if len(target_columns) < 3:
                    logger.warning(f'Sheet "{sheet_name}" 缺少必要字段，跳过此sheet')
                    # 如果第一个sheet没有这三个字段，直接跳过
                    if sheet_name == sheet_names[0]:
                        logger.info(f'第一个sheet "{sheet_name}" 缺少字段，已跳过')
                        continue
                    else:
                        continue

                # 构建新的DataFrame，只保留需要的列
                new_columns = [tc[1] for tc in target_columns]
                new_df = df[new_columns].copy()

                # 重命名列
                column_mapping = {tc[1]: tc[0] for tc in target_columns}
                new_df.rename(columns=column_mapping, inplace=True)

                # 去除空值行（可选）
                new_df.dropna(how='all', inplace=True)

                # 写入新Excel文件
                new_df.to_excel(writer, sheet_name=sheet_name, index=False)

                # 设置表头样式
                ws = writer.sheets[sheet_name]
                header_font = Font(bold=True, color='FFFFFF', size=11)
                header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                header_alignment = Alignment(horizontal='center', vertical='center')

                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment

                # 自动调整列宽
                for column in ws.columns:
                    max_length = 0
                    col_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

                logger.info(f'Sheet "{sheet_name}" 处理完成，共 {len(new_df)} 行数据')

        logger.info(f'数据提取完成，已保存到: {output_file}')
        return True

    except Exception as e:
        logger.error(f'数据处理失败: {e}')
        import traceback
        logger.error(traceback.format_exc())
        return False


def login():
    """京东登录"""
    logger.info('开始登录京东....')
    url = 'https://www.jd.com/'
    ele = '//li[@id="ttbar-login-2024"]/div[1]'
    key = 'login.jd_cookie6'
    Playwright_.login(url, ele, key)
    logger.info('京东登录成功')

def read_data(file = 'd:/_code/spider_jd/标题.xlsx'):

    file = os.path.abspath(file)
    try:

        header = 0
        # 获取所有sheet名称
        excel_file = pandas.ExcelFile(file)
        all_sheets_data = {}

        for sheet_name in excel_file.sheet_names:
            # 读取每个sheet，指定表头行
            df = pandas.read_excel(file, sheet_name=sheet_name, header=header, keep_default_na=False).astype(str)

            if len(df) == 0:  # 如果sheet为空
                all_sheets_data[sheet_name] = []
                continue

            col_keys = df.columns.tolist()
            sheet_data = []

            # 遍历每一行数据（跳过表头）
            for idx in range(len(df)):
                row_data = {col: df[col].iloc[idx] for col in col_keys}
                # 行号从1开始，因为idx从0开始但我们要排除表头，所以实际数据行号是idx+1
                row_number = idx + 1
                sheet_data.append((row_number, row_data))

            all_sheets_data[sheet_name] = sheet_data

        logger.info('%s文件所有sheet按行读取成功' % file)
        return all_sheets_data
    except Exception as e:
        logger.error('%s文件所有sheet按行读取失败：%s' % (file, e))
        return None


def search(title, price, url):
    Playwright_.input('//input[@aria-label="搜索"]', title, enter=True)
    time.sleep(random.randint(30, 50))
    ele = '(//div[contains(@class,"goodsContainer")]/div[@data-point-id])[1]'
    new_title = Playwright_.get_attribute(f'{ele}//div[contains(@class,"goods_title")]/span', 'title')

    price_ele = f'{ele}//span[contains(@class,"price")]/span'
    new_price = ''
    for price_ in range(1, Playwright_.get_count(price_ele)+1):
        new_price += Playwright_.get_text(f'({price_ele})[{price_}]')
    new_url = Playwright_.get_attribute(ele, 'data-sku')
    new_url = 'https://item.jd.com/' + new_url + '.html'

    return new_title, new_price, new_url
def main():
    file = 'd:/_code/spider_jd/标题.xlsx'
    from openpyxl import load_workbook
    login()
    all_sheets_data = read_data(file)
    wb = load_workbook(file)
    Playwright_.goto('https://search.jd.com/Search?keyword=%E8%8B%B9%E6%9E%9C%E6%89%8B%E6%9C%BA')
    time.sleep(5)
    for sheet_name, sheet_data in all_sheets_data.items():  # , '个护清洁'
        if sheet_name  not in ['美妆护肤']:
            continue
        ws = wb[sheet_name]
        for row in sheet_data:
            row_id = row[0]
            title = row[1].get('商品标准名称')
            price = row[1].get('京东零售价金额')
            url = row[1].get('链接')
            if price and url:
                continue
            else:
                logger.info(f'{sheet_name}:第{row_id}行数据处理开始：{title}')
                new_title, new_price, new_url = search(title, price, url)
                ws.cell(row=row_id+1, column=2, value=new_price)
                ws.cell(row=row_id+1, column=3, value=new_url)
                wb.save(file)
                logger.info(f'{sheet_name}:第{row_id}行数据处理完成：{title}， 新数据：{new_title}')




if __name__ == '__main__':
    # 配置输入输出文件路径
    # input_file = r'D:\software\wechat_file\xwechat_files\wxid_tye1h8rj4god29_27e0\msg\file\2026-05\(已瘦身)链动蔚来全品类商品明细2026版-ljy.xlsx'  # 替换为你的输入文件路径
    # output_file = 'd:/_code/spider_jd/标题.xlsx'  # 替换为你的输出文件路径
    #
    # extract_jd_data(input_file, output_file)
    main()