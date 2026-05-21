# -*- coding: utf-8 -*-
import sys
from pathlib import Path

sys.path.append(str(Path(__file__).parent.parent))


from ReadFile import ReadData
import openpyxl
from Logger import logger
from PlayWright import Playwright_
import re
import time
import random


file = 'd:/_code/spider_jd/京东手机评论爬取.xlsx'
exist_data = ReadData.read_xlsx_col(file)
column = exist_data['评论人']
column_ = exist_data['机型']
comtent = exist_data['评论内容']
exist_data = [comtent[i][:10]+ column[i]+ column_[i] for i in range(len(column))]

exist_count = len(exist_data)  # 已爬取数据行数
current_row_id = exist_count + 1  # 当前行号
current_valid_id = 0  # 当前有效数据条数

wb = openpyxl.load_workbook(file)
ws = wb['Sheet1']


def login():
    """京东登录"""
    logger.info('登录京东....')
    url = 'https://www.jd.com/'
    ele = '//li[@id="ttbar-login-2024"]/div[1]'
    key = 'login.jd_cookie'
    Playwright_.login(url, ele, key)
    logger.info('京东登录成功')


def get_info(phone_id, url_text):
    global exist_data  # 已爬取数据
    global current_row_id  # 当前行号
    global current_valid_id  # 当前有效行数
    global file

    flag = False  # 标记是否有新数据

    # 获取评论数量
    comment_count_ele = '//div[@data-testid="virtuoso-item-list"]/div'
    comment_count = Playwright_.get_count(comment_count_ele)

    comment_count = min(comment_count, 6)  # 每次最多爬取5条数据
    for comement_ in range(1, comment_count):
        # 商品型号
        # phone_type_ele = f'({comment_count_ele})[{comement_}]//span[@class="info"]'
        # phone_type = Playwright_.get_text(phone_type_ele)
        
        # 评论人名
        comment_name_ele = f'({comment_count_ele})[{comement_}]//span[@class="jdc-pc-rate-card-nick"]'
        comment_name = Playwright_.get_text(comment_name_ele)

        # 评论时间
        comment_time_ele = f'({comment_count_ele})[{comement_}]//span[@class="date list"]'
        comment_time = Playwright_.get_text(comment_time_ele)

        # 评论内容
        comment_text_ele = f'({comment_count_ele})[{comement_}]//span[@class="jdc-pc-rate-card-main-desc"]'
        comment_text = Playwright_.get_text(comment_text_ele)

        # 评论ID：评论人名+型号作为唯一标识，避免重复爬取，
        comment_id = comment_text[:10] + comment_name
        if comment_id in exist_data:
            logger.info(f'已爬取过该条数据，评论ID：{comment_id}')
            continue

        # 写入数据
        row_number = current_row_id + 1
        ws.cell(row=row_number, column=1, value=comment_time)
        ws.cell(row=row_number, column=2, value=comment_name)
        # ws.cell(row=row_number, column=3, value=phone_type)
        ws.cell(row=row_number, column=4, value=comment_text)
        ws.cell(row=row_number, column=5, value=phone_id)
        ws.cell(row=row_number, column=6, value=url_text)
        wb.save(file)

        current_valid_id += 1
        current_row_id += 1
        logger.info(f'第{current_valid_id}条有效数据，已保存第 {current_row_id} 行，合计条数：{current_row_id-1}，评论ID：{comment_id}')

        flag = True
        exist_data.append(comment_id)
    return 1 if flag else 0
        
        
def spider_url(url, url_text):
    # 登录
    login()
    
    # 访问商品详情页
    Playwright_.goto(url)
    phone_id = re.findall(r'\d+', url)[0]

    logger.info(f'开始{url}爬取数据，初始数据数量：{exist_count}')
    
    # 滚动页面，查看评价
    Playwright_.page.keyboard.press('PageDown')
    time.sleep(2)
    logger.info('点击查看更多')
    Playwright_.click('//div[@class="applause-rate golden"]')
    time.sleep(5)
    
    flag = 0  # 是否有新数据
    roll_time = 0  # 滚动次数
    limit_roll_time = 6
    while True:
        # 爬取数据足够就退出，获取连续6次未获取到新数据就退出
        if exist_count == 1000:
            logger.info('已爬取所有评论')
            return True

        # 获取评论
        flag += get_info(phone_id, url_text)
        # 滚动页面
        down_size = random.randint(900, 1500)
        Playwright_.page.mouse.wheel(0, down_size)  # 向下滚动
        roll_time += 1


        if roll_time % limit_roll_time == 0 and roll_time >200:
            if flag == 0:
                logger.info(f'已连续滚动{limit_roll_time}次未获取到新数据，退出')
                return False
            flag = 0

        # 睡眠
        if flag != 0:

            sleep_sec = random.randint(10, 15)
        else:
            sleep_sec = random.randint(5, 8)
        logger.info(f'已滚动{roll_time}次，睡眠{sleep_sec}秒')
        time.sleep(sleep_sec)
        flag = 0


if __name__ == '__main__':

    url_dict = {
        # 'https://item.jd.com/10122849689934.html': '352 H301加湿器空气净化器无雾加湿净化一体机【滤芯配件】 H301加湿滤网',
        # 'https://item.jd.com/10116339176210.html': '352空气净化加湿器 超大加湿量 无雾加湿器+专业空气净化器一机多能 四季可用H301 H301',
        # 'https://item.jd.com/10101252435118.html': '352H300/H301加湿器空气净化器无雾加湿净化一体机 高效除醛过滤器 配件',
        # 'https://item.jd.com/10089842490086.html': '352 标配抑菌模块 无雾加湿器空气净化器专用配件 适用于H70/H80/H300/H301/H500',
        # 'https://item.jd.com/10089841142845.html': '352加湿器无雾加湿器标配滤网 适用于H70/H80 配件',
        # 'https://item.jd.com/10066903506375.html': '352 H300加湿器空气净化器无雾加湿净化一体机【滤芯配件】 H300加湿滤网',
        # 'https://item.jd.com/10066903506376.html': '352 H300加湿器空气净化器无雾加湿净化一体机【滤芯配件】 H300/H301空净滤网',
        # 'https://item.jd.com/10090762282714.html': '352X63 Pet宠物净化器滤芯 家用杀菌消毒吸猫毛宠物除尘螨杀菌消毒除臭家用滤芯 高效除味过滤器（内：黑）',
        # 'https://item.jd.com/10090762282715.html': '352X63 Pet宠物净化器滤芯 家用杀菌消毒吸猫毛宠物除尘螨杀菌消毒除臭家用滤芯 高效空气过滤器（外：白）',
        # 'https://item.jd.com/10088090061995.html': '352X50滤芯pro标准版滤芯【适用于X50/X50S】 50pro滤芯',
        # 'https://item.jd.com/10051465849921.html': '352 X66C/X63C滤芯 空气净化器原装滤芯 除甲醛滤芯 除雾霾灰尘异味 滤网',
        # 'https://item.jd.com/10033095845995.html': '352 Y106C家用空气净化器滤芯 2片装 Y106C过滤器套装',
        # 'https://item.jd.com/10033095433012.html': '352 Y106/Y100家用空气净化器滤芯 2片装 Y106/Y100过滤器套装',
        # 'https://item.jd.com/68068751599.html': '352 Y100C空气净化器滤芯套装 2片装 专业除醛款 Y100C滤芯',
        'https://item.jd.com/47779981996.html': '352 X83C Plus顶层高效复合过滤器Ⅰ空气净化器家用 除甲醛 除雾霾 滤芯'
    }

    for url, url_text in url_dict.items():
        spider_url(url, url_text)
        # time.sleep(120)

