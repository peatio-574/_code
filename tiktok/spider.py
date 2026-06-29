# -*- coding: utf-8 -*-
import sys
from pathlib import Path

sys.path.append(str(Path(__file__).parent.parent))

from PlayWright import Playwright_, logger
import time
import os
import requests
from ReadFile import ReadData
from openpyxl import Workbook, load_workbook
from resolve_tiktok_desktop_link import resolve_tiktok_link

picture_dir = os.path.join(os.path.dirname(__file__), 'pictures')
os.makedirs(picture_dir, exist_ok=True)

proxy = {
    "server": 'http://127.0.0.1:7892',
    "username": '2953515027',
    "password": '121920.pS',
}
http_proxy = {
    "http": proxy["server"],
    "https": proxy["server"],
}

headers = {
    'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
    'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/149.0.0.0 Safari/537.36'
}

save_file = os.path.join(os.path.dirname(__file__), '商品数据.xlsx')
if os.path.exists(save_file):
    # 如果文件存在，加载现有工作簿
    wb = load_workbook(save_file)
    ws = wb.active
else:
    xlsx_headers = ['商品id', '商品名称', '商品原始链接', '商品链接', '一级类目',
                    '类目', '商品主图', '商品图片', '商品详情', '商品描述',
                    'SKU名称', 'SKU图片', 'SKU价格']
    wb = Workbook()
    ws = wb.active
    ws.title = '数据'
    ws.append(xlsx_headers)
    wb.save(save_file)

def save_picture(picture_url, filename):
    picture_file = os.path.join(picture_dir, filename)
    try:
        with open(picture_file, 'wb') as f:
            f.write(requests.get(picture_url, headers=headers, proxies=http_proxy).content)
        # logger.info(f'图片保存成功：{filename}')
        return filename
    except Exception as e:
        logger.error(f'链接：{picture_url}')
        logger.error(f'图片下载失败: {e}')
        return picture_url


def get_main_info(url, product_id):
    Playwright_.goto(url, proxy=proxy)

    verify_ele = '//img[@id="captcha-verify-image"]'
    if Playwright_.get_count(verify_ele) == 1:
        exit('captcha required, please verify manually')

    Playwright_.wait_for_selector('//div[text()="登录"]', timeout=60 * 1000)

    product_name = Playwright_.get_text('//h1/span')

    category_ele = '//ol/li/a/span'
    category_count = Playwright_.get_count(category_ele)
    category_list = [Playwright_.get_text(f'({category_ele})[{i}]') for i in range(1, category_count+1)]
    first_category = category_list[0]
    category = '>'.join(category_list)

    main_picture_ele = '//div[contains(@class, "relative overflow-visible")]/div/img'
    main_picture_count = Playwright_.get_count(main_picture_ele)
    main_picture_list = [Playwright_.get_attribute(f'({main_picture_ele})[{i}]', 'src') for i in range(1, main_picture_count+1)]
    real_main_pictures = []
    for idx, main_picture in enumerate(main_picture_list, start=1):
        real_main_picture = save_picture(main_picture, f'{product_id}_main_{idx}.webp')
        real_main_pictures.append(real_main_picture)

    first_main_picture = real_main_pictures[1]
    real_main_pictures = '\n'.join(real_main_pictures)

    more_ele = '//div[text()="View more"]'
    if Playwright_.get_count(more_ele) == 1:
        Playwright_.click(more_ele)
        time.sleep(1)
    if Playwright_.get_count(more_ele) == 1:
        Playwright_.click(more_ele)
        time.sleep(1)

    detail_ele = '//tbody/tr'
    detail_count = Playwright_.get_count(detail_ele)
    detail = '\n'.join([f'{Playwright_.get_text(f"({detail_ele})[{i}]/td[1]/span")}: {Playwright_.get_text(f"({detail_ele})[{i}]/td[2]/span")}' for i in range(1, detail_count+1)])

    describe_ele = '//div[contains(@class, "whitespace-normal")]'
    describe_count = Playwright_.get_count(describe_ele)
    describe = '\n'.join([Playwright_.get_text(f'({describe_ele})[{i}]') for i in range(1, describe_count+1)])
    return {
        'product_id': product_id,
        'product_name': product_name,
        'category': category,
        'first_category': first_category,
        'real_main_pictures': real_main_pictures,
        'first_main_picture': first_main_picture,
        'detail': detail,
        'describe': describe,
    }

def get_sku_info(product_id, main_info):
    """获取sku数据"""
    sku_ele = '//div[contains(@class, "flex flex-row overflow-x-auto")]/div'
    sku_count = Playwright_.get_count(sku_ele)
    price_ele = '//span[@class="flex flex-row items-baseline"]/span/span'
    valid_sku_count = 0
    for i in range(1, sku_count+1):
        current_ele = f'({sku_ele})[{i}]'
        Playwright_.click(f'{current_ele}/span')
        time.sleep(2)

        sku_picture_ele = f'{current_ele}/img'
        if Playwright_.get_count(sku_picture_ele) == 0:
            continue
        sku_picture = Playwright_.get_attribute(sku_picture_ele, 'src')
        real_sku_picture = save_picture(sku_picture, f'{product_id}_sku_{i}.webp')

        sku_name = Playwright_.get_text(f'{current_ele}/span')

        price_count = Playwright_.get_count(price_ele)
        sku_price = [Playwright_.get_text(f'({price_ele})[{j}]') for j in range(1, price_count+1)]
        sku_price = ''.join(sku_price)
        sku_price = ''.join(sku_price)

        ws.append([
            product_id,
            main_info['product_name'],
            main_info['source_url'],
            main_info['product_url'],
            main_info['first_category'],
            main_info['category'],
            main_info['first_main_picture'],
            main_info['real_main_pictures'],
            main_info['detail'],
            main_info['describe'],
            sku_name,
            real_sku_picture,
            sku_price,
        ])
        logger.info(f'{product_id}商品第{i}条sku信息获取成功')
        wb.save(save_file)
        valid_sku_count += 1
    logger.info(f'{product_id}商品sku信息保存成功，共{valid_sku_count}条\n\n\n')



def resolve_short_link(short_url):
    """
    返回商品id
    """
    try:
        result = resolve_tiktok_link(short_url, proxies=http_proxy)
    except Exception as e:
        logger.error(f"解析失败：{e}", file=sys.stderr)
        sys.exit(1)
    return result.product_id if result.product_id else None


def run():
    data_file = os.path.join(os.path.dirname(__file__), '链接.xlsx')
    links = ReadData.read_xlsx_col(data_file)['链接']
    for link in links:
        logger.info(f'开始处理链接：{link}')
        product_id = resolve_short_link(link)
        logger.info(f'获取到对应product_id为：{product_id}')
        if product_id is None:
            continue

        product_url = f'https://shop.tiktok.com/my/pdp/x/{product_id}'

        main_info = get_main_info(product_url, product_id)
        main_info['source_url'] = link
        main_info['product_url'] = product_url
        logger.info(f'{product_url}对应的商品主信息获取成功')
        get_sku_info(product_id, main_info)


if __name__ == '__main__':
    run()