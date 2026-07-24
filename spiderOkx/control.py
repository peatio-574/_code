# coding='utf-8'
import sys
import os

if getattr(sys, 'frozen', False):
    BUNDLE_DIR = sys._MEIPASS
    BASE_DIR = os.path.dirname(sys.executable)
else:
    BUNDLE_DIR = os.path.dirname(os.path.abspath(__file__))
    BASE_DIR = BUNDLE_DIR

sys.path.insert(0, BUNDLE_DIR)

from PlayWright import Playwright_, logger, get_config_value
import time
from ReadFile import ReadData
from openpyxl import load_workbook, Workbook
import requests
import re


configFile = os.path.join(BASE_DIR, 'config.ini')
fileName = os.path.join(BASE_DIR, '欧易数据.xlsx')


def okxLogin():
    logger.info('开始登录欧易....')
    url = 'https://www.btmzhjpdgxc.com/zh-hans/orbit/feed'
    ele = '//a[text()="充币" and @class="deposit-btn"]'
    key = 'login.okxCookie'
    Playwright_.login(url, ele, key, file=configFile)
    logger.info('✅️ 欧易登录成功')


def getDatas():
    okxLogin()
    Playwright_.close()

    if os.path.exists(fileName):
        # 如果文件存在，加载现有工作簿
        wb = load_workbook(fileName)
        ws = wb.active
    else:
        headers = ['发布时间', '作者', '帖子内容']
        wb = Workbook()
        ws = wb.active
        ws.title = '数据'
        ws.append(headers)
        wb.save(fileName)


    existData = ReadData.read_xlsx_col(fileName)['发布时间']
    cursor = ''
    while True:
        if cursor == '0':
            logger.info('没有更多数据了')
            break
        cursor = getSinglePageInfo(ws, wb, fileName, existData, cursor)
        time.sleep(10)


def getSinglePageInfo(ws, wb, fileName, existData, cursor=''):
    logger.info(f'\n开始获取数据....cursor：{cursor}')
    ms_timestamp = int(time.time() * 1000)
    cursor = f'&cursor={cursor}' if cursor else ''

    url = f'https://www.btmzhjpdgxc.com/priapi/v5/rubik/public/content/feed/1?size=10{cursor}&t={ms_timestamp}'

    token = ''
    cookies = get_config_value('login', 'okxCookie', file=configFile)
    for cookie in eval(cookies):
        if cookie.get('name') == 'token':
            token = cookie.get('value')
    headers = {
        "Accept": "application/json",
        "content-type": "application/json;charset=UTF-8",
        "referer": "https://www.btmzhjpdgxc.com/zh-hans/orbit/feed",
        "app-type": "web",
        "x-id-group": "2130948145125530008-c-14",
        "devid": "afa0643a-cfee-4783-a5a2-29b321fd234d",
        "cookie": get_config_value('login', 'okxCookie_api', file=configFile),
        "authorization": token,
        # "x-fptoken-signature": "{P1363}QZeatxR9NsjI2VRNDzWgvlj9MZ+EaVifI5UXBex5fmrJu4+uBM1QLEVVtCijIELcW6H2Fd2ikZoVmsiqg2Rd+Q==",
        # "x-fptoken": "eyJraWQiOiIxNjgzMzgiLCJhbGciOiJFUzI1NiJ9.eyJpYXQiOjE3ODQ4MTQ1MTksImVmcCI6IkRKbm5VQ2h3MmJCWW9oTWlGdVVHUUpseHZJbTVUKys2eTliRGtISHBkdXBkSkRNZFRTK2hsRGZGMUptZDk1VFoiLCJkaWQiOiJhZmEwNjQzYS1jZmVlLTQ3ODMtYTVhMi0yOWIzMjFmZDIzNGQiLCJjcGsiOiJNRmt3RXdZSEtvWkl6ajBDQVFZSUtvWkl6ajBEQVFjRFFnQUUwU2k1d1BnVkhpaDBNSytqU3ZkUFVmaWs1MnpDcTBaZlVoNHhWY3lDc2dFbHNlODk5Sjl6Y1Fub3VMWEZRWmFEb1hQMmJTK3huS1oxWC9sNUgzT01rdz09In0.SyCAS4n9sFcxsOVth_JnU_t5ZcDOJVu38AfPn-Lve_xz_tjy_pB3EdzmAVsdPBaSHGEO5koEngBtzcl5rLXAGA",
        "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/150.0.0.0 Safari/537.36 Edg/150.0.0.0",

    }
    response = requests.get(url, headers=headers).json()['data']

    nextCursor = response['nextCursor']
    count = 0
    for row in response['dataList']:
        pubulishTime = row['contentData']['publishTime']

        if int(time.time() * 1000) - int(pubulishTime) > 3 * 24 * 60 * 60 * 1000:
            logger.info(f'数据不在三天之内，退出程序')
            exit()

        pubulishTime = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(int(pubulishTime)/1000))

        if pubulishTime in existData:
            continue

        author = row['contentData']['author']['nickName']
        content = row['contentData']['content']
        logger.info(f'发布时间：{pubulishTime} 作者：{author} 内容：{content[:10]}')
        ws.append([pubulishTime, author, content])
        existData.append(pubulishTime)
        count += 1
    wb.save(fileName)
    logger.info(f'本次保存{count}条数据成功')
    return nextCursor


def analyzeCoins(text):
    """提取帖子中提到的币种"""
    coin_patterns = {
        'BTC': r'\$BTC|#BTC|比特币(?![^)]*现货)|BTC(?!\w)',
        'ETH': r'\$ETH|#ETH|以太坊(?![^)]*现货)|ETH(?!\w)',
        'SOL': r'\$SOL|#SOL|SOL(?!\w)',
        'XRP': r'\$XRP|#XRP|XRP(?!\w)',
        'SNDK': r'\$SNDK|闪迪|SNDK(?!\w)',
        'MU': r'\$MU|美光|MU(?!\w)',
        'NVDA': r'\$NVDA|英伟达|NVDA(?!\w)',
        'TSLA': r'\$TSLA|特斯拉|TSLA(?!\w)',
        'GOOGL': r'\$GOOGL|谷歌(?!云)|GOOGL(?!\w)',
        'SPCX': r'\$SPCX|SpaceX|SPCX(?!\w)',
        'HYPE': r'\$HYPE|HYPE(?!\w)',
        'CL': r'\$CL|原油|CL(?!\w)',
        'BZ': r'\$BZ|BZ(?!\w)',
        'SKHY': r'\$SKHY|海力士|SKHY(?!\w)',
        'WLFI': r'\$WLFI|WLFI(?!\w)',
        'TRUMP': r'\$TRUMP|TRUMP(?!\w)',
        'KAITO': r'\$KAITO|KAITO(?!\w)',
        'LAB': r'\$LAB|LAB(?!\w)',
        'UNI': r'\$UNI|UNI(?!\w)',
        'LDO': r'\$LDO|LDO(?!\w)',
        'ENA': r'\$ENA|ENA(?!\w)',
        'INTC': r'\$INTC|英特尔|INTC(?!\w)',
        'PLTR': r'\$PLTR|PLTR(?!\w)',
        'WBTC': r'\$WBTC|WBTC(?!\w)',
        'AMD': r'\$AMD|AMD(?!\w)',
        'MRVL': r'\$MRVL|MRVL(?!\w)',
        'AVGO': r'\$AVGO|AVGO(?!\w)',
        'TSM': r'\$TSM|TSM(?!\w)',
        'ORCL': r'\$ORCL|ORCL(?!\w)',
        'COIN': r'\$COIN|COIN(?!\w)',
        'SATS': r'\$SATS|SATS(?!\w)',
        'CORE': r'\$CORE|CORE(?!\w)',
        'BMEX': r'\$BMEX|BMEX(?!\w)',
    }

    found = []
    for coin, pattern in coin_patterns.items():
        if re.search(pattern, text, re.IGNORECASE):
            found.append(coin)
    return list(set(found))


def analyzePrice(text, coins):
    """提取开仓点位并与币种配对"""
    entries = []

    # 提取所有价格
    price_patterns = [
        r'(\d+[.,]?\d*)[kK]?\s*(?:美元|U|USDT|刀|点)',
        r'(\d+[.,]?\d*)[kK]?\s*(?:附近|左右|上方|下方)',
        r'(\d+[.,]?\d*)[kK]?[-\s到]{1,3}(\d+[.,]?\d*)[kK]?',
        r'开仓[^0-9]*(\d+[.,]?\d*)',
        r'入场[^0-9]*(\d+[.,]?\d*)',
        r'均价[^0-9]*(\d+[.,]?\d*)',
        r'成本[^0-9]*(\d+[.,]?\d*)',
        r'(\d+[.,]?\d*)[kK]?\s*做多',
        r'(\d+[.,]?\d*)[kK]?\s*做空',
    ]

    all_prices = []
    for pattern in price_patterns:
        matches = re.findall(pattern, text)
        for m in matches:
            if isinstance(m, tuple):
                for val in m:
                    num = safe_float(val)
                    if num is not None and 1 <= num <= 1000000:
                        all_prices.append(num)
            else:
                num = safe_float(m)
                if num is not None and 1 <= num <= 1000000:
                    all_prices.append(num)

    all_prices = sorted(list(set(all_prices)))[:5]

    if not all_prices or not coins:
        return entries

    # 按价格区间配对币种
    high_price_coins = ['BTC', 'WBTC']
    mid_price_coins = ['ETH', 'SNDK', 'MU', 'NVDA', 'TSLA', 'GOOGL', 'SPCX', 'INTC', 'PLTR', 'AMD', 'MRVL', 'AVGO',
                       'TSM', 'ORCL']
    low_price_coins = ['CL', 'BZ', 'LAB', 'KAITO', 'WLFI', 'TRUMP', 'CORE']

    for p in all_prices:
        matched = False

        if p > 10000:
            for c in coins:
                if c in high_price_coins:
                    entries.append({'币种': c, '点位': p})
                    matched = True
                    break
        elif 1000 <= p <= 10000:
            for c in coins:
                if c in mid_price_coins:
                    entries.append({'币种': c, '点位': p})
                    matched = True
                    break
        else:
            for c in coins:
                if c in low_price_coins:
                    entries.append({'币种': c, '点位': p})
                    matched = True
                    break

        if not matched and coins:
            entries.append({'币种': coins[0], '点位': p})

    # 去重
    seen = set()
    unique = []
    for e in entries:
        key = f"{e['币种']}_{e['点位']}"
        if key not in seen:
            seen.add(key)
            unique.append(e)
    return unique


def safe_float(val):
    """安全转换为浮点数"""
    try:
        return float(str(val).replace(',', ''))
    except (ValueError, TypeError):
        return None


def analyzePoint(text):
    """提取观点方向：看多/看空/中性"""
    long_patterns = [
        r'做多|看多|多单|买入|买进|抄底|多头|看涨|必须买入|强势关注|上车|博弈多单|低买',
        r'支撑.*?做多|企稳.*?买入|分批.*?买入',
        r'目标.*?\d+[kK]?.*?(?:美元|U|USDT)',
    ]
    short_patterns = [
        r'做空|看空|空单|卖出|卖空|空头|看跌|追空|开空|布空|顺势空',
        r'压力.*?做空|反弹.*?做空|跌破.*?做空',
    ]

    is_long = any(re.search(p, text, re.IGNORECASE) for p in long_patterns)
    is_short = any(re.search(p, text, re.IGNORECASE) for p in short_patterns)

    if re.search(r'不(?:是)?做多|不做多|不看多', text, re.IGNORECASE):
        is_long = False
    if re.search(r'不(?:是)?做空|不做空|不看空', text, re.IGNORECASE):
        is_short = False

    if is_long and is_short:
        long_cnt = sum(1 for p in long_patterns if re.search(p, text, re.IGNORECASE))
        short_cnt = sum(1 for p in short_patterns if re.search(p, text, re.IGNORECASE))
        return '看多' if long_cnt >= short_cnt else '看空'
    elif is_long:
        return '看多'
    elif is_short:
        return '看空'
    return '中性'


def analyze():
    wb = load_workbook(fileName)
    ws = wb.active
    existsData = ReadData.read_xlsx_col(fileName)['帖子内容']
    ws.cell(row=1, column=4, value='涉及币种')
    ws.cell(row=1, column=5, value='开仓点位')
    ws.cell(row=1, column=6, value='帖子观点')
    wb.save(fileName)
    for rowIdx, content in enumerate(existsData, start=2):
        coins = analyzeCoins(content)
        price = analyzePrice(content, coins)
        point = analyzePoint(content)
        if coins:
            ws.cell(row=rowIdx, column=4, value=str(coins))
        if price:
            ws.cell(row=rowIdx, column=5, value=str(price))
        ws.cell(row=rowIdx, column=6, value=point)
    wb.save(fileName)
    logger.info('✅️ 数据处理完成')


if __name__ == '__main__':
    step = input('请输入操作步骤（1.爬取数据，2.数据分析汇总）：')
    if step == '1':
        getDatas()
    elif step == '2':
        analyze()