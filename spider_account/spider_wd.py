# -*- coding: utf-8 -*-
import sys
from pathlib import Path

sys.path.append(str(Path(__file__).parent.parent))

from PlayWright import Playwright_, logger, get_config_value
import time
import os
import pandas
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import warnings


warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

config_file = os.path.join(os.path.dirname(__file__), 'config.ini')

def wd_login(shop_id=1):
    try:
        logger.info('开始登录微店....')
        url = 'https://d.weidian.com/weidian-pc/login/?spider_token=35ed#/shopSelect'
        ele = '//div[@class="nick-name"]'
        key = f'login.wd_cookie{shop_id}'
        shop_id_ = (shop_id - 1) % 5 + 1
        extra = f'(//div[@data-spider-mode="trackAction"])[{shop_id_}]/div[1]'
        Playwright_.login(url, ele, key, extra=extra, file=config_file)
        logger.info('微店登录成功....')
        return True
    except Exception as e:
        logger.error(f'微店登录失败：{e[:50]}')
        return False



def wd_search(start, end):
    shop_name = Playwright_.get_text('//div[@class="user-name"]')
    shop_name = shop_name.strip()
    logger.info(f'当前店铺名称：{shop_name}')
    Playwright_.goto('https://d.weidian.com/weidian-pc/weidian-loader/#/pc-vue-balance/overview')
    time.sleep(8)

    Playwright_.input('//input[@placeholder="开始日期"]', start)
    Playwright_.input('//input[@placeholder="结束日期"]', end, enter=True)
    Playwright_.click('//span[text()="筛选"]')
    return shop_name


def wd_save(start, end, filename):
    """直接导出"""
    try:
        Playwright_.goto('https://d.weidian.com/weidian-pc/weidian-loader/#/pc-vue-balance/exportList?shopType=master')
        time.sleep(5)
        date_info = Playwright_.get_text('(//div[@class="record-item"])[1]//div[@class="item-bd"]/div[2]')

        end = time.strptime(end, '%Y-%m-%d')
        end = time.mktime(end) + 86400  # 加一天
        end = time.strftime('%Y-%m-%d', time.localtime(end))

        if start in date_info and end in date_info:
            # Playwright_.click('(//div[@class="record-item"])[1]//span[text()="下载报表"]')
            # 直接导出
            with Playwright_.page.expect_download(timeout=30000) as download_info:
                Playwright_.click('(//div[@class="record-item"])[1]//span[text()="下载报表"]')
                pass  # 等待下载触发
            download = download_info.value
            # 获取文件名并保存
            download.save_as(filename)
            return True
    except Exception as e:
        logger.info(f'{filename}下载异常：{e}')
    return False


def wd_deal_data(shopname, file):
    """
    使用pandas处理微店数据：
    1. 按提现单独一列
    收入-手续费=净收入
    """
    try:
        # 读取Excel文件
        df = pandas.read_excel(file)

        # 将数值列转换为数字类型（处理可能的文本格式数字）
        df['收入(元)'] = df['收入(元)'].astype(str).str.replace(',', '').str.strip()
        df['收入(元)'] = pandas.to_numeric(df['收入(元)'].replace('', '0'), errors='coerce').fillna(0)

        # 提取日期（从入账时间）
        df['日期'] = pandas.to_datetime(df['时间']).dt.date

        # 根据账单类型分类
        # 1. 提现数据
        df_withdraw = df[df['账单类型'] == '提现']

        # 2. 货款收入数据
        df_goods_income = df[df['账单类型'] == '货款收入']

        # 3. 交易手续费数据
        df_fee = df[df['账单类型'] == '交易手续费']


        # 按日汇总提现金额
        if len(df_withdraw) > 0:
            df_withdraw_summary = df_withdraw.groupby('日期').agg({
                '收入(元)': 'sum'
            }).reset_index()
            df_withdraw_summary.columns = ['日期', '提现金额']
        else:
            df_withdraw_summary = pandas.DataFrame(columns=['日期', '提现金额'])

        # 按日汇总货款收入
        if len(df_goods_income) > 0:
            df_goods_summary = df_goods_income.groupby('日期').agg({
                '收入(元)': 'sum' }).reset_index()
            df_goods_summary.columns = ['日期', '货款收入']
        else:
            df_goods_summary = pandas.DataFrame(columns=['日期', '货款收入'])

        # 按日汇总交易手续费
        if len(df_fee) > 0:
            df_fee_summary = df_fee.groupby('日期').agg({
                '收入(元)': 'sum'
            }).reset_index()
            df_fee_summary.columns = ['日期', '交易手续费']
        else:
            df_fee_summary = pandas.DataFrame(columns=['日期', '交易手续费'])

        # 合并三个汇总表
        df_summary = df_goods_summary.copy()
        df_summary = pandas.merge(df_summary, df_withdraw_summary, on='日期', how='outer')
        df_summary = pandas.merge(df_summary, df_fee_summary, on='日期', how='outer')


        # 填充空值为0
        df_summary = df_summary.fillna(0)


        # 计算每日总净收入
        df_summary['日净收入'] = df_summary['货款收入'] + df_summary['交易手续费']

        # 按日期排序
        df_summary = df_summary.sort_values('日期', ascending=True).reset_index(drop=True)

        # 添加汇总行
        total_row = pandas.DataFrame({
            '日期': ['所有汇总'],
            '货款收入': [df_summary['货款收入'].sum()],
            '提现金额': [df_summary['提现金额'].sum()],
            '交易手续费': [df_summary['交易手续费'].sum()],
            '日净收入': [df_summary['日净收入'].sum()]
        })
        df_summary = pandas.concat([df_summary, total_row], ignore_index=True)

        # 重新排列列顺序
        columns_order = ['日期', '货款收入', '提现金额', '交易手续费', '日净收入']
        df_summary = df_summary[columns_order]

        # 使用 ExcelWriter 追加到现有文件，保留原有的 Sheet
        with pandas.ExcelWriter(file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            sheet_name = '汇总数据'
            # 先删除可能已存在的同名 Sheet
            if sheet_name in writer.book.sheetnames:
                del writer.book[sheet_name]

            # 写入数据
            df_summary.to_excel(writer, sheet_name=sheet_name, index=False)

            # 获取工作表对象并设置样式
            ws = writer.sheets[sheet_name]
            header_font = Font(bold=True, color='FFFFFF', size=11)
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center')

            # 遍历第一行所有单元格设置样式
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # 自动调整列宽
            for column in ws.columns:
                max_length = 0
                col_letter = column[0].column_letter
                for cell in column:
                    if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))


                ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

            # 消除 "Workbook contains no default style" 警告
            workbook = writer.book
            if not workbook.style_names:
                default_font = Font(name='Calibri', size=11, bold=False, italic=False)
                default_style = openpyxl.styles.NamedStyle(name='Normal', font=default_font)
                workbook.add_named_style(default_style)

        logger.info(f'数据汇总完成，共汇总{len(df_summary) - 1}天的数据，已保存到: {file}')

        logger.info(f'总收入: {df_summary["货款收入"].sum()/2:.2f}，交易手续费: {df_summary["交易手续费"].sum()/2:.2f} '
                    f'总提现: {df_summary["提现金额"].sum()/2:.2f}，总净收入: {df_summary["日净收入"].sum()/2:.2f}\n')

        return df_summary

    except Exception as e:
        logger.error(f'{shopname}店铺数据处理失败: {e}\n')
        return None


def main_(account_id=1):
    title = f'========================开始爬取微店第{account_id}个店铺======================='
    logger.info(title)
    end = time.strftime("%Y-%m-%d", time.localtime(time.time() - 86400))
    start = f"{end[:-3]}-01"

    login = wd_login(account_id)
    if not login:
        return False
    shop_name = wd_search(start, end)

    time.sleep(10)

    logger.info(f'{shop_name}店铺开始导出明细....')
    status = False
    st_time = time.time()
    for roll in range(1, 6):
        logger.info(f'第{roll}次导出明细....')
        Playwright_.click('//span[text()="导出报表"]')
        Playwright_.click('//span[text()="生成报表"]')
        time.sleep(5)
        date_ = Playwright_.get_text('(//div[@class="record-item"])[1]/div[1]/div[1]')[9:]
        time_struct = time.strptime(date_, "%Y-%m-%d %H:%M:%S")
        timestamp = int(time.mktime(time_struct))
        if timestamp >= st_time:
            status = True
            break
        else:
            Playwright_.reload()

    text = f'✅️ {shop_name}店铺明细数据预生成中....' if status else f'❌️ {shop_name}店铺明细数据预生成失败'
    logger.info(text)
    Playwright_.clear_cookie()
    return [shop_name, start, end] if status else None



def main_save_deal_data(shop_id, shop_name, start, end, filename):
    logger.info(f'{shop_name}店铺开始导出明细....')
    wd_login(shop_id)
    save_flag = False
    for roll in range(1, 6):
        logger.info(f'第{roll}次导出明细....')
        save_flag = wd_save(start, end, filename)
        if save_flag:
            break
    if save_flag:
        logger.info(f'✅️ {shop_name}店铺明细数据下载成功')
        wd_deal_data(shop_name, filename)
    else:
        logger.info(f'❌️ {shop_name}店铺明细数据下载失败')



if __name__ == '__main__':
    # wd_deal_data('测试', './数据/测试微店.xlsx')
    shop_count_ = get_config_value('login', 'wd_shop_count', file=config_file)
    shop_flag = input('请输入查询店铺序号（0默认查询全部）：')
    if shop_flag == '0':
        start_id = 1
        end_id = int(shop_count_) + 1
    elif ':' in shop_flag or '：' in shop_flag:
        shop_flag = shop_flag.replace('：', '').replace(':', '')
        start_id = int(shop_flag)
        end_id = int(shop_count_) + 1
    else:
        start_id = int(shop_flag)
        end_id = start_id + 1
    current_info = []
    for account_id in range(start_id, end_id):
        shop_reuslt = main_(account_id)
        current_info.append(shop_reuslt)
    wait_mins = 10
    logger.info(f'等待{wait_mins}分钟，正在预生成报表，请勿关闭程序')
    time.sleep(wait_mins*60)

    for shop_id, shopinfo in enumerate(current_info, start=1):
        if shopinfo:
            shop_name, start, end = shopinfo
            dirname = 'd:/_code/spider_account/数据'
            filename = f'微店-{shop_name}店铺{end}明细.xlsx'
            filename = os.path.join(dirname, filename)
            main_save_deal_data(shop_id, shop_name, start, end, filename)
        else:
            logger.info(f'❌️ 第{shop_id}个店铺明细数据预生成失败')