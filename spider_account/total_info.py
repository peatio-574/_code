# -*- coding: utf-8 -*-
import sys
from pathlib import Path

sys.path.append(str(Path(__file__).parent.parent))

import pandas
import os
from ReadFile import ReadData
from openpyxl import load_workbook

dir_path = os.path.join(os.path.dirname(__file__), '数据')



def get_xlsx_info():
    total_info = {}
    for file in os.listdir(dir_path):
        if '每日' in file:
            continue
        file_name = os.path.join(dir_path, file)
        data = ReadData.read_xlsx_col(file_name, sheetname=1, header=0, index_col=None)

        date = data['日期'][:-1]

        platform = file.split('-')[0]
        shop_name = file.split('-')[1].split('店铺')[0]
        full_shop_name = f'{shop_name}-{platform}'

        incom = data['日净收入'][:-1]

        total_info[full_shop_name] = [[date[i], incom[i]]for i in range(len(date))]
    return total_info


def run():
    total_info = get_xlsx_info()
    file = os.path.join(dir_path, '每日店铺到账提现资金明细.xlsx')

    wb = load_workbook(file)
    ws = wb.worksheets[3]

    df = pandas.read_excel(file, sheet_name=3, header=[0, 1])
    date_list = df[('  店铺 \n \n\n日期 ', '各平台可提现金')]
    date_list = [str(i) for i in date_list]
    for shop, info in total_info.items():
        shop_ = shop.split('-')[0]
        shop_ = ''.join([c.lower() if c.isupper() else c for c in shop_])
        platform_ = shop.split('-')[1]

        for col_idx, k in enumerate(df.columns):
            k = list(k)
            k[0] = ''.join([c.lower() if c.isupper() else c for c in k[0]])
            if [shop_, platform_] == k:
                print(shop)
                for row in info:
                    row_id = date_list.index(row[0])
                    ws.cell(row=row_id+3, column=col_idx+1, value=float(row[1]))
                wb.save(file)



if __name__ == '__main__':
    run()








