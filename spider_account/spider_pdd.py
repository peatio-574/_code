# -*- coding: utf-8 -*-
import sys
from pathlib import Path

sys.path.append(str(Path(__file__).parent.parent))


from PlayWright import Playwright_, logger, get_config_value
from openpyxl.styles import Font, Alignment, PatternFill
import openpyxl
import warnings
import os
import pandas
import requests
import time
import zipfile

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
pandas.set_option('future.no_silent_downcasting', True)


config_file = os.path.join(os.path.dirname(__file__), 'config.ini')


def pdd_login(account_id=1):
    try:
        logger.info('开始登录拼多多....')
        url = 'https://mms.pinduoduo.com/home'
        ele = '//div[@class="user-name-name"]/span'

        key = f'login.pdd_cookie_{account_id}'
        Playwright_.login(url, ele, key, file=config_file)
        logger.info('拼多多登录成功....')
        return True
    except Exception as e:
        logger.error(f'拼多多登录失败：{e[:50]}')
        return False


def pdd_search(start, end):
    Playwright_.page.mouse.click(700, 500, button='left')
    Playwright_.click('//span[text()="资金中心"]')
    Playwright_.switch_to_page()
    time.sleep(5)
    shop_name = Playwright_.get_text('//span[contains(@class,"Header_header_name")]')
    logger.info(f'当前店铺名称：{shop_name}')

    start = start[-1:] if start[-2] == '0' else start[-2:]
    end = end[-1:] if end[-2] == '0' else end[-2:]

    Playwright_.click('//a[text()="收支明细"]')
    Playwright_.click('//input[@placeholder="开始日期-结束日期"]')
    Playwright_.click(f'(//td[not (contains(@class,"outOfMonth"))]/div[text()="{start}"])[last()]')
    Playwright_.click(f'(//td[not (contains(@class,"outOfMonth"))]/div[text()="{end}"])[last()]')
    Playwright_.mouse_wheel(80)
    sure_ele = '//span[text()="确认"]'
    Playwright_.click(sure_ele)
    if Playwright_.get_count(sure_ele):
        Playwright_.click(sure_ele)
    if Playwright_.get_count(sure_ele):
        Playwright_.click(sure_ele)
    if Playwright_.get_count(sure_ele):
        Playwright_.click(sure_ele)
    Playwright_.click('//span[text()="查询"]')
    return shop_name


def pdd_deal_data(shopname, file):
    """
    使用pandas处理数据：
    1. 按每日汇总交易类型描述不为"提现"的收入和支出
    2. 按每日汇总交易类型描述为"提现"的提现金额
    """
    try:
        # 读取Excel文件
        df = pandas.read_excel(file)

        # 将数值列转换为数字类型（处理可能的文本格式数字）
        df['收入金额'] = pandas.to_numeric(df['收入金额（+元）'], errors='coerce').fillna(0)
        df['支出金额'] = pandas.to_numeric(df['支出金额（-元）'].replace('-', '0'), errors='coerce').fillna(0)
        # 提取日期
        df['日期'] = pandas.to_datetime(df['发生时间']).dt.date

        # 分离提现和非提现数据
        df_withdraw = df[df['账务类型'].str.contains('提现', na=False)]
        df_normal = df[~df['账务类型'].str.contains('提现', na=False)]

        # 汇总非提现的收入和支出
        df_normal_summary = df_normal.groupby('日期').agg({
            '收入金额': 'sum',
            '支出金额': 'sum'
        }).reset_index()
        df_normal_summary.columns = ['日期', '日收入', '日支出']

        # 汇总提现数据（提现金额在支出列）
        df_withdraw_summary = df_withdraw.groupby('日期').agg({
            '支出金额': 'sum'
        }).reset_index()
        df_withdraw_summary.columns = ['日期', '日提现金额']

        # 合并两个汇总表
        df_summary = pandas.merge(df_normal_summary, df_withdraw_summary, on='日期', how='outer')

        # 填充空值为0
        df_summary = df_summary.fillna(0)

        # 计算日净收入（收入 - 支出 - 提现）
        df_summary['日净收入'] = df_summary['日收入'] + df_summary['日支出']

        # 按日期排序
        df_summary = df_summary.sort_values('日期', ascending=True).reset_index(drop=True)

        # 汇总数据
        total_row = pandas.DataFrame({
            '日期': ['所有汇总'],
            '日收入': [df_summary['日收入'].sum()],
            '日支出': [df_summary['日支出'].sum()],
            '日提现金额': [df_summary['日提现金额'].sum()],
            '日净收入': [df_summary['日净收入'].sum()]
        })
        df_summary = pandas.concat([df_summary, total_row], ignore_index=True)

        # 使用 ExcelWriter 追加到现有文件，保留原有的 Sheet1
        with pandas.ExcelWriter(file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            sheet_name = '汇总数据'
            # 先删除可能已存在的同名 Sheet
            if sheet_name in writer.book.sheetnames:
                del writer.book[sheet_name]

            # 写入数据
            df_summary.to_excel(writer, sheet_name=sheet_name, index=False)

            # 获取工作表对象并设置样式
            ws = writer.sheets[sheet_name]
            header_font = Font(bold=True, color='FFFFFF', size=11)
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center')

            # 遍历第一行所有单元格设置样式
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # 自动调整列宽
            for column in ws.columns:
                max_length = 0
                col_letter = column[0].column_letter
                for cell in column:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))

                ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

            # 消除 "Workbook contains no default style" 警告
            workbook = writer.book
            if not workbook.style_names:
                default_font = Font(name='Calibri', size=11, bold=False, italic=False)
                default_style = openpyxl.styles.NamedStyle(name='Normal', font=default_font)
                workbook.add_named_style(default_style)

        logger.info(f'数据汇总完成，共汇总{len(df_summary)}天的数据，已保存到: {file}')

        # 打印汇总统计
        logger.info(f'总收入: {df_summary["日收入"].sum()/2:.2f}，总支出: {df_summary["日支出"].sum()/2:.2f} '
                    f'总提现: {df_summary["日提现金额"].sum()/2:.2f}，总净收入: {df_summary["日净收入"].sum()/2:.2f}\n')
        return df_summary

    except Exception as e:
        logger.error(f'{shopname}数据处理失败: {e}\n')
        return None



def pdd_save(filename):
    """直接导出"""
    try:
        # 直接导出
        zip_name = f'{filename[:-5]}.zip'
        # 保存当前页面引用，避免新页面开关导致 context 变化
        current_page = Playwright_.page
        with current_page.expect_download(timeout=15000) as download_info:
            Playwright_.click('(//div[@class="export-history-bills-card"])[1]//span[text()="下载账单"]')
            pass  # 等待下载触发
        download = download_info.value
        # 获取文件名并保存
        download.save_as(zip_name)

        # 解压zip文件
        with zipfile.ZipFile(zip_name, 'r') as zip_ref:
            # 获取压缩包内的文件名
            file_list = zip_ref.namelist()
            if len(file_list) == 0:
                logger.error('压缩包内没有文件')
                return False

            # 获取第一个文件（通常只有一个）
            inner_filename = file_list[0]

            # 解压到临时目录
            zip_ref.extractall(os.path.dirname(zip_name))

        # 构建解压后的文件路径
        extracted_file = os.path.join(os.path.dirname(zip_name), inner_filename)

        inner_filename = os.path.join(os.path.dirname(filename), inner_filename)
        # 如果解压出来的是csv文件，直接重命名为目标文件名
        if inner_filename.endswith('.csv'):
            df = pandas.read_csv(inner_filename, encoding="gbk", skiprows=4)
            if len(df) > 4:
                df = df.iloc[:-4]
            df.to_excel(filename, index=False, engine="openpyxl")

        # 删除zip文件

        os.remove(zip_name)
        os.remove(inner_filename)

        return True
    except Exception as e:
        logger.info(f'{filename}临时下载异常：{e}')
        return False


def main_(account_id=1):
    title = f'========================开始爬取第{account_id}个店铺======================='
    logger.info(title)
    end = time.strftime("%Y-%m-%d", time.localtime(time.time() - 86400))
    start = f"{end[:-3]}-01"
    dirname = 'd:/_code/spider_account/数据'

    login = pdd_login(account_id)
    if not login:
        return False
    shop_name = pdd_search(start, end)

    filename = f'拼多多-{shop_name}店铺{end}明细.xlsx'
    filename = os.path.join(dirname, filename)

    time.sleep(10)

    logger.info(f'{shop_name}店铺开始导出明细....')
    status = 0
    for roll in range(1, 6):
        Playwright_.click('//span[text()="导出"]')
        Playwright_.switch_to_page()
        logger.info(f'第{roll}次导出明细....')
        status = pdd_save(filename)
        if status:
            break

    text = f'✅️ {shop_name}明细数据下载成功：{filename}' if status else f'❌️ {shop_name}明细数据下载失败'
    logger.info(text)
    if status:
        pdd_deal_data(shop_name, filename)
    Playwright_.clear_cookie()
    return True if status else False


if __name__ == '__main__':
    # pdd_deal_data('甜心花栗店铺', file='./数据/meow meow店铺2026-05-27明细.xlsx')
    shop_count_ = get_config_value('login', 'pdd_shop_count', file=config_file)
    shop_flag = input('请输入查询店铺序号（0默认查询全部）：')
    if shop_flag == '0':
        start_id = 1
        end_id = int(shop_count_) + 1
    else:
        start_id = int(shop_flag)
        end_id = start_id + 1
    for account_id in range(start_id, end_id):
        main_(account_id)

