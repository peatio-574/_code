# -*- coding: utf-8 -*-
import sys
from pathlib import Path

sys.path.append(str(Path(__file__).parent.parent))


from PlayWright import Playwright_, logger, get_config_value
from openpyxl.styles import Font, Alignment, PatternFill
import openpyxl
import warnings
import os
import pandas
import time


warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')


config_file = os.path.join(os.path.dirname(__file__), 'config.ini')


def dd_login(account_id=1):
    try:
        logger.info('开始登录抖店....')
        url = 'https://fxg.jinritemai.com/login/common?channel=zhaoshang'
        ele = '//div[text()="请选择店铺"]'

        key = f'login.dd_cookie_{account_id}'
        sub_account = f'(//div[contains(@class,"index_roleItem")])[{account_id}]//img[2]'
        Playwright_.login(url, ele, key, extra=sub_account, file=config_file)
        logger.info('抖店登录成功....')
        return True
    except Exception as e:
        logger.error(f'抖店登录失败：{e[:50]}')
        return False


def dd_search(start, end):
    Playwright_.goto('https://fxg.jinritemai.com/ffa/fxg-bill/fund-detail-bill')
    shop_name = Playwright_.get_text('//div[contains(@class,"index_userName")]')
    logger.info(f'当前店铺名称：{shop_name}')
    time.sleep(8)
    start_ele = '//input[@placeholder="开始日期"]'
    end_ele = '//input[@placeholder="结束日期"]'
    Playwright_.click(start_ele)
    Playwright_.input(start_ele, start, enter=True)
    Playwright_.click(end_ele)
    Playwright_.input(end_ele, end, enter=True)
    Playwright_.click('//span[text()="查询"]')
    return shop_name


def dd_deal_data(shopname, file):
    """
    使用pandas处理数据：
    1. 提取动账场景为充值保证金、提现的数据
    2. 统计每日净收入
    """
    try:
        # 读取Excel文件
        print(1)
        df = pandas.read_excel(file)

        # 将数值列转换为数字类型（处理可能的文本格式数字）
        df['动账金额'] = df['动账金额'].astype(str).str.replace(',', '').str.strip()
        df['动账金额'] = pandas.to_numeric(df['动账金额'], errors='coerce').fillna(0)
        # 提取日期
        df['日期'] = pandas.to_datetime(df['动账时间']).dt.date

        # 分类数据
        # 1. 充值保证金
        df_recharge = df[df['动账场景'] == '充值保证金']

        # 2. 提现
        df_withdraw = df[df['动账场景'] == '提现']

        # 3. 其他交易（排除充值保证金和提现）
        df_other = df[~df['动账场景'].apply(lambda x: '充值保证金' in x or '提现' in x)]

        # 按日汇总充值保证金（仅展示）
        if len(df_recharge) > 0:
            df_recharge_summary = df_recharge.groupby('日期').agg({
                '动账金额': 'sum'
            }).reset_index()
            df_recharge_summary.columns = ['日期', '充值保证金']
        else:
            df_recharge_summary = pandas.DataFrame(columns=['日期', '充值保证金'])

        # 按日汇总提现（仅展示）
        if len(df_withdraw) > 0:
            df_withdraw_summary = df_withdraw.groupby('日期').agg({
                '动账金额': 'sum'
            }).reset_index()
            df_withdraw_summary.columns = ['日期', '提现金额']
        else:
            df_withdraw_summary = pandas.DataFrame(columns=['日期', '提现金额'])

        # 按日汇总其他交易，区分入账和出账
        if len(df_other) > 0:
            # 入账户
            df_income = df_other[df_other['动账方向'] == '入账']
            df_income_summary = df_income.groupby('日期').agg({
                '动账金额': 'sum'
            }).reset_index()
            df_income_summary.columns = ['日期', '入账金额']

            # 出账户
            df_expense = df_other[df_other['动账方向'] == '出账']
            df_expense_summary = df_expense.groupby('日期').agg({
                '动账金额': 'sum'
            }).reset_index()
            df_expense_summary.columns = ['日期', '出账金额']

            # 合并入账和出账
            df_other_summary = pandas.merge(df_income_summary, df_expense_summary, on='日期', how='outer')
            df_other_summary = df_other_summary.fillna(0)
            df_other_summary['日净收入'] = df_other_summary['入账金额'] - df_other_summary['出账金额']
        else:
            df_other_summary = pandas.DataFrame(columns=['日期', '入账金额', '出账金额', '日净收入'])

        # 合并所有汇总表
        df_summary = df_other_summary.copy()
        df_summary = pandas.merge(df_summary, df_recharge_summary, on='日期', how='outer')
        df_summary = pandas.merge(df_summary, df_withdraw_summary, on='日期', how='outer')

        # 填充空值为0
        df_summary = df_summary.fillna(0)

        # 按日期排序
        df_summary = df_summary.sort_values('日期', ascending=False).reset_index(drop=True)
        # 汇总数据
        total_row = pandas.DataFrame({
            '日期': ['所有汇总'],
            '入账金额': [df_summary['入账金额'].sum()],
            '出账金额': [df_summary['出账金额'].sum()],
            '日净收入': [df_summary['日净收入'].sum()],
            '充值保证金': [df_summary['充值保证金'].sum()],
            '提现金额': [df_summary['提现金额'].sum()]
        })
        df_summary = pandas.concat([df_summary, total_row], ignore_index=True)

        # 使用 ExcelWriter 追加到现有文件，保留原有的 Sheet1
        with pandas.ExcelWriter(file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            sheet_name = '汇总数据'
            # 先删除可能已存在的同名 Sheet
            if sheet_name in writer.book.sheetnames:
                del writer.book[sheet_name]

            # 写入数据
            df_summary.to_excel(writer, sheet_name=sheet_name, index=False)

            # 获取工作表对象并设置样式
            ws = writer.sheets[sheet_name]
            header_font = Font(bold=True, color='FFFFFF', size=11)
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center')

            # 遍历第一行所有单元格设置样式
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # 自动调整列宽
            for column in ws.columns:
                max_length = 0
                col_letter = column[0].column_letter
                for cell in column:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))

                ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

            # 消除 "Workbook contains no default style" 警告
            workbook = writer.book
            if not workbook.style_names:
                default_font = Font(name='Calibri', size=11, bold=False, italic=False)
                default_style = openpyxl.styles.NamedStyle(name='Normal', font=default_font)
                workbook.add_named_style(default_style)

        logger.info(f'数据汇总完成，共汇总{len(df_summary)}天的数据，已保存到: {file}')

        # 打印汇总统计
        logger.info(f'总收入: {df_summary["日收入"].sum()/2:.2f}，总支出: {df_summary["日支出"].sum()/2:.2f} '
                    f'总提现: {df_summary["日提现金额"].sum()/2:.2f}，总净收入: {df_summary["日净收入"].sum()/2:.2f}\n')
        return df_summary

    except Exception as e:
        logger.error(f'{shopname}店铺数据处理失败: {e}\n')
        return None



def dd_save(end, shop_name):
    """直接导出"""
    dirname = 'd:/_code/spider_account/数据'
    filename = f'抖店-{shop_name}店铺{end}明细.xlsx'
    filename = os.path.join(dirname, filename)
    try:
        Playwright_.click('//span[text()="生成报表"]')
        Playwright_.click('//span[text()="生成"]')
        down_ele = '(//div[contains(@class,"cardList")])[1]//span[text()="下载"]'
        Playwright_.wait_for_selector(down_ele)
        Playwright_.click(down_ele)
        with Playwright_.page.expect_download(timeout=15000) as download_info:
            Playwright_.click(down_ele)
            pass  # 等待下载触发
        download = download_info.value
        # 获取文件名并保存
        download.save_as(filename)
        return filename
    except Exception as e:
        logger.info(f'{filename}临时下载异常：{e}')
        return False


def main_(account_id=1):
    title = f'========================开始爬取抖店第{account_id}个店铺======================='
    logger.info(title)
    end = time.strftime("%Y-%m-%d", time.localtime(time.time() - 86400))
    start = f"{end[:-3]}-01"


    login = dd_login(account_id)
    if not login:
        return False

    status = 0
    shop_name = ''
    for roll in range(1, 6):
        shop_name = dd_search(start, end)
        if not shop_name:
            continue
        logger.info(f'{shop_name}店铺第{roll}次导出明细....')
        status = dd_save(end, shop_name)
        if status:
            break

    text = f'✅️ {shop_name}明细数据下载成功：{status}' if status else f'❌️ {shop_name}明细数据下载失败'
    logger.info(text)
    if status:
        dd_deal_data(shop_name, status)
    Playwright_.clear_cookie()
    return True if status else False


if __name__ == '__main__':
    dd_deal_data('甜心花栗店铺', file='./数据/抖店-甜心花栗店铺2026-05-27明细.xlsx')
    # shop_count_ = get_config_value('login', 'dd_shop_count', file=config_file)
    # shop_flag = input('请输入查询店铺序号（0默认查询全部）：')
    # if shop_flag == '0':
    #     start_id = 1
    #     end_id = int(shop_count_) + 1
    # else:
    #     start_id = int(shop_flag)
    #     end_id = start_id + 1
    # for account_id in range(start_id, end_id):
    #     main_(account_id)

