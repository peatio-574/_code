# coding='utf-8'
# noinspection PyProtectedMember
import sys
from pathlib import Path

sys.path.append(str(Path(__file__).parent.parent))

import requests
import os
from newPlayWright import PlayWright, logger, get_config_value, write_config_value
import time
import pandas
from openpyxl.styles import Font, Alignment, PatternFill
import openpyxl
import warnings
import zipfile

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
config_file = os.path.join(os.path.dirname(__file__), 'config.ini')
dataDir = os.path.join(os.path.dirname(__file__), '数据')
os.makedirs(dataDir, exist_ok=True)


class XHS(object):

    @classmethod
    def xhsLogin(cls, account_id):
        """登录小红书"""
        logger.info('开始登录小红书....')
        url = 'https://ark.xiaohongshu.com/app-system/home'
        ele = '(//a[text()="学习中心"])[1]'
        key = f'login.xhs_cookie_{account_id}'
        idx = (account_id - 1) % 10 + 1
        sub_account = f'(//span[text()="子账号"])[{idx}]/../div/div[1]'
        loginStatus = PlayWright.login(url, ele, key, extra=sub_account, file=config_file)
        if loginStatus:
            shopName = PlayWright.get_text('//div[@class="avatar-wrapper"]/div')
            logger.info(f'✅️ 【店铺：{shopName}】小红书登录成功....')
            return shopName
        else:
            logger.error(f'❌️ 小红书登录失败')
            return False

    @classmethod
    def fundsSearch(cls, startTime, endTime):
        """访问账号资金明细页面，进行搜索"""
        PlayWright.goto('https://ark.xiaohongshu.com/app-merchant/third-settle/account')
        time.sleep(8)

        # 循环关闭弹窗
        know_ele = '(//span[text()="知道了" or text()="跳过"])[last()]'
        for roll in range(10):
            if PlayWright.get_count(know_ele):
                PlayWright.click(know_ele)
                time.sleep(1)

        PlayWright.slow_input('//input[@placeholder="开始时间"]', startTime)
        PlayWright.slow_input('//input[@placeholder="结束时间"]', endTime)
        PlayWright.click('//span[text()="查询"]')

    @classmethod
    def fundsHtmlSave(cls, fileName):
        """账号资金明细-页面导出"""
        try:
            with PlayWright.page.expect_download(timeout=60000) as download_info:
                PlayWright.click('//span[text()="导出"]')
                time.sleep(1)
                confirm_btn = '//span[text()="确定"]'
                if PlayWright.get_count(confirm_btn):
                    PlayWright.click(confirm_btn)
            download = download_info.value
            download.save_as(fileName)
            return True
        except Exception as e:
            logger.error(f'❌️ {fileName}-页面导出-临时下载异常：{e}')
            return False

    @classmethod
    def ApiSave(cls, fileName):
        """api导出"""
        try:
            PlayWright.click('//span[text()="消息"]')
            PlayWright.click('(//span[text()="店铺"])[last()]')

            errorInfo = f'❌️ {fileName}-api导出失败-未获取到文件链接'

            # 第一行时间是否存在
            firstRowTimeEle = '(//div[@class="list-item-wrapper"])[1]//div[@class="date-wrap unread"]/span'
            if not PlayWright.get_count(firstRowTimeEle):
                logger.info(errorInfo)
                return False

            # 第一行时间是否符合条件
            firstRowTime = PlayWright.get_text(firstRowTimeEle)
            if firstRowTime not in ('1分钟前', '刚刚'):
                logger.info(errorInfo)
                return False

            # 第一行链接元素是否存在
            firstRowLinkEle = '(//div[@class="list-item-wrapper"])[1]//a'
            if not PlayWright.get_count(firstRowLinkEle):
                logger.info(errorInfo)
                return False

            # 第一行链接元素href是否存在
            href = PlayWright.get_attribute(firstRowLinkEle, 'href')
            with open(fileName, mode='wb') as f:
                f.write(requests.get(href).content)

            # 关闭消息弹窗
            PlayWright.click('//div[@class="ark-message-title-wrap"]/span[2]')
            return True
        except Exception as e:
            logger.error(f'❌️ {fileName}-api导出-临时下载异常：{e}')
            return False

    @classmethod
    def fundsDataDeal(cls, shopName, fileName):
        """汇总账号资金详细数据"""
        try:
            # 读取Excel文件
            df = pandas.read_excel(fileName)

            # 将数值列转换为数字类型（处理可能的文本格式数字）
            df['收入（元）'] = pandas.to_numeric(df['收入（元）'], errors='coerce').fillna(0)
            df['支出（元）'] = pandas.to_numeric(df['支出（元）'].replace('-', '0'), errors='coerce').fillna(0)

            # 提取日期
            df['日期'] = pandas.to_datetime(df['创建时间']).dt.date

            # 分离提现和非提现数据
            df_withdraw = df[df['交易类型描述'].str.contains('提现', na=False)]
            df_normal = df[~df['交易类型描述'].str.contains('提现', na=False)]

            logger.info(f'总数据量: {len(df)}, 提现数据: {len(df_withdraw)}, 非提现数据: {len(df_normal)}')

            # 汇总非提现的收入和支出
            df_normal_summary = df_normal.groupby('日期').agg({
                '收入（元）': 'sum',
                '支出（元）': 'sum'
            }).reset_index()
            df_normal_summary.columns = ['日期', '日收入', '日支出']

            # 汇总提现数据（提现金额在支出列）
            df_withdraw_summary = df_withdraw.groupby('日期').agg({
                '支出（元）': 'sum'
            }).reset_index()
            df_withdraw_summary.columns = ['日期', '日提现金额']

            # 合并两个汇总表
            df_summary = pandas.merge(df_normal_summary, df_withdraw_summary, on='日期', how='outer')

            # 填充空值为0
            df_summary = df_summary.fillna(0)

            # 计算日净收入（收入 - 支出 - 提现）
            df_summary['日净收入'] = df_summary['日收入'] - df_summary['日支出']

            # 按日期排序
            df_summary = df_summary.sort_values('日期', ascending=True).reset_index(drop=True)

            # 汇总数据
            total_row = pandas.DataFrame({
                '日期': ['所有汇总'],
                '日收入': [df_summary['日收入'].sum()],
                '日支出': [df_summary['日支出'].sum()],
                '日提现金额': [df_summary['日提现金额'].sum()],
                '日净收入': [df_summary['日净收入'].sum()]
            })
            df_summary = pandas.concat([df_summary, total_row], ignore_index=True)

            # 使用 ExcelWriter 追加到现有文件，保留原有的 Sheet1
            with pandas.ExcelWriter(fileName, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                sheet_name = '汇总数据'
                # 先删除可能已存在的同名 Sheet
                if sheet_name in writer.book.sheetnames:
                    del writer.book[sheet_name]

                # 写入数据
                df_summary.to_excel(writer, sheet_name=sheet_name, index=False)

                # 获取工作表对象并设置样式
                ws = writer.sheets[sheet_name]
                header_font = Font(bold=True, color='FFFFFF', size=11)
                header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                header_alignment = Alignment(horizontal='center', vertical='center')

                # 遍历第一行所有单元格设置样式
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment

                # 自动调整列宽
                for column in ws.columns:
                    max_length = 0
                    col_letter = column[0].column_letter
                    for cell in column:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))

                    ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

                # 消除 "Workbook contains no default style" 警告
                workbook = writer.book
                if not workbook.style_names:
                    default_font = Font(name='Calibri', size=11, bold=False, italic=False)
                    default_style = openpyxl.styles.NamedStyle(name='Normal', font=default_font)
                    workbook.add_named_style(default_style)

            logger.info(f'数据汇总完成，共汇总{len(df_summary)}天的数据，已保存到: {fileName}')

            # 打印汇总统计
            logger.info(f'总收入: {df_summary["日收入"].sum()/2:.2f}，总支出: {df_summary["日支出"].sum()/2:.2f} '
                        f'总提现: {df_summary["日提现金额"].sum()/2:.2f}，总净收入: {df_summary["日净收入"].sum()/2:.2f}\n')
            return df_summary

        except Exception as e:
            logger.error(f'{shopName}数据处理失败: {e}\n')
            return None

    @classmethod
    def fundsSingleRun(cls, account_id):
        """账户资金明细-单个店铺运行"""
        title = f'========================开始爬取小红书第{account_id}个店铺账号资金详情======================='
        logger.info(title)

        # 结束时间为昨天，开始时间为结束时间的当月第一天
        endTime = time.strftime("%Y-%m-%d", time.localtime(time.time() - 86400))
        startTime = endTime[:-2] + '01'

        shopName = cls.xhsLogin(account_id)
        if not shopName:
            logger.error(f'小红书第{account_id}个店铺登录异常')
            return False

        fileName = f'小红书-{shopName}店铺{endTime}账户资金明细.xlsx'
        fileName = os.path.join(dataDir, fileName)

        saveStatus = False
        for roll in range(1, 6):
            logger.info(f'开始第{roll}次尝试导出明细')
            # 搜索
            cls.fundsSearch(startTime, endTime)
            # 页面导出
            saveStatus = cls.fundsHtmlSave(fileName)
            if saveStatus:
                break
            # api导出
            saveStatus = cls.ApiSave(fileName)
            if saveStatus:
                break

        text = f'✅️ {shopName}明细数据下载成功：{fileName}' if saveStatus else f'❌️ {shopName}明细数据下载失败'
        logger.info(text)
        if saveStatus:
            cls.fundsDataDeal(shopName, fileName)
        PlayWright.clear_cookie()
        return True if saveStatus else False

    @classmethod
    def fundsRun(cls, startId, endId):
        """统筹运行账号资金明细"""
        for account_id in range(startId, endId):
            try:
                cls.fundsSingleRun(account_id)
            except Exception as e:
                logger.error(f'第{account_id}个店铺查询【账号资金明细】操作流程失败：{e}')

    @classmethod
    def salesSearch(cls, startTime, endTime):
        """访问销量明细页面，进行搜索"""
        PlayWright.goto('https://ark.xiaohongshu.com/app-order/order/query')
        time.sleep(8)

        # 循环关闭弹窗
        know_ele = '(//span[text()="知道了" or text()="跳过"])[last()]'
        for roll in range(10):
            if PlayWright.get_count(know_ele):
                PlayWright.click(know_ele)
                time.sleep(1)

        PlayWright.slow_input('//div[@class="d-daterangepicker-content"]/div[1]/input', startTime)
        PlayWright.slow_input('//div[@class="d-daterangepicker-content"]/div[3]/input', endTime)
        PlayWright.click('//span[text()="查询"]')

    @classmethod
    def salesHtmlSave(cls, fileName):
        try:
            # 直接导出
            with PlayWright.page.expect_download(timeout=15000) as download_info:
                PlayWright.click('//span[text()="全部导出"]')
                PlayWright.click('//span[text()="确定"]')

                pass  # 等待下载触发
            download = download_info.value
            # 获取文件名并保存
            download.save_as(fileName)
            return True
        except Exception as e:
            logger.error(f'❌️ {fileName}-页面导出-临时下载异常：{e}')
            return False

    @classmethod
    def salesSingleRun(cls, account_id):
        """销量明细-单个店铺运行"""
        title = f'========================开始爬取小红书第{account_id}个店铺销量详情======================='
        logger.info(title)

        # 结束时间为昨天，开始时间为结束时间的当月第一天
        endDate = time.strftime("%Y-%m-%d", time.localtime(time.time() - 86400))
        endTime = endDate + ' 23:59:59'
        startTime = endDate[:-2] + '01 00:00:00'

        shopName = cls.xhsLogin(account_id)
        if not shopName:
            logger.error(f'小红书第{account_id}个店铺登录异常')
            return False

        fileName = f'小红书-{shopName}店铺{endDate}销量明细.xlsx'
        fileName = os.path.join(dataDir, fileName)

        saveStatus = False
        for roll in range(1, 6):
            logger.info(f'开始第{roll}次尝试导出明细')
            # 搜索
            cls.salesSearch(startTime, endTime)
            # 页面导出
            saveStatus = cls.salesHtmlSave(fileName)
            if saveStatus:
                break
            # api导出
            saveStatus = cls.ApiSave(fileName)
            if saveStatus:
                break

        text = f'✅️ {shopName}明细数据下载成功：{fileName}' if saveStatus else f'❌️ {shopName}明细数据下载失败'
        logger.info(text)
        # if saveStatus:
        #     cls.fundsDataDeal(shopName, fileName)
        PlayWright.clear_cookie()
        return True if saveStatus else False

    @classmethod
    def salesRun(cls, startId, endId):
        """统筹运行销量明细"""
        for account_id in range(startId, endId):
            try:
                cls.salesSingleRun(account_id)
            except Exception as e:
                logger.error(f'第{account_id}个店铺查询【销量明细】操作流程失败：{e}')


class TB(object):

    @classmethod
    def tbLogin(cls, account_id):
        """登录淘宝千牛"""
        logger.info('开始登录淘宝千牛....')
        url = 'https://myseller.taobao.com/home.htm/QnworkbenchHome/'
        ele = '//span[contains(text(),"首页")]'
        key = f'login.tb_cookie_{account_id}'
        loginStatus = PlayWright.login(url, ele, key, file=config_file)
        if loginStatus:
            shopName = PlayWright.get_text('//div[@class="user-area-pop-up-panel"]/div[1]/div/div[1]')
            logger.info(f'✅️ 【店铺：{shopName}】淘宝千牛登录成功....')
            return shopName
        else:
            logger.error(f'❌️ 淘宝登录失败')
            return False

    @classmethod
    def fundsSearch(cls, startTime, endTime):
        """访问账号资金明细页面，进行搜索"""
        PlayWright.goto('https://qn.taobao.com/home.htm/whale-accountant/pay/capital/home?active=fund_detail')
        time.sleep(8)

        # 循环关闭弹窗
        know_ele = '(//button[text()="知道了" or text()="跳过" or text()="完成"])[last()]'
        for roll in range(10):
            if PlayWright.get_count(know_ele):
                PlayWright.click(know_ele)
                time.sleep(1)

        PlayWright.input('//input[@placeholder="起始日期"]', startTime)
        PlayWright.input('//input[@placeholder="结束日期"]', endTime, enter=True)
        PlayWright.click('//span[text()="搜索"]')

    @classmethod
    def fundsHtmlSave(cls, fileName):
        """账号资金明细-页面导出"""
        try:
            # 直接导出
            with PlayWright.page.expect_download(timeout=15000) as download_info:
                PlayWright.click('//span[text()="导出"]')
                time.sleep(3)
            download = download_info.value
            download.save_as(fileName)
            return True
        except Exception as e:
            logger.error(f'❌️ {fileName}-页面导出-临时下载异常：{e}')
            return False

    @classmethod
    def fundsDataDeal(cls, shopName, fileName):
        """汇总账号资金详细数据"""
        try:
            # 读取Excel文件
            df = pandas.read_excel(fileName)

            # 将数值列转换为数字类型（处理可能的文本格式数字）
            df['收入金额（元）'] = df['收入金额（元）'].astype(str).str.replace(',', '').str.strip()
            df['收入金额（元）'] = pandas.to_numeric(df['收入金额（元）'].replace('', '0'), errors='coerce').fillna(0)

            df['支出金额'] = df['支出金额'].astype(str).str.replace(',', '').str.strip()
            df['支出金额'] = pandas.to_numeric(df['支出金额'].replace('', '0'), errors='coerce').fillna(0)

            # 提取日期（从入账时间）
            df['日期'] = pandas.to_datetime(df['入账时间']).dt.date
            # 分离不同类型的交易
            df_transfer = df[df['入账类型'].apply(lambda x: '转账' in x)]
            df_withdraw = df[df['入账类型'].apply(lambda x: '提现' in x)]
            df_other = df[~df['入账类型'].apply(lambda x: '转账' in x or '提现' in x)]

            logger.info(f'总数据量: {len(df)}, 提现数据: {len(df_withdraw)}, 转账数据: {len(df_transfer)}, '
                        f'日数据: {len(df_other)}')

            # 汇总转账数据（按日）
            df_transfer_summary = df_transfer.groupby('日期').agg({
                '收入金额（元）': 'sum',
                '支出金额': 'sum'
            }).reset_index()
            df_transfer_summary.columns = ['日期', '转账收入', '转账支出']
            df_transfer_summary['转账净额'] = df_transfer_summary['转账收入'] - df_transfer_summary['转账支出']

            # 汇总提现数据（按日）
            df_withdraw_summary = df_withdraw.groupby('日期').agg({
                '支出金额': 'sum'
            }).reset_index()
            df_withdraw_summary.columns = ['日期', '提现支出']
            df_withdraw_summary['提现金额'] = df_withdraw_summary['提现支出']

            # 汇总其他交易数据（按日）
            df_other_summary = df_other.groupby('日期').agg({
                '收入金额（元）': 'sum',
                '支出金额': 'sum'
            }).reset_index()
            df_other_summary.columns = ['日期', '日收入', '日支出']
            df_other_summary['日净收入'] = df_other_summary['日收入'] - df_other_summary['日支出']

            # 合并三个汇总表
            df_summary = pandas.merge(df_other_summary, df_transfer_summary, on='日期', how='outer')
            df_summary = pandas.merge(df_summary, df_withdraw_summary, on='日期', how='outer')

            # 填充空值为0
            df_summary = df_summary.fillna(0)

            # 计算每日总净收入
            df_summary['日净收入'] = df_summary['日净收入']

            # 按日期排序
            df_summary = df_summary.sort_values('日期', ascending=True).reset_index(drop=True)

            # 添加汇总行
            total_row = pandas.DataFrame({
                '日期': ['所有汇总'],
                '日收入': [df_summary['日收入'].sum()],
                '日支出': [df_summary['日支出'].sum()],
                '提现金额': [df_summary['提现金额'].sum()],
                '转账收入': [df_summary['转账收入'].sum()],
                '转账支出': [df_summary['转账支出'].sum()],
                '转账净额': [df_summary['转账净额'].sum()],
                '日净收入': [df_summary['日净收入'].sum()]
            })
            df_summary = pandas.concat([df_summary, total_row], ignore_index=True)

            # 重新排列列顺序
            columns_order = ['日期', '日收入', '日支出', '提现金额', '转账收入', '转账支出', '转账净额', '日净收入']
            df_summary = df_summary[columns_order]

            # 使用 ExcelWriter 追加到现有文件，保留原有的 Sheet
            with pandas.ExcelWriter(fileName, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                sheet_name = '汇总数据'
                # 先删除可能已存在的同名 Sheet
                if sheet_name in writer.book.sheetnames:
                    del writer.book[sheet_name]

                # 写入数据
                df_summary.to_excel(writer, sheet_name=sheet_name, index=False)

                # 获取工作表对象并设置样式
                ws = writer.sheets[sheet_name]
                header_font = Font(bold=True, color='FFFFFF', size=11)
                header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                header_alignment = Alignment(horizontal='center', vertical='center')

                # 遍历第一行所有单元格设置样式
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment

                # 自动调整列宽
                for column in ws.columns:
                    max_length = 0
                    col_letter = column[0].column_letter
                    for cell in column:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))

                    ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

                # 消除 "Workbook contains no default style" 警告
                workbook = writer.book
                if not workbook.style_names:
                    default_font = Font(name='Calibri', size=11, bold=False, italic=False)
                    default_style = openpyxl.styles.NamedStyle(name='Normal', font=default_font)
                    workbook.add_named_style(default_style)

            logger.info(f'数据汇总完成，共汇总{len(df_summary) - 1}天的数据，已保存到: {fileName}')

            logger.info(f'总收入: {df_summary["日收入"].sum() / 2:.2f}，总支出: {df_summary["日支出"].sum() / 2:.2f} '
                        f'总提现: {df_summary["提现金额"].sum() / 2:.2f}，，总转账: {df_summary["转账净额"].sum() / 2:.2f}，'
                        f'总净收入: {df_summary["日净收入"].sum() / 2:.2f}\n')
            return df_summary
        except Exception as e:
            logger.error(f'{shopName}数据处理失败: {e}\n')
            return None

    @classmethod
    def fundsSingleRun(cls, account_id):
        """账户资金明细-单个店铺运行"""
        title = f'========================开始爬取淘宝千牛第{account_id}个店铺账号资金详情======================='
        logger.info(title)

        # 结束时间为昨天，开始时间为结束时间的当月第一天
        endTime = time.strftime("%Y-%m-%d", time.localtime(time.time() - 86400))
        startTime = endTime[:-2] + '01'

        shopName = cls.tbLogin(account_id)
        if not shopName:
            logger.error(f'淘宝第{account_id}个店铺登录异常')
            return False


        fileName = f'淘宝-{shopName}店铺{endTime}账户资金明细.xlsx'
        fileName = os.path.join(dataDir, fileName)

        saveStatus = False
        for roll in range(1, 6):
            logger.info(f'开始第{roll}次尝试导出明细')
            # 搜索
            cls.fundsSearch(startTime, endTime)
            time.sleep(3)
            if PlayWright.get_count('//div[text()="没有数据"]'):
                logger.info(f'{shopName}店铺报表暂无数据')
                PlayWright.clear_cookie()
                return False
            # 页面导出
            saveStatus = cls.fundsHtmlSave(fileName)
            if saveStatus:
                break

        text = f'✅️ {shopName}明细数据下载成功：{fileName}' if saveStatus else f'❌️ {shopName}明细数据下载失败'
        logger.info(text)
        if saveStatus:
            cls.fundsDataDeal(shopName, fileName)
        PlayWright.clear_cookie()
        return True if saveStatus else False

    @classmethod
    def fundsRun(cls, startId, endId):
        """统筹运行账号资金明细"""
        for account_id in range(startId, endId):
            try:
                cls.fundsSingleRun(account_id)
            except Exception as e:
                logger.error(f'第{account_id}个店铺查询【账号资金明细】操作流程失败：{e}')

    @classmethod
    def salesSearch(cls, startDate, endDate):
        """访问销量明细页面，进行搜索"""
        PlayWright.goto('https://qn.taobao.com/home.htm/trade-platform/tp/sold')
        time.sleep(8)

        # 循环关闭弹窗
        know_ele = '(//button[text()="知道了" or text()="跳过" or text()="完成"])[last()]'
        for roll in range(10):
            if PlayWright.get_count(know_ele):
                PlayWright.click(know_ele)
                time.sleep(1)

        PlayWright.click('(//input[@placeholder="起始日期"])[1]')
        PlayWright.slow_input('(//input[@placeholder="YYYY-MM-DD"])[1]', startDate, enter=True)
        PlayWright.slow_input('(//input[@placeholder="HH:mm:ss"])[1]', '00:00:00', enter=True)
        PlayWright.slow_input('(//input[@placeholder="YYYY-MM-DD"])[2]', endDate, enter=True)
        PlayWright.slow_input('(//input[@placeholder="HH:mm:ss"])[2]', '23:59:59', enter=True)
        PlayWright.click('//div[@class="next-date-picker-panel-footer"]//span[text()="确定"]')
        PlayWright.click('//span[text()="搜索订单"]')

    @classmethod
    def salesHtmlSave(cls, fileName):
        try:
            # 勾选对应字段
            PlayWright.click('//span[text()="批量导出"]')
            time.sleep(3)

            # 存在两种导出方式
            timeFlag = time.time()
            extraEle = '//div[@popupcontainer="qn-worbench-container"]//span[text()="生成报表"]'
            if PlayWright.get_count(extraEle):  # 文案提示类型
                PlayWright.click(extraEle)
                PlayWright.click_catch_new_page('//span[text()="确定"]')
                rowTimeELe = '//ul[@class="sheet-list"]/li[1]/h2'
                downloadEle = '//ul[@class="sheet-list"]/li[1]//a[text()="下载宝贝报表"]'
            else:
                PlayWright.click('//span[text()="宝贝销售明细报表"]')
                choose1Ele = '//span[text()="商品标题"]/..//input[@aria-checked="false"]'
                choose2Ele = '//span[text()="外部系统编号"]/..//input[@aria-checked="false"]'
                if PlayWright.get_count(choose1Ele):
                    PlayWright.click(choose1Ele)
                if PlayWright.get_count(choose2Ele):
                    PlayWright.click(choose2Ele)
                # 点击生成报表，并切换至新页面
                PlayWright.click('//span[text()="生成报表"]')
                PlayWright.click_catch_new_page('//span[text()="确认"]')
                rowTimeELe = '(//div[contains(@class, "order-export_order-block")])[1]//div[contains(text(), "申请")]'
                downloadEle = '(//div[contains(@class, "order-export_order-block")])[1]//span[text()="下载宝贝报表"]'

            time.sleep(5)
            PlayWright.switch_page(close=False)

            # 判断第一条数据时间是否符合
            rowTime = PlayWright.get_text(f'{rowTimeELe}')[-19:]
            rowTime = time.mktime(time.strptime(rowTime, "%Y-%m-%d %H:%M:%S"))
            if rowTime < timeFlag:
                logger.info(f'未找到符合时间的数据，列表第一条时间为：{rowTime}')
                return False

            # 判断第一条数据下载按钮是否存在
            for buttonRoll in range(1, 6):
                time.sleep(5)
                if PlayWright.get_count(downloadEle):
                    break
                PlayWright.reload()
            if not PlayWright.get_count(downloadEle):
                logger.info('报表未生成成功')
                return False

            # 导出
            with PlayWright.page.expect_download(timeout=15000) as download_info:
                PlayWright.click(downloadEle)
                pass  # 等待下载触发
            download = download_info.value
            # 获取文件名并保存
            download.save_as(fileName)
            return True
        except Exception as e:
            logger.error(f'❌️ {fileName}-页面导出-临时下载异常：{e}')
            return False

    @classmethod
    def salesSingleRun(cls, account_id):
        """销量明细-单个店铺运行"""
        title = f'========================开始爬取淘宝第{account_id}个店铺销量详情======================='
        logger.info(title)

        # 结束时间为昨天，开始时间为结束时间的当月第一天
        endDate = time.strftime("%Y-%m-%d", time.localtime(time.time() - 86400))
        startDate = endDate[:-2] + '01'

        shopName = cls.tbLogin(account_id)
        if not shopName:
            logger.error(f'淘宝第{account_id}个店铺登录异常')
            return False



        fileName = f'淘宝-{shopName}店铺{endDate}销量明细.xlsx'
        fileName = os.path.join(dataDir, fileName)

        saveStatus = False
        for roll in range(1, 6):
            logger.info(f'开始第{roll}次尝试导出明细')
            # 搜索
            cls.salesSearch(startDate, endDate)
            # 页面导出
            saveStatus = cls.salesHtmlSave(fileName)
            time.sleep(10)
            PlayWright.switch_page('old')
            if saveStatus:
                break
            logger.info('休息5*60秒再次尝试导出')
            time.sleep(5*60)

        text = f'✅️ {shopName}明细数据下载成功：{fileName}' if saveStatus else f'❌️ {shopName}明细数据下载失败'
        logger.info(text)

        PlayWright.clear_cookie()
        return True if saveStatus else False

    @classmethod
    def salesRun(cls, startId, endId):
        """统筹运行销量明细"""
        for account_id in range(startId, endId):
            try:
                cls.salesSingleRun(account_id)
            except Exception as e:
                logger.error(f'第{account_id}个店铺查询【销量明细】操作流程失败：{e}')


class WeiDian(object):

    @classmethod
    def wdLogin(cls, account_id):
        logger.info('开始登录微店....')
        url = 'https://d.weidian.com/weidian-pc/login/#/shopSelect'
        ele = '//div[@class="nick-name"]'
        key = f'login.wd_cookie_1' if account_id <= 5 else f'login.wd_cookie_2' if account_id <= 10 else f'login.wd_cookie_3'

        idx = (account_id - 1) % 5 + 1
        extra = f'(//div[text()="子账号店铺:"]/../../div[@data-spider-mode="trackAction"])[{idx}]/div[1]'
        loginStatus = PlayWright.login(url, ele, key, extra=extra, file=config_file)

        if loginStatus:
            shopName = PlayWright.get_text('//div[@class="user-name"]')
            logger.info(f'✅️ 【店铺：{shopName}】微店登录成功....')
            return shopName
        else:
            logger.error(f'❌️ 微店登录失败')
            return False

    @classmethod
    def fundsSearch(cls, startDate, endDate):
        """访问账号资金明细页面，进行搜索"""
        PlayWright.goto('https://d.weidian.com/weidian-pc/weidian-loader/#/pc-vue-balance/overview')
        time.sleep(8)

        # 循环关闭弹窗
        know_ele = '(//button[text()="知道了" or text()="跳过" or text()="完成"])[last()]'
        for roll in range(10):
            if PlayWright.get_count(know_ele):
                PlayWright.click(know_ele)
                time.sleep(1)

        PlayWright.input('//input[@placeholder="开始日期"]', startDate)
        PlayWright.input('//input[@placeholder="结束日期"]', endDate, enter=True)
        PlayWright.click('//span[text()="筛选"]')

    @classmethod
    def fundsPreHtmlSave(cls):
        """账号资金明细-页面预导出"""
        timeFlag = time.time()
        PlayWright.click('(//span[text()="导出报表"])[1]')
        if PlayWright.get_count('//span[text()="暂无数据"]'):
            return False
        PlayWright.click('//span[text()="生成报表"]')
        return timeFlag

    @classmethod
    def fundsHtmlSave(cls, fileName, timeFlag):
        """账号资金明细-页面导出"""
        try:
            PlayWright.goto('https://d.weidian.com/weidian-pc/weidian-loader/#/pc-vue-balance/exportList?shopType=master')
            time.sleep(5)

            # 判断第一条数据时间是否符合
            firstRowEle ='(//div[@class="record-item"])[1]'
            rowTime = PlayWright.get_text(f'{firstRowEle}/div[1]/div[1]')[-19:]
            rowTime = time.mktime(time.strptime(rowTime, "%Y-%m-%d %H:%M:%S"))
            if rowTime < timeFlag:
                logger.info(f'未找到符合时间的数据，列表第一条时间为：{rowTime}')
                return False

            # 直接导出
            with PlayWright.page.expect_download(timeout=15000) as download_info:
                PlayWright.click(f'{firstRowEle}//span[text()="下载报表"]')
                time.sleep(3)
            download = download_info.value
            download.save_as(fileName)
            return True
        except Exception as e:
            logger.error(f'❌️ {fileName}-页面导出-临时下载异常：{e}')
            return False

    @classmethod
    def fundsDataDeal(cls, shopName, fileName):
        """汇总账号资金详细数据"""
        try:
            # 读取Excel文件
            df = pandas.read_excel(fileName)

            # 将数值列转换为数字类型（处理可能的文本格式数字）
            df['收入(元)'] = df['收入(元)'].astype(str).str.replace(',', '').str.strip()
            df['收入(元)'] = pandas.to_numeric(df['收入(元)'].replace('', '0'), errors='coerce').fillna(0)

            # 提取日期（从入账时间）
            df['日期'] = pandas.to_datetime(df['时间']).dt.date

            # 根据账单类型分类
            # 1. 提现数据
            df_withdraw = df[df['账单类型'] == '提现']

            # 2. 货款收入数据
            df_goods_income = df[df['账单类型'] == '货款收入']

            # 3. 交易手续费数据
            df_fee = df[df['账单类型'] == '交易手续费']

            # 按日汇总提现金额
            if len(df_withdraw) > 0:
                df_withdraw_summary = df_withdraw.groupby('日期').agg({
                    '收入(元)': 'sum'
                }).reset_index()
                df_withdraw_summary.columns = ['日期', '提现金额']
            else:
                df_withdraw_summary = pandas.DataFrame(columns=['日期', '提现金额'])

            # 按日汇总货款收入
            if len(df_goods_income) > 0:
                df_goods_summary = df_goods_income.groupby('日期').agg({
                    '收入(元)': 'sum'}).reset_index()
                df_goods_summary.columns = ['日期', '货款收入']
            else:
                df_goods_summary = pandas.DataFrame(columns=['日期', '货款收入'])

            # 按日汇总交易手续费
            if len(df_fee) > 0:
                df_fee_summary = df_fee.groupby('日期').agg({
                    '收入(元)': 'sum'
                }).reset_index()
                df_fee_summary.columns = ['日期', '交易手续费']
            else:
                df_fee_summary = pandas.DataFrame(columns=['日期', '交易手续费'])

            # 合并三个汇总表
            df_summary = df_goods_summary.copy()
            df_summary = pandas.merge(df_summary, df_withdraw_summary, on='日期', how='outer')
            df_summary = pandas.merge(df_summary, df_fee_summary, on='日期', how='outer')

            # 填充空值为0
            df_summary = df_summary.fillna(0)

            # 计算每日总净收入
            df_summary['日净收入'] = df_summary['货款收入'] + df_summary['交易手续费']

            # 按日期排序
            df_summary = df_summary.sort_values('日期', ascending=True).reset_index(drop=True)

            # 添加汇总行
            total_row = pandas.DataFrame({
                '日期': ['所有汇总'],
                '货款收入': [df_summary['货款收入'].sum()],
                '提现金额': [df_summary['提现金额'].sum()],
                '交易手续费': [df_summary['交易手续费'].sum()],
                '日净收入': [df_summary['日净收入'].sum()]
            })

            df_summary = pandas.concat([df_summary, total_row], ignore_index=True)

            # 重新排列列顺序
            columns_order = ['日期', '货款收入', '提现金额', '交易手续费', '日净收入']
            df_summary = df_summary[columns_order]

            # 使用 ExcelWriter 追加到现有文件，保留原有的 Sheet
            with pandas.ExcelWriter(fileName, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                sheet_name = '汇总数据'
                # 先删除可能已存在的同名 Sheet
                if sheet_name in writer.book.sheetnames:
                    del writer.book[sheet_name]

                # 写入数据
                df_summary.to_excel(writer, sheet_name=sheet_name, index=False)

                # 获取工作表对象并设置样式
                ws = writer.sheets[sheet_name]
                header_font = Font(bold=True, color='FFFFFF', size=11)
                header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                header_alignment = Alignment(horizontal='center', vertical='center')

                # 遍历第一行所有单元格设置样式
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment

                # 自动调整列宽
                for column in ws.columns:
                    max_length = 0
                    col_letter = column[0].column_letter
                    for cell in column:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))

                    ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

                # 消除 "Workbook contains no default style" 警告
                workbook = writer.book
                if not workbook.style_names:
                    default_font = Font(name='Calibri', size=11, bold=False, italic=False)
                    default_style = openpyxl.styles.NamedStyle(name='Normal', font=default_font)
                    workbook.add_named_style(default_style)

            logger.info(f'数据汇总完成，共汇总{len(df_summary) - 1}天的数据，已保存到: {fileName}')

            logger.info(
                f'总收入: {df_summary["货款收入"].sum() / 2:.2f}，交易手续费: {df_summary["交易手续费"].sum() / 2:.2f} '
                f'总提现: {df_summary["提现金额"].sum() / 2:.2f}，总净收入: {df_summary["日净收入"].sum() / 2:.2f}\n')

            return df_summary

        except Exception as e:
            logger.error(f'{shopName}店铺数据处理失败: {e}\n')
            return None

    @classmethod
    def fundsSingleRun(cls, account_id):
        """账户资金明细-单个店铺运行"""
        title = f'========================开始爬取微店第{account_id}个店铺账号资金详情======================='
        logger.info(title)

        # 结束时间为昨天，开始时间为结束时间的当月第一天
        endTime = time.strftime("%Y-%m-%d", time.localtime(time.time() - 86400))
        startTime = endTime[:-2] + '01'

        shopName = cls.wdLogin(account_id)
        if not shopName:
            logger.error(f'微店第{account_id}个店铺登录异常')
            return False

        fileName = f'微店-{shopName}店铺{endTime}账户资金明细.xlsx'
        fileName = os.path.join(dataDir, fileName)

        saveStatus = False
        for roll in range(1, 6):
            logger.info(f'开始第{roll}次尝试导出明细')
            # 搜索
            cls.fundsSearch(startTime, endTime)
            time.sleep(3)
            # 预生成
            timeFlag = cls.fundsPreHtmlSave()
            if not timeFlag:
                continue
            logger.info('正在预生成，休息60秒导出')
            time.sleep(60)
            # 页面导出
            saveStatus = cls.fundsHtmlSave(fileName, timeFlag)
            if saveStatus:
                break

        text = f'✅️ {shopName}明细数据下载成功：{fileName}' if saveStatus else f'❌️ {shopName}明细数据下载失败'
        logger.info(text)
        if saveStatus:
            cls.fundsDataDeal(shopName, fileName)
        PlayWright.clear_cookie()
        return True if saveStatus else False

    @classmethod
    def fundsRun(cls, startId, endId):
        """统筹运行账号资金明细"""
        for account_id in range(startId, endId):
            try:
                cls.fundsSingleRun(account_id)
            except Exception as e:
                logger.error(f'第{account_id}个店铺查询【账号资金明细】操作流程失败：{e}')

    @classmethod
    def salesSearch(cls, startDate, endDate):
        """访问销量明细页面，进行搜索"""
        PlayWright.goto('https://d.weidian.com/weidian-pc/weidian-loader/#/pc-vue-order/orderList')
        time.sleep(8)

        # 循环关闭弹窗
        know_ele = '(//button[text()="知道了" or text()="跳过" or text()="完成"])[last()]'
        for roll in range(10):
            if PlayWright.get_count(know_ele):
                PlayWright.click(know_ele)
                time.sleep(1)

        PlayWright.click('//input[@placeholder="请输入起始下单时间"]')
        PlayWright.input('//input[@placeholder="选择日期"]', startDate)
        PlayWright.input('//input[@placeholder="选择时间"]', '00:00:00', enter=True)
        PlayWright.page.press('//input[@placeholder="选择时间"]', 'Enter')

        PlayWright.click('//input[@placeholder="请输入结束下单时间"]')
        PlayWright.input('(//input[@placeholder="选择日期"])[2]', endDate)
        PlayWright.input('(//input[@placeholder="选择时间"])[2]', '23:59:59', enter=True)
        PlayWright.page.press('(//input[@placeholder="选择时间"])[2]', 'Enter')

        PlayWright.click('(//span[contains(text(),"筛选")])[1]')
        PlayWright.click('//div[@class="tab-list"]/div[1]')
        time.sleep(5)
        if PlayWright.get_count('//div[@class="no-order-list"]'):
            return False
        PlayWright.click('(//span[contains(text(),"批量导出")])[1]')
        PlayWright.mouse_wheel(200)
        return True

    @classmethod
    def salesHtmlSave(cls, fileName):
        try:
            # 勾选对应字段
            timeFlag = time.time()

            PlayWright.click_catch_new_page('(//span[contains(text(),"确认导出")])[1]')

            # 切换新页面
            PlayWright.switch_page(close=False)

            # 判断第一条数据时间是否符合
            firstRowEle = '//div[@class="report-card"][1]'
            rowTime = PlayWright.get_text(f'{firstRowEle}//div[@class="down-time"]')[7:26]
            rowTime = time.mktime(time.strptime(rowTime, "%Y-%m-%d %H:%M:%S"))
            if rowTime < timeFlag:
                logger.info(f'未找到符合时间的数据，列表第一条时间为：{rowTime}')
                return False

            # 判断第一条数据下载按钮是否存在
            downloadEle = f'{firstRowEle}//span[text()="下载报表"]'
            for buttonRoll in range(1, 6):
                time.sleep(5)
                if PlayWright.get_count(downloadEle):
                    break
                PlayWright.reload()
            if not PlayWright.get_count(downloadEle):
                logger.info('报表未生成成功')
                return False

            # 导出
            with PlayWright.page.expect_download(timeout=15000) as download_info:
                PlayWright.click(downloadEle)
                pass  # 等待下载触发
            download = download_info.value
            # 获取文件名并保存
            download.save_as(fileName)
            return True
        except Exception as e:
            logger.error(f'❌️ {fileName}-页面导出-临时下载异常：{e}')
            PlayWright.switch_page('old')
            return False

    @classmethod
    def salesSingleRun(cls, account_id):
        """销量明细-单个店铺运行"""
        title = f'========================开始爬取微店第{account_id}个店铺销量详情======================='
        logger.info(title)

        # 结束时间为昨天，开始时间为结束时间的当月第一天
        endDate = time.strftime("%Y-%m-%d", time.localtime(time.time() - 86400))
        startDate = endDate[:-2] + '01'

        shopName = cls.wdLogin(account_id)
        if not shopName:
            logger.error(f'微店第{account_id}个店铺登录异常')
            return False

        fileName = f'微店-{shopName}店铺{endDate}销量明细.xlsx'
        fileName = os.path.join(dataDir, fileName)

        saveStatus = False
        for roll in range(1, 6):
            logger.info(f'开始第{roll}次尝试导出明细')
            # 搜索
            existData = cls.salesSearch(startDate, endDate)
            if not existData:
                logger.info('当前账号暂无销量数据')
                return False
            # 页面导出
            saveStatus = cls.salesHtmlSave(fileName)
            time.sleep(10)
            PlayWright.switch_page('old')
            if saveStatus:
                break
            logger.info('休息5*60秒再次尝试导出')
            time.sleep(5*60)

        text = f'✅️ {shopName}明细数据下载成功：{fileName}' if saveStatus else f'❌️ {shopName}明细数据下载失败'
        logger.info(text)

        PlayWright.clear_cookie()
        return True if saveStatus else False

    @classmethod
    def salesRun(cls, startId, endId):
        """统筹运行销量明细"""
        for account_id in range(startId, endId):
            try:
                cls.salesSingleRun(account_id)
            except Exception as e:
                logger.error(f'第{account_id}个店铺查询【销量明细】操作流程失败：{e}')


class DouDian(object):
    shop_names = [
        '来口大米饭',
        'wawa同人社',
        '甜心花栗',
        'meow meow',
        'Meow市集',
        '海苔不睡',
        'wawa小厨'
    ]

    @classmethod
    def ddLogin(cls, account_id):
        logger.info('开始登录抖店....')
        try:
            cookie = get_config_value('login', f'dd_cookie_{account_id}', file=config_file)
            if cookie:
                PlayWright.add_cookie(eval(cookie))

            url = 'https://fxg.jinritemai.com/login/common?channel=zhaoshang'
            PlayWright.goto(url)
            time.sleep(30)

            login_ele = '//div[contains(@class,"index_userName")]'
            element = PlayWright.wait_for_selector(login_ele, timeout=20 * 1000)

            # 未登录
            if not element:
                logger.info('请登录......')
                choose = ((account_id - 1) % 6)
                sub_account = f'//div[text()="{cls.shop_names[choose]}"]'
                element = PlayWright.wait_for_selector(sub_account, timeout=3 * 60 * 1000)
                if not element:
                    logger.error(f'抖店登录失败')
                    return False
                PlayWright.click(sub_account)
                time.sleep(20)

            choose_shop = PlayWright.get_text(login_ele)
            logger.info(f'当前店铺名称：{choose_shop}')

            # 页面cookie
            cookie_list = PlayWright.context.cookies()

            # api_cookie
            api_cookie = "; ".join([f"{cookie['name']}={cookie['value']}" for cookie in cookie_list])
            config_info = {
                f'dd_cookie_{account_id}': cookie_list,
                f'dd_cookie_{account_id}_api': api_cookie
            }
            write_config_value('login', config_info, file=config_file)

            logger.info('抖店登录成功....')
            return choose_shop

        except Exception as e:
            logger.error(f'抖店登录失败：{e}')
            return False

    @classmethod
    def fundsSearch(cls, startDate, endDate):
        """访问账号资金明细页面，进行搜索"""
        PlayWright.goto('https://fxg.jinritemai.com/ffa/fxg-bill/fund-detail-bill')
        time.sleep(8)

        # 循环关闭弹窗
        know_ele = '(//button[text()="知道了" or text()="跳过" or text()="完成"])[last()]'
        for roll in range(10):
            if PlayWright.get_count(know_ele):
                PlayWright.click(know_ele)
                time.sleep(1)

        start_ele = '//input[@placeholder="开始日期"]'
        end_ele = '//input[@placeholder="结束日期"]'
        PlayWright.click(start_ele)
        PlayWright.input(start_ele, startDate, enter=True)
        PlayWright.click(end_ele)
        PlayWright.input(end_ele, endDate, enter=True)
        PlayWright.click('//span[text()="查询"]')

    @classmethod
    def fundsHtmlSave(cls, fileName):
        """账号资金明细-页面导出"""
        try:
            PlayWright.click('//span[text()="生成报表"]')
            PlayWright.click('//span[text()="生成"]')
            time.sleep(5)
            wait_ele = '(//div[contains(@class,"cardList")]/li)[1]//div[text()="生成中"]'
            if PlayWright.get_count(wait_ele):
                logger.info('正在生成报表，等待30秒....')
                time.sleep(30)
                PlayWright.reload()
                time.sleep(3)

            with PlayWright.page.expect_download(timeout=15000) as download_info:
                PlayWright.click('(//div[contains(@class,"cardList")]/li)[1]//span[text()="下载"]')
                pass  # 等待下载触发
            download = download_info.value
            # 获取文件名并保存
            tmpFile = fileName[:-5] + 'csv'
            download.save_as(tmpFile)
            df = pandas.read_csv(tmpFile, encoding="utf-8")
            df.to_excel(fileName, index=False, engine="openpyxl")
            os.remove(tmpFile)
            return True
        except Exception as e:
            logger.error(f'❌️ {fileName}-页面导出-临时下载异常：{e}')
            return False

    @classmethod
    def fundsDataDeal(cls, shopName, fileName):
        """汇总账号资金详细数据"""
        try:

            # 读取Excel文件
            df = pandas.read_excel(fileName)

            # 将数值列转换为数字类型（处理可能的文本格式数字）
            df['动账金额'] = df['动账金额'].astype(str).str.replace(',', '').str.strip()
            df['动账金额'] = pandas.to_numeric(df['动账金额'], errors='coerce').fillna(0)
            # 提取日期
            df['日期'] = pandas.to_datetime(df['动账时间']).dt.date

            # 分类数据
            # 1. 充值保证金
            df_recharge = df[df['动账场景'] == '充值保证金']

            # 2. 提现
            df_withdraw = df[df['动账场景'] == '提现']

            # 3. 其他交易（排除提现）
            df_other = df[~df['动账场景'].apply(lambda x: '提现' in x)]

            # 按日汇总充值保证金（仅展示）
            if len(df_recharge) > 0:
                df_recharge_summary = df_recharge.groupby('日期').agg({
                    '动账金额': 'sum'
                }).reset_index()
                df_recharge_summary.columns = ['日期', '充值保证金']
            else:
                df_recharge_summary = pandas.DataFrame(columns=['日期', '充值保证金'])

            # 按日汇总提现（仅展示）
            if len(df_withdraw) > 0:
                df_withdraw_summary = df_withdraw.groupby('日期').agg({
                    '动账金额': 'sum'
                }).reset_index()
                df_withdraw_summary.columns = ['日期', '提现金额']
            else:
                df_withdraw_summary = pandas.DataFrame(columns=['日期', '提现金额'])

            # 按日汇总其他交易，区分入账和出账
            if len(df_other) > 0:
                # 入账户
                df_income = df_other[df_other['动账方向'] == '入账']
                df_income_summary = df_income.groupby('日期').agg({
                    '动账金额': 'sum'
                }).reset_index()
                df_income_summary.columns = ['日期', '入账金额']

                # 出账户
                df_expense = df_other[df_other['动账方向'] == '出账']
                df_expense_summary = df_expense.groupby('日期').agg({
                    '动账金额': 'sum'
                }).reset_index()
                df_expense_summary.columns = ['日期', '出账金额']

                # 合并入账和出账
                df_other_summary = pandas.merge(df_income_summary, df_expense_summary, on='日期', how='outer')
                df_other_summary = df_other_summary.fillna(0)
                df_other_summary['日净收入'] = df_other_summary['入账金额'] - df_other_summary['出账金额']
            else:
                df_other_summary = pandas.DataFrame(columns=['日期', '入账金额', '出账金额', '日净收入'])

            # 合并所有汇总表
            df_summary = df_other_summary.copy()
            df_summary = pandas.merge(df_summary, df_recharge_summary, on='日期', how='outer')
            df_summary = pandas.merge(df_summary, df_withdraw_summary, on='日期', how='outer')

            # 填充空值为0
            df_summary = df_summary.fillna(0)

            # 按日期排序
            df_summary = df_summary.sort_values('日期', ascending=True).reset_index(drop=True)
            # 汇总数据
            total_row = pandas.DataFrame({
                '日期': ['所有汇总'],
                '入账金额': [df_summary['入账金额'].sum()],
                '出账金额': [df_summary['出账金额'].sum()],
                '充值保证金': [df_summary['充值保证金'].sum()],
                '日净收入': [df_summary['日净收入'].sum()],
                '提现金额': [df_summary['提现金额'].sum()]
            })
            df_summary = pandas.concat([df_summary, total_row], ignore_index=True)

            # 使用 ExcelWriter 追加到现有文件，保留原有的 Sheet1
            with pandas.ExcelWriter(fileName, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                sheet_name = '汇总数据'
                # 先删除可能已存在的同名 Sheet
                if sheet_name in writer.book.sheetnames:
                    del writer.book[sheet_name]

                # 写入数据
                df_summary.to_excel(writer, sheet_name=sheet_name, index=False)

                # 获取工作表对象并设置样式
                ws = writer.sheets[sheet_name]
                header_font = Font(bold=True, color='FFFFFF', size=11)
                header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                header_alignment = Alignment(horizontal='center', vertical='center')

                # 遍历第一行所有单元格设置样式
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment

                # 自动调整列宽
                for column in ws.columns:
                    max_length = 0
                    col_letter = column[0].column_letter
                    for cell in column:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))

                    ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

                # 消除 "Workbook contains no default style" 警告
                workbook = writer.book
                if not workbook.style_names:
                    default_font = Font(name='Calibri', size=11, bold=False, italic=False)
                    default_style = openpyxl.styles.NamedStyle(name='Normal', font=default_font)
                    workbook.add_named_style(default_style)

            logger.info(f'数据汇总完成，共汇总{len(df_summary)}天的数据，已保存到: {fileName}')

            # 打印汇总统计
            logger.info(
                f'总收入: {df_summary["入账金额"].sum() / 2:.2f}，总支出: {df_summary["出账金额"].sum() / 2:.2f} '
                f'总提现: {df_summary["提现金额"].sum() / 2:.2f}，总净收入: {df_summary["日净收入"].sum() / 2:.2f}\n')
            return df_summary

        except Exception as e:
            logger.error(f'{shopName}店铺数据处理失败: {e}\n')
            return None

    @classmethod
    def fundsSingleRun(cls, account_id):
        """账户资金明细-单个店铺运行"""
        title = f'========================开始爬取抖店第{account_id}个店铺账号资金详情======================='
        logger.info(title)

        # 结束时间为昨天，开始时间为结束时间的当月第一天
        endTime = time.strftime("%Y-%m-%d", time.localtime(time.time() - 86400))
        startTime = endTime[:-2] + '01'

        shopName = cls.ddLogin(account_id)
        if not shopName:
            logger.error(f'抖店第{account_id}个店铺登录异常')
            return False

        fileName = f'抖店-{shopName}店铺{endTime}账户资金明细.xlsx'
        fileName = os.path.join(dataDir, fileName)

        saveStatus = False
        for roll in range(1, 6):
            logger.info(f'开始第{roll}次尝试导出明细')
            # 搜索
            cls.fundsSearch(startTime, endTime)
            time.sleep(3)

            # 页面导出
            saveStatus = cls.fundsHtmlSave(fileName)
            if saveStatus:
                break

        text = f'✅️ {shopName}明细数据下载成功：{fileName}' if saveStatus else f'❌️ {shopName}明细数据下载失败'
        logger.info(text)
        if saveStatus:
            cls.fundsDataDeal(shopName, fileName)
        PlayWright.clear_cookie()
        return True if saveStatus else False

    @classmethod
    def fundsRun(cls, startId, endId):
        """统筹运行账号资金明细"""
        for account_id in range(startId, endId):
            try:
                cls.fundsSingleRun(account_id)
            except Exception as e:
                logger.error(f'第{account_id}个店铺查询【账号资金明细】操作流程失败：{e}')

    @classmethod
    def salesSearch(cls, startDate, endDate):
        """访问销量明细页面，进行搜索"""
        PlayWright.goto('https://fxg.jinritemai.com/ffa/morder/order/list?btm_ppre=a2427.b76571.c902327.d871297&btm_p'
                        're=a2427.b76571.c902327.d871297')
        time.sleep(8)

        startMonth = startDate.split('-')[1]
        startMonth = startMonth[1] if startMonth[0] == '0' else startMonth
        startDay = startDate.split('-')[2]
        startDay = startDay[1] if startDay[0] == '0' else startDay
        endMonth = endDate.split('-')[1]
        endMonth = endMonth[1] if endMonth[0] == '0' else endMonth
        endDay = endDate.split('-')[2]
        endDay = endDay[1] if endDate[0] == '0' else endDay

        # 循环关闭弹窗
        know_ele = '(//button[text()="知道了" or text()="跳过" or text()="完成"])[last()]'
        for roll in range(10):
            if PlayWright.get_count(know_ele):
                PlayWright.click(know_ele)
                time.sleep(1)

        PlayWright.click('//input[@placeholder="开始时间"]')
        PlayWright.click(f'//button[text()="{startMonth}月"]/../../../div[2]//td[contains(@class, "view")]/div[text()="{startDay}"]')
        PlayWright.click(f'//button[text()="{endMonth}月"]/../../../div[2]//td[contains(@class, "view")]/div[text()="{endDay}"]')
        PlayWright.click('//span[text()="确定"]')
        PlayWright.click('//span[text()="查询"]')
        time.sleep(3)
        if PlayWright.get_count('//div[text()="暂无数据"]'):
            return False
        return True

    @classmethod
    def salesHtmlSave(cls, fileName):
        try:
            # 勾选对应字段
            timeFlag = time.time()
            PlayWright.click('//span[text()="导出订单"]')
            PlayWright.click('//span[text()="导出"]')
            PlayWright.click('//span[text()="确认"]')

            time.sleep(10)
            PlayWright.reload()

            # 判断第一条数据时间是否符合
            firstRowEle = '//li[contains(@class,"cardItem")][1]'
            rowTime = PlayWright.get_text(f'{firstRowEle}//span[text()="生成时间"]/../span[2]')
            rowTime = time.mktime(time.strptime(rowTime, "%Y/%m/%d %H:%M:%S"))
            if rowTime < timeFlag:
                logger.info(f'未找到符合时间的数据，列表第一条时间为：{rowTime}')
                return False

            # 判断第一条数据下载按钮是否存在
            downloadEle = f'{firstRowEle}//span[text()="下载报表"]'
            verifyEle = '//span[text()="获取验证码"]'
            if PlayWright.get_count(verifyEle):
                PlayWright.click(verifyEle)
                logger.info('请在60s内手动输入验证码，注：无需点击确定')
                time.sleep(60)
                downloadEle = '//span[text()="确定"]'

            # 导出
            with PlayWright.page.expect_download(timeout=15000) as download_info:
                PlayWright.click(downloadEle)
                pass  # 等待下载触发
            download = download_info.value
            # 获取文件名并保存
            download.save_as(fileName)
            PlayWright.switch_page('old')
            return True
        except Exception as e:
            logger.error(f'❌️ {fileName}-页面导出-临时下载异常：{e}')
            PlayWright.switch_page('old')
            return False

    @classmethod
    def salesSingleRun(cls, account_id):
        """销量明细-单个店铺运行"""
        title = f'========================开始爬取抖店第{account_id}个店铺销量详情======================='
        logger.info(title)

        # 结束时间为昨天，开始时间为结束时间的当月第一天
        endDate = time.strftime("%Y-%m-%d", time.localtime(time.time() - 86400))
        startDate = endDate[:-2] + '01'

        shopName = cls.ddLogin(account_id)
        if not shopName:
            logger.error(f'抖店第{account_id}个店铺登录异常')
            return False

        fileName = f'抖店-{shopName}店铺{endDate}销量明细.xlsx'
        fileName = os.path.join(dataDir, fileName)

        saveStatus = False
        for roll in range(1, 6):
            logger.info(f'开始第{roll}次尝试导出明细')
            # 搜索
            existData = cls.salesSearch(startDate, endDate)
            if not existData:
                logger.info('当前账号暂无销量数据')
                return False
            # 页面导出
            saveStatus = cls.salesHtmlSave(fileName)
            if saveStatus:
                break
            logger.info('休息5*60秒再次尝试导出')
            time.sleep(5*60)

        text = f'✅️ {shopName}明细数据下载成功：{fileName}' if saveStatus else f'❌️ {shopName}明细数据下载失败'
        logger.info(text)

        PlayWright.clear_cookie()
        return True if saveStatus else False

    @classmethod
    def salesRun(cls, startId, endId):
        """统筹运行销量明细"""
        for account_id in range(startId, endId):
            try:
                cls.salesSingleRun(account_id)
            except Exception as e:
                logger.error(f'第{account_id}个店铺查询【销量明细】操作流程失败：{e}')


class PDD(object):

    @classmethod
    def pddLogin(cls, account_id):
        logger.info('开始登录拼多多....')
        url = 'https://mms.pinduoduo.com/home'
        ele = '//div[@class="user-name-name"]/span'
        key = f'login.pdd_cookie_{account_id}'
        loginStatus = PlayWright.login(url, ele, key, file=config_file)
        if loginStatus:
            shopName = PlayWright.get_text('//span[contains(@class,"Header_header_name")]')
            logger.info(f'✅️ 【店铺：{shopName}】拼多多登录成功....')
            return shopName
        else:
            logger.error(f'❌️ 拼多多登录失败')
            return False

    @classmethod
    def fundsSearch(cls, startDate, endDate):
        """访问账号资金明细页面，进行搜索"""
        PlayWright.page.mouse.click(700, 500, button='left')
        PlayWright.click('//span[text()="资金中心"]')
        PlayWright.switch_to_page()
        time.sleep(5)

        start = startDate[-1:] if startDate[-2] == '0' else startDate[-2:]
        end = endDate[-1:] if endDate[-2] == '0' else endDate[-2:]

        PlayWright.click('//a[text()="收支明细"]')
        PlayWright.click('//input[@placeholder="开始日期-结束日期"]')
        PlayWright.click(f'(//td[not (contains(@class,"outOfMonth"))]/div[text()="{start}"])[last()]')
        PlayWright.click(f'(//td[not (contains(@class,"outOfMonth"))]/div[text()="{end}"])[last()]')
        PlayWright.mouse_wheel(80)
        sure_ele = '//span[text()="确认"]'

        for sureRoll in range(1, 6):
            if PlayWright.get_count(sure_ele):
                PlayWright.click(sure_ele)
        PlayWright.click('//span[text()="查询"]')

    @classmethod
    def fundsHtmlSave(cls, fileName):
        """账号资金明细-页面导出"""
        try:
            with PlayWright.page.expect_download(timeout=15000) as download_info:
                PlayWright.click('(//div[@class="export-history-bills-card"])[1]//span[text()="下载账单"]')
                pass  # 等待下载触发
            download = download_info.value
            # 获取文件名并保存
            tmpFile = fileName[:-5] + 'zip'
            # 解压zip文件
            with zipfile.ZipFile(tmpFile, 'r') as zip_ref:
                # 获取压缩包内的文件名
                file_list = zip_ref.namelist()
                if len(file_list) == 0:
                    logger.error('压缩包内没有文件')
                    return False

                # 获取第一个文件（通常只有一个）
                inner_filename = file_list[0]

                # 解压到临时目录
                zip_ref.extractall(os.path.dirname(tmpFile))

            # 构建解压后的文件路径
            extracted_file = os.path.join(os.path.dirname(tmpFile), inner_filename)

            inner_filename = os.path.join(os.path.dirname(tmpFile), inner_filename)
            # 如果解压出来的是csv文件，直接重命名为目标文件名
            if inner_filename.endswith('.csv'):
                df = pandas.read_csv(inner_filename, encoding="gbk", skiprows=4)
                if len(df) > 4:
                    df = df.iloc[:-4]
                df.to_excel(fileName, index=False, engine="openpyxl")

            # 删除zip文件

            os.remove(tmpFile)
            os.remove(inner_filename)

            return True
        except Exception as e:
            logger.error(f'❌️ {fileName}-页面导出-临时下载异常：{e}')
            return False

    @classmethod
    def fundsDataDeal(cls, shopName, fileName):
        """汇总账号资金详细数据"""
        try:
            # 读取Excel文件
            df = pandas.read_excel(fileName)

            # 将数值列转换为数字类型（处理可能的文本格式数字）
            df['收入金额'] = pandas.to_numeric(df['收入金额（+元）'], errors='coerce').fillna(0)
            df['支出金额'] = pandas.to_numeric(df['支出金额（-元）'].replace('-', '0'), errors='coerce').fillna(0)
            # 提取日期
            df['日期'] = pandas.to_datetime(df['发生时间']).dt.date

            # 分离提现和非提现数据
            df_withdraw = df[df['账务类型'].str.contains('提现', na=False)]
            df_normal = df[~df['账务类型'].str.contains('提现', na=False)]

            # 汇总非提现的收入和支出
            df_normal_summary = df_normal.groupby('日期').agg({
                '收入金额': 'sum',
                '支出金额': 'sum'
            }).reset_index()
            df_normal_summary.columns = ['日期', '日收入', '日支出']

            # 汇总提现数据（提现金额在支出列）
            df_withdraw_summary = df_withdraw.groupby('日期').agg({
                '支出金额': 'sum'
            }).reset_index()
            df_withdraw_summary.columns = ['日期', '日提现金额']

            # 合并两个汇总表
            df_summary = pandas.merge(df_normal_summary, df_withdraw_summary, on='日期', how='outer')

            # 填充空值为0
            df_summary = df_summary.fillna(0)

            # 计算日净收入（收入 - 支出 - 提现）
            df_summary['日净收入'] = df_summary['日收入'] + df_summary['日支出']

            # 按日期排序
            df_summary = df_summary.sort_values('日期', ascending=True).reset_index(drop=True)

            # 汇总数据
            total_row = pandas.DataFrame({
                '日期': ['所有汇总'],
                '日收入': [df_summary['日收入'].sum()],
                '日支出': [df_summary['日支出'].sum()],
                '日提现金额': [df_summary['日提现金额'].sum()],
                '日净收入': [df_summary['日净收入'].sum()]
            })
            df_summary = pandas.concat([df_summary, total_row], ignore_index=True)

            # 使用 ExcelWriter 追加到现有文件，保留原有的 Sheet1
            with pandas.ExcelWriter(fileName, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                sheet_name = '汇总数据'
                # 先删除可能已存在的同名 Sheet
                if sheet_name in writer.book.sheetnames:
                    del writer.book[sheet_name]

                # 写入数据
                df_summary.to_excel(writer, sheet_name=sheet_name, index=False)

                # 获取工作表对象并设置样式
                ws = writer.sheets[sheet_name]
                header_font = Font(bold=True, color='FFFFFF', size=11)
                header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                header_alignment = Alignment(horizontal='center', vertical='center')

                # 遍历第一行所有单元格设置样式
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment

                # 自动调整列宽
                for column in ws.columns:
                    max_length = 0
                    col_letter = column[0].column_letter
                    for cell in column:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))

                    ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

                # 消除 "Workbook contains no default style" 警告
                workbook = writer.book
                if not workbook.style_names:
                    default_font = Font(name='Calibri', size=11, bold=False, italic=False)
                    default_style = openpyxl.styles.NamedStyle(name='Normal', font=default_font)
                    workbook.add_named_style(default_style)

            logger.info(f'数据汇总完成，共汇总{len(df_summary)}天的数据，已保存到: {fileName}')

            # 打印汇总统计
            logger.info(f'总收入: {df_summary["日收入"].sum() / 2:.2f}，总支出: {df_summary["日支出"].sum() / 2:.2f} '
                        f'总提现: {df_summary["日提现金额"].sum() / 2:.2f}，总净收入: {df_summary["日净收入"].sum() / 2:.2f}\n')
            return df_summary

        except Exception as e:
            logger.error(f'{shopName}数据处理失败: {e}\n')
            return None

    @classmethod
    def fundsSingleRun(cls, account_id):
        """账户资金明细-单个店铺运行"""
        title = f'========================开始爬取拼多多第{account_id}个店铺账号资金详情======================='
        logger.info(title)

        # 结束时间为昨天，开始时间为结束时间的当月第一天
        endTime = time.strftime("%Y-%m-%d", time.localtime(time.time() - 86400))
        startTime = endTime[:-2] + '01'

        shopName = cls.pddLogin(account_id)
        if not shopName:
            logger.error(f'拼多多第{account_id}个店铺登录异常')
            return False

        fileName = f'拼多多-{shopName}店铺{endTime}账户资金明细.xlsx'
        fileName = os.path.join(dataDir, fileName)

        saveStatus = False
        for roll in range(1, 6):
            logger.info(f'开始第{roll}次尝试导出明细')
            cls.fundsSearch(startTime, endTime)
            time.sleep(3)

            # 页面导出
            saveStatus = cls.fundsHtmlSave(fileName)
            if saveStatus:
                break

        text = f'✅️ {shopName}明细数据下载成功：{fileName}' if saveStatus else f'❌️ {shopName}明细数据下载失败'
        logger.info(text)
        if saveStatus:
            cls.fundsDataDeal(shopName, fileName)
        PlayWright.clear_cookie()
        return True if saveStatus else False

    @classmethod
    def fundsRun(cls, startId, endId):
        """统筹运行账号资金明细"""
        for account_id in range(startId, endId):
            try:
                cls.fundsSingleRun(account_id)
            except Exception as e:
                logger.error(f'第{account_id}个店铺查询【账号资金明细】操作流程失败：{e}')


def deleteAccount(platform, account_ids):
    """删除指定账号"""
    platformDict = {
        '1': 'xhs_cookie_',
        '2': 'tb_cookie_',
        '3': 'wd_cookie_',
        '4': 'dd_cookie_',
        '5': 'pdd_cookie_',
    }

    account_ids = account_ids.split(' ')
    account_ids = [i for i in account_ids if i]

    for account_id in account_ids:
        new = {
            platformDict[platform] + account_id: None,
            platformDict[platform] + account_id + '_api': None,
        }
        write_config_value('login', new, file=config_file)
        logger.info(f'✅️ 第{account_id}个店铺删除成功')


if __name__ == '__main__':
    while True:
        step = input('请输入操作步骤（1.查看账号资金明细，2.查询销量明细，3.删除账号）：')

        platform = input('请输入操作平台（1.小红书，2.淘宝，3.微店，4.抖店，5.拼多多）：')

        if step == '3':
            account_ids = input('请输入删除店铺序号（0默认全部，带空格可删除多个）：')
            deleteAccount(platform, account_ids)
            continue

        startId = input('请输入查询店铺序号（0默认查询全部，序号+：可查询多个）：')

        keywordsDict = {
            '1': 'xhs_shop_count',
            '2': 'tb_shop_count',
            '3': 'wd_shop_count',
            '4': 'dd_shop_count',
            '5': 'pdd_shop_count',
        }

        # 获取店铺数量，处理店铺索引
        shopCount = get_config_value('login', keywordsDict[platform], file=config_file)

        if startId == '0':
            startId = 1
            endId = int(shopCount) + 1
        elif ':' in startId or '：' in startId:
            startId = startId.replace('：', '').replace(':', '')
            startId = int(startId)
            endId = int(shopCount) + 1
        else:
            startId = int(startId)
            endId = startId + 1

        # 根据输入，执行对应步骤
        if step == '1' and platform == '1':
            XHS.fundsRun(startId, endId)
        elif step == '2' and platform == '1':
            XHS.salesRun(startId, endId)
        elif step == '1' and platform == '2':
            TB.fundsRun(startId, endId)
        elif step == '2' and platform == '2':
            TB.salesRun(startId, endId)
        elif step == '1' and platform == '3':
            WeiDian.fundsRun(startId, endId)
        elif step == '2' and platform == '3':
            WeiDian.salesRun(startId, endId)
        elif step == '1' and platform == '4':
            DouDian.fundsRun(startId, endId)
        elif step == '2' and platform == '4':
            DouDian.salesRun(startId, endId)
        elif step == '1' and platform == '5':
            PDD.fundsRun(startId, endId)




