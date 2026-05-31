# -*- coding: utf-8 -*-
import sys
from pathlib import Path

sys.path.append(str(Path(__file__).parent.parent))

from PlayWright import Playwright_, logger, get_config_value
import time
import os
import pandas
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import warnings

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

config_file = os.path.join(os.path.dirname(__file__), 'config.ini')


def tb_login(shop_id=1):
    try:
        logger.info('开始登录千牛....')
        url = 'https://myseller.taobao.com/home.htm/QnworkbenchHome/'
        ele = '//span[contains(text(),"首页")]'
        key = f'login.tb_cookie{shop_id}'

        Playwright_.login(url, ele, key, file=config_file)
        logger.info('千牛登录成功....')
        return True
    except Exception as e:
        logger.error(f'千牛登录失败：{e}')
        return False


def tb_save(filename):
    """直接导出"""
    try:
        # 直接导出
        with Playwright_.page.expect_download(timeout=15000) as download_info:
            pass  # 等待下载触发
        download = download_info.value
        # 获取文件名并保存
        download.save_as(filename)
        return True
    except Exception as e:
        logger.info(f'{filename}临时下载异常：{e}')
        return False


def tb_search(start, end):
    Playwright_.goto('https://qn.taobao.com/home.htm/whale-accountant/pay/capital/home?active=fund_detail')
    shop_name = Playwright_.get_text('//div[@class="user-area-pop-up-panel"]/div[1]/div/div[1]')
    shop_name = shop_name.replace('账号信息店铺信息页面设置退出当前账号', '')
    logger.info(f'当前店铺名称：{shop_name}')
    time.sleep(8)

    Playwright_.input('//input[@placeholder="起始日期"]', start)
    Playwright_.input('//input[@placeholder="结束日期"]', end, enter=True)
    Playwright_.click('//span[text()="搜索"]')
    return shop_name


def tb_deal_data(shopname, file):
    """
    使用pandas处理千牛数据：
    1. 按转账、提现单独一列
    最后进行统计
    """
    try:
        # 读取Excel文件
        df = pandas.read_excel(file)

        # 将数值列转换为数字类型（处理可能的文本格式数字）
        df['收入金额（元）'] = df['收入金额（元）'].astype(str).str.replace(',', '').str.strip()
        df['收入金额（元）'] = pandas.to_numeric(df['收入金额（元）'].replace('', '0'), errors='coerce').fillna(0)

        df['支出金额'] = df['支出金额'].astype(str).str.replace(',', '').str.strip()
        df['支出金额'] = pandas.to_numeric(df['支出金额'].replace('', '0'), errors='coerce').fillna(0)

        # 提取日期（从入账时间）
        df['日期'] = pandas.to_datetime(df['入账时间']).dt.date
        # 分离不同类型的交易
        df_transfer = df[df['入账类型'].apply(lambda x: '转账' in x)]
        df_withdraw = df[df['入账类型'].apply(lambda x: '提现' in x)]
        df_other = df[~df['入账类型'].apply(lambda x: '转账' in x or '提现' in x)]

        logger.info(f'总数据量: {len(df)}, 提现数据: {len(df_withdraw)}, 转账数据: {len(df_transfer)}, '
                    f'日数据: {len(df_other)}')

        # 汇总转账数据（按日）
        df_transfer_summary = df_transfer.groupby('日期').agg({
            '收入金额（元）': 'sum',
            '支出金额': 'sum'
        }).reset_index()
        df_transfer_summary.columns = ['日期', '转账收入', '转账支出']
        df_transfer_summary['转账净额'] = df_transfer_summary['转账收入'] - df_transfer_summary['转账支出']

        # 汇总提现数据（按日）
        df_withdraw_summary = df_withdraw.groupby('日期').agg({
            '支出金额': 'sum'
        }).reset_index()
        df_withdraw_summary.columns = ['日期', '提现支出']
        df_withdraw_summary['提现金额'] = df_withdraw_summary['提现支出']

        # 汇总其他交易数据（按日）
        df_other_summary = df_other.groupby('日期').agg({
            '收入金额（元）': 'sum',
            '支出金额': 'sum'
        }).reset_index()
        df_other_summary.columns = ['日期', '日收入', '日支出']
        df_other_summary['日净收入'] = df_other_summary['日收入'] - df_other_summary['日支出']

        # 合并三个汇总表
        df_summary = pandas.merge(df_other_summary, df_transfer_summary, on='日期', how='outer')
        df_summary = pandas.merge(df_summary, df_withdraw_summary, on='日期', how='outer')

        # 填充空值为0
        df_summary = df_summary.fillna(0)

        # 计算每日总净收入
        df_summary['日净收入'] = df_summary['日净收入']

        # 按日期排序
        df_summary = df_summary.sort_values('日期', ascending=True).reset_index(drop=True)

        # 添加汇总行
        total_row = pandas.DataFrame({
            '日期': ['所有汇总'],
            '日收入': [df_summary['日收入'].sum()],
            '日支出': [df_summary['日支出'].sum()],
            '提现金额': [df_summary['提现金额'].sum()],
            '转账收入': [df_summary['转账收入'].sum()],
            '转账支出': [df_summary['转账支出'].sum()],
            '转账净额': [df_summary['转账净额'].sum()],
            '日净收入': [df_summary['日净收入'].sum()]
        })
        df_summary = pandas.concat([df_summary, total_row], ignore_index=True)

        # 重新排列列顺序
        columns_order = ['日期', '日收入', '日支出', '提现金额', '转账收入', '转账支出', '转账净额', '日净收入']
        df_summary = df_summary[columns_order]

        # 使用 ExcelWriter 追加到现有文件，保留原有的 Sheet
        with pandas.ExcelWriter(file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            sheet_name = '汇总数据'
            # 先删除可能已存在的同名 Sheet
            if sheet_name in writer.book.sheetnames:
                del writer.book[sheet_name]

            # 写入数据
            df_summary.to_excel(writer, sheet_name=sheet_name, index=False)

            # 获取工作表对象并设置样式
            ws = writer.sheets[sheet_name]
            header_font = Font(bold=True, color='FFFFFF', size=11)
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center')

            # 遍历第一行所有单元格设置样式
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # 自动调整列宽
            for column in ws.columns:
                max_length = 0
                col_letter = column[0].column_letter
                for cell in column:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))

                ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

            # 消除 "Workbook contains no default style" 警告
            workbook = writer.book
            if not workbook.style_names:
                default_font = Font(name='Calibri', size=11, bold=False, italic=False)
                default_style = openpyxl.styles.NamedStyle(name='Normal', font=default_font)
                workbook.add_named_style(default_style)

        logger.info(f'数据汇总完成，共汇总{len(df_summary) - 1}天的数据，已保存到: {file}')

        logger.info(f'总收入: {df_summary["日收入"].sum() / 2:.2f}，总支出: {df_summary["日支出"].sum() / 2:.2f} '
                    f'总提现: {df_summary["提现金额"].sum() / 2:.2f}，，总转账: {df_summary["转账净额"].sum() / 2:.2f}，'
                    f'总净收入: {df_summary["日净收入"].sum() / 2:.2f}\n')

        return df_summary

    except Exception as e:
        logger.error(f'{shopname}数据处理失败: {e}\n')
        return None


def main_(account_id=1):
    title = f'========================开始爬取千牛第{account_id}个店铺======================='
    logger.info(title)
    end = time.strftime("%Y-%m-%d", time.localtime(time.time() - 86400))
    start = f"{end[:-3]}-01"
    dirname = 'd:/_code/spider_account/数据'

    login = tb_login(account_id)
    if not login:
        return False
    shop_name = tb_search(start, end)
    time.sleep(3)
    if Playwright_.get_count('//div[text()="没有数据"]'):
        logger.info(f'{shop_name}店铺报表暂无数据')
        return False

    filename = f'千牛-{shop_name}店铺{end}明细.xlsx'
    filename = os.path.join(dirname, filename)

    time.sleep(10)

    logger.info(f'{shop_name}店铺开始导出明细....')
    status = 0
    for roll in range(1, 6):
        logger.info(f'第{roll}次导出明细....')
        Playwright_.click('//span[text()="导出"]')
        status = tb_save(filename)
        if status:
            break
    text = f'✅️ {shop_name}明细数据下载成功：{filename}' if status else f'❌️ {shop_name}明细数据下载失败'
    logger.info(text)
    if status:
        tb_deal_data(shop_name, filename)
    Playwright_.clear_cookie()
    return True


if __name__ == '__main__':
    # tb_deal_data('测试', './数据/测试2.xlsx')
    shop_count_ = get_config_value('login', 'tb_shop_count', file=config_file)
    shop_flag = input('请输入查询店铺序号（0默认查询全部）：')
    if shop_flag == '0':
        start_id = 1
        end_id = int(shop_count_) + 1
    elif ':' in shop_flag or '：' in shop_flag:
        shop_flag = shop_flag.replace('：', '').replace(':', '')
        start_id = int(shop_flag)
        end_id = int(shop_count_) + 1
    else:
        start_id = int(shop_flag)
        end_id = start_id + 1
    for account_id in range(start_id, end_id):
        main_(account_id)
