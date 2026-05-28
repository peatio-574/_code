# -*- coding: utf-8 -*-
import sys
from pathlib import Path

sys.path.append(str(Path(__file__).parent.parent))


from PlayWright import Playwright_, logger, get_config_value
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import openpyxl
import warnings
import os
import pandas
import requests
import time


warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')


config_file = os.path.join(os.path.dirname(__file__), 'config.ini')


def xhs_login(account_id=1):
    try:
        logger.info('开始登录小红书....')
        url = 'https://ark.xiaohongshu.com/app-system/home'
        ele = '(//a[text()="学习中心"])[1]'

        key = f'login.xhs_cookie_{account_id}'
        sub_account = f'(//span[text()="子账号"])[{account_id}]/../div/div[1]'
        Playwright_.login(url, ele, key, extra=sub_account, file=config_file)
        logger.info('小红书登录成功....')
        return True
    except Exception as e:
        logger.error(f'小红书登录失败：{e[:50]}')
        return False


def xhs_search(start, end):
    Playwright_.goto('https://ark.xiaohongshu.com/app-merchant/third-settle/account')
    shop_name = Playwright_.get_text('//div[@class="avatar-wrapper"]/div')[:-2]
    logger.info(f'当前店铺名称：{shop_name}')
    time.sleep(8)
    know_ele = '(//span[text()="知道了"])[last()]'
    if Playwright_.get_count(know_ele):
        Playwright_.click(know_ele)
        time.sleep(1)
    if Playwright_.get_count(know_ele):
        Playwright_.click(know_ele)
        time.sleep(1)
    if Playwright_.get_count(know_ele):
        Playwright_.click(know_ele)
        time.sleep(1)
    if Playwright_.get_count(know_ele):
        Playwright_.click(know_ele)
        time.sleep(1)
    if Playwright_.get_count(know_ele):
        Playwright_.click(know_ele)
        time.sleep(1)
    if Playwright_.get_count(know_ele):
        Playwright_.click(know_ele)
        time.sleep(1)
    if Playwright_.get_count(know_ele):
        Playwright_.click(know_ele)
        time.sleep(1)
    if Playwright_.get_count(know_ele):
        Playwright_.click(know_ele)
        time.sleep(1)

    Playwright_.input('//input[@placeholder="开始时间"]', start)
    Playwright_.input('//input[@placeholder="结束时间"]', end)
    Playwright_.click('//span[text()="查询"]')
    return shop_name


def xhs_deal_data(shopname, file):
    """
    使用pandas处理数据：
    1. 按每日汇总交易类型描述不为"提现"的收入和支出
    2. 按每日汇总交易类型描述为"提现"的提现金额
    """
    try:
        # 读取Excel文件
        df = pandas.read_excel(file)

        # 将数值列转换为数字类型（处理可能的文本格式数字）
        df['收入（元）'] = pandas.to_numeric(df['收入（元）'], errors='coerce').fillna(0)
        df['支出（元）'] = pandas.to_numeric(df['支出（元）'].replace('-', '0'), errors='coerce').fillna(0)

        # 提取日期
        df['日期'] = pandas.to_datetime(df['创建时间']).dt.date

        # 分离提现和非提现数据
        df_withdraw = df[df['交易类型描述'].str.contains('提现', na=False)]
        df_normal = df[~df['交易类型描述'].str.contains('提现', na=False)]

        logger.info(f'总数据量: {len(df)}, 提现数据: {len(df_withdraw)}, 非提现数据: {len(df_normal)}')

        # 汇总非提现的收入和支出
        df_normal_summary = df_normal.groupby('日期').agg({
            '收入（元）': 'sum',
            '支出（元）': 'sum'
        }).reset_index()
        df_normal_summary.columns = ['日期', '日收入', '日支出']

        # 汇总提现数据（提现金额在支出列）
        df_withdraw_summary = df_withdraw.groupby('日期').agg({
            '支出（元）': 'sum'
        }).reset_index()
        df_withdraw_summary.columns = ['日期', '日提现金额']

        # 合并两个汇总表
        df_summary = pandas.merge(df_normal_summary, df_withdraw_summary, on='日期', how='outer')

        # 填充空值为0
        df_summary = df_summary.fillna(0)

        # 计算日净收入（收入 - 支出 - 提现）
        df_summary['日净收入'] = df_summary['日收入'] - df_summary['日支出']

        # 按日期排序
        df_summary = df_summary.sort_values('日期', ascending=False).reset_index(drop=True)

        # 汇总数据
        total_row = pandas.DataFrame({
            '日期': ['所有汇总'],
            '日收入': [df_summary['日收入'].sum()],
            '日支出': [df_summary['日支出'].sum()],
            '日提现金额': [df_summary['日提现金额'].sum()],
            '日净收入': [df_summary['日净收入'].sum()]
        })
        df_summary = pandas.concat([df_summary, total_row], ignore_index=True)

        # 使用 ExcelWriter 追加到现有文件，保留原有的 Sheet1
        with pandas.ExcelWriter(file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            sheet_name = '汇总数据'
            # 先删除可能已存在的同名 Sheet
            if sheet_name in writer.book.sheetnames:
                del writer.book[sheet_name]

            # 写入数据
            df_summary.to_excel(writer, sheet_name=sheet_name, index=False)

            # 获取工作表对象并设置样式
            ws = writer.sheets[sheet_name]
            header_font = Font(bold=True, color='FFFFFF', size=11)
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center')

            # 遍历第一行所有单元格设置样式
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # 自动调整列宽
            for column in ws.columns:
                max_length = 0
                col_letter = column[0].column_letter
                for cell in column:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))

                ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

            # 消除 "Workbook contains no default style" 警告
            workbook = writer.book
            if not workbook.style_names:
                default_font = Font(name='Calibri', size=11, bold=False, italic=False)
                default_style = openpyxl.styles.NamedStyle(name='Normal', font=default_font)
                workbook.add_named_style(default_style)

        logger.info(f'数据汇总完成，共汇总{len(df_summary)}天的数据，已保存到: {file}')

        # 打印汇总统计
        logger.info(f'总收入: {df_summary["日收入"].sum()/2:.2f}，总支出: {df_summary["日支出"].sum()/2:.2f} '
                    f'总提现: {df_summary["日提现金额"].sum()/2:.2f}，总净收入: {df_summary["日净收入"].sum()/2:.2f}\n')
        return df_summary

    except Exception as e:
        logger.error(f'{shopname}数据处理失败: {e}\n')
        return None



def xhs_save(filename):
    """直接导出"""
    try:
        # 直接导出
        with Playwright_.page.expect_download(timeout=15000) as download_info:
            pass  # 等待下载触发
        download = download_info.value
        # 获取文件名并保存
        download.save_as(filename)
        return True
    except Exception as e:
        logger.info(f'{filename}临时下载异常：{e}')
        return False


def xhs_download(filename):
    """接口导出"""
    Playwright_.click('//span[text()="消息"]')
    Playwright_.click('(//span[text()="店铺"])[last()]')
    judge_ele = '(//div[@class="list-item-wrapper"])[1]//div[@class="date-wrap unread"]/span'
    if Playwright_.get_count(judge_ele):
        judge = Playwright_.get_text(judge_ele)
        if judge in ('1分钟前', '刚刚'):
            ele = '(//div[@class="list-item-wrapper"])[1]//a'
            if Playwright_.get_count(ele):
                href = Playwright_.get_attribute(ele, 'href')
                with open(filename, mode='wb') as f:
                    f.write(requests.get(href).content)
                    return True
        Playwright_.click('//div[@class="ark-message-title-wrap"]/span[2]')
    return False


def main_(account_id=1):
    title = f'========================开始爬取第{account_id}个店铺======================='
    logger.info(title)
    end = time.strftime("%Y-%m-%d", time.localtime(time.time() - 86400))
    start = f"{end[:-3]}-01"
    dirname = 'd:/_code/spider_account/数据'

    login = xhs_login(account_id)
    if not login:
        return False
    shop_name = xhs_search(start, end)

    filename = f'小红书-{shop_name}店铺{end}明细.xlsx'
    filename = os.path.join(dirname, filename)

    time.sleep(10)

    logger.info(f'{shop_name}店铺开始导出明细....')
    status = 0
    for roll in range(1, 6):
        logger.info(f'第{roll}次导出明细....')
        Playwright_.click('//span[text()="导出"]')
        status = xhs_save(filename)
        if status:
            break

        status = xhs_download(filename)
        if status:
            break
    text = f'✅️ {shop_name}明细数据下载成功：{filename}' if status else f'❌️ {shop_name}明细数据下载失败'
    logger.info(text)
    if status:
        xhs_deal_data(shop_name, filename)
    Playwright_.clear_cookie()
    return True if status else False


if __name__ == '__main__':
    # xhs_deal_data('甜心花栗店铺', file='./数据/meow meow店铺2026-05-27明细.xlsx')
    shop_count_ = get_config_value('login', 'xhs_shop_count', file=config_file)
    shop_flag = input('请输入查询店铺序号（0默认查询全部）：')
    if shop_flag == '0':
        start_id = 1
        end_id = int(shop_count_) + 1
    else:
        start_id = int(shop_flag)
        end_id = start_id + 1
    for account_id in range(start_id, end_id):
        main_(account_id)

