# coding='utf-8'
import sys
import os

if getattr(sys, 'frozen', False):
    bundleDir = sys._MEIPASS
    baseDir = os.path.dirname(sys.executable)
else:
    bundleDir = os.path.dirname(os.path.abspath(__file__))
    baseDir = bundleDir

sys.path.insert(0, bundleDir)

import time
from PlayWright import Playwright_, logger
import requests
from openpyxl import load_workbook, Workbook
from ReadFile import ReadData

configFile = os.path.join(baseDir, 'config.ini')
pictureDir = os.path.join(baseDir, 'pictures')
os.makedirs(pictureDir, exist_ok=True)


def vipLogin():
    logger.info("开始登录....")
    url = 'https://passport.vip.com/login'
    ele = '//a[text()="我的订单"]'
    key = 'login.vipCookie'
    Playwright_.login(url, ele, key, file=configFile)
    logger.info('✅️ 登录成功')


def download(url, file):
    file = file.replace('/', '-').replace(' ', '-').replace('.', '-').replace(r'\\', '-') + '.jpg'
    url = 'https:' + url if 'https' not in url else url
    for roll in range(1, 4):
        try:
            logger.info(f'开始第{roll}次尝试下载{file}....')
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
            }
            fileName = os.path.join(pictureDir, file)
            with open(fileName, 'wb') as f:
                f.write(requests.get(url, headers=headers).content)
                logger.info(f'{file}保存成功\n')
                break
        except Exception as e:
            logger.error(f'{file}下载失败：{e}\n')
        finally:
            time.sleep(2)


def getPageInfo(styleName, styleCode, pageId):
    logger.info(f'开始爬取[款式 {styleName}] 第{pageId}页数据....')
    url = f'https://category.vip.com/suggest.php?keyword=韩系女装&props={styleCode}&page={pageId}&orderId=undefined&price_start=200&price_end='
    Playwright_.goto(url)
    time.sleep(3)

    global existsData

    for roll in range(10):
        logger.info(f'第{roll + 1}次向下滑动1000xp')
        Playwright_.mouse_wheel(1000)
        time.sleep(2)

    rowEle = '//div[@class="c-goods-item  J-goods-item c-goods-item--auto-width"]'
    rowCount = Playwright_.get_count(rowEle)

    for rowId in range(1, rowCount + 1):
        detailEle = f'({rowEle})[{rowId}]/a'
        detailUrl = Playwright_.get_attribute(detailEle, 'href')
        detailUrl = 'https:' + detailUrl if 'https' not in detailUrl else detailUrl
        if detailUrl in existsData:
            continue
        if len(existsData) >= limitCount:
            logger.info(f'已爬取足量数据：{limitCount}条')
            exit()

        titleEle = f'({rowEle})[{rowId}]//div[contains(@class, "c-goods-item__name")]'
        title = Playwright_.get_text(titleEle)

        priceEle = f'({rowEle})[{rowId}]//div[contains(@class, "c-goods-item__sale-price J-goods-item__sale-price")]'
        price = Playwright_.get_text(priceEle)
        title += f'【{price[1:]}】'

        # imgEle = f'({rowEle})[{rowId}]//div[contains(@class, "c-goods-item__img")]/img'
        # mainPhoto = Playwright_.get_attribute(imgEle, 'src')
        # mainPhoto = 'https:' + mainPhoto if 'https' not in mainPhoto else mainPhoto
        ws.append([title, detailUrl])
        existsData.append(detailUrl)
        logger.info(f'{title}：{detailUrl}')
    wb.save(fileName)
    logger.info(f'[款式 {styleName}] 第{pageId}页数据保存成功，当前共{len(existsData)}条数据')

    totalEle = '//span[@class="total"]'
    total = Playwright_.get_text(totalEle)
    total = total.split('/')[-1]
    return int(total)


def getStyleInfo():
    """获取款式codes"""
    styleInfo = {'A字裙': '21528', '百褶裙': '21526', '灯笼裤': '33763', '短裤': '19866', '工装裤': '32082', '哈伦裤': '32078',
                 '开衫': '18518', '阔腿裤': '32084', '喇叭裤': '32076', '连帽': '32091', '牛仔裤': '19872', '铅笔裤': '32079',
                 '日常便服': '21940', '沙滩裤': '19868', '束腿裤': '32087', '套头': '18520', '通勤套装': '32094', '小脚裤': '18599',
                 '休闲套装': '33762', '运动裤': '32081', '针织裤': '19870', '直筒裤': '18595', '锥形裤': '32086', '连衣裙': '42707',
                 '紧身健美裤': '42712', '休闲裤': '42737', '梭织裤': '50555', '萝卜裤': '51336', '裤裙': '51337', '直筒裙': '51349',
                 '修身裤': '67897', '宽松垂感裤': '71050', '轻户外短裤': '71055', '轻户外长裤': '71056', '拖地裤': '71057',
                 '伞兵裤': '71058', '微喇叭': '71063'}
    # logger.info('开始获取款式数据....')
    # url = 'https://category.vip.com/suggest.php?keyword=韩系女装'
    # Playwright_.goto(url)
    # time.sleep(3)
    #
    # styleInfo = {}
    # styleEle = '//h3[text()="款式"]/../div[1]/div/ul/li'
    # styleCount = Playwright_.get_count(styleEle)
    # for styleId in range(1, styleCount + 1):
    #     styleName = Playwright_.get_text(f'({styleEle})[{styleId}]/a')
    #     styleCode = Playwright_.get_attribute(f'({styleEle})[{styleId}]', 'data-id')
    #     styleInfo[styleName] = styleCode
    # logger.info(f'款式信息：{styleInfo}')
    return styleInfo


def getData():
    vipLogin()

    styleInfo = getStyleInfo()  # 款式

    for styleName, styleCode in styleInfo.items():  # 遍历款式
        total = getPageInfo(styleName, styleCode, pageId=1)
        if total == 1:  # 页码=1
            continue
        for pageId in range(2, total + 1):  # 遍历页码
            getPageInfo(styleName, styleCode, pageId)


def getPhoto(title, url):
    Playwright_.goto(url)
    time.sleep(2)
    imgEle = '//div[@class="pic-slider-items J-picSlider-items"]/img'
    imgCount = Playwright_.get_count(imgEle)
    imgCount = max(2, imgCount)
    for imgId in range(1, imgCount + 1):
        img = Playwright_.get_attribute(f'({imgEle})[{imgId}]', 'src')
        fileName = os.path.join(pictureDir, f'{title}_{imgId}')
        download(img, fileName)


def getAllPhoto():
    vipLogin()
    titleData = ReadData.read_xlsx_col(fileName)['标题']
    logger.info(f'共计{len(titleData)}条连接')
    for rowId, detailUrl in enumerate(existsData):
        logger.info(f'开始处理第{rowId + 1}条链接：{detailUrl}')
        title = titleData[rowId]
        getPhoto(title, detailUrl)


fileName = os.path.join(baseDir, '基础数据.xlsx')
if os.path.exists(fileName):
    # 如果文件存在，加载现有工作簿
    wb = load_workbook(fileName)
    ws = wb.active
else:
    headers = ['标题', '详情链接', '主图', '产品图']
    wb = Workbook()
    ws = wb.active
    ws.title = '数据'
    ws.append(headers)
    wb.save(fileName)
existsData = ReadData.read_xlsx_col(fileName)['详情链接']
limitCount = 5000  # 限制总数

if __name__ == '__main__':
    while True:
        step = input('请选择操作步骤（1.爬取链接，2.下载图片）：')
        if step == '1':
            getData()
        elif step == '2':
            getAllPhoto()
