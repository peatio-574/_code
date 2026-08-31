# coding='utf-8'
import sys
from pathlib import Path

# 把项目根目录加入Python路径
sys.path.append(str(Path(__file__).parent.parent))

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import threading
import queue
import os
from openpyxl import load_workbook

# 导入 run.py 中的函数（请确保 run.py 在当前目录或 Python 路径中）
from run import (
    checkAccountStatus,
    addEdgeProfile,
    edge_is_running,
    close_edge,
    is_blank_value,
    list_profiles,
    delete_profile,
)

PAGE_SIZE = 100


class AccountGUI:
    def __init__(self, root):
        self.page_label = None
        self.btn_next = None
        self.status_var = None
        self.step_var = None
        self.tree = None
        self.delete_tree = None
        self.btn_start = None
        self.btn_pause = None
        self.btn_prev = None
        self.btn_save = None
        self.btn_choose = None
        self.btn_refresh_delete = None
        self.progress = None
        self.progress_text_var = None
        self.file_label = None
        self.root = root
        self.root.title("Edge 账号批量处理工具")
        self.root.geometry("1050x760")

        # 数据
        self.df = None              # 完整 DataFrame
        self.file_path = None       # 当前文件路径
        self.current_page = 0       # 当前页码（从0开始）
        self.total_pages = 0
        self.processing = False     # 是否正在处理
        self.paused = False         # 是否暂停
        self.current_step = None    # 当前处理步骤
        self.queue = queue.Queue()  # 线程通信队列
        self.thread = None          # 处理线程
        self.pause_event = threading.Event()
        self.pause_event.set()
        self.delete_profiles = []   # 删除列表缓存

        # 创建界面
        self.create_widgets()
        self.refresh_delete_list(show_message=False)

        # 定时检查队列
        self.root.after(100, self.process_queue)

    def create_widgets(self):
        # 顶部控制区
        control_frame = ttk.LabelFrame(self.root, text="控制面板", padding=10)
        control_frame.pack(fill=tk.X, padx=10, pady=5)

        # 文件选择行
        file_frame = ttk.Frame(control_frame)
        file_frame.pack(fill=tk.X, pady=2)

        self.file_label = ttk.Label(file_frame, text="未选择文件", width=60)
        self.file_label.pack(side=tk.LEFT, padx=5)

        self.btn_choose = ttk.Button(file_frame, text="选择文件", command=self.choose_file)
        self.btn_choose.pack(side=tk.LEFT, padx=5)

        self.btn_refresh_delete = ttk.Button(file_frame, text="刷新删除列表", command=self.refresh_delete_list)
        self.btn_refresh_delete.pack(side=tk.LEFT, padx=5)

        # 操作选项行
        option_frame = ttk.Frame(control_frame)
        option_frame.pack(fill=tk.X, pady=5)

        ttk.Label(option_frame, text="操作步骤：").pack(side=tk.LEFT)
        self.step_var = tk.StringVar(value="1")
        ttk.Radiobutton(option_frame, text="1. 仅查询状态", variable=self.step_var, value="1").pack(side=tk.LEFT, padx=5)
        ttk.Radiobutton(option_frame, text="2. 查询并添加配置", variable=self.step_var, value="2").pack(side=tk.LEFT, padx=5)
        ttk.Radiobutton(option_frame, text="3. 删除浏览器已有用户配置", variable=self.step_var, value="3").pack(side=tk.LEFT, padx=5)
        self.step_var.trace_add("write", self.on_step_changed)

        # 按钮行
        btn_frame = ttk.Frame(control_frame)
        btn_frame.pack(fill=tk.X, pady=5)

        self.btn_start = ttk.Button(btn_frame, text="开始处理", command=self.start_processing, state=tk.NORMAL)
        self.btn_start.pack(side=tk.LEFT, padx=5)

        self.btn_pause = ttk.Button(btn_frame, text="暂停", command=self.toggle_pause, state=tk.DISABLED)
        self.btn_pause.pack(side=tk.LEFT, padx=5)

        self.btn_save = ttk.Button(btn_frame, text="手动保存", command=self.save_file, state=tk.DISABLED)
        self.btn_save.pack(side=tk.LEFT, padx=5)

        self.progress = ttk.Progressbar(btn_frame, mode='determinate')
        self.progress.pack(side=tk.LEFT, padx=10, fill=tk.X, expand=True)

        self.progress_text_var = tk.StringVar(value="0/0  0%")
        ttk.Label(btn_frame, textvariable=self.progress_text_var, width=18, anchor=tk.E).pack(side=tk.LEFT, padx=5)

        # 列表区域
        notebook = ttk.Notebook(self.root)
        notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        account_frame = ttk.LabelFrame(notebook, text="账号列表", padding=10)
        notebook.add(account_frame, text="账号列表")

        # 创建账号 Treeview
        self.tree = ttk.Treeview(account_frame, columns=("row_id", "account", "status"), show="headings", height=20)
        self.tree.heading("row_id", text="序号")
        self.tree.heading("account", text="账号")
        self.tree.heading("status", text="状态")
        self.tree.column("row_id", width=60, anchor=tk.CENTER)
        self.tree.column("account", width=360, anchor=tk.W)
        self.tree.column("status", width=360, anchor=tk.W)
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # 账号列表垂直滚动条
        account_vsb = ttk.Scrollbar(account_frame, orient="vertical", command=self.tree.yview)
        account_vsb.pack(side=tk.RIGHT, fill=tk.Y)
        self.tree.configure(yscrollcommand=account_vsb.set)

        delete_frame = ttk.LabelFrame(notebook, text="删除列表（默认保留第一个配置）", padding=10)
        notebook.add(delete_frame, text="删除列表")

        # 创建删除 Treeview
        self.delete_tree = ttk.Treeview(
            delete_frame,
            columns=("profile", "name", "user_name", "gaia_name", "path", "result"),
            show="headings",
            height=20,
        )
        self.delete_tree.heading("profile", text="配置目录")
        self.delete_tree.heading("name", text="显示名")
        self.delete_tree.heading("user_name", text="账号")
        self.delete_tree.heading("gaia_name", text="昵称")
        self.delete_tree.heading("path", text="路径")
        self.delete_tree.heading("result", text="处理结果")
        self.delete_tree.column("profile", width=100, anchor=tk.W)
        self.delete_tree.column("name", width=180, anchor=tk.W)
        self.delete_tree.column("user_name", width=220, anchor=tk.W)
        self.delete_tree.column("gaia_name", width=160, anchor=tk.W)
        self.delete_tree.column("path", width=260, anchor=tk.W)
        self.delete_tree.column("result", width=120, anchor=tk.W)
        self.delete_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        delete_vsb = ttk.Scrollbar(delete_frame, orient="vertical", command=self.delete_tree.yview)
        delete_vsb.pack(side=tk.RIGHT, fill=tk.Y)
        self.delete_tree.configure(yscrollcommand=delete_vsb.set)

        # 分页控制
        page_frame = ttk.Frame(self.root)
        page_frame.pack(fill=tk.X, padx=10, pady=5)

        self.btn_prev = ttk.Button(page_frame, text="上一页", command=self.prev_page, state=tk.DISABLED)
        self.btn_prev.pack(side=tk.LEFT, padx=5)

        self.page_label = ttk.Label(page_frame, text="第 0/0 页")
        self.page_label.pack(side=tk.LEFT, padx=10)

        self.btn_next = ttk.Button(page_frame, text="下一页", command=self.next_page, state=tk.DISABLED)
        self.btn_next.pack(side=tk.LEFT, padx=5)

        # 状态栏
        self.status_var = tk.StringVar(value="就绪")
        status_bar = ttk.Label(self.root, textvariable=self.status_var, relief=tk.SUNKEN, anchor=tk.W)
        status_bar.pack(side=tk.BOTTOM, fill=tk.X)

    def on_step_changed(self, *_):
        """切换步骤时刷新相关 UI。"""
        if self.processing:
            return
        if self.step_var.get() == "3":
            self.refresh_delete_list(show_message=False)
            self.progress['maximum'] = max(len(self.delete_profiles), 1)
            self.update_progress(0, len(self.delete_profiles), "删除列表已刷新")
        elif self.df is not None:
            self.progress['maximum'] = len(self.df)
            self.update_progress(0, len(self.df), "已切换处理步骤")
        else:
            self.update_progress(0, 0, "就绪")

    def refresh_delete_list(self, show_message=True):
        """刷新浏览器用户配置删除列表。"""
        try:
            profiles = list_profiles()
            self.delete_profiles = profiles[1:] if len(profiles) > 1 else []

            if self.delete_tree is not None:
                for item in self.delete_tree.get_children():
                    self.delete_tree.delete(item)

                for index, item in enumerate(self.delete_profiles, start=1):
                    profile_name = item.get("profile", "")
                    values = (
                        profile_name,
                        item.get("name", ""),
                        item.get("user_name", ""),
                        item.get("gaia_name", ""),
                        item.get("path", ""),
                        "待删除",
                    )
                    self.delete_tree.insert("", tk.END, iid=profile_name or f"profile_{index}", values=values)

            if show_message:
                self.status_var.set(f"删除列表已刷新，共 {len(self.delete_profiles)} 个待删除配置")
        except Exception as e:
            if show_message:
                messagebox.showerror("错误", f"刷新删除列表失败：{e}")

    def update_progress(self, current, total, text=None):
        """实时更新进度条和进度文字。"""
        total = int(total or 0)
        current = int(current or 0)
        maximum = max(total, 1)
        current = min(current, maximum)
        percent = int(current * 100 / maximum) if total else 0
        self.progress['maximum'] = maximum
        self.progress['value'] = current
        self.progress_text_var.set(f"{current}/{total}  {percent}%")
        if text is not None:
            self.status_var.set(text)

    def wait_if_paused(self):
        """后台线程暂停点。暂停只能在单条任务之间生效。"""
        while self.processing and self.paused:
            self.pause_event.wait(0.2)

    def choose_file(self):
        """选择 Excel 文件并加载数据"""
        file_path = filedialog.askopenfilename(
            title="选择 Excel 文件",
            filetypes=[("Excel 文件", "*.xlsx"), ("所有文件", "*.*")]
        )
        if not file_path:
            return

        try:
            # 使用 pandas 读取，假设第一列是账号，第二列是状态（如果没有则新建）
            df = pd.read_excel(file_path, sheet_name=0, engine='openpyxl', keep_default_na=False)
            if df.empty:
                messagebox.showerror("错误", "Excel 文件为空")
                return

            # 确保第一列是账号，重命名为'账号'
            df.columns = [str(col).strip() for col in df.columns]
            if '账号' not in df.columns:
                # 如果第一列不是账号，假设第一列就是账号
                df.rename(columns={df.columns[0]: '账号'}, inplace=True)

            # 确保有'状态'列，没有则创建
            if '状态' not in df.columns:
                df['状态'] = ''

            df['账号'] = df['账号'].apply(lambda value: '' if is_blank_value(value) else str(value).strip())
            df['状态'] = df['状态'].apply(lambda value: '' if is_blank_value(value) else str(value).strip())

            self.df = df
            self.file_path = file_path
            self.current_page = 0
            self.total_pages = (len(self.df) + PAGE_SIZE - 1) // PAGE_SIZE

            self.file_label.config(text=os.path.basename(file_path))
            self.btn_start.config(state=tk.NORMAL)
            self.btn_save.config(state=tk.NORMAL)
            self.btn_prev.config(state=tk.NORMAL if self.total_pages > 1 else tk.DISABLED)
            self.btn_next.config(state=tk.NORMAL if self.total_pages > 1 else tk.DISABLED)

            self.update_table()
            self.update_page_label()
            self.status_var.set(f"已加载 {len(self.df)} 条记录，共 {self.total_pages} 页")
            self.update_progress(0, len(self.df), f"已加载 {len(self.df)} 条记录，共 {self.total_pages} 页")

        except Exception as e:
            messagebox.showerror("错误", f"读取文件失败：{e}")

    def update_table(self):
        """刷新当前页的表格内容"""
        if self.df is None:
            return

        # 清空旧数据
        for item in self.tree.get_children():
            self.tree.delete(item)

        # 计算当前页数据范围
        start = self.current_page * PAGE_SIZE
        end = min(start + PAGE_SIZE, len(self.df))
        page_data = self.df.iloc[start:end]

        for idx, row in page_data.iterrows():
            # 全局行号从1开始
            global_row = idx + 1
            account_value = row.get('账号', '')
            status_value = row.get('状态', '')
            account = '' if is_blank_value(account_value) else str(account_value).strip()
            status = '' if is_blank_value(status_value) else str(status_value).strip()
            self.tree.insert("", tk.END, values=(global_row, account, status))

        self.update_page_label()

    def update_page_label(self):
        if self.df is not None:
            self.page_label.config(text=f"第 {self.current_page + 1}/{self.total_pages} 页")
        else:
            self.page_label.config(text="第 0/0 页")

    def prev_page(self):
        if self.current_page > 0:
            self.current_page -= 1
            self.update_table()

    def next_page(self):
        if self.current_page < self.total_pages - 1:
            self.current_page += 1
            self.update_table()

    def start_processing(self):
        """启动后台处理线程"""
        if self.step_var.get() != "3" and (self.df is None or self.file_path is None):
            messagebox.showwarning("警告", "请先选择 Excel 文件")
            return

        if self.processing:
            messagebox.showinfo("提示", "正在处理中，请先暂停或等待完成。")
            return

        # 确认操作（步骤2/3可能会关闭 Edge 或修改本地配置）
        if self.step_var.get() == "2":
            if not messagebox.askyesno("确认", "步骤2将关闭 Edge 并修改本地配置，是否继续？"):
                return
        elif self.step_var.get() == "3":
            self.refresh_delete_list(show_message=False)
            if not self.delete_profiles:
                messagebox.showinfo("提示", "没有需要删除的浏览器用户配置。")
                return
            if not messagebox.askyesno("确认", f"步骤3将关闭 Edge 并删除 {len(self.delete_profiles)} 个浏览器用户配置，是否继续？"):
                return

        self.processing = True
        self.paused = False
        self.current_step = self.step_var.get()
        self.pause_event.set()
        self.btn_start.config(state=tk.DISABLED)
        self.btn_pause.config(text="暂停", state=tk.NORMAL)
        self.btn_choose.config(state=tk.DISABLED)
        self.btn_refresh_delete.config(state=tk.DISABLED)
        self.btn_save.config(state=tk.DISABLED)
        self.status_var.set("处理中...")

        if self.current_step == "3":
            total = len(self.delete_profiles)
        else:
            total = len(self.df)
        self.update_progress(0, total, "处理中...")

        # 如果步骤2/3，先关闭 Edge
        if self.current_step in {"2", "3"}:
            if edge_is_running():
                self.status_var.set("关闭 Edge 中...")
                self.root.update()
                close_edge()

        # 启动线程
        self.thread = threading.Thread(target=self.process_worker, args=(self.current_step,), daemon=True)
        self.thread.start()

    def toggle_pause(self):
        """暂停/继续当前处理任务。"""
        if not self.processing:
            return

        if self.paused:
            self.paused = False
            self.pause_event.set()
            self.btn_pause.config(text="暂停")
            self.status_var.set("继续处理中...")
        else:
            self.paused = True
            self.pause_event.clear()
            self.btn_pause.config(text="继续")
            self.status_var.set("已暂停，点击继续恢复处理")

    def process_worker(self, step):
        """后台处理线程：遍历账号，更新状态，定期保存"""
        try:
            if step == '3':
                profiles = list(self.delete_profiles)
                total = len(profiles)
                if not profiles:
                    self.queue.put(('log', "没有需要删除的浏览器用户配置"))
                for i, item in enumerate(profiles):
                    if not self.processing:
                        break
                    self.wait_if_paused()
                    if not self.processing:
                        break

                    profile_name = item.get("profile", "")
                    display_name = item.get("name", "") or profile_name
                    self.queue.put(('log', f"正在删除 [{display_name}] ..."))
                    self.queue.put(('delete_update', profile_name, "删除中"))
                    result = delete_profile(profile_name, close_browser=False)
                    status = result.get("message", "已删除") if isinstance(result, dict) else "删除失败"
                    self.queue.put(('delete_update', profile_name, status))
                    self.queue.put(('progress', i + 1, total, f"已删除 {i + 1}/{total}：{display_name}，{status}"))
                self.queue.put(('done',))
                return

            # 加载工作簿用于保存
            wb = load_workbook(self.file_path)
            ws = wb.active
            # 确保第二列表头为'状态'
            ws.cell(row=1, column=2, value='状态')
            wb.save(self.file_path)

            accounts = self.df['账号'].tolist()
            statuses = self.df['状态'].tolist()
            total = len(accounts)

            for i, account in enumerate(accounts):
                if not self.processing:
                    break
                self.wait_if_paused()
                if not self.processing:
                    break

                account = '' if is_blank_value(account) else str(account).strip()
                status_value = statuses[i] if i < len(statuses) else ''

                if is_blank_value(account):
                    self.queue.put(('log', f"第 {i + 2} 行账号为空，跳过"))
                    self.queue.put(('progress', i + 1, total, f"已处理 {i + 1}/{total}"))
                    continue

                # 如果状态已有值，跳过（空、NaN、'nan' 不算已有状态）
                if not is_blank_value(status_value):
                    self.queue.put(('log', f"账号 [{account}] 已有状态，跳过：{status_value}"))
                    self.queue.put(('progress', i + 1, total, f"已处理 {i + 1}/{total}"))
                    continue

                # 检查状态
                self.queue.put(('log', f"正在处理 [{account}] ..."))
                status = checkAccountStatus(account)

                # 如果是步骤2且状态正常，添加配置
                if step == '2' and status == '正常':
                    self.wait_if_paused()
                    if not self.processing:
                        break
                    result = addEdgeProfile(account, force_new=False)
                    if result:
                        status += '，创建用户配置成功'
                    else:
                        status += '，创建用户配置失败'

                # 更新状态
                self.df.at[i, '状态'] = status
                statuses[i] = status

                # 发送更新到主线程
                self.queue.put(('update', i, status))
                self.queue.put(('progress', i + 1, total, f"已处理 {i + 1}/{total}：{account}"))

                # 每10条保存一次
                if (i + 1) % 10 == 0 or (i + 1) == total:
                    self.queue.put(('save',))

            # 处理完成
            self.queue.put(('done',))

        except Exception as e:
            self.queue.put(('error', str(e)))

    def process_queue(self):
        """处理队列中的消息，更新 UI"""
        try:
            while True:
                msg = self.queue.get_nowait()
                if msg[0] == 'update':
                    _, row_idx, status = msg
                    # 如果当前页包含该行，刷新表格
                    start = self.current_page * PAGE_SIZE
                    end = min(start + PAGE_SIZE, len(self.df))
                    if start <= row_idx < end:
                        self.update_table()
                elif msg[0] == 'progress':
                    _, current, total, text = msg
                    self.update_progress(current, total, text)
                elif msg[0] == 'delete_update':
                    _, profile_name, result_text = msg
                    if self.delete_tree is not None and self.delete_tree.exists(profile_name):
                        values = list(self.delete_tree.item(profile_name, 'values'))
                        if values:
                            values[-1] = result_text
                            self.delete_tree.item(profile_name, values=values)
                elif msg[0] == 'save':
                    self.save_file(show_message=False)
                elif msg[0] == 'log':
                    self.status_var.set(msg[1])
                elif msg[0] == 'done':
                    self.finish_processing(success=True)
                elif msg[0] == 'error':
                    self.finish_processing(success=False)
                    self.status_var.set("处理出错")
                    messagebox.showerror("错误", msg[1])
        except queue.Empty:
            pass
        finally:
            if self.processing:
                self.root.after(100, self.process_queue)
            else:
                # 如果不在处理中，也保持循环检查
                self.root.after(200, self.process_queue)

    def finish_processing(self, success=True):
        """恢复按钮状态并做收尾保存/提示。"""
        finished_step = self.current_step
        self.processing = False
        self.paused = False
        self.current_step = None
        self.pause_event.set()
        self.btn_start.config(state=tk.NORMAL)
        self.btn_pause.config(text="暂停", state=tk.DISABLED)
        self.btn_choose.config(state=tk.NORMAL)
        self.btn_refresh_delete.config(state=tk.NORMAL)
        self.btn_save.config(state=tk.NORMAL if self.df is not None else tk.DISABLED)

        if not success:
            return

        self.status_var.set("处理完成")
        if finished_step in {"1", "2"} and self.df is not None and self.file_path is not None:
            self.save_file(show_message=True)
            messagebox.showinfo("完成", "所有账号处理完毕并已保存。")
        elif finished_step == "3":
            self.refresh_delete_list(show_message=False)
            messagebox.showinfo("完成", "浏览器用户配置删除处理完成。")

    def save_file(self, show_message=True):
        """保存 DataFrame 到原文件（更新状态列）"""
        if self.file_path is None or self.df is None:
            return

        try:
            # 使用 openpyxl 直接修改，保留原格式
            wb = load_workbook(self.file_path)
            ws = wb.active
            # 确保第二列表头
            ws.cell(row=1, column=2, value='状态')

            # 更新状态列（从第2行开始）
            for i, status in enumerate(self.df['状态']):
                ws.cell(row=i + 2, column=2, value='' if is_blank_value(status) else status)

            wb.save(self.file_path)
            if show_message:
                self.status_var.set("文件已保存")
        except Exception as e:
            if show_message:
                messagebox.showerror("错误", f"保存失败：{e}")


if __name__ == "__main__":
    root = tk.Tk()
    app = AccountGUI(root)
    root.mainloop()
