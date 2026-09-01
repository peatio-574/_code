# coding='utf-8'
"""检测账号状态，正常则新增 Microsoft Edge 本地用户配置。

说明：
1. 从指定 txt 文件读取账号，每行一个。
2. 本脚本只创建 Edge 本地配置，不会打开浏览器页面。
3. 本脚本不会自动登录 Microsoft 账号，不会处理密码、Cookie、Token。
4. 新增的配置会显示在 edge://settings/profiles 中，显示名默认为账号。
5. 写入 Edge 配置前会自动关闭 Edge。
"""

from __future__ import annotations

import json
import os
import shutil
import subprocess
import sys
import time
from pathlib import Path
from ReadFile import ReadData

import requests
import urllib3

sys.path.append(str(Path(__file__).parent.parent))
from Logger import logger
from openpyxl import load_workbook
import random

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)  # 消除 HTTPS 证书告警


# Edge 默认用户数据目录。新增配置会写入该目录下的 Local State。
USER_DATA_DIR = Path(os.environ.get("LOCALAPPDATA", r"C:\Users\Administrator\AppData\Local")) / "Microsoft" / "Edge" / "User Data"
LOCAL_STATE = USER_DATA_DIR / "Local State"

# ==================== 运行配置 ====================




def checkAccountStatus(account: str = "yumi-ufu@f2.dion.ne.jp") -> str:
    """检查账号状态。

    接口返回 info == 'Both' 时按当前业务规则视为正常。
    其他返回值会原样返回，调用方据此决定是否跳过。
    """
    try:
        time.sleep(random.uniform(0.5, 1.5))
        url = "https://odc.officeapps.live.com/odc/v2.1/idp"
        user_agent = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/143.0.0.0 Safari/537.36 Edg/143.0.0.0"
        headers = {
            "User-Agent": user_agent,
            "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6",
        }
        params = {
            "hm": "0",
            "emailAddress": account,
            "_": int(time.time() * 1000),
        }
        response = requests.get(url, headers=headers, params=params, verify=False, timeout=20).json()
        info = response.get("account")
        if info == "Both" or info == 'MSAccount':
            return "正常"
        return info or "未知"
    except Exception as e:
        logger.error(f"{account}账号接口请求异常：{e}")
        return "请求异常"


def edge_is_running() -> bool:
    """检测 Edge 是否正在运行。"""
    result = subprocess.run(
        ["tasklist", "/FI", "IMAGENAME eq msedge.exe"],
        capture_output=True,
        text=True,
        encoding="mbcs",
        errors="ignore",
    )
    return "msedge.exe" in result.stdout.lower()


def close_edge() -> None:
    """关闭所有 Edge 进程。"""
    subprocess.run(
        ["taskkill", "/F", "/IM", "msedge.exe"],
        capture_output=True,
        encoding="mbcs",
        errors="ignore",
    )
    time.sleep(2)


def load_local_state() -> dict:
    """读取 Edge Local State；不存在时返回最小结构。"""
    if not LOCAL_STATE.exists():
        USER_DATA_DIR.mkdir(parents=True, exist_ok=True)
        return {"profile": {"info_cache": {}}}
    return json.loads(LOCAL_STATE.read_text(encoding="utf-8"))


def save_local_state(data: dict) -> Path:
    """保存 Local State，并在写入前生成备份文件。"""
    USER_DATA_DIR.mkdir(parents=True, exist_ok=True)
    backup = USER_DATA_DIR / ("Local State.bak." + time.strftime("%Y%m%d%H%M%S"))
    if LOCAL_STATE.exists():
        shutil.copy2(LOCAL_STATE, backup)
    LOCAL_STATE.write_text(json.dumps(data, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    return backup


def profile_info(data: dict) -> dict:
    """获取 Local State 中的 profile.info_cache 字典。"""
    return data.setdefault("profile", {}).setdefault("info_cache", {})


def find_existing_profile(data: dict, account: str) -> str | None:
    """按账号或显示名查找已存在配置，避免重复创建。"""
    account_lower = account.lower()
    for profile_dir, meta in profile_info(data).items():
        if (meta.get("user_name") or "").lower() == account_lower:
            return profile_dir
        if (meta.get("name") or "").lower() == account_lower:
            return profile_dir
    return None


def match_profile(profile_dir: str, meta: dict, keyword: str) -> bool:
    """判断指定配置是否匹配关键字。

    支持匹配：
    1. Profile 目录名，例如 Profile 18。
    2. 配置显示名 name。
    3. 已登录账号 user_name。
    4. Microsoft 昵称 gaia_name。
    """
    keyword = str(keyword or "").strip().lower()
    if not keyword:
        return False
    values = [
        profile_dir,
        meta.get("name", ""),
        meta.get("user_name", ""),
        meta.get("gaia_name", ""),
    ]
    return any(str(value or "").strip().lower() == keyword for value in values)


def next_profile_dir(data: dict) -> str:
    """获取下一个可用 Profile 目录名，例如 Profile 32。"""
    used = set(profile_info(data).keys())
    used.update(path.name for path in USER_DATA_DIR.glob("Profile *") if path.is_dir())
    index = 1
    while f"Profile {index}" in used:
        index += 1
    return f"Profile {index}"


def addEdgeProfile(account: str, force_new: bool = False) -> dict:
    """新增 Edge 本地用户配置。

    注意：这里不会把账号写成已登录状态，只是创建一个显示名为账号的本地配置。
    后续如果需要登录，需要用户手动在 Edge 中完成。
    """
    try:
        data = load_local_state()

        existing = find_existing_profile(data, account)
        if existing and not force_new:
            logger.info(f'{account}用户配置已存在')
            return {"account": account, "profile": existing, "created": False, "message": "配置已存在"}

        profile_dir = next_profile_dir(data)

        info = profile_info(data)
        info[profile_dir] = {
            "name": account,
            "user_name": "",
            "gaia_name": "",
            "is_using_default_name": False,
            "is_using_default_avatar": True,
            "avatar_icon": "chrome://theme/IDR_PROFILE_AVATAR_26",
        }

        profile_root = data.setdefault("profile", {})
        for key in ("profiles_order", "last_active_profiles"):
            value = profile_root.setdefault(key, [])
            if type(value) is list and profile_dir not in value:
                value.append(profile_dir)
        profile_root["last_used"] = profile_dir

        profile_path = USER_DATA_DIR / profile_dir
        profile_path.mkdir(parents=True, exist_ok=True)
        preferences = {
            "profile": {
                "name": account,
                "is_using_default_name": False,
            }
        }
        (profile_path / "Preferences").write_text(
            json.dumps(preferences, ensure_ascii=False, separators=(",", ":")),
            encoding="utf-8",
        )

        backup = save_local_state(data)
        return {
            "account": account,
            "profile": profile_dir,
            "created": True,
            "backup": str(backup),
        }
    except Exception as e:
        logger.error(f'{account}创建用户配置失败：{e}')
        return False


def list_profiles() -> list[dict]:
    """获取当前所有 Edge 用户配置列表。"""
    data = load_local_state()
    result = []
    for profile_dir, meta in profile_info(data).items():
        result.append({
            "profile": profile_dir,
            "name": meta.get("name", ""),
            "user_name": meta.get("user_name", ""),
            "gaia_name": meta.get("gaia_name", ""),
            "path": str(USER_DATA_DIR / profile_dir),
            "exists": (USER_DATA_DIR / profile_dir).exists(),
        })
    return result


def delete_profile(profile_name: str, close_browser: bool = True) -> dict:
    """删除指定 Edge 用户配置。

    按 profile 目录名、显示名、已登录账号或 Microsoft 昵称匹配。
    """
    if close_browser and edge_is_running():
        logger.info("Edge 正在运行，删除前先关闭 Edge...")
        close_edge()

    data = load_local_state()
    info = profile_info(data)

    targets = []
    for profile_dir, meta in list(info.items()):
        if match_profile(profile_dir, meta, profile_name):
            targets.append(profile_dir)

    if not targets:
        logger.warning(f"未找到配置：{profile_name}")
        return {"profile": profile_name, "deleted": False, "message": "未找到该配置"}

    for target in targets:
        info.pop(target, None)

    profile_root = data.get("profile", {})
    for key in ("profiles_order", "last_active_profiles"):
        value = profile_root.get(key, [])
        if type(value) is list:
            profile_root[key] = [item for item in value if item not in targets]

    if profile_root.get("last_used") in targets:
        remaining = list(info.keys())
        profile_root["last_used"] = remaining[0] if remaining else ""

    # 先保存 Local State，确保 Edge 配置列表中不再引用待删除 profile。
    backup = save_local_state(data)

    deleted_paths = []
    failed_paths = []
    for target in targets:
        profile_path = USER_DATA_DIR / target
        if not profile_path.exists():
            continue
        try:
            shutil.rmtree(profile_path)
            deleted_paths.append(str(profile_path))
        except Exception as e:
            failed_paths.append({"path": str(profile_path), "error": str(e)})
            logger.error(f"删除目录失败：{profile_path}，原因：{e}")

    return {
        "profile": profile_name,
        "deleted": not failed_paths,
        "targets": targets,
        "deleted_paths": deleted_paths,
        "failed_paths": failed_paths,
        "backup": str(backup),
        "message": "已删除" if not failed_paths else "部分删除失败",
    }


def delete_all_profiles(close_browser: bool = True) -> list[dict]:
    """删除当前所有 Edge 用户配置。"""
    profiles = list_profiles()
    results = []
    for profile in profiles:
        results.append(delete_profile(profile["profile"], close_browser=close_browser and not results))
    return results


def is_blank_value(value) -> bool:
    """判断 Excel 单元格值是否应视为空。

    pandas/openpyxl 在读取空单元格时可能返回 None、NaN，或者字符串
    'nan'。这些值都不能被当作“已有状态”，否则会导致空状态被跳过。
    """
    if value is None:
        return True
    try:
        if value != value:  # NaN 不等于自身
            return True
    except Exception:
        pass
    return str(value).strip().lower() in {"", "nan", "none", "null", "nat"}


def main(accounts_file: str, force_new=False, step='1') -> None:
    """按 Excel 第一列账号、第二列状态批量处理账号。

    逻辑与 execute.py 保持一致：
    1. 默认使用活动工作表，不强依赖 Sheet1。
    2. 第一列作为账号列，第二列作为状态列。
    3. 状态为空、None、NaN、'nan' 时继续处理；其他状态跳过。
    4. 每处理 10 条自动保存一次。
    """
    wb = load_workbook(accounts_file)
    ws = wb.active

    if ws.max_row < 2:
        raise SystemExit(f"账号文件为空或无有效账号：{accounts_file}")

    # 对齐 execute.py 的读取规则：第一列是账号，第二列是状态。
    ws.cell(row=1, column=1, value='账号')
    ws.cell(row=1, column=2, value='状态')
    wb.save(accounts_file)

    datas = ReadData.read_xlsx_col(accounts_file)
    if not datas or '账号' not in datas:
        raise SystemExit(f"账号文件为空或无有效账号：{accounts_file}")

    accounts = datas.get('账号', [])
    allStatus = datas.get('状态', [''] * len(accounts))

    if not accounts:
        raise SystemExit(f"账号文件为空或无有效账号：{accounts_file}")

    if step == '2' and edge_is_running():
        logger.info("Edge 正在运行，正在关闭...")
        close_edge()

    for index, account in enumerate(accounts):
        rowId = index + 2
        account = str(account or '').strip()
        status_value = allStatus[index] if index < len(allStatus) else ''

        if is_blank_value(account):
            logger.info(f'第 {rowId} 行账号为空，跳过')
            continue

        if not is_blank_value(status_value):
            logger.info(f'账号【{account}】已备注：{status_value}')
            continue

        logger.info(f'开始执行账号【{account}】')

        status = checkAccountStatus(account)
        if status == '正常' and step == '2':
            status += '，创建用户配置成功' if addEdgeProfile(account, force_new=force_new) else '，创建用户配置失败'

        ws.cell(row=rowId, column=2, value=status)
        logger.info(f'{account}状态：{status}')
        if rowId % 10 == 0 or rowId == len(accounts) + 1:
            wb.save(accounts_file)

    wb.save(accounts_file)


if __name__ == "__main__":
    # 删除示例：删除指定配置。支持 Profile 目录名、显示名、已登录账号。
    # print(delete_profile("Profile 7"))
    # print(delete_profile("acheng6@126.com"))

    step = input('请输入操作步骤：1.查询账号状态，2.添加用户配置，3.删除浏览器已有用户配置：')

    if step == '3':
        for item in list_profiles()[1:]:
            print(item)
            delete_profile(item["profile"])
    else:
        force_new = False  # 账号已存在时是否仍创建新配置。
        file = r"C:\Users\Administrator\Desktop\data.xlsx"
        main(file, force_new=force_new, step=step)
