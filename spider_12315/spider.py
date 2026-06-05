# -*- coding: utf-8 -*-
import sys
from pathlib import Path

sys.path.append(str(Path(__file__).parent.parent))

from PlayWright import Playwright_
import os
from openpyxl import load_workbook, Workbook
import time

from wordcloud import WordCloud
import matplotlib.pyplot as plt
import jieba
from collections import Counter
import re

filename = os.path.join(os.path.dirname(__file__), '评价数据.xlsx')


def get_info():
    url = 'https://tsgs.12315.cn/#/viewport'
    Playwright_.goto(url)
    time.sleep(20)
    row_ele = '//div[@class="response-list-item"]'

    if os.path.exists(filename):
        # 如果文件存在，加载现有工作簿
        wb = load_workbook(filename)
        ws = wb.active
    else:
        headers = ['评论内容']
        wb = Workbook()
        ws = wb.active
        ws.title = '评论'
        ws.append(headers)
    wb.save(filename)

    contents = []
    for page_id in range(2, 10):
        row_count = Playwright_.get_count(row_ele)  # 当页数据量
        for row in range(1, row_count + 1):
            content_ele = f'({row_ele})[{row}]/div[2]/div[1]/div'
            content = Playwright_.get_text(content_ele)
            if content in contents:
                continue

            ws.append([content])
        wb.save(filename)
        Playwright_.page.keyboard.press('PageDown')
        Playwright_.page.keyboard.press('PageDown')

        # 点击下一页
        Playwright_.click(f'//li[@class="number"and text()="{page_id}"]')


def analyze():
    """读取文件对第一列评论数据进行词云分析"""

    if not os.path.exists(filename):
        print(f'文件不存在: {filename}')
        return

    # 加载工作簿
    wb = load_workbook(filename)
    ws = wb.active

    # 读取第一列所有评论数据（从第2行开始，跳过表头）
    comments = []
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True):
        if row[0] is not None and str(row[0]).strip():  # 跳过空值
            comments.append(str(row[0]))

    print(f'共读取 {len(comments)} 条评论数据')

    if not comments:
        print('没有评论数据可分析')
        return

    # 合并所有评论
    all_text = ' '.join(comments)

    # 使用jieba分词
    words = jieba.cut(all_text)

    # 过滤停用词和无效字符
    stop_words = {
        '的', '了', '在', '是', '我', '有', '和', '就', '不', '人', '都', '一',
        '一个', '上', '也', '很', '到', '说', '要', '去', '你', '会', '着', '没有',
        '看', '好', '自己', '这', '他', '她', '它', '们', '那', '些', '什么', '怎么',
        '吗', '呢', '吧', '啊', '哦', '嗯', '这个', '那个', '可以', '可以', '已经',
        '但是', '因为', '所以', '如果', '虽然', '然后', '而且', '或者', '不是',
        '\n', '\r', '\t', ' ', ''
    }

    # 统计词频
    word_counts = Counter()
    for word in words:
        word = word.strip()
        # 过滤条件：不在停用词表中、长度大于1、只包含中文或英文
        if (word not in stop_words and
                len(word) > 1 and
                re.match(r'^[\u4e00-\u9fa5a-zA-Z]+$', word)):
            word_counts[word] += 1

    # 获取出现频率最高的前50个词
    top_words = word_counts.most_common(50)

    print('\n=== 词频分析结果（前50个高频词）===')
    print(f'{"排名":<6}{"词语":<15}{"出现次数":<10}')
    print('-' * 35)
    for rank, (word, count) in enumerate(top_words, 1):
        print(f'{rank:<6}{word:<15}{count:<10}')

    # 保存词频结果到Excel
    result_filename = os.path.join(os.path.dirname(__file__), '词频分析结果.xlsx')
    result_wb = Workbook()
    result_ws = result_wb.active
    result_ws.title = '词频统计'
    result_ws.append(['排名', '词语', '出现次数'])

    for rank, (word, count) in enumerate(top_words, 1):
        result_ws.append([rank, word, count])

    result_wb.save(result_filename)
    print(f'\n词频分析结果已保存到: {result_filename}')

    # 尝试生成词云图（需要安装wordcloud库）
    try:
        # 构建词云数据
        wordcloud_data = dict(top_words[:30])  # 取前30个词

        # 创建词云
        wc = WordCloud(
            font_path='C:/Windows/Fonts/simhei.ttf',  # 设置中文字体
            width=800,
            height=600,
            background_color='white',
            max_words=30,
            colormap='viridis'
        ).generate_from_frequencies(wordcloud_data)

        # 显示词云
        plt.figure(figsize=(10, 8))
        plt.imshow(wc, interpolation='bilinear')
        plt.axis('off')
        plt.title('评论数据词云图', fontsize=16)

        # 保存词云图
        cloud_filename = os.path.join(os.path.dirname(__file__), '词云图.png')
        plt.savefig(cloud_filename, dpi=300, bbox_inches='tight')
        plt.show()

        print(f'词云图已保存到: {cloud_filename}')

    except ImportError:
        print('\n提示: 如需生成词云图，请安装以下库:')
        print('pip install wordcloud matplotlib')
search = []

if __name__ == '__main__':
    # get_info()
    # deal()
    analyze()

