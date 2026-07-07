# -*- coding: utf-8 -*-
"""
处理 商品数据导出.xlsx —— 每个商品拆为两行
  第1行: 商品编号 | 商品名称 | 价格 | 商品规格 | 主图1...主图N
  第2行: 商品编号 | 商品名称 | 价格 | 商品规格 | 详情图1...详情图N
"""
import os
import copy
from io import BytesIO

import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

# ============================================================================
# 配置
# ============================================================================
SRC_FILE = os.path.join(os.path.dirname(__file__), "商品数据导出.xlsx")
DST_FILE = os.path.join(os.path.dirname(__file__), "商品数据导出_两行版.xlsx")

IMG_WIDTH = 120
IMG_HEIGHT = 120
ROW_HEIGHT = 100

FIXED_COUNT = 4  # 商品编号、商品名称、价格、商品规格

# ============================================================================
# 主流程
# ============================================================================
def main():
    # ---- 读取源文件 ----
    print(f"读取: {SRC_FILE}")
    src_wb = openpyxl.load_workbook(SRC_FILE)
    src_ws = src_wb.active

    total_rows = src_ws.max_row
    total_cols = src_ws.max_column
    headers = [src_ws.cell(row=1, column=c).value for c in range(1, total_cols + 1)]
    print(f"源: {total_rows}行 x {total_cols}列, 图片{len(src_ws._images)}张")

    # ---- 分析列结构 ----
    main_img_count = 0
    detail_img_count = 0
    main_start = None
    detail_start = None
    for idx, h in enumerate(headers):
        if h and str(h).startswith("主图"):
            main_img_count += 1
            if main_start is None:
                main_start = idx + 1  # Excel列号从1开始
        elif h and str(h).startswith("详情图"):
            detail_img_count += 1
            if detail_start is None:
                detail_start = idx + 1

    print(f"固定列={FIXED_COUNT}, 主图={main_img_count}列(起始{main_start}), 详情图={detail_img_count}列(起始{detail_start})")

    # ---- 创建目标工作簿 ----
    dst_wb = openpyxl.Workbook()
    dst_ws = dst_wb.active
    dst_ws.title = src_ws.title

    # 目标列数: 固定列 + max(主图, 详情图)
    img_col_count = max(main_img_count, detail_img_count)
    dst_cols = FIXED_COUNT + img_col_count

    # ---- 写表头 ----
    dst_headers = headers[:FIXED_COUNT]
    for i in range(1, img_col_count + 1):
        dst_headers.append(f"图片{i}")

    header_font = openpyxl.styles.Font(bold=True, size=11)
    header_fill = openpyxl.styles.PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_align = openpyxl.styles.Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, h in enumerate(dst_headers, 1):
        cell = dst_ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # 列宽
    dst_ws.column_dimensions[get_column_letter(1)].width = 35
    dst_ws.column_dimensions[get_column_letter(2)].width = 45
    dst_ws.column_dimensions[get_column_letter(3)].width = 10
    dst_ws.column_dimensions[get_column_letter(4)].width = 50
    for col in range(FIXED_COUNT + 1, dst_cols + 1):
        dst_ws.column_dimensions[get_column_letter(col)].width = 18

    # ---- 逐行处理 ----
    data_align = openpyxl.styles.Alignment(horizontal="left", vertical="top", wrap_text=True)

    for src_row in range(2, total_rows + 1):  # 跳过表头
        # 目标行号: 商品1→行2/3, 商品2→行4/5, ...
        dst_row_a = (src_row - 2) * 2 + 2  # 主图行
        dst_row_b = dst_row_a + 1           # 详情图行

        print(f"\n商品{src_row-1}: 源行{src_row} → 目标行{dst_row_a}/{dst_row_b}")

        # ---- 复制商品信息文本 ----
        for col_idx in range(1, FIXED_COUNT + 1):
            src_cell = src_ws.cell(row=src_row, column=col_idx)
            # 写入主图行
            c = dst_ws.cell(row=dst_row_a, column=col_idx, value=src_cell.value)
            c.alignment = data_align
            if col_idx == 3:  # 价格列
                c.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="top")
                c.number_format = "0.00"
            # 写入详情图行
            c = dst_ws.cell(row=dst_row_b, column=col_idx, value=src_cell.value)
            c.alignment = data_align
            if col_idx == 3:
                c.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="top")
                c.number_format = "0.00"

        # ---- 复制主图 → 目标行A ----
        main_copied = 0
        for img in src_ws._images:
            # openpyxl图片锚点格式: OneCellAnchor 或 TwoCellAnchor
            anchor = img.anchor
            try:
                img_row = anchor._from.row + 1  # openpyxl行号从0开始
                img_col = anchor._from.col + 1  # openpyxl列号从0开始
            except AttributeError:
                continue

            # 判断是否属于当前行且是主图列
            if img_row != src_row:
                continue
            if img_col < main_start or img_col >= main_start + main_img_count:
                continue

            img_idx = img_col - main_start  # 0-based
            dst_col = FIXED_COUNT + 1 + img_idx

            # 读取图片数据并创建新图片对象
            img_data = BytesIO(img._data())
            new_img = XLImage(img_data)
            new_img.width = IMG_WIDTH
            new_img.height = IMG_HEIGHT
            dst_ws.add_image(new_img, f"{get_column_letter(dst_col)}{dst_row_a}")
            main_copied += 1

        # ---- 复制详情图 → 目标行B ----
        detail_copied = 0
        for img in src_ws._images:
            anchor = img.anchor
            try:
                img_row = anchor._from.row + 1
                img_col = anchor._from.col + 1
            except AttributeError:
                continue

            if img_row != src_row:
                continue
            if img_col < detail_start or img_col >= detail_start + detail_img_count:
                continue

            img_idx = img_col - detail_start
            dst_col = FIXED_COUNT + 1 + img_idx

            img_data = BytesIO(img._data())
            new_img = XLImage(img_data)
            new_img.width = IMG_WIDTH
            new_img.height = IMG_HEIGHT
            dst_ws.add_image(new_img, f"{get_column_letter(dst_col)}{dst_row_b}")
            detail_copied += 1

        # 行高
        dst_ws.row_dimensions[dst_row_a].height = ROW_HEIGHT
        dst_ws.row_dimensions[dst_row_b].height = ROW_HEIGHT

        print(f"  主图{main_copied}张, 详情图{detail_copied}张")

    # ---- 保存 ----
    dst_ws.freeze_panes = "A2"
    dst_wb.save(DST_FILE)
    print(f"\n[OK] 已保存: {DST_FILE}")
    print(f"  总行数: {dst_ws.max_row} (含表头)")
    print(f"  总列数: {dst_cols}")
    print(f"  图片数: {len(dst_ws._images)}")
    print(f"  商品数: {(dst_ws.max_row - 1) // 2}")


if __name__ == "__main__":
    main()
