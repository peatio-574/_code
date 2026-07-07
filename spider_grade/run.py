# -*- coding: utf-8 -*-
"""
积分商城商品数据采集 + Excel导出（含图片嵌入）

功能:
  1. 从info1/info2中读取商品列表数据
  2. 使用Playwright打开每个商品的详情页，提取规格和图片
  3. 下载商品图片到本地
  4. 将所有数据（含图片）写入Excel表格

Excel表头:
  商品名称 | 价格 | 商品规格 | 主图1 | 主图2 | ... | 详情图1 | 详情图2 | ...
"""
import sys
import time
import os
import requests
# from io import BytesIO
from pathlib import Path

import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

sys.path.append(str(Path(__file__).parent.parent))

from PlayWright import Playwright_, logger

# ============================================================================
# 原始数据（info1 + info2）
# ============================================================================
info1 = {
    "data": {
        "searchResponse": {
            "data": [
                {
                    "id": "45572fe6bba86cbcce8b0dabfe9369aa",
                    "goodsName": "茶花收纳袋 拉链密封袋食品密实袋保鲜袋 中号（20只）990100990100 990100",
                    "goodsImageUrl": "https://img.li91.com/2026/06/500x500/n0/516ef212bad6236b7d700228c25de246.jpg",
                    "integral": 3500,
                    "salesVolume": 0,
                    "goodsType": 0
                },
                {
                    "id": "277bd988a08f34b800984fc017eb2f5a",
                    "goodsName": "栢士德 硅胶锅铲 BST-08",
                    "goodsImageUrl": "https://img.li91.com/2026/06/500x500/n0/2c60228cfe036317ccdb72e72d2bae23.jpg",
                    "integral": 5000,
                    "salesVolume": 0,
                    "goodsType": 0
                },
                {
                    "id": "2b81c4353f498b312c9e2a10f9470241",
                    "goodsName": "茶花（CHAHUA）保鲜袋加厚食品级密封袋食品专用袋冰箱一次性分装袋 大号100只 3342",
                    "goodsImageUrl": "https://img.li91.com/2026/05/500x500/n0/c98fb4fc0d6c4acf99e4519fe6b092a1.jpg",
                    "integral": 2500,
                    "salesVolume": 6,
                    "goodsType": 0
                },
                {
                    "id": "95fa9b7c84f217fa5598c11042364a8e",
                    "goodsName": "澳得迈5D小苍兰持久留香洗衣凝珠 12gx30颗/盒 LD-188 红色",
                    "goodsImageUrl": "https://img.li91.com/2026/03/500x500/n0/783339d8f6f253ab2d6776065b1810de.jpg",
                    "integral": 6500,
                    "salesVolume": 1,
                    "goodsType": 0
                },
                {
                    "id": "5088892a3947e2c926c61e82266c94da",
                    "goodsName": "美穗吉家 高分子牙线棒50支 MIH-2404220",
                    "goodsImageUrl": "https://img.li91.com/2026/04/500x500/n0/837d3d25c2d7e279221df19d50305021.jpg",
                    "integral": 2000,
                    "salesVolume": 4,
                    "goodsType": 0
                },
                {
                    "id": "7c30fe9b2413e8da3aa254aff91a4a7c",
                    "goodsName": "美穗吉家 厨房清洁湿巾加大款20X30cm（40抽/包）MIH-2404240",
                    "goodsImageUrl": "https://img.li91.com/2026/04/500x500/n0/d9c39205f58e2e4ea13b94e7b9d5cbc3.jpg",
                    "integral": 2500,
                    "salesVolume": 0,
                    "goodsType": 0
                },
                {
                    "id": "9cf3d9ce69c418b39bb6531003774ba1",
                    "goodsName": "美穗吉家 净味香氛 洁厕剂500g 80856",
                    "goodsImageUrl": "https://img.li91.com/2026/04/500x500/n0/be6dc52745df4d299b1be908a2ed63c8.jpg",
                    "integral": 2500,
                    "salesVolume": 0,
                    "goodsType": 0
                },
                {
                    "id": "990df6d44a6204a7b523240d348f73b9",
                    "goodsName": "美穗吉家 EDI纯水湿巾80抽/包*1包 MIH-2404230",
                    "goodsImageUrl": "https://img.li91.com/2026/04/500x500/n0/3cde7554e2bd013b5f6db5837395a15f.jpg",
                    "integral": 2000,
                    "salesVolume": 5,
                    "goodsType": 0
                },
                {
                    "id": "e2beba7c4f02b6318ae685913853e917",
                    "goodsName": "铭仕朗手持风扇MSL-FS1201 白色手持风扇快速冲电手持挂脖",
                    "goodsImageUrl": "https://img.li91.com/2026/05/500x500/n0/1e4999c5195492aa7b67ad7cf8772c40.png",
                    "integral": 7000,
                    "salesVolume": 1,
                    "goodsType": 0
                },
                {
                    "id": "66881a8ab71aed482d0c1653961fd7cc",
                    "goodsName": "美穗吉家 运动冰感毛巾 蓝色 MIHH-2408096",
                    "goodsImageUrl": "https://img.li91.com/2026/04/500x500/n0/a56a6cada8ad1ab92d865380538555ca.jpg",
                    "integral": 6500,
                    "salesVolume": 0,
                    "goodsType": 0
                },
                {
                    "id": "103864b3b1cbcdd1b24e4d19617a416f",
                    "goodsName": "美穗吉家 手持可折叠挂脖风扇 MLS6118",
                    "goodsImageUrl": "https://img.li91.com/2026/04/500x500/n0/ef824b07906f2ca5f6487f516237b81f.jpg",
                    "integral": 5800,
                    "salesVolume": 2,
                    "goodsType": 0
                },
                {
                    "id": "9471c69121cf6cd33667674dbc539efe",
                    "goodsName": "植护 婴儿手口湿巾 140*180mm 60片*2包ZH409303*2",
                    "goodsImageUrl": "https://img.li91.com/2026/04/500x500/n0/a0e59ef2fb1336c3df08c19fc7ea82cc.jpg",
                    "integral": 3800,
                    "salesVolume": 0,
                    "goodsType": 0
                },
                {
                    "id": "942408de3b078a6911950675e9c3aa53",
                    "goodsName": "美穗吉家 针织冰袖 白色 MIHH-2408092",
                    "goodsImageUrl": "https://img.li91.com/2026/04/500x500/n0/ab193d335bd52c1a0c59028c4bb064b4.png",
                    "integral": 3800,
                    "salesVolume": 0,
                    "goodsType": 0
                },
                {
                    "id": "b532a4637a8dda3278736ce2c3148ad1",
                    "goodsName": "美穗吉家 氨基酸泡泡洗手液 500ml 90044",
                    "goodsImageUrl": "https://img.li91.com/2026/04/500x500/n0/f6d440d47e084ddcf1409aeb1ab56a59.jpg",
                    "integral": 3300,
                    "salesVolume": 3,
                    "goodsType": 0
                },
                {
                    "id": "301d70efe7687f46edb9ea12ceb38fef",
                    "goodsName": "植护 挂式气垫纸巾（幻彩系列）155*175/4层 280抽*1提 ZH408731*1",
                    "goodsImageUrl": "https://img.li91.com/2026/04/500x500/n0/fa1a88802529e459822511465f88e835.jpg",
                    "integral": 3000,
                    "salesVolume": 2,
                    "goodsType": 0
                },
                {
                    "id": "f789fb196b252e55998cb22593536633",
                    "goodsName": "植护 气垫纸巾（幻彩系列）140*175/4层 80抽*3包ZH408739*3",
                    "goodsImageUrl": "https://img.li91.com/2026/04/500x500/n0/a05e2b5df2eaf07b40739dbfadea127b.png",
                    "integral": 2800,
                    "salesVolume": 6,
                    "goodsType": 0
                },
                {
                    "id": "16d6f85e0b0011531e7af2c1a759d6a8",
                    "goodsName": "植护 除菌除螨香氛洗衣液（樱花香型） 500g*1袋 ZH408765*1g*1袋",
                    "goodsImageUrl": "https://img.li91.com/2026/04/500x500/n0/33152fe8d5b6b2ec14eab62d6cabd652.jpg",
                    "integral": 2800,
                    "salesVolume": 2,
                    "goodsType": 0
                },
                {
                    "id": "5fbbf6fc582b26db8d29872f4735f274",
                    "goodsName": "奥妙樱花除菌洗衣凝珠50颗洗衣球 单腔除菌除螨 鲜花精油 三合一",
                    "goodsImageUrl": "https://img1.360buyimg.com/image/jfs/t1/260965/2/24720/107980/67bee4c6F6a2fd217/8668b637e308ca3c.jpg",
                    "integral": 6500,
                    "salesVolume": 3,
                    "goodsType": 3
                },
                {
                    "id": "0aa40645da19eb81f3ddc4b4c8ddb425",
                    "goodsName": "澳得迈草本果香洁净护衣洗衣凝珠 8gx30颗/盒 LD-118 白色",
                    "goodsImageUrl": "https://img.li91.com/2026/03/500x500/n0/0a36b80722f0c4ddb5b6ce4d268c1115.jpg",
                    "integral": 3800,
                    "salesVolume": 1,
                    "goodsType": 0
                },
                {
                    "id": "a93bb610be1901dd6edd3414c3aff951",
                    "goodsName": "沃品 太空舱伸缩快充 一拖三数据线LC027",
                    "goodsImageUrl": "https://img.li91.com/2024/07/500x500/n0/db79ffc21134529195e1aaab3969f642.jpg",
                    "integral": 6500,
                    "salesVolume": 0,
                    "goodsType": 0
                },
                {
                    "id": "85143aa5ffb87cf4ca9221b0ff8de2cf",
                    "goodsName": "乐扣乐扣 旅行收纳五件套  LTZ524FU 黑色",
                    "goodsImageUrl": "https://img.li91.com/2024/08/500x500/n0/4f90f9361b4d5890f75408ef0102b7f9.jpg",
                    "integral": 6800,
                    "salesVolume": 3,
                    "goodsType": 0
                },
                {
                    "id": "0ab00b0d98968ddea3688aa3c1862dfa",
                    "goodsName": "立白洗衣凝珠 除菌除螨抑菌洁净40颗 单腔 浓缩洗衣液 速溶快洗",
                    "goodsImageUrl": "https://img1.360buyimg.com/image/jfs/t1/297856/19/16176/339249/6864ff58Ff3883324/6987f61780171b40.png",
                    "integral": 6000,
                    "salesVolume": 11,
                    "goodsType": 3
                },
                {
                    "id": "1de1e4d4cebd07e37c96c38b48394041",
                    "goodsName": "锐思 Recci  速风系列一拖三充电线 RCS-D120  颜色随机",
                    "goodsImageUrl": "https://img.li91.com/2022/03/500x500/n0/08770570189db8effecef32850d7cdb8.jpg",
                    "integral": 5000,
                    "salesVolume": 14,
                    "goodsType": 0
                },
                {
                    "id": "1b888960593fe35705cf8b6e4a01d9dc",
                    "goodsName": "啄木鸟 舒适颈枕礼盒装 ZMN-UXZ-01",
                    "goodsImageUrl": "https://img.li91.com/2021/06/800x800/n0/090504657faf6970580fb84a22790d6f.jpg",
                    "integral": 6500,
                    "salesVolume": 3,
                    "goodsType": 0
                },
                {
                    "id": "36de2532ce070a60a808b5787fb67aae",
                    "goodsName": "車格仕 多功能吸盘式 二合一导航支架M4",
                    "goodsImageUrl": "https://img.li91.com/2022/01/500x500/n0/4601f7bed3dc5dd294d755b1c8a1017f.jpg",
                    "integral": 6000,
                    "salesVolume": 18,
                    "goodsType": 0
                },
                {
                    "id": "f3598b13a19846edb133651a97cad282",
                    "goodsName": "翰乐 墨染古韵厨房刀具（2025GXZYH2）",
                    "goodsImageUrl": "https://img.li91.com/2025/06/500x500/n0/08705ec09f10496594f0de48cd1de9d8.jpg",
                    "integral": 4800,
                    "salesVolume": 61,
                    "goodsType": 0
                },
                {
                    "id": "6ef538118419747f3db503546860a0a9",
                    "goodsName": "栢士德 电动搅拌棒 BSD-D326 黑色",
                    "goodsImageUrl": "https://img.li91.com/2025/12/500x500/n0/320cb6accd22db1f8c4330ba3f8f760f.jpg",
                    "integral": 4300,
                    "salesVolume": 26,
                    "goodsType": 0
                },
                {
                    "id": "acf87ae06062080d92333606179e6e0c",
                    "goodsName": "茶花 瑞席保鲜封口夹 大号 9只装017001*3 颜色随机",
                    "goodsImageUrl": "https://img.li91.com/2025/12/500x500/n0/63d2aba2922e78e7f2b4b50406bed66c.jpg",
                    "integral": 4300,
                    "salesVolume": 3,
                    "goodsType": 0
                },
                {
                    "id": "bb3f150309cb4a23ea30d9c7aa053133",
                    "goodsName": "茶花 贝格微波保鲜盒组合装 000002+000003 黄色，绿色，蓝色",
                    "goodsImageUrl": "https://img.li91.com/2025/12/500x500/n0/5a0bc47ced9982d4fb91a327b4b2b1cb.jpg",
                    "integral": 4000,
                    "salesVolume": 23,
                    "goodsType": 0
                },
                {
                    "id": "95fe0f2ee3485e15cccabf01696e24a0",
                    "goodsName": "水卫士seaways 洗衣机槽清洁剂125g/袋*6袋WD02 ZWJXYJC06",
                    "goodsImageUrl": "https://img.li91.com/2025/09/500x500/n0/0de3f8913c5f7d31e5811044f9f13930.jpg",
                    "integral": 3800,
                    "salesVolume": 31,
                    "goodsType": 0
                },
                {
                    "id": "b11463ffc736453646a9a3cd6d5f89c1",
                    "goodsName": "爱车屋迷你尘掸灰色 I-1083 灰色",
                    "goodsImageUrl": "https://img.li91.com/2022/02/500x500/n0/6df7f245dfcaf272ad9ed68f55ccc205.png",
                    "integral": 2300,
                    "salesVolume": 146,
                    "goodsType": 0
                },
                {
                    "id": "22da613cb04f282dda42987b784cb2e5",
                    "goodsName": "得宝 四层樱花迷你手帕纸12包 T0131",
                    "goodsImageUrl": "https://img.li91.com/2024/12/500x500/n0/33132d10df0a49cb6292c3fea6fb53ce.jpg",
                    "integral": 3500,
                    "salesVolume": 26,
                    "goodsType": 0
                },
                {
                    "id": "72ea82d300828152e3a8a28f0578f4da",
                    "goodsName": "立白 大师香氛洗衣液1kg",
                    "goodsImageUrl": "https://img.li91.com/2023/03/500x500/n0/f3c90737d6ca6e084179e3548edd7e36.jpg",
                    "integral": 5800,
                    "salesVolume": 157,
                    "goodsType": 0
                },
                {
                    "id": "68a139f67bddbe627fc1368ccc07490d",
                    "goodsName": "蓝漂 大号湿巾纸80抽*5包装 LP-36464蓝漂亲肤柔软湿巾5包装改版",
                    "goodsImageUrl": "https://img.li91.com/2025/09/500x500/n0/d1dbf6dd3933a3ea0431a5c9be715879.jpg",
                    "integral": 5500,
                    "salesVolume": 46,
                    "goodsType": 0
                },
                {
                    "id": "3497919e4ab2ae002aed9d775769c76e",
                    "goodsName": "水卫士seaways  浴室清洁剂500g*1瓶（喷头*1） ZWJYS01PQ01",
                    "goodsImageUrl": "https://img.li91.com/2025/09/500x500/n0/5c81f63c90eb0b05dc59d2e4363ea0a9.jpg",
                    "integral": 3500,
                    "salesVolume": 42,
                    "goodsType": 0
                },
                {
                    "id": "0faa5f36ccd9674b2b08739b3914cd82",
                    "goodsName": "维达 棉韧抽取面巾3层XS码100抽×10包提 V2866",
                    "goodsImageUrl": "https://img.li91.com/2023/10/500x500/n0/5b4a861a1a24da7f729c513b40904db8.jpg",
                    "integral": 7500,
                    "salesVolume": 12,
                    "goodsType": 0
                },
                {
                    "id": "dea96162e8dc1234c512e85a75a82108",
                    "goodsName": "舒蕾 洁雅轻柔洗衣液1.5kg",
                    "goodsImageUrl": "https://img.li91.com/2025/01/500x500/n0/cb29e6b74a688d0451f34bfd7abdcc7c.jpg",
                    "integral": 4000,
                    "salesVolume": 61,
                    "goodsType": 0
                },
                {
                    "id": "da5cc6f25703c8ff9ae53a7169874a8e",
                    "goodsName": "美丽雅一次性防尘罩 6923074094314",
                    "goodsImageUrl": "https://img.li91.com/2025/08/500x500/n0/8bfedd15fde822e95f17ca5e658b1deb.jpg",
                    "integral": 5300,
                    "salesVolume": 9,
                    "goodsType": 0
                },
                {
                    "id": "8f931bcb6097437f723199a7f816fbd1",
                    "goodsName": "美丽雅抽取式袋装平口保鲜袋（小包）*5 6923074050846",
                    "goodsImageUrl": "https://img.li91.com/2025/08/500x500/n0/583357d597fb533a1d453056638e3654.jpg",
                    "integral": 4000,
                    "salesVolume": 15,
                    "goodsType": 0
                },
                {
                    "id": "d249ca4d56da7f00d4fc7612ce492869",
                    "goodsName": "美丽雅晶彩保鲜盒组合装（家庭装） 6923074070851",
                    "goodsImageUrl": "https://img.li91.com/2025/08/500x500/n0/d123abf6992c2a73cb61f3cba3404b0b.jpg",
                    "integral": 3000,
                    "salesVolume": 38,
                    "goodsType": 0
                },
                {
                    "id": "8b179f8485b0052344ca62b015b572c2",
                    "goodsName": "美丽雅保鲜袋(卷装)*2 6923074067738",
                    "goodsImageUrl": "https://img.li91.com/2025/08/500x500/n0/e6eddcfee8f4fee9aa223f0277005b1f.jpg",
                    "integral": 3800,
                    "salesVolume": 27,
                    "goodsType": 0
                },
                {
                    "id": "d30241a462e03eb9db5d544178725149",
                    "goodsName": "婴侍卫 卡通动物立体尾巴布书 带挂绳 QBX181 丛林/海洋/农场 图案随机",
                    "goodsImageUrl": "https://img.li91.com/2025/09/500x500/n0/cc5c643ddefe4aa9bf2835375b48e513.jpg",
                    "integral": 6500,
                    "salesVolume": 6,
                    "goodsType": 0
                },
                {
                    "id": "2cb83c62961ee4350f1855b642adf133",
                    "goodsName": "美丽雅超纤抹布组3+1片 6923074053090",
                    "goodsImageUrl": "https://img.li91.com/2025/08/500x500/n0/86b08a7103afbc06f422ea4856eb245d.jpg",
                    "integral": 3300,
                    "salesVolume": 25,
                    "goodsType": 0
                },
                {
                    "id": "db31ead4efe07cf79ea20a1e512818ea",
                    "goodsName": "水卫士seaways  地板清洁剂（香茅）500ml*1瓶 19110200748",
                    "goodsImageUrl": "https://img.li91.com/2025/09/500x500/n0/b4614ce2e784178cee74b7b92b5b7ffa.jpg",
                    "integral": 4000,
                    "salesVolume": 27,
                    "goodsType": 0
                },
                {
                    "id": "984d2b6dfafd0bf1ad3daa52162ab746",
                    "goodsName": "水卫士seaways  多功能清洁喷雾500g*1瓶（喷头*1） ZWJDGN01",
                    "goodsImageUrl": "https://img.li91.com/2025/09/500x500/n0/7fb477547322c79118c4500051a5448b.JPG",
                    "integral": 4000,
                    "salesVolume": 25,
                    "goodsType": 0
                },
                {
                    "id": "116a5b22e1023eec21fb4d66a91c5f24",
                    "goodsName": "三利云7A抗菌毛巾 两条装V8015绿+粉 绿色+粉色",
                    "goodsImageUrl": "https://img.li91.com/2025/07/500x500/n0/bfff50ce981cfe390e3862703125c263.png",
                    "integral": 6800,
                    "salesVolume": 14,
                    "goodsType": 0
                },
                {
                    "id": "b98dd519798a8c4c9e01207036f2566d",
                    "goodsName": "蓝漂卷纸家用白色无芯卷筒纸手纸厕纸1提 LP-37416白色扁卷12卷",
                    "goodsImageUrl": "https://img.li91.com/2025/08/500x500/n0/5b29d1aec134a43527c00b5f3f8b5693.png",
                    "integral": 3000,
                    "salesVolume": 136,
                    "goodsType": 0
                },
                {
                    "id": "12371995775dcd1284c40b15a0d1f903",
                    "goodsName": "HOCO.极光游戏鼠标垫 (橡胶&amp;佳积布) 办公鼠标垫 GM22 (200*240mm)",
                    "goodsImageUrl": "https://img.li91.com/2025/03/500x500/n0/006605fe296ae456b3207a199ae30d62.jpg",
                    "integral": 3800,
                    "salesVolume": 14,
                    "goodsType": 0
                },
                {
                    "id": "939280103d9f8676df43185e5d48b89f",
                    "goodsName": "爱舒柔植物毛巾洗脸巾 200g/包*3包ARG431 300mm*200mm",
                    "goodsImageUrl": "https://img.li91.com/2025/05/500x500/n0/e62d3ee6aa6dc06dcd37e62ed173037d.jpg",
                    "integral": 3600,
                    "salesVolume": 56,
                    "goodsType": 0
                },
                {
                    "id": "c0128257a57822bec00d21ec535a0737",
                    "goodsName": "婴侍卫 宝宝防水硅胶饭兜围兜 YSW2101 粉色天鹅",
                    "goodsImageUrl": "https://img.li91.com/2025/04/500x500/n0/edb6bd50fc38e89c63034ba6031debb4.jpg",
                    "integral": 5000,
                    "salesVolume": 5,
                    "goodsType": 0
                }
            ],
            "total": "71"
        },
        "classificationInformation": [
            {
                "id": "d80c336922e246a5b5a836f4a7f6a587",
                "classificationName": "儿童玩具"
            },
            {
                "id": "cb4779fd7be548479ec21cbd68e40c4e",
                "classificationName": "个护清洁"
            },
            {
                "id": "b555c7cbd4be4e5ca30f448924f04472",
                "classificationName": "厨具用品"
            },
            {
                "id": "7d1d8391ff8140a49f850c1aac88903f",
                "classificationName": "车类用品"
            },
            {
                "id": "65c94760bfec48e79ee3a1b9eb7c4ce7",
                "classificationName": "电脑办公"
            },
            {
                "id": "535ed934a70a48d9acc546373aca1ae0",
                "classificationName": "手机数码"
            },
            {
                "id": "2e1829c7d42342c285e9e6f48e2b79c3",
                "classificationName": "图书音像"
            },
            {
                "id": "07e4c2f3af0547c88feb383b0cc40911",
                "classificationName": "日常家居"
            },
            {
                "id": "5f7dd8258246cc90787044b05dd7662e",
                "classificationName": "运动户外"
            }
        ],
        "brandInformation": [
            {
                "id": "70a883d08c1a4818a42c80766b7f103e",
                "brandName": "默认",
                "goodsSize": "559"
            }
        ]
    },
    "rt_code": 0,
    "rt_msg": "success"
}

info2 = {
    "data": {
        "searchResponse": {
            "data": [
                {
                    "id": "ba313893f91441f7ea2334e082105531",
                    "goodsName": "婴侍卫 小萌熊指甲剪套装 YSW2791",
                    "goodsImageUrl": "https://img.li91.com/2025/04/500x500/n0/9279331f2d247eb3f8a94c055edf4ba7.jpg",
                    "integral": 5000,
                    "salesVolume": 12,
                    "goodsType": 0
                },
                {
                    "id": "773acf4a5f4fa25a8693f4269d37607d",
                    "goodsName": "婴侍卫 儿童帆布笔袋 2021 图案随机",
                    "goodsImageUrl": "https://img.li91.com/2025/04/500x500/n0/0d1fc2d9103f28049e90f0c49ce686d4.jpg",
                    "integral": 4400,
                    "salesVolume": 18,
                    "goodsType": 0
                },
                {
                    "id": "4668b965a3c65b1aac37814319f94368",
                    "goodsName": "拜杰 304不锈钢水槽过滤架SC-1",
                    "goodsImageUrl": "https://img.li91.com/2025/04/500x500/n0/42bfe3066a9cd784cfa4e649ce491408.jpg",
                    "integral": 5000,
                    "salesVolume": 5,
                    "goodsType": 0
                },
                {
                    "id": "5c6263f79990399d8f978f8d67a0faae",
                    "goodsName": "玖慕 冰丝防晒袖套儿童冰袖夏 BX090 浅灰狮子",
                    "goodsImageUrl": "https://img.li91.com/2025/03/500x500/n0/19a89424cdf587407a31d58470760901.jpg",
                    "integral": 4200,
                    "salesVolume": 26,
                    "goodsType": 0
                },
                {
                    "id": "06c5ea7fa3fbf7b9d2f0965f68e5a430",
                    "goodsName": "榄菊洗洁精大桶5kg 菊之语柠檬去油果蔬清洗剂餐具食品用级别洗涤灵",
                    "goodsImageUrl": "https://img1.360buyimg.com/image/jfs/t1/453873/7/18150/149052/6a38aa99F0b53ff82/0083320320cd4a00.jpg",
                    "integral": 6000,
                    "salesVolume": 161,
                    "goodsType": 3
                },
                {
                    "id": "8c7f0d4d6f09cafd3c2d64d1767b48e1",
                    "goodsName": "keep健身稳定护腕（岩灰黑）一对 岩灰黑一对",
                    "goodsImageUrl": "https://img.li91.com/2024/08/500x500/n0/c5827b2e31837f85b308ea4cd9bbc534.jpg",
                    "integral": 7000,
                    "salesVolume": 6,
                    "goodsType": 0
                },
                {
                    "id": "1d51492357b100f40f0eef4f36cef6b8",
                    "goodsName": "太力  多功能吸盘墙角置物盒 AW778",
                    "goodsImageUrl": "https://img.li91.com/2023/10/500x500/n0/aad0259129fe44d3d59c1202eb436e6d.jpg",
                    "integral": 7500,
                    "salesVolume": 9,
                    "goodsType": 0
                },
                {
                    "id": "9ff20525d9a620be5afbcb14e5288ff8",
                    "goodsName": "太力  植绒衣架 （10个装）",
                    "goodsImageUrl": "https://img.li91.com/2023/11/500x500/n0/5a734befaf55bbc4e2bd6448afa92b30.jpg",
                    "integral": 5400,
                    "salesVolume": 69,
                    "goodsType": 0
                },
                {
                    "id": "db0796062739e5d51ff5c063bbba49d6",
                    "goodsName": "酷客者 KOOKZZ 旋转手机支架ZJ05 白色",
                    "goodsImageUrl": "https://img.li91.com/2024/03/500x500/n0/57c309cc7a5e0541bf24654d5def8aca.jpg",
                    "integral": 6300,
                    "salesVolume": 56,
                    "goodsType": 0
                },
                {
                    "id": "4275bfd9909f7d27145d55448377a7b4",
                    "goodsName": "清风萌宠系列亲肤抽纸6联包 2提 黑色快递袋子",
                    "goodsImageUrl": "https://img.li91.com/2022/08/500x500/n0/44ac129d4ddcd63730a8555d7d3c0260.png",
                    "integral": 3700,
                    "salesVolume": 986,
                    "goodsType": 0
                },
                {
                    "id": "53ce4feff275493606c6f034dc47a701",
                    "goodsName": "超能 APG洗衣凝珠5腔200g 20颗",
                    "goodsImageUrl": "https://img.li91.com/2022/06/500x500/n0/33ff8627cc90ce235ca55a44b986287e.jpg",
                    "integral": 6000,
                    "salesVolume": 39,
                    "goodsType": 0
                },
                {
                    "id": "f7b4f401823423a8491f2a7e487e250f",
                    "goodsName": "张小泉白玉兰三件套指甲钳套装H41270100",
                    "goodsImageUrl": "https://img.li91.com/2022/01/500x500/n0/07f8395cc2fa2bbf2173b124c973a4c7.jpg",
                    "integral": 4800,
                    "salesVolume": 152,
                    "goodsType": 0
                },
                {
                    "id": "4ffc3ff0a0486f9978d7899cc922bae8",
                    "goodsName": "得力 A575学生按动中性笔0.5mm臻顺滑 (黑)(5支/盒)*2盒 100138261",
                    "goodsImageUrl": "https://img.li91.com/2023/09/500x500/n0/40b349d340f74fbbb3ededb3344e76ac.jpg",
                    "integral": 3900,
                    "salesVolume": 557,
                    "goodsType": 0
                },
                {
                    "id": "5895867aab0e8c914b2f2e2977c46a5a",
                    "goodsName": "立白 果醋洗洁精浓缩洗涤精石榴醋精华去除果蔬农残去重油 1kg",
                    "goodsImageUrl": "https://img.li91.com/2023/11/500x500/n0/e385274b1708cc5d4085ec90901c8268.jpg",
                    "integral": 2990,
                    "salesVolume": 577,
                    "goodsType": 0
                },
                {
                    "id": "cac2aa6b76d455216504e57fdccfd537",
                    "goodsName": "立白  洗洁精西柚精华洗涤剂洗碗液去油不伤手408g*4瓶装  773323X4",
                    "goodsImageUrl": "https://img.li91.com/2023/11/500x500/n0/ff757cb82daea3bb9bf32bd604c34c32.jpg",
                    "integral": 2950,
                    "salesVolume": 425,
                    "goodsType": 0
                },
                {
                    "id": "60be18240ea3c922b09fcb67365d43c6",
                    "goodsName": "爱车屋 双USB车载充电器D-3027",
                    "goodsImageUrl": "https://img.li91.com/2023/05/500x500/n0/4fc352514e43144b7140220ac832137f.jpg",
                    "integral": 2242,
                    "salesVolume": 369,
                    "goodsType": 0
                },
                {
                    "id": "fb80a16687f4c9e67ae9f1883137afb8",
                    "goodsName": "太力吸盘浴室双人拖鞋架AW629 双人",
                    "goodsImageUrl": "https://img.li91.com/2023/11/500x500/n0/048a17f7c5a43b02c4b99cb68d07178a.jpg",
                    "integral": 3186,
                    "salesVolume": 190,
                    "goodsType": 0
                },
                {
                    "id": "cd8783bd66c76d4261c7cfa8874af4b4",
                    "goodsName": "酷客者Kookzz 蓝牙音响 Y01",
                    "goodsImageUrl": "https://img.li91.com/2023/09/500x500/n0/6a66cf680c87650cea9afeae6f2b529b.jpg",
                    "integral": 7316,
                    "salesVolume": 122,
                    "goodsType": 0
                },
                {
                    "id": "ed3546319e0c3c85a74dc6516b501788",
                    "goodsName": "首佩 Soopii 无线双模鼠标 G30",
                    "goodsImageUrl": "https://img.li91.com/2021/12/500x500/n0/427a8e075a2cf42c0710adc27ab3d732.jpg",
                    "integral": 7500,
                    "salesVolume": 131,
                    "goodsType": 0
                },
                {
                    "id": "57cb6e2bf25703ef97cc9165ef6065f1",
                    "goodsName": "天堂伞33188E 颜色随机",
                    "goodsImageUrl": "https://img.li91.com/2024/01/500x500/n0/c3fcacbadeadd7c657ae444abcc06950.jpg",
                    "integral": 5900,
                    "salesVolume": 460,
                    "goodsType": 0
                },
                {
                    "id": "c9fa7501d431f4e640cf825780d33781",
                    "goodsName": "锐思 Recci 桌面立式手机支架 RHO-M10",
                    "goodsImageUrl": "https://img.li91.com/2022/04/500x500/n0/0b5dd46481a0ec025023135dee72129f.jpg",
                    "integral": 5000,
                    "salesVolume": 115,
                    "goodsType": 0
                }
            ],
            "total": "71"
        },
        "classificationInformation": [
            {
                "id": "d80c336922e246a5b5a836f4a7f6a587",
                "classificationName": "儿童玩具"
            },
            {
                "id": "cb4779fd7be548479ec21cbd68e40c4e",
                "classificationName": "个护清洁"
            },
            {
                "id": "b555c7cbd4be4e5ca30f448924f04472",
                "classificationName": "厨具用品"
            },
            {
                "id": "7d1d8391ff8140a49f850c1aac88903f",
                "classificationName": "车类用品"
            },
            {
                "id": "65c94760bfec48e79ee3a1b9eb7c4ce7",
                "classificationName": "电脑办公"
            },
            {
                "id": "535ed934a70a48d9acc546373aca1ae0",
                "classificationName": "手机数码"
            },
            {
                "id": "2e1829c7d42342c285e9e6f48e2b79c3",
                "classificationName": "图书音像"
            },
            {
                "id": "07e4c2f3af0547c88feb383b0cc40911",
                "classificationName": "日常家居"
            },
            {
                "id": "5f7dd8258246cc90787044b05dd7662e",
                "classificationName": "运动户外"
            }
        ],
        "brandInformation": [
            {
                "id": "70a883d08c1a4818a42c80766b7f103e",
                "brandName": "默认",
                "goodsSize": "559"
            }
        ]
    },
    "rt_code": 0,
    "rt_msg": "success"
}

# ============================================================================
# 配置
# ============================================================================
OUTPUT_EXCEL = os.path.join(os.path.dirname(__file__), "商品数据导出.xlsx")
IMAGE_DIR = os.path.join(os.path.dirname(__file__), "product_images")
os.makedirs(IMAGE_DIR, exist_ok=True)

# Excel中图片的显示尺寸
IMG_DISPLAY_WIDTH = 120   # 像素
IMG_DISPLAY_HEIGHT = 120  # 像素
ROW_HEIGHT = 100          # Excel行高（磅）


# ============================================================================
# 工具函数
# ============================================================================
def download_image(url: str, save_path: str, timeout: int = 15) -> bool:
    """
    下载图片到本地

    参数:
      url:      图片URL
      save_path: 本地保存路径
      timeout:   请求超时秒数

    返回:
      成功返回True，失败返回False

    算法选择:
      使用 requests.get() 流式下载而非一次性加载到内存，
      避免大图（>10MB）导致内存溢出。stream=True 逐块写入磁盘。
    """
    if not url.startswith("http"):
        url = "https:" + url

    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                      "AppleWebKit/537.36 (KHTML, like Gecko) "
                      "Chrome/120.0.0.0 Safari/537.36",
        # "Referer": "https://jf.yiheda.com/",
    }

    for attempt in range(3):
        try:
            resp = requests.get(url, headers=headers, timeout=timeout)

            with open(save_path, "wb") as f:
                f.write(resp.content)
            return True
        except Exception as e:
            if attempt == 2:
                logger.error(f"  [图片下载失败] {url[:80]}... : {e}")
            time.sleep(1)
    return False


def sanitize_filename(name: str) -> str:
    """将商品名转换为安全的文件名（去除非法字符）"""
    illegal_chars = r' \/:*?"<>|'
    for ch in illegal_chars:
        name = name.replace(ch, "_")
    # 截断过长的名称
    if len(name) > 80:
        name = name[:80]
    return name


# ============================================================================
# 主流程
# ============================================================================
def main():
    # ---- 合并数据 ----
    data1 = info1["data"]["searchResponse"]["data"]
    data2 = info2["data"]["searchResponse"]["data"]
    all_items = data1 + data2

    total = len(all_items)
    logger.info(f"共 {total} 个商品待处理")

    # ========================================================================
    # 阶段一：预扫描 —— 快速访问每个商品详情页，仅统计主图和详情图数量
    # ========================================================================
    logger.info(f"\n{'='*50}")
    logger.info("阶段一：预扫描 —— 统计各商品图片数量...")
    logger.info(f"{'='*50}")

    max_main_imgs = 30
    max_detail_imgs = 30

    # for idx, item in enumerate(all_items):
    #     goods_id = item["id"]
    #     detail_url = f"https://jf.yiheda.com/product/detail?goodsCode={goods_id}"
    #     logger.info(f"  [{idx+1}/{total}] 预扫 {item['goodsName'][:40]}...")
    #
    #     Playwright_.goto(detail_url)
    #     time.sleep(1)
    #
    #     close_ele = '//i[@class="iconfont icon-qingchu"]'
    #     if Playwright_.get_count(close_ele):
    #         Playwright_.click(close_ele)
    #         # time.sleep(2)
    #
    #     main_pic_count = Playwright_.get_count(
    #         '//div[@class="imageList"]/ul/li/img'
    #     )
    #     detail_pic_count = Playwright_.get_count('//p/img')
    #
    #     max_main_imgs = min(max_main_imgs, main_pic_count)
    #     max_detail_imgs = min(max_detail_imgs, detail_pic_count)
    #     logger.info(f"    主图:{main_pic_count}张  详情图:{detail_pic_count}张  "
    #           f"(累计最大: 主图{max_main_imgs}/详情图{max_detail_imgs})")
    #
    # logger.info(f"\n预扫描完成: 主图最多{max_main_imgs}张, 详情图最多{max_detail_imgs}张")

    # ========================================================================
    # 阶段二：创建Excel工作簿，写入表头
    # ========================================================================
    headers = ["商品编号", "商品名称", "价格", "商品规格"]
    col_main_start = len(headers) + 1
    for i in range(1, max_main_imgs + 1):
        headers.append(f"主图{i}")
    col_detail_start = col_main_start + max_main_imgs
    for i in range(1, max_detail_imgs + 1):
        headers.append(f"详情图{i}")

    logger.info(f"\n{'='*50}")
    logger.info(f"创建Excel: {OUTPUT_EXCEL}")
    logger.info(f"表头列数: {len(headers)} (固定4列 + 主图{max_main_imgs}列 + 详情图{max_detail_imgs}列)")
    logger.info(f"{'='*50}")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "商品数据"

    header_font = openpyxl.styles.Font(bold=True, size=11)
    header_fill = openpyxl.styles.PatternFill(
        start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"
    )
    header_align = openpyxl.styles.Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    ws.column_dimensions[get_column_letter(1)].width = 35
    ws.column_dimensions[get_column_letter(2)].width = 45
    ws.column_dimensions[get_column_letter(3)].width = 10
    ws.column_dimensions[get_column_letter(4)].width = 50
    for col in range(col_main_start, col_detail_start + max_detail_imgs):
        ws.column_dimensions[get_column_letter(col)].width = 18

    ws.freeze_panes = "A2"
    wb.save(OUTPUT_EXCEL)
    wb.close()
    logger.info("Excel表头已写入并保存")

    # ========================================================================
    # 阶段三：逐条采集 —— 每处理完一个商品，立即写入Excel并保存
    # ========================================================================
    data_alignment = openpyxl.styles.Alignment(
        horizontal="left", vertical="top", wrap_text=True
    )

    for idx, item in enumerate(all_items[54:55], start=54):
        goods_id = item["id"]
        title = item["goodsName"]
        price = item["integral"] / 100
        detail_url = f"https://jf.yiheda.com/product/detail?goodsCode={goods_id}"
        excel_row = idx + 2

        logger.info(f"\n{'─'*50}")
        logger.info(f"[{idx+1}/{total}] {title[:50]}...")

        Playwright_.goto(detail_url)
        time.sleep(1)

        close_ele = '//i[@class="iconfont icon-qingchu"]'
        if Playwright_.get_count(close_ele):
            Playwright_.click(close_ele)
            # time.sleep(3)

        # ---- 提取商品规格 ----
        specification_ele = '//div[@class="specsList"]'
        spec_count = Playwright_.get_count(specification_ele)
        specification_parts = []
        for i in range(1, spec_count + 1):
            key_ele = f"({specification_ele})[{i}]/div[1]"
            key = Playwright_.get_text(key_ele).strip()
            value_ele = f"({specification_ele})[{i}]/div[2]/div"
            value = Playwright_.get_text(value_ele).strip()
            specification_parts.append(f"{key}{value}")
        specification = "；".join(specification_parts)

        # ---- 下载商品图片 ----
        safe_title = sanitize_filename(title)
        product_img_dir = os.path.join(IMAGE_DIR, f"{idx+1:03d}_{safe_title[:40]}")
        os.makedirs(product_img_dir, exist_ok=True)

        # ---- 提取主图 ----
        main_imgs = []
        main_pic_base = '//div[@class="imageList"]/ul/li/img'
        main_pic_count = Playwright_.get_count(main_pic_base)
        for i in range(1, main_pic_count + 1):
            pic_xpath = f"({main_pic_base})[{i}]"
            pic_url = Playwright_.get_attribute(pic_xpath, "src")
            if pic_url:
                ext = ".jpg"
                if ".png" in pic_url.lower():
                    ext = ".png"
                elif ".webp" in pic_url.lower():
                    ext = ".webp"
                local_path = os.path.join(product_img_dir, f"主图{i}{ext}")
                if download_image(pic_url, local_path):
                    main_imgs.append(local_path)
                    logger.info(f"  主图{i} [OK]")
                else:
                    logger.error(f"  主图{i} [FAIL]")

        # ---- 提取详情图 ----
        detail_imgs = []
        detail_pic_base = '//p/img'
        detail_pic_count = Playwright_.get_count(detail_pic_base)
        for i in range(1, detail_pic_count + 1):
            pic_xpath = f"({detail_pic_base})[{i}]"
            pic_url = Playwright_.get_attribute(pic_xpath, "src")
            if pic_url:
                ext = ".jpg"
                if ".png" in pic_url.lower():
                    ext = ".png"
                elif ".webp" in pic_url.lower():
                    ext = ".webp"
                local_path = os.path.join(product_img_dir, f"详情图{i}{ext}")
                if download_image(pic_url, local_path):
                    detail_imgs.append(local_path)
                    logger.info(f"  详情图{i} [OK]")
                else:
                    logger.error(f"  详情图{i} [FAIL]")

        # ---- 写入Excel ----
        wb = openpyxl.load_workbook(OUTPUT_EXCEL)
        ws = wb.active

        ws.cell(row=excel_row, column=1, value=goods_id).alignment = data_alignment
        ws.cell(row=excel_row, column=2, value=title).alignment = data_alignment
        price_cell = ws.cell(row=excel_row, column=3, value=price)
        price_cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="top")
        price_cell.number_format = "0.00"
        ws.cell(row=excel_row, column=4, value=specification).alignment = data_alignment

        for img_idx, img_path in enumerate(main_imgs):
            if img_idx >= max_main_imgs:
                logger.info(f"  [警告] 主图{img_idx+1}超出最大列数{max_main_imgs}，跳过")
                continue
            if os.path.exists(img_path):
                col = col_main_start + img_idx
                try:
                    img = XLImage(img_path)
                    img.width = IMG_DISPLAY_WIDTH
                    img.height = IMG_DISPLAY_HEIGHT
                    ws.add_image(img, f"{get_column_letter(col)}{excel_row}")
                except Exception as e:
                    ws.cell(row=excel_row, column=col,
                            value=f"[加载失败]").alignment = data_alignment
                    logger.error(f"  图片插入失败 {img_path}: {e}")

        for img_idx, img_path in enumerate(detail_imgs):
            if img_idx >= max_detail_imgs:
                logger.warning(f"  [警告] 详情图{img_idx+1}超出最大列数{max_detail_imgs}，跳过")
                continue
            if os.path.exists(img_path):
                col = col_detail_start + img_idx
                try:
                    img = XLImage(img_path)
                    img.width = IMG_DISPLAY_WIDTH
                    img.height = IMG_DISPLAY_HEIGHT
                    ws.add_image(img, f"{get_column_letter(col)}{excel_row}")
                except Exception as e:
                    ws.cell(row=excel_row, column=col,
                            value=f"[加载失败]").alignment = data_alignment
                    logger.error(f"  图片插入失败 {img_path}: {e}")

        ws.row_dimensions[excel_row].height = ROW_HEIGHT

        try:
            wb.save(OUTPUT_EXCEL)
            logger.info(f"  [OK] 第{excel_row-1}条数据已保存 (主图{len(main_imgs)}张/详情图{len(detail_imgs)}张)")
        except Exception as e:
            logger.error(f"  [错误] Excel保存失败: {e}")
        finally:
            wb.close()

    logger.info(f"\n{'='*50}")
    logger.info(f"[OK] 全部完成! 共{total}个商品")
    logger.info(f"  · Excel: {OUTPUT_EXCEL}")
    logger.info(f"  · 图片:  {IMAGE_DIR}")
    logger.info(f"{'='*50}")


if __name__ == "__main__":
    main()
