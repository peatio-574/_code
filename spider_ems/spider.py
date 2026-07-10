# -*- coding: utf-8 -*-
import sys
import time

import pandas
from pathlib import Path

sys.path.append(str(Path(__file__).parent.parent))

from PlayWright import Playwright_, logger, get_config_value
from ReadFile import ReadData
import os
from openpyxl import load_workbook

configFile = os.path.join(os.path.dirname(__file__), 'config.ini')
dataFile = os.path.join(os.path.dirname(__file__), '邮寄单号汇总表.xlsx')
pictureDir = os.path.join(os.path.dirname(__file__), 'picture')
os.makedirs(pictureDir, exist_ok=True)

def emsLogin():
    """EMS"""
    logger.info('=' * 80)
    logger.info('开始登录EMS....')
    url = 'https://www.ems.com.cn/personal_center'
    ele = '//span[text()="我的快递"]'
    key = 'login.ems_cookie'
    try:
        Playwright_.login(url, ele, key, file=configFile)
        logger.info('✓ EMS登录成功')
        return True
    except Exception as e:
        logger.error(f'✗ EMS登录失败：{e}')
        return False

def getPageInfo():
    """获取订单信息"""
    Playwright_.goto('https://www.ems.com.cn/query_express_delivery')
    Playwright_.click('//span[text()=" 筛选"]')
    Playwright_.click('//span[text()=" 已签收 "]')
    Playwright_.click('//span[text()=" 退件 "]')
    Playwright_.click('//span[text()="确定"]')
    time.sleep(3)

    pageCount = Playwright_.get_text('(//ul/li[@class="number"])[last()]')
    pageCount = int(pageCount)
    info = {}
    for pageId in range(1, pageCount + 1):
        if pageId != 1:
            Playwright_.click(f'//ul/li[@class="number" and text()="{pageId}"]')
            time.sleep(2)
        logger.info(f'开始获取第{pageId}页订单信息')

        currentPage = Playwright_.get_text('//ul/li[@class="number active"]')
        if currentPage != str(pageId):
            logger.error(f'获取订单信息失败：当前页码{currentPage}与期望页码{pageId}不一致，请重新运行')
            exit()

        rowEle = '//div[@role="tabpanel"]/div[@class="orderItem"]'
        rowCount = Playwright_.get_count(rowEle)
        for rowId in range(1, rowCount + 1):
            mailNo = Playwright_.get_text(f'({rowEle})[{rowId}]//span[@class="order-number"][1]').split('/')[0]
            mailNo = mailNo.split('：')[1].strip()
            orderNo = Playwright_.get_text(f'({rowEle})[{rowId}]//span[@class="order-number"][2]')
            orderNo = orderNo.split('：')[1].strip()
            info[mailNo] = orderNo
            logger.info(f'获取订单信息成功：{mailNo} {orderNo}')
    currentData = ReadData.read_xlsx_col(dataFile, header=1)['物流单号']
    wb = load_workbook(dataFile)
    ws = wb.active
    ws.cell(row=2, column=4, value='订单号')
    wb.save(dataFile)
    for rowIdx, mailNo in enumerate(currentData, start=3):
        code = info.get(mailNo)
        if code:
            ws.cell(row=rowIdx, column=4, value=code)
    wb.save(dataFile)
    logger.info('✅️ 获取订单信息成功')

def getOrderPicture(idx, orderNo, mailNo, interval):
    """获取订单图片"""
    url = f'https://www.ems.com.cn/express_detail?type=0&status=1&orderNo={orderNo}&mailNo={mailNo}&listType=0&isSubAccount=false&userId=&userPhone='
    Playwright_.goto(url)
    time.sleep(int(interval))
    fileName = os.path.join(pictureDir, f'{idx}_{mailNo}.png')
    Playwright_.screenshot(fileName)

def getAllPicture(interval):
    """获取所有图片"""
    allData = ReadData.read_xlsx_row(dataFile, header=1)
    for row in allData:
        idx = row['序号']
        orderNo = row['订单号']
        mailNo = row['物流单号']
        if orderNo == '':
            logger.info(f'【{idx}_{mailNo}】无对应订单号')
            continue
        logger.info(f'开始获取{idx}-{orderNo}-{mailNo}图片')
        getOrderPicture(idx, orderNo, mailNo, interval)
        logger.info(f'✅️ {idx}_{mailNo}图片保存成功')

if __name__ == '__main__':
    interval = get_config_value('login', 'interval', file=configFile)
    while True:
        run = input('请输入（1.获取订单列表 2.获取快递信息）：')
        emsLogin()
        if run == '1':
            getPageInfo()
        elif run == '2':
            getAllPicture(interval)
