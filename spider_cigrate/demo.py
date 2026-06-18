

from PlayWright import Playwright_, logger
import time
import os
import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.drawing.image import Image
filename = os.path.join(os.path.dirname(__file__), '数据.xlsx')

if os.path.exists(filename):
    # 如果文件存在，加载现有工作簿
    wb = load_workbook(filename)
    ws = wb.active
else:
    headers = ['产品名称', '批发价', '产品编码', '盒码']
    wb = Workbook()
    ws = wb.active
    ws.title = '数据'
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    wb.save(filename)

picture_dir = os.path.join(os.path.dirname(__file__), 'pictures')

os.makedirs(picture_dir, exist_ok=True)


def login():
    logger.info('开始登录')
    url = 'https://www.jinye.cn/marketing-orderplatform/ui/pages/list.html?funCode=MOHD0418&moreCondition=isRecommend'
    ele = '//a[text()="退出"]'
    Playwright_.login(url, ele, 'login.cookie')


def get_page_info(page=1):
    logger.info(f'开始获取第{page}页链接')
    if page != 1:
        Playwright_.click(f'//a[@title="第{page}页"]')
    item_ids = []
    row_count_ele = '//div[@class="productitem-shoppinglist  productitem"]'
    row_count = Playwright_.get_count(row_count_ele)
    for row in range(1, row_count + 1):
        ele = f'({row_count_ele})[{row}]'
        item_id = Playwright_.get_attribute(ele, 'item-id')
        item_id = f'https://www.jinye.cn/marketing-orderplatform/ui/pages/detail.html?pk_cigarette={item_id}'
        item_ids.append(item_id)
    logger.info(f'获取到第{page}页数据：{len(item_ids)}条链接')
    return item_ids


def get_itmes():
    login()
    time.sleep(30)
    logger.info('开始获取商品信息')
    pages = 62
    items = []
    for page in range(1, pages + 1):
        item_ids = get_page_info(page)
        items += item_ids
    info = '\n'.join(items)
    with open('./items.txt', 'w', encoding='utf-8') as f:
        f.write(info)
        logger.info('商品信息保存成功')



def get_detail(url):
    logger.info(f'开始处理产品：{url}')
    Playwright_.goto(url)
    time.sleep(3)


    name_ele = '//h1[@class="productname"]'
    name = Playwright_.get_text(name_ele)

    src = Playwright_.get_attribute('(//li/img)[1]', 'src')
    picture_file = os.path.join(picture_dir, f'{name}.jpg')
    picture = requests.get(src).content
    with open(picture_file, mode='wb') as f:
        f.write(picture)
        logger.info('')

    price_ele = '//span[@class=" single-price"]'
    price = Playwright_.get_text(price_ele)

    product_code_ele = '//td[@cgtinfo="vcgtcode"]'
    product_code = Playwright_.get_text(product_code_ele)

    box_code_ele = '//td[@cgtinfo="vcgtboxcode"]'
    box_code = Playwright_.get_text(box_code_ele)

    row_data = [name, price, product_code, box_code]

    ws.append(row_data)
    wb.save(filename)

    logger.info(f'保存数据成功：{row_data}')


def run():
    while True:
        need_data = input('请输入需要爬取的页码，中间用-分割（例如：1-30）：')
        if need_data == '0':
            exit()
        login()
        start_page, end_page = need_data.split('-')
        start_page = int(start_page)
        end_page = int(end_page)

        items_per_page = 12
        start_index = (start_page - 1) * items_per_page
        end_index = end_page * items_per_page
        logger.info(f"页码范围: {start_page}-{end_page}")
        logger.info(f"数据索引范围: {start_index}-{end_index}")
        logger.info(f"预计获取数据量: {end_index - start_index} 条")


        with open('./items.txt', mode='r', encoding='utf-8') as f:
            items = f.readlines()
        real_items = items[start_index:end_index]
        for url in real_items:
            get_detail(url)
            # break
        logger.info(f'第{start_page}-{end_page}页数据处理完成')


def save_pic():
    data_file = os.path.join(os.path.dirname(__file__), '图片.xlsx')

    if os.path.exists(data_file):
        # 如果文件存在，加载现有工作簿
        wb_ = load_workbook(data_file)
        ws_ = wb_.active
    else:
        headers_ = ['产品名称', '图片']
        wb_ = Workbook()
        ws_ = wb_.active
        ws_.title = '数据'
        ws_.append(headers_)

        for cell in ws_[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')

        wb_.save(filename)


    pictures = os.listdir(picture_dir)
    picnames = [i.split('.')[0] for i in pictures]

    for row_id, picture in enumerate(pictures, start=2):
        ws_.cell(row=row_id, column=1, value=picnames[row_id-2])
    wb_.save(data_file)


    for row_id, picture in enumerate(pictures, start=2):
        pic = os.path.join(picture_dir, picture)
        logger.info(f'开始处理图片：{picture}')
        try:
            img = Image(pic)
            img.width = 60
            img.height = 60
            ws_.add_image(img, f'B{row_id}')
            ws_.row_dimensions[row_id].height = 45
        except Exception as e:
            logger.error(f'{picture}添加图片失败: {e}')
        wb_.save(data_file)



if __name__ == '__main__':
    # get_itmes()           #获取连接
    run()                 #保存数据
    # save_pic()              # 保存图片





