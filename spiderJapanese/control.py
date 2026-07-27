# coding='utf-8'
import sys
from pathlib import Path

# 把项目根目录加入Python路径
sys.path.append(str(Path(__file__).parent.parent))

import re
from newPlayWright import SpecialPlayWright, logger
from ReadFile import ReadData
import os, time
from openpyxl import load_workbook

config_file = os.path.join(os.path.dirname(__file__), 'config.ini')

SpecialPlayWright = SpecialPlayWright(config_file=config_file)

def input_info(email,  password):
    """访问地址输入账号密码并提交"""
    try:
        url = 'https://www.pokemoncenter-online.com/login/'
        SpecialPlayWright.goto(url)
        time.sleep(5)
        SpecialPlayWright.input('//input[@type="email" and @id="login-form-email"]', email)
        SpecialPlayWright.input('//input[@type="password" and @id="current-password"]', password)
        SpecialPlayWright.click('//button[@type="submit" and @class="btn btn-block btn-primary"]')
        enterVerify = SpecialPlayWright.wait_for_selector('//input[@id="authCode"]', timeout=10*1000)
        return enterVerify
    except Exception as e:
        logger.error(f'输入账号密码异常：{e}')
        return False


def get_verify_code(apiUrl):
    """获取邮箱验证码"""
    try:
        SpecialPlayWright.new_goto(apiUrl, close=False)
        time.sleep(5)
        verifyCodeEle = '//div[@class="email-content"]/div[1]/div/p[3]'
        if SpecialPlayWright.get_count(verifyCodeEle):
            text = SpecialPlayWright.get_text(verifyCodeEle)
            verifyCode = re.findall(r'\d+', text)[0]
            SpecialPlayWright.switch_page('old')
            return verifyCode
        SpecialPlayWright.switch_page('old')
        return False
    except Exception as e:
        logger.error(f'获取邮箱验证码异常：{e}')
        return False


def login(accountCOde, email, password, emailUrl):
    try:
        logger.info(f'开始登录账号：{accountCOde}')
        enterVerify = False

        # 循环3次输入账号密码，进入验证码页面
        for roll in range(3):
            logger.info(f'{accountCOde}账号：开始尝试第{roll+1}次输入账号密码')
            enterVerify = input_info(email, password)
            if enterVerify:
                break
        if not enterVerify:
            logger.error(f'{accountCOde}账号登录失败：输入账号密码后未进入验证码页面')
            return False

        # 循环10次获取邮箱验证码
        verifyCode = False
        time.sleep(10)
        for roll in range(10):
            logger.info(f'{accountCOde}账号：开始尝试第{roll+1}次获取邮箱验证码')
            verifyCode = get_verify_code(emailUrl)
            if verifyCode:
                break
        if not verifyCode:
            logger.error(f'{accountCOde}账号登录失败：获取邮箱验证码失败')
            return False

        # 输入验证码，勾选同意并提交
        logger.info(f'{accountCOde}账号：开始输入验证码【{verifyCode}】')
        SpecialPlayWright.input('//input[@id="authCode"]', verifyCode)
        SpecialPlayWright.click('//input[@id="rememberMe"]')
        SpecialPlayWright.click('//a[@id="authBtn"]')
        status = SpecialPlayWright.wait_for_selector('//ul[@class="tabUl flex"]', timeout=20*1000)
        if status:
            logger.info(f'{accountCOde}账号登录成功')
            return True
        else:
            logger.error(f'{accountCOde}账号登录失败')
            return False
    except Exception as e:
        logger.error(f'{accountCOde}账号登录异常：{e}')
        return False


def run():
    fileName = os.path.join(os.path.dirname(__file__), '登录数据.xlsx')
    data = ReadData.read_xlsx_row(fileName)
    wb = load_workbook(fileName)
    ws = wb.active
    for rowId, row in enumerate(data, start=2):
        email = row['邮箱']
        accountCOde = row['编号']
        accountCOde += '_' + email
        emailUrl = row['邮箱验证地址']
        password = row['密码']
        status = login(accountCOde, email, password, emailUrl)
        SpecialPlayWright.clear_cookie()
        ws.cell(row=rowId, column=5, value=1 if status else 0)
        wb.save(fileName)
        break


if __name__ == '__main__':
    run()
