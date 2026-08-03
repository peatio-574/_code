# coding='utf-8'
import sys
import os

if getattr(sys, 'frozen', False):
    BUNDLE_DIR = sys._MEIPASS
    BASE_DIR = os.path.dirname(sys.executable)
else:
    BUNDLE_DIR = os.path.dirname(os.path.abspath(__file__))
    BASE_DIR = BUNDLE_DIR

parent_dir = os.path.join(BASE_DIR, '..')
sys.path.insert(0, BUNDLE_DIR)
sys.path.insert(0, parent_dir)

import tkinter as tk
from tkinter import filedialog, ttk, scrolledtext, messagebox
import threading

from newPlayWright import SpecialPlayWright, logger
from ReadFile import ReadData
import re
import time
import asyncio
import requests
import queue
from openpyxl import load_workbook


class App:
    # ==================== 初始化 ====================

    def __init__(self, root):
        # 内部状态
        self._status_update_scheduled = False
        self._pending_status = {}
        self._last_display = []

        # 窗口
        self.root = root
        self.root.title("宝可梦 登录工具")
        self.root.geometry("900x700")
        self.root.resizable(True, True)

        # 文件/数据路径
        self.file_path = None

        # 运行控制
        self.is_running = False
        self.is_paused = False
        self.stop_event = threading.Event()
        self.fetch_stop_event = threading.Event()
        self.fetch_stop_event.set()
        self.pause_event = threading.Event()
        self.task_epoch = 0

        # 电话号码队列
        self.phone_queue = queue.Queue()
        self.phone_queue_list = []
        self.phone_status = {}
        self.phone_queue_lock = threading.Lock()

        # 日志与线程
        self.log_queue = queue.Queue()
        self.fetch_thread = None

        # UI 组件
        self.file_label = None
        self.thread_var = None
        self.start_btn = None
        self.fetch_btn = None
        self.pause_btn = None
        self.stop_btn = None
        self.clear_btn = None
        self.queue_canvas = None
        self.queue_check_frame = None
        self.queue_scrollbar = None
        self.queue_vars = {}
        self.account_tree = None
        self.progress = None
        self.status_label = None
        self.log_text = None

        # Token
        self.start_time = time.time()
        self.end_time = None
        self.token = None

        self.create_widgets()
        self.poll_log_queue()
        self.poll_queue_display()

    # ==================== UI 构建 ====================

    def create_widgets(self):
        self._build_config_frame()
        self._build_btn_frame()
        self._build_status_frame()
        self._build_progress_and_log()

    def _build_config_frame(self):
        frame = tk.LabelFrame(self.root, text="配置信息", padx=10, pady=10)
        frame.pack(fill=tk.X, padx=10, pady=5)

        tk.Label(frame, text="账号文件:").grid(row=0, column=0, sticky=tk.W, padx=5, pady=3)
        self.file_label = tk.Label(frame, text="未选择文件", anchor=tk.W)
        self.file_label.grid(row=0, column=1, padx=5, pady=3, sticky=tk.W+tk.E)
        tk.Button(frame, text="选择", command=self.select_file, width=8).grid(row=0, column=2, padx=5, pady=3)

        tk.Label(frame, text="并发数:").grid(row=1, column=0, sticky=tk.W, padx=5, pady=3)
        spin_frame = tk.Frame(frame, bd=1, relief=tk.SOLID)
        spin_frame.grid(row=1, column=1, padx=5, pady=3, sticky=tk.W)
        self.thread_var = tk.StringVar(value="3")
        tk.Button(spin_frame, text="\u2212", width=2, command=self._spin_down, relief=tk.FLAT, font=("", 10, "bold")).pack(side=tk.LEFT)
        tk.Label(spin_frame, textvariable=self.thread_var, width=3, anchor=tk.CENTER, font=("", 10, "bold")).pack(side=tk.LEFT, padx=2)
        tk.Button(spin_frame, text="+", width=2, command=self._spin_up, relief=tk.FLAT, font=("", 10, "bold")).pack(side=tk.LEFT)

        self.fetch_btn = tk.Button(frame, text="获取号码", command=self.toggle_fetch, bg="#2196F3", fg="white", width=10)
        self.fetch_btn.grid(row=1, column=2, padx=5, pady=3)
        frame.columnconfigure(1, weight=1)

    def _build_btn_frame(self):
        frame = tk.Frame(self.root)
        frame.pack(fill=tk.X, padx=10, pady=5)

        self.start_btn = tk.Button(frame, text="开始", command=self.start_task, bg="#4CAF50", fg="white", width=10)
        self.start_btn.pack(side=tk.LEFT, padx=5)

        self.pause_btn = tk.Button(frame, text="暂停(不可用)", command=self.toggle_pause, bg="#FF9800", fg="white", width=12, state=tk.DISABLED)
        self.pause_btn.pack(side=tk.LEFT, padx=5)

        self.stop_btn = tk.Button(frame, text="停止(不可用)", command=self.stop_task, bg="#f44336", fg="white", width=12, state=tk.DISABLED)
        self.stop_btn.pack(side=tk.LEFT, padx=5)

        tk.Button(frame, text="清空日志", command=self.clear_log, width=10).pack(side=tk.RIGHT, padx=5)

    def _build_status_frame(self):
        frame = tk.Frame(self.root)
        frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=2)

        # 电话号码队列（左）
        qframe = tk.LabelFrame(frame, text="电话号码队列", padx=5, pady=5)
        qframe.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 5))
        qtop = tk.Frame(qframe)
        qtop.pack(fill=tk.X)
        tk.Button(qtop, text="删除选中", command=self._delete_selected_phones, width=8).pack(side=tk.LEFT)
        tk.Button(qtop, text="清空", command=self.clear_phone_queue, width=6).pack(side=tk.RIGHT)

        self.queue_canvas = tk.Canvas(qframe, height=150)
        self.queue_scrollbar = tk.Scrollbar(qframe, orient=tk.VERTICAL, command=self.queue_canvas.yview)
        self.queue_check_frame = tk.Frame(self.queue_canvas)
        self.queue_check_frame.bind('<Configure>', lambda e: self.queue_canvas.configure(scrollregion=self.queue_canvas.bbox('all')))
        self.queue_canvas.create_window((0, 0), window=self.queue_check_frame, anchor=tk.NW)
        self.queue_canvas.configure(yscrollcommand=self.queue_scrollbar.set)
        self.queue_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.queue_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.queue_vars = {}

        # 账号状态（右）
        aframe = tk.LabelFrame(frame, text="账号状态", padx=5, pady=5)
        aframe.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=(5, 0))
        atop = tk.Frame(aframe)
        atop.pack(fill=tk.X)
        tk.Button(atop, text="清空", command=self.clear_accounts, width=6).pack(side=tk.RIGHT)

        columns = ('account', 'email', 'status')
        self.account_tree = ttk.Treeview(aframe, columns=columns, show='headings', height=8)
        self.account_tree.heading('account', text='编号')
        self.account_tree.heading('email', text='邮箱')
        self.account_tree.heading('status', text='状态')
        self.account_tree.column('account', width=80)
        self.account_tree.column('email', width=160)
        self.account_tree.column('status', width=120, anchor=tk.CENTER)
        self.account_tree.pack(fill=tk.BOTH, expand=True)
        ts = tk.Scrollbar(self.account_tree, orient=tk.VERTICAL, command=self.account_tree.yview)
        ts.pack(side=tk.RIGHT, fill=tk.Y)
        self.account_tree.configure(yscrollcommand=ts.set)

        frame.columnconfigure(0, weight=1)
        frame.columnconfigure(1, weight=1)

    def _build_progress_and_log(self):
        self.progress = ttk.Progressbar(self.root, mode='determinate')
        self.progress.pack(fill=tk.X, padx=10, pady=2)

        self.status_label = tk.Label(self.root, text="就绪")
        self.status_label.pack(padx=10, anchor=tk.W)

        lf = tk.LabelFrame(self.root, text="执行日志", padx=10, pady=10)
        lf.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        self.log_text = scrolledtext.ScrolledText(lf, wrap=tk.WORD, height=8)
        self.log_text.pack(fill=tk.BOTH, expand=True)
        self.log_text.config(state=tk.DISABLED)

    # ==================== 控件交互 ====================

    def select_file(self):
        file_path = filedialog.askopenfilename(
            title="选择 xlsx 文件",
            filetypes=[("Excel 文件", "*.xlsx"), ("所有文件", "*.*")]
        )
        if file_path:
            self.file_path = file_path
            self.file_label.config(text=os.path.basename(file_path))
            self.log(f"已选择文件: {file_path}")
            self.populate_accounts()

    def _spin_up(self):
        self.thread_var.set(str(min(int(self.thread_var.get()) + 1, 10)))

    def _spin_down(self):
        self.thread_var.set(str(max(int(self.thread_var.get()) - 1, 1)))

    # ==================== 获取号码 ====================

    def toggle_fetch(self):
        if self.fetch_stop_event.is_set():
            self.fetch_stop_event.clear()
            self._start_fetch_thread()
            self.fetch_btn.config(text="停止获取", bg="#f44336")
            self.log("开始获取电话号码")
        else:
            self.fetch_stop_event.set()
            self.fetch_btn.config(text="获取号码", bg="#2196F3")
            self.log("停止获取电话号码")

    def _start_fetch_thread(self):
        if self.fetch_thread and self.fetch_thread.is_alive():
            return

        def _fetch():
            asyncio.set_event_loop(asyncio.new_event_loop())
            while not self.stop_event.is_set() and not self.fetch_stop_event.is_set():
                try:
                    phones = self.get_phone()
                    self.log(f'获取号码数据：{phones}')
                    if phones:
                        with self.phone_queue_lock:
                            existing = set(self.phone_queue_list)
                        for phone in phones:
                            if phone['data'] in existing:
                                self.log(f"队列已存在该数据: {phone['data']}")
                                continue
                            existing.add(phone['data'])
                            self.phone_queue.put(phone)
                            with self.phone_queue_lock:
                                self.phone_queue_list.append(phone['data'])
                                self.phone_status[phone['data']] = '待处理'
                            self.log(f"加入队列: {phone['data']}")
                except Exception as e:
                    self.log(f"获取电话异常: {e}")
                for _ in range(20):
                    if self.stop_event.is_set() or self.fetch_stop_event.is_set():
                        return
                    time.sleep(0.5)
            self.log("电话号码获取已停止")

        self.fetch_thread = threading.Thread(target=_fetch, daemon=True)
        self.fetch_thread.start()

    def clear_phone_queue(self):
        self._reset_phone_state()
        self.log("已清空电话号码队列")

    def _delete_selected_phones(self):
        selected = set(p for p, v in list(self.queue_vars.items()) if v.get())
        if not selected:
            return
        removed, processing = set(), set()
        with self.phone_queue_lock:
            for p in selected:
                if self.phone_status.get(p) == '处理中':
                    processing.add(p)
                    self.queue_vars[p].set(False)
                else:
                    removed.add(p)
                    self.phone_queue_list.remove(p) if p in self.phone_queue_list else None
                    self.phone_status.pop(p, None)
                    self.queue_vars.pop(p, None)
        if processing:
            self.log(f"当前存在号码{','.join(processing)}正在执行，已取消勾选")
        if removed:
            self._drain_phone_queue(removed)
            self.log(f"已删除{len(removed)}个号码")

    def _drain_phone_queue(self, removed=None):
        items = []
        while not self.phone_queue.empty():
            try:
                items.append(self.phone_queue.get_nowait())
            except queue.Empty:
                break
        for item in items:
            if not removed or item['data'] not in removed:
                self.phone_queue.put(item)

    def _reset_phone_state(self):
        with self.phone_queue_lock:
            self.phone_queue_list.clear()
            self.phone_status.clear()
        self._drain_phone_queue()
        for w in self.queue_check_frame.winfo_children():
            w.destroy()
        self.queue_vars.clear()
        self._last_display = []

    # ==================== 暂停 / 日志 / 显示 ====================

    def toggle_pause(self):
        if self.is_paused:
            self.is_paused = False
            self.pause_event.clear()
            self.pause_btn.config(text="暂停")
            self.log("任务继续")
        else:
            self.is_paused = True
            self.pause_event.set()
            self.pause_btn.config(text="继续")
            self.log("任务暂停")

    def poll_queue_display(self):
        with self.phone_queue_lock:
            current = list(self.phone_queue_list[-100:])
            display = [(p, f"{p} [{self.phone_status.get(p, '')}]" if self.phone_status.get(p) else p) for p in current]
        if self._last_display != display:
            self._last_display = display
            for w in self.queue_check_frame.winfo_children():
                w.destroy()
            cur_phones = set()
            for phone, text in display:
                cur_phones.add(phone)
                if phone not in self.queue_vars:
                    self.queue_vars[phone] = tk.BooleanVar()
                tk.Checkbutton(self.queue_check_frame, text=text, variable=self.queue_vars[phone],
                               anchor=tk.W, font=("Consolas", 9), padx=2).pack(anchor=tk.W, fill=tk.X)
            for p in list(self.queue_vars.keys()):
                if p not in cur_phones:
                    del self.queue_vars[p]
        self.root.after(1000, self.poll_queue_display)

    def update_account_status(self, account_code, status):
        self._pending_status[account_code] = status
        if not self._status_update_scheduled:
            self._status_update_scheduled = True
            def _flush():
                self._status_update_scheduled = False
                pending, self._pending_status = self._pending_status, {}
                for item in self.account_tree.get_children():
                    v = self.account_tree.item(item, 'values')
                    if v and v[0] in pending:
                        self.account_tree.item(item, values=(v[0], v[1], pending.pop(v[0])))
            self.root.after(200, _flush)

    def _after(self, fn):
        self.root.after(0, fn)

    def populate_accounts(self):
        if not self.file_path:
            return
        for item in self.account_tree.get_children():
            self.account_tree.delete(item)
        data = ReadData.read_xlsx_row(self.file_path)
        if data:
            for row in data:
                email = row.get('邮箱', '')
                code = row.get('编号', '')
                code = f"{code}_{email}" if code else email
                self.account_tree.insert('', tk.END, values=(code, email, '等待'))
            self.log(f"已加载 {len(data)} 个账号")

    def clear_accounts(self):
        for item in self.account_tree.get_children():
            self.account_tree.delete(item)
        self.log("已清空账号列表")

    def log(self, message):
        logger.info(message)
        self.log_queue.put(message)

    def poll_log_queue(self):
        while not self.log_queue.empty():
            msg = self.log_queue.get_nowait()
            self.log_text.config(state=tk.NORMAL)
            self.log_text.insert(tk.END, f"{time.strftime('%H:%M:%S')} - {msg}\n")
            self.log_text.config(state=tk.DISABLED)
        if self.log_text.yview()[1] >= 0.99:
            self.log_text.see(tk.END)
        self.root.after(100, self.poll_log_queue)

    def clear_log(self):
        self.log_text.config(state=tk.NORMAL)
        self.log_text.delete(1.0, tk.END)
        self.log_text.config(state=tk.DISABLED)

    # ==================== 任务控制 ====================

    def start_task(self):
        if not self.file_path:
            messagebox.showwarning("提示", "请先选择账号文件")
            return
        if self.is_running:
            self.log("任务正在运行中")
            return

        self.is_running = True
        self.is_paused = False
        self.stop_event.clear()
        self.fetch_stop_event.clear()
        self.pause_event.clear()
        self.task_epoch += 1
        self._start_fetch_thread()

        self.start_btn.config(state=tk.DISABLED, text="运行中...", bg="#9E9E9E")
        self.fetch_btn.config(state=tk.NORMAL, text="停止获取", bg="#f44336")
        self.pause_btn.config(state=tk.NORMAL, text="暂停")
        self.stop_btn.config(state=tk.NORMAL, text="停止")

        self.log(f"配置信息: 并发: {min(int(self.thread_var.get()), 10)}个线程")
        self.populate_accounts()
        self._reset_phone_state()
        self._pending_status = {}
        self._status_update_scheduled = False

        t = threading.Thread(target=self.run_task, daemon=True)
        t.start()

    def stop_task(self):
        self.stop_event.set()
        self.pause_event.clear()
        self.log("正在停止任务...")
        self.start_btn.config(state=tk.NORMAL, text="开始", bg="#4CAF50")
        self.stop_btn.config(state=tk.DISABLED, text="停止(不可用)")
        self.pause_btn.config(state=tk.DISABLED, text="暂停(不可用)")

    def run_task(self):
        try:
            self.log("开始执行任务...")
            data = ReadData.read_xlsx_row(self.file_path)
            if not data:
                self.log("错误: 读取数据文件失败")
                self.reset_ui(); return

            self.log(f"共读取{len(data)}条账户数据")
            wb, ws, row_map = self._init_xlsx(data)

            lock = threading.Lock()
            xlsx_lock = threading.Lock()
            processed_count = [0]
            failed_count = [0]
            account_index = [0]
            max_workers = int(self.thread_var.get())
            self.log(f"启动{max_workers}个并发线程")

            def write_xlsx(account_code, status, phone=''):
                rid = row_map.get(account_code)
                if rid:
                    with xlsx_lock:
                        ws.cell(row=rid, column=6, value=status)
                        if phone:
                            ws.cell(row=rid, column=7, value=phone)
                        wb.save(self.file_path)

            def run_worker(user_id, epoch):
                asyncio.set_event_loop(asyncio.new_event_loop())
                while not self.stop_event.is_set() and self.task_epoch == epoch:
                    if self._pause_check(epoch): break

                    phone_success = False
                    try:
                        sp_wrapper = [SpecialPlayWright()]
                        account_code, login_ok = self._try_login(sp_wrapper, data, lock, account_index, write_xlsx, user_id, epoch)
                        sp = sp_wrapper[0]
                        if not login_ok:
                            continue
                        self.update_account_status(account_code, '登录成功')
                        phone_success, success_phone = self._handle_phone(sp_wrapper[0], user_id, account_code, write_xlsx)
                    except Exception as e:
                        self.log(f"[{user_id}] 异常: {str(e)}")
                    finally:
                        if sp_wrapper[0]:
                            try: sp_wrapper[0].delete_enviroment()
                            except: pass
                            try: sp_wrapper[0].close()
                            except: pass

                    with lock:
                        if phone_success: processed_count[0] += 1
                        else: failed_count[0] += 1
                    self._after(lambda p=processed_count[0], f=failed_count[0]: (
                        self.progress.__setitem__('value', p % 100),
                        self.status_label.config(text=f"电话成功: {p} | 电话失败: {f} | 总计: {p + f}")))

            epoch = self.task_epoch
            for i in range(max_workers):
                threading.Thread(target=run_worker, args=(i + 1, epoch), daemon=True).start()

            while not self.stop_event.is_set():
                self.stop_event.wait(1)

            self.log(f"任务停止成功: {processed_count[0]}, 失败: {failed_count[0]}")
        except Exception as e:
            self.log(f"错误: {str(e)}")
        finally:
            self.reset_ui()

    def _init_xlsx(self, data):
        wb = load_workbook(self.file_path)
        ws = wb.active
        row_map = {}
        for idx, row in enumerate(data, start=2):
            email = row.get('邮箱', '')
            code = row.get('编号', '')
            code = f"{code}_{email}" if code else email
            row_map[code] = idx
        return wb, ws, row_map

    def _pause_check(self, epoch):
        if self.stop_event.is_set() or self.task_epoch != epoch:
            return True
        while self.pause_event.is_set() and not self.stop_event.is_set():
            time.sleep(0.5)
        return self.stop_event.is_set()

    def _try_login(self, sp_wrapper, data, lock, account_index, write_xlsx, user_id, epoch):
        login_ok, attempts, account_code = False, 0, ''
        while not login_ok and not self.stop_event.is_set():
            if self._pause_check(epoch): break
            with lock:
                if account_index[0] >= len(data): break
                row = data[account_index[0] % len(data)]
                account_index[0] += 1
            if attempts > 0:
                sp_wrapper[0].delete_enviroment()
                sp_wrapper[0].close()
                time.sleep(2)
                sp_wrapper[0] = SpecialPlayWright()
            attempts += 1
            account_code = f"{row.get('编号', '')}_{row.get('邮箱', '')}" if row.get('编号') else row.get('邮箱', '')
            self.update_account_status(account_code, '登录中')
            self.log(f"[{user_id}] 尝试登录({attempts}): {account_code}")
            login_ok = self.login_account(sp_wrapper[0], account_code, row.get('邮箱', ''), row.get('密码', ''),
                                          row.get('邮箱验证地址', ''))
            if not login_ok:
                self.update_account_status(account_code, '登录失败')
                write_xlsx(account_code, '登录失败')
        if not login_ok:
            try:
                sp_wrapper[0].delete_enviroment()
                sp_wrapper[0].close()
            except: pass
        return account_code, login_ok

    def _handle_phone(self, sp, user_id, account_code, write_xlsx):
        """已登录成功，取号码进行修改+验证。无号码时阻塞等待。返回 (success, success_phone)"""
        while not self.stop_event.is_set():
            try:
                phone_item = self.phone_queue.get(timeout=1)
                break
            except queue.Empty:
                self.log(f"[{user_id}] 队列无可用电话号码，等待中...")
                self.stop_event.wait(3)
        if self.stop_event.is_set():
            return False, ''

        phone_code, phone_id = phone_item['data'], phone_item['id']
        with self.phone_queue_lock:
            self.phone_status[phone_code] = '处理中'
            if phone_code in self.phone_queue_list:
                self.phone_queue_list.remove(phone_code)

        for try_round in range(2):
            if try_round == 0:
                cur_code, cur_id = phone_code, phone_id
            else:
                try:
                    next_item = self.phone_queue.get(timeout=1)
                    cur_code, cur_id = next_item['data'], next_item['id']
                    with self.phone_queue_lock:
                        self.phone_status[cur_code] = '处理中'
                        if cur_code in self.phone_queue_list:
                            self.phone_queue_list.remove(cur_code)
                    self.log(f"[{user_id}] 替换号码: {cur_code}")
                except queue.Empty:
                    self.log(f"[{user_id}] 队列无可用替换号码")
                    break

            self.log(f"[{user_id}] {account_code} 登录成功, 处理号码: {cur_code}")
            if not self.modify_phone(sp, cur_code):
                self.log(f"[{user_id}] {cur_code} 修改电话失败")
                with self.phone_queue_lock:
                    self.phone_status[cur_code] = '修改失败'
                    self.phone_queue_list.append(cur_code)
                continue

            if self.verify_phone(sp, cur_id):
                with self.phone_queue_lock:
                    self.phone_status[cur_code] = '验证成功'
                    self.phone_queue_list.append(cur_code)
                self.log(f"[{user_id}] {cur_code} 验证成功")
                self.update_account_status(account_code, '登录成功，验证成功')
                write_xlsx(account_code, '登录成功，验证成功', cur_code)
                return True, cur_code

            with self.phone_queue_lock:
                self.phone_status[cur_code] = '验证失败'
                self.phone_queue_list.append(cur_code)
            self.log(f"[{user_id}] {cur_code} 验证失败")

        self.update_account_status(account_code, '登录成功，验证失败')
        write_xlsx(account_code, '登录成功，验证失败')
        return False, ''

    def reset_ui(self):
        self.is_running = False
        self.is_paused = False
        self.pause_event.clear()
        self.start_btn.config(state=tk.NORMAL, text="开始", bg="#4CAF50")
        self.pause_btn.config(state=tk.DISABLED, text="暂停(不可用)")
        self.stop_btn.config(state=tk.DISABLED, text="停止(不可用)")

    # ==================== 登录流程 ====================

    def login_account(self, sp, account_code, email, password, email_url):
        try:
            self.log(f"{account_code}: 开始登录...")
            enter_verify = False
            for roll in range(2):
                if self.stop_event.is_set(): return False
                self.log(f"{account_code}: 第{roll+1}次尝试输入账号密码")
                enter_verify = self._do_input(sp, email, password)
                if enter_verify: break

            if not enter_verify:
                self.log(f"{account_code}: 登录失败 - 未进入邮箱验证码页面")
                return False

            if not self.stop_event.is_set():
                self.stop_event.wait(10)

            verify_code = False
            for roll in range(10):
                if self.stop_event.is_set(): return False
                self.log(f"{account_code}: 第{roll+1}次尝试获取邮箱验证码")
                verify_code = self._do_get_email_code(sp, email_url)
                if verify_code: break

            if not verify_code:
                self.log(f"{account_code}: 登录失败 - 获取邮箱验证码失败")
                return False

            self.log(f"{account_code}: 验证码: {verify_code}")
            sp.slow_input('//input[@id="authCode"]', verify_code)
            sp.click('//input[@id="rememberMe"]')
            sp.click('//a[@id="authBtn"]', no_wait=True)
            if not sp.wait_for_selector('//ul[@class="tabUl flex"]', timeout=15*1000):
                self.log(f"{account_code}: 输入邮箱验证码后，未登录成功")
                return False

            self.log(f"{account_code}: 登录成功")
            return True
        except Exception as e:
            self.log(f"{account_code}: 登录异常: {str(e)}")
            return False

    def _do_input(self, sp, email, password):
        try:
            if not sp.goto('https://www.pokemoncenter-online.com/login/'):
                self.log('页面访问失败'); return False
            sp.wait_for_selector('//input[@type="email" and @id="login-form-email"]', timeout=10*1000)
            sp.slow_input('//input[@type="email" and @id="login-form-email"]', email)
            sp.slow_input('//input[@type="password" and @id="current-password"]', password)
            submi_ele = '//button[@type="submit" and @class="btn btn-block btn-primary"]'
            sp.click(submi_ele)
            if sp.get_count(submi_ele):
                sp.click(submi_ele)
            if sp.get_count(submi_ele):
                sp.click(submi_ele)

            if sp.wait_for_selector('//div[@class="comErrorBox" and contains(text(), "reCA")]', timeout=5*1000):
                sp.click(submi_ele, timeout=5*1000)
                time.sleep(5)
            return sp.wait_for_selector('//input[@id="authCode"]', timeout=3*1000)
        except Exception as e:
            self.log(f"输入账号密码流程异常: {str(e)}")
            return False

    def _do_get_email_code(self, sp, api_url):
        try:
            sp.new_goto(api_url, close=False)
            sel = '//div[@class="email-content"]/div[1]/div/p[3]'
            if sp.wait_for_selector(sel, timeout=10*1000):
                text = sp.get_text(sel)
                code = re.findall(r'\d+', text)[0]
                sp.switch_page('old')
                return code
            return False
        except Exception as e:
            self.log(f"获取邮箱验证码流程异常: {str(e)}")
            return False

    # ==================== 修改 & 验证电话 ====================

    def modify_phone(self, sp, phone):
        self.log('进行电话号码修改')
        try:
            sp.click('//div[@class="topBox"]//a[contains(@href, "mypage")]')
            time.sleep(3)
            sp.click('//a[@class="editProfile "]')
            sel = '//input[@class="js-validate telNumber form-control"]'
            if not sp.wait_for_selector(sel, timeout=10*1000):
                self.log('电话号码修改页面跳转失败'); return False
            sp.mouse_wheel(1000)
            sp.slow_input(sel, phone)
            confirm = '//button[@class="submitButton"]'
            sp.click(confirm)
            if not sp.wait_for_selector(confirm, timeout=6*1000):
                self.log('点击确认修改电话号码，跳转失败'); return False
            sp.mouse_wheel(1000)
            sp.click(confirm)
            sp.wait_for_selector('//ul[@class="linkList"]/li/a[@href="/"]', timeout=10*1000)
            return True
        except Exception as e:
            self.log(f"修改手机号流程异常：{str(e)}")
            return False

    def verify_phone(self, sp, phone_id):
        self.log('进行电话号码验证')
        tried = set()
        for attempt in range(2):
            try:
                sp.click('//div[@class="topBox"]//a[contains(@href, "mypage")]')
                time.sleep(5)
                sp.click('//form[@class="sendCertification-form"]/a')
                btn = '//a[@name="smsSubmit"]'
                if not sp.wait_for_selector(btn, timeout=15*1000):
                    self.log('电话号码验证页面跳转失败'); continue
                sp.click(btn)
                inp = '//input[@name="auth_code"]'
                if not sp.wait_for_selector(inp, timeout=15*1000):
                    self.log('未出现验证码输入框'); continue

                timeCount = 50
                self.log(f'等待{timeCount}s....')
                time.sleep(timeCount)
                code = self.get_phone_code(phone_id)
                if not code:
                    if attempt == 0:
                        self.log(f'{timeCount}秒内未获取到验证码，再次重新获取'); continue
                    else:
                        self.log(f'{timeCount}秒内未获取到验证码，尝试更换号码'); return False
                if code in tried:
                    self.log(f'验证码{code}已尝试过，换下一个号码'); return False
                tried.add(code)
                sp.slow_input(inp, code)
                sp.click('//button[@class="CertificationCodesubmit"]')
                if sp.wait_for_selector('//div[@class="comBtn"]', timeout=10*1000):
                    return True
                self.log(f'验证码{code}未通过')
            except Exception as e:
                self.log(f'验证电话号码流程异常：{e}')
                if attempt == 0: continue
        return False

    # ==================== API 调用 ====================

    def get_phone_code(self, phone_id):
        try:
            headers = {'Authorization': f'Bearer {self.get_token()}'}
            url = f'https://geyuehui.com/ajax/getYzm?id={phone_id}'
            r = requests.post(url=url, headers=headers, proxies={'http': '', 'https': ''})
            data = r.json()
            phone_code = data.get('yzm') or data.get('code') or data.get('data')
            if not phone_code:
                self.log(f'【{phone_id}】验证码接口返回异常: {r.text[:200]}')
                return False
            return str(phone_code)
        except Exception as e:
            self.log(f'【{phone_id}】电话验证码获取失败：{e}')
            return False

    def get_phone(self):
        try:
            url = 'https://geyuehui.com/simadmin/sqlb_data?page=1&limit=50'
            headers = {'Authorization': f'Bearer {self.get_token()}'}
            r = requests.get(url=url, headers=headers, proxies={'http': '', 'https': ''}).json()['data']
            return [{'time': i['time'], 'data': eval(i['data1'])['username'], 'id': i['id']} for i in r]
        except Exception as e:
            self.log(f'电话获取失败：{e}')
            return False

    def get_token(self):
        self.end_time = time.time()
        if self.end_time - self.start_time <= 110 * 60 and self.token:
            return self.token
        try:
            url = 'https://geyuehui.com/simlogin/login'
            r = requests.post(url=url, data={'username': 'admin', 'password': 'admin123'},
                              proxies={'http': '', 'https': ''})
            self.token = r.json()['token']
            self.start_time = time.time()
            return self.token
        except Exception as e:
            self.log(f'token获取异常，退出运行：{e}')
            self.stop_event.set()
            return False


if __name__ == '__main__':
    root = tk.Tk()
    app = App(root)
    root.mainloop()
