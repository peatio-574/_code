# -*- coding: utf-8 -*-
import sys
from pathlib import Path

sys.path.append(str(Path(__file__).parent.parent))

from PlayWright import Playwright_, logger
import os
import time
from openpyxl import load_workbook
import json
from xhs import deal_str, deal_date, get_page_comment
import requests

dirName = os.path.join(os.path.dirname(__file__), '数据')
os.makedirs(dirName, exist_ok=True)
config_file = os.path.join(os.path.dirname(__file__), 'config.ini')

host = 'https://www.xiaohongshu.com'


def login():
    """小红薯登录"""
    logger.info('=' * 80)
    logger.info('开始登录小红书....')
    ele = '//li/div/a//span[text()="我"]'
    key = 'login.xiaohongshu1'
    try:
        Playwright_.login(host, ele, key, file=config_file)
        logger.info('✓ 小红书登录成功')
        return True
    except Exception as e:
        logger.error(f'✗ 小红书登录失败：{e}')
        return False




def get_product_info(title, imageDir):
    """获取作品的详情"""
    try:
        error_ele = '//div[text()="访问频繁，请稍后再试"]'
        if Playwright_.get_count(error_ele):
            logger.error('❌️ 访问频繁，请稍后再试')
            exit()

        close_ele = '//div[contains(@class, "close-button")]'
        if Playwright_.get_count(close_ele):
            Playwright_.click(close_ele)

        # 正文
        content_ele = '//div[@class="desc"]/span/span[text()]'
        content_count = Playwright_.get_count(content_ele)
        content = ''
        for content_ in range(1, content_count + 1):
            content += Playwright_.get_text(f'({content_ele})[{content_}]')

        # 标签
        tag_ele = '//a[@class="tag"]'
        tag_count = Playwright_.get_count(tag_ele)
        tags = list()
        for tag_ in range(1, tag_count + 1):
            tag = Playwright_.get_text(f'({tag_ele})[{tag_}]')
            tags.append(tag)
        tags = '，'.join(tags)

        # 发布时间
        publish_time = Playwright_.get_text('//span[@class="date"]', timeout=5 * 1000)
        publish_time = deal_date(publish_time)

        # 点赞数
        like_ele = '//div[@class="left"]/span[1]/span[2]'
        like_count = Playwright_.get_text(like_ele) if Playwright_.get_count(like_ele) else '0'
        like_count = like_count if like_count != '点赞' else '0'

        # 评论数
        comment_count = Playwright_.get_text('//div[@class="left"]/span[3]/span')
        comment_count = comment_count if comment_count != '评论' else '0'

        # 作品图片
        pictures = list()
        picture_ele = '//meta[@name="og:url"]/preceding-sibling::meta[@name="og:image"]'
        picture_count = Playwright_.get_count(picture_ele)
        picture_count = min(picture_count, 3)
        for picture_ in range(1, picture_count + 1):
            picture = Playwright_.get_attribute(f'({picture_ele})[{picture_}]', 'content')
            pictures_date = requests.get(picture).content

            title_text = deal_str(title)
            pic = f'{title_text[:18]}_{publish_time[:10]}_{picture_}.jpg'
            pic_name = os.path.join(imageDir, pic)
            with open(pic_name, 'wb') as f:
                f.write(pictures_date)
            pictures.append(pic)
        pictures = ','.join(pictures)

        product_info = {
            'content': content,
            'tags': tags,
            'publish_time': publish_time,
            'like_count': like_count,
            'comment_count': comment_count,
            'pictures': pictures
        }
        return product_info
    except Exception as e:
        logger.error(f'获取作品详情失败：{e}')
        # time.sleep(1000)
        return dict()


def run(keyword):
    """运行单个博主的数据采集"""
    logger.info('\n' + '=' * 80)
    logger.info(f'开始处理博主：{keyword}')
    logger.info('=' * 80)

    if not login():
        logger.error('登录失败，终止执行')
        return None

    json_file = os.path.join(dirName, f'{keyword}.json')


    if not os.path.exists(json_file):
        logger.info(f'✗ 博主 {keyword}：博主不存在')
        return None

    with open(json_file, 'r', encoding='utf-8') as f:
        text_json = json.load(f)

    author_url = text_json['博主链接']
    Playwright_.goto(author_url)
    time.sleep(8)

    author_info = text_json['博主详情']

    userCode = author_info['user_code']
    userName = author_info['user_name']
    userName = deal_str(userName)
    userIp = author_info['user_ip']
    userDescription = author_info['user_description']
    userTag = author_info['user_tag']

    logger.info(f'博主信息：{userName} ({userCode})')
    logger.info(f'IP属地：{userIp}')


    products = text_json['作品'][:80]
    total = len(products)

    imageDir = text_json['图片保存目录']
    filename = text_json['信息保存目录']

    wb = load_workbook(filename)
    ws = wb.active


    exists_urls = []
    for row in range(2, ws.max_row + 1):
        value = ws.cell(row=row, column=14).value
        exists_urls.append(value)

    close_ele = '//div[@class="close close-mask-dark"]'

    for product_id, product_ in enumerate(products, start=1):
        logger.info(f'\n--- 当前关键词：{keyword}，博主：{userName}，处理第 {product_id}/{total} 个作品 ---')
        title, productUrl = product_
        if productUrl in exists_urls:
            logger.info(f'作品已存在，跳过')
            continue
        logger.info(f'作品标题：{title}')
        try:

            product_info = dict()
            for roll in range(5):
                Playwright_.click(f'//div/a[2][@href="{productUrl}"]')
                time.sleep(3)
                product_info = get_product_info(title, imageDir)
                if product_info:
                    logger.info(f'✓ 作品详情获取成功（尝试第{roll + 1}次）')
                    break
                logger.warning(f'作品详情获取失败，重试第{roll + 1}次...')
                Playwright_.click(close_ele)


            if not product_info:
                logger.error(f'✗ 作品 "{title}" 获取详情失败')
                if Playwright_.get_count(close_ele):
                    Playwright_.click(close_ele)
                continue

            logger.info(f'正在获取评论数据...')
            coments = get_page_comment()
            logger.info(f'✓ 获取到 {len(coments)} 条评论')

            row_data = [
                userCode,
                userName,
                userIp,
                userDescription,
                userTag,
                title,
                product_info['publish_time'],
                product_info['content'],
                product_info['tags'],
                product_info['like_count'],
                product_info['comment_count'],
                '\n'.join(coments),
                product_info['pictures'],
                productUrl
            ]
            ws.append(row_data)
            wb.save(filename)
            exists_urls.append(title)
            logger.info(f'✓ 数据已保存，当前共 {len(exists_urls)} 个作品')


        except Exception as e:
            logger.error(f'处理作品失败：{e}')
            Playwright_.page.mouse.click(60, 300, button='left', click_count=1)
            continue

        Playwright_.page.mouse.click(60, 300, button='left', click_count=1)



        # logger.info('向下滚动页面加载更多作品...')
        # Playwright_.page.keyboard.press('PageDown')
        # Playwright_.page.keyboard.press('PageDown')

    logger.info(f'\n✓ 博主 {userName} 数据采集完成，共 {len(exists_urls)} 个作品')
    if len(exists_urls) == len(products):
        os.remove(json_file)
    return True


if __name__ == '__main__':
    import pandas
    # from prepare import prepare

    data_file = os.path.join(os.path.dirname(__file__), '第一批200用户2026.6.3.xlsx')
    data_ids = pandas.read_excel(data_file, sheet_name=0)['user_id']
    # prepare(data_ids)

    logger.info('=' * 80)
    logger.info(f'开始批量处理 {len(data_ids)} 个博主')
    logger.info('=' * 80)

    success_count = 0
    fail_count = 0

    for idx, user_id in enumerate(data_ids, 1):
        logger.info('\n' + '=' * 80)
        logger.info(f'总体进度：[{idx}/{len(data_ids)}]')
        logger.info('=' * 80)

        try:
            result = run(user_id)
            if result:
                success_count += 1
            else:
                fail_count += 1
        except Exception as e:
            logger.error(f'✗ 博主 {user_id} 处理异常：{e}')
            fail_count += 1

    logger.info('\n' + '=' * 80)
    logger.info(f'全部处理完成！')
    logger.info(f'成功：{success_count} 个博主')
    logger.info(f'失败：{fail_count} 个博主')
    logger.info('=' * 80)
