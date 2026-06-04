# -*- coding: utf-8 -*-

"""
小红书封装函数
"""


import sys
from pathlib import Path

sys.path.append(str(Path(__file__).parent.parent))


from PlayWright import Playwright_, logger
import time
import requests
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment



config_file = os.path.join(os.path.dirname(__file__), 'config.ini')

host = 'https://www.xiaohongshu.com'

dirName = os.path.join(os.path.dirname(__file__), '数据')
os. makedirs(dirName, exist_ok=True)


def login():
    """小红薯登录"""
    logger.info('登录小红书....')
    ele = '//li/div/a//span[text()="我"]'
    key = 'login.xiaohongshu1'
    Playwright_.login(host, ele, key, file=config_file)
    logger.info('小红书登录成功')


def search(keyword):
    """搜索商品"""
    Playwright_.input('(//div[@class="textarea-wrapper"])[2]/textarea', keyword, enter=True)
    time.sleep(5)


def get_comment(comment_ele, comment_):
    """获取单条评论信息（不含回复信息），返回字典"""
    comment_str = ''
    try:
        # 评论内容
        comment_msg_ele = f'({comment_ele})[{comment_}]/div[1]//div[@class="content"][1]/span/span'
        comment_msg_count = Playwright_.get_count(comment_msg_ele)
        comment_text = ''
        for comment_msg_ in range(1, comment_msg_count + 1):
            comment_msg = Playwright_.get_text(f'({comment_msg_ele})[{comment_msg_}]')
            comment_text += comment_msg
        # 评论人
        comment_author = Playwright_.get_text(f'({comment_ele})[{comment_}]/div[1]//div[@class="author"]/a')

        # 评论时间
        comment_time = Playwright_.get_text(f'({comment_ele})[{comment_}]/div[1]//div[@class="date"]/span[1]')
        comment_time = deal_date(comment_time)
        # 评论地点
        comment_addr = Playwright_.get_text(f'({comment_ele})[{comment_}]/div[1]//div[@class="date"]/span[2]')
        comment_str = f'用户：{comment_author}|内容：{comment_text}|时间：{comment_time}|IP：{comment_addr}'
        return comment_str
    except Exception as e:
        logger.error(f'第{comment_}条评论信息获取失败：{e}')
        return comment_str


def get_page_comment():
    """获取20条评论"""
    comment_ele = '//div[@class="parent-comment"]'
    comments = []
    record_count = 0
    while len(comments) < 20:
        comment_count = Playwright_.get_count(comment_ele)
        for comment_ in range(1, comment_count + 1):
            # 获取评论信息
            comment_info = get_comment(comment_ele, comment_)

            # 跳过已存在的评论或获取失败的评论
            if comment_info is None:
                continue
            if comment_info not in comments:
                logger.info(comment_info)
                comments.append(comment_info)

        if record_count == len(comments):
            # logger.info('当前没有新评论了，已获取所有评论')
            break

        record_count = len(comments)
        for i in range(5):
            Playwright_.page.keyboard.press('PageDown')
    return comments


def get_author_url():
    """获取博主url"""
    author_ele = '//div[@class="onebox"]/a'
    if Playwright_.get_count(author_ele) == 0:
        return ''
    author_url = Playwright_.get_attribute(author_ele, 'href')
    author_url = host + author_url
    return author_url


def get_author_info():
    """获取博主信息"""
    user_name = Playwright_.get_text('//div[@class="user-name"]')
    user_name = user_name.replace(r'\\', '').replace(r'\n', '').replace(r'\r', '').replace(r'\t', '').strip()

    user_code = Playwright_.get_text('//span[@class="user-redId"]')

    user_ip_ele = '//span[@class="user-IP"]'
    user_ip = Playwright_.get_text(user_ip_ele) if Playwright_.get_count(user_ip_ele) else ''
    user_ip = user_ip.strip()

    user_description_ele = '//div[@class="user-desc"]'
    user_description = Playwright_.get_text(user_description_ele) if Playwright_.get_count(user_description_ele) else ''

    tag_ele = '//div[@class="tag-item"]/div[not(contains(@class,"gender"))]'
    tag_count = Playwright_.get_count(tag_ele)
    user_tag = ''
    if tag_count:
        for tag_ in range(1, Playwright_.get_count(tag_ele) + 1):
            user_tag += Playwright_.get_text(f'({tag_ele})[{tag_}]') + ','
        user_tag = user_tag[:-1]
    return {
        'user_name': user_name,
        'user_code': user_code,
        'user_ip': user_ip,
        'user_description': user_description,
        'user_tag': user_tag,
    }


def get_author_title_urls():
    urls = []
    record_count = 0
    while len(urls) < 100:
        time.sleep(3)
        product_ele = '//section[@data-height]'
        product_count = Playwright_.get_count(product_ele)
        if product_count == 0:
            break


        for product_ in range(1, product_count + 1):
            if len(urls) == 100:
                break
            title = Playwright_.get_text(f'({product_ele})[{product_}]/div/div/a[1]/span')
            product_url = Playwright_.get_attribute(f'({product_ele})[{product_}]/div/a[2]', 'href')
            product_url = host + product_url
            if [title, product_url] not in urls:
                urls.append([title, product_url])
        if record_count == len(urls):
            # logger.info('当前没有新作品了，已获取所有作品链接')
            break

        record_count = len(urls)
        if len(urls) >= 100:
            break
        Playwright_.page.keyboard.press('PageDown')
        Playwright_.page.keyboard.press('PageDown')
        Playwright_.page.keyboard.press('PageDown')
    return urls


def deal_date(date):
    """处理发布、评论、回复时间"""
    today = time.strftime('%Y-%m-%d', time.localtime())
    yesterday = time.strftime('%Y-%m-%d', time.localtime(time.time() - 86400))
    for i in range(1, 20):
        date = date.replace(f'{i}天前', time.strftime('%Y-%m-%d', time.localtime(time.time() - i * 86400)))
        date = date.replace(f'前{i}天', time.strftime('%Y-%m-%d', time.localtime(time.time() - i * 86400)))
    date = date.replace('昨天', yesterday)
    date = date.replace('今天', today)
    return date

def deal_str(title_text):
    import re
    pattern = r'[^\u4e00-\u9fa5a-zA-Z0-9_\-=\+\.,，。&~！! ]'
    title_text = re.sub(pattern, '', title_text)
    return title_text


def get_product_info(url, title, currentDir):
    """获取作品的详情"""
    try:
        Playwright_.goto(url)
        time.sleep(3)

        # 正文
        content_ele = '//div[@class="desc"]/span/span'
        content_count = Playwright_.get_count(content_ele)
        content = ''
        for content_ in range(1, content_count + 1):
            content += Playwright_.get_text(f'({content_ele})[{content_}]')

        # 标签
        tag_ele = '//a[@class="tag"]'
        tag_count = Playwright_.get_count(tag_ele)
        tags = list()
        for tag_ in range(1, tag_count + 1):
            tag = Playwright_.get_text(f'({tag_ele})[{tag_}]')
            tags.append(tag)
        tags = '，'.join(tags)

        # 发布时间
        publish_time = Playwright_.get_text('//span[@class="date"]')
        publish_time = deal_date(publish_time)

        # 点赞数
        like_count = Playwright_.get_text('//div[@class="left"]/span[1]/span[2]')
        like_count = like_count if like_count != '点赞' else '0'

        # 评论数
        comment_count = Playwright_.get_text('//div[@class="left"]/span[3]/span')
        comment_count = comment_count if comment_count != '评论' else '0'

        # 作品图片
        pictures = list()
        picture_ele = '//meta[@name="og:url"]/preceding-sibling::meta[@name="og:image"]'
        picture_count = Playwright_.get_count(picture_ele)
        picture_count = min(picture_count, 5)
        for picture_ in range(1, picture_count + 1):
            picture = Playwright_.get_attribute(f'({picture_ele})[{picture_}]', 'content')
            pictures_date = requests.get(picture).content

            title_text = deal_str(title)
            pic = f'{title_text[:18]}_{publish_time[:10]}_{picture_}.jpg'
            pic_name = os.path.join(currentDir, pic)
            with open(pic_name, 'wb') as f:
                f.write(pictures_date)
            pictures.append(pic)
        pictures = ','.join(pictures)

        product_info = {
            'content': content,
            'tags': tags,
            'publish_time': publish_time,
            'like_count': like_count,
            'comment_count': comment_count,
            'pictures': pictures
        }
        return product_info
    except Exception as e:
        logger.error(f'获取作品详情失败：{e}')
        return dict()


def run(keyword='909030373'):
    search(keyword)
    author_url = get_author_url()
    if not author_url:
        logger.info(f'{keyword}：作者不存在')
        return False
    logger.info(f'作者链接：{author_url}')
    Playwright_.goto(author_url)
    author_info = get_author_info()
    logger.info(f'作者信息：{author_info}')
    urls = get_author_title_urls()
    logger.info(f'共{len(urls)}条作品链接')

    userCode = author_info['user_code']
    userName = author_info['user_name']
    uerIp = author_info['user_ip']
    userDescription = author_info['user_description']
    userTag = author_info['user_tag']

    currentDate = time.strftime('%Y-%m-%d', time.localtime())
    currentDir = os.path.join(dirName, f'{userCode}_{userName}_{currentDate}')
    os.makedirs(currentDir, exist_ok=True)

    imageDir = os.path.join(currentDir, 'images')
    os.makedirs(imageDir, exist_ok=True)

    filename = os.path.join(currentDir, f'{userCode}_{userName}_{currentDate}.xlsx')

    if os.path.exists(filename):
        # 如果文件存在，加载现有工作簿
        from openpyxl import load_workbook
        wb = load_workbook(filename)
        ws = wb.active
    else:
        headers = ['用户ID', '用户名称', 'IP属地', '用户简介', '用户标签',
                   '笔记标题', '编辑时间', '正文文本', '标签', '点赞数',
                   '评论数', '评论列表', '笔记图片']
        wb = Workbook()
        ws = wb.active
        ws.title = '小红书笔记数据'
        ws.append(headers)

    # 设置表头样式
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    start_id = 0
    if keyword == '26567723394':
        start_id = 41
    for idx_, title_url in enumerate(urls[start_id:], start=start_id+1):
        title, url = title_url
        logger.info('\n')
        logger.info('='*80)
        logger.info(f'共{len(urls)}条链接，第{idx_}条作品链接：{url}')
        product_info = get_product_info(url, title, imageDir)
        logger.info(f'作品详情：{product_info}')
        coments = get_page_comment()

        row_data = [
            userCode,
            userName,
            uerIp,
            userDescription,
            userTag,
            title,
            product_info['publish_time'],
            product_info['content'],
            product_info['tags'],
            product_info['like_count'],
            product_info['comment_count'],
            '\n'.join(coments),
            product_info['pictures']
        ]
        ws.append(row_data)
        wb.save(filename)
        logger.info(f'保存数据成功：{row_data}')
        logger.info('等待20秒')
        time.sleep(20)

if __name__ == '__main__':

    string = """
26567723394
261899829
95189528582
KIRAjia1125
9746420503
95063864829
5609867723
664058835
6739556701
11520351332
27470970266
1804669058
388233831
630301923
8012070792
9567548993
1119387905
95690338543
2831365419
6808323931
267810408
631136658
9502256244
"""
    userids = string.strip().split('\n')

    for user_id_id in userids:
        login()
        run(user_id_id)