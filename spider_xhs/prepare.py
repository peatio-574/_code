# -*- coding: utf-8 -*-
import sys
from pathlib import Path

sys.path.append(str(Path(__file__).parent.parent))

from PlayWright import Playwright_, logger
import os
import time
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
import json
from xhs import deal_str

dirName = os.path.join(os.path.dirname(__file__), '数据')
os. makedirs(dirName, exist_ok=True)
config_file = os.path.join(os.path.dirname(__file__), 'config.ini')


host = 'https://www.xiaohongshu.com'


def login():
    """小红薯登录"""
    logger.info('登录小红书....')
    ele = '//li/div/a//span[text()="我"]'
    key = 'login.xiaohongshu1'
    Playwright_.login(host, ele, key, file=config_file)
    logger.info('小红书登录成功')


def search(keyword):
    """搜索商品"""
    keyword = str(keyword)
    input_ele = '(//input[@placeholder="搜索或输入任何问题"])[2]'
    if Playwright_.get_count(input_ele):
        Playwright_.input(input_ele, keyword, enter=True)
    input_ele = '(//div[@class="textarea-wrapper"])[2]/textarea'
    if Playwright_.get_count(input_ele):
        Playwright_.input(input_ele, keyword, enter=True)
    time.sleep(5)


def get_author_url():
    """获取博主url"""
    author_ele = '//div[@class="onebox"]/a'
    if Playwright_.get_count(author_ele) == 0:
        return ''
    author_url = Playwright_.get_attribute(author_ele, 'href')
    author_url = host + author_url
    return author_url


def get_author_info():
    """获取博主信息"""
    user_name = Playwright_.get_text('//div[@class="user-name"]')
    user_name = user_name.replace(r'\\', '').replace(r'\n', '').replace(r'\r', '').replace(r'\t', '').strip()

    user_code = Playwright_.get_text('//span[@class="user-redId"]')

    user_ip_ele = '//span[@class="user-IP"]'
    user_ip = Playwright_.get_text(user_ip_ele) if Playwright_.get_count(user_ip_ele) else ''
    user_ip = user_ip.strip()

    user_description_ele = '//div[@class="user-desc"]'
    user_description = Playwright_.get_text(user_description_ele) if Playwright_.get_count(user_description_ele) else ''

    tag_ele = '//div[@class="tag-item"]/div[not(contains(@class,"gender"))]'
    tag_count = Playwright_.get_count(tag_ele)
    user_tag = ''
    if tag_count:
        for tag_ in range(1, Playwright_.get_count(tag_ele) + 1):
            user_tag += Playwright_.get_text(f'({tag_ele})[{tag_}]') + ','
        user_tag = user_tag[:-1]
    return {
        'user_name': user_name,
        'user_code': user_code,
        'user_ip': user_ip,
        'user_description': user_description,
        'user_tag': user_tag,
    }


def get_author_title_urls():
    urls = []
    record_count = 0
    while len(urls) < 80:
        time.sleep(3)
        product_ele = '//section[@data-height]'
        product_count = Playwright_.get_count(product_ele)
        if product_count == 0:
            break


        for product_ in range(1, product_count + 1):
            if len(urls) == 100:
                break
            title = Playwright_.get_text(f'({product_ele})[{product_}]/div/div/a[1]/span')
            product_url = Playwright_.get_attribute(f'({product_ele})[{product_}]/div/a[2]', 'href')
            product_url = host + product_url
            if [title, product_url] not in urls:
                urls.append([title, product_url])
        if record_count == len(urls):
            # logger.info('当前没有新作品了，已获取所有作品链接')
            break

        record_count = len(urls)
        if len(urls) >= 80:
            break
        Playwright_.page.keyboard.press('PageDown')
        Playwright_.page.keyboard.press('PageDown')
        Playwright_.page.keyboard.press('PageDown')
    return urls


def prepare(keywords):
    login()
    for keyword in keywords:
        logger.info(f'开始处理：{keyword}')
        search(keyword)
        author_url = get_author_url()
        if not author_url:
            logger.info(f'{keyword}：作者不存在')
            continue
        Playwright_.goto(author_url)
        author_info = get_author_info()
        product_info = get_author_title_urls()

        userCode = author_info['user_code']
        userName = author_info['user_name']
        userName = deal_str(userName)

        currentDate = time.strftime('%Y-%m-%d', time.localtime())
        currentDir = os.path.join(dirName, f'{userCode}_{userName}_{currentDate}')
        # if os.path.exists(currentDir):
        #     logger.info(f'{keyword}：已处理过')
        #     continue
        os.makedirs(currentDir, exist_ok=True)

        imageDir = os.path.join(currentDir, 'images')
        os.makedirs(imageDir, exist_ok=True)

        filename = os.path.join(currentDir, f'{userCode}_{userName}_{currentDate}.xlsx')

        if os.path.exists(filename):
            # 如果文件存在，加载现有工作簿
            wb = load_workbook(filename)
            ws = wb.active
        else:
            headers = ['用户ID', '用户名称', 'IP属地', '用户简介', '用户标签',
                       '笔记标题', '编辑时间', '正文文本', '标签', '点赞数',
                       '评论数', '评论列表', '笔记图片']
            wb = Workbook()
            ws = wb.active
            ws.title = '小红书笔记数据'
            ws.append(headers)
        wb.save(filename)

        # 设置表头样式
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')


        infoText = os.path.join(dirName, f'{keyword}.json')
        info_dict = {
            '关键词': keyword,
            '博主': author_info['user_name'],
            '博主id': author_info['user_code'],
            '博主链接': author_url,
            '博主详情': author_info,
            '作品': product_info,
            '图片保存目录': imageDir,
            '信息保存目录': filename,
        }

        with open(infoText, 'w', encoding='utf-8') as f:
            json.dump(info_dict, f, ensure_ascii=False, indent=4)
        time.sleep(20)



if __name__ == '__main__':
    import pandas
    data_file = './第一批200用户2026.6.3.xlsx'
    data_ids = pandas.read_excel(data_file, sheet_name=0)['user_id']
    prepare(data_ids)
