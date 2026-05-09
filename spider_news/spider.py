# -*- coding: utf-8 -*-
import sys
from pathlib import Path

# 把项目根目录 D:\robot 加入Python路径
sys.path.append(str(Path(__file__).parent.parent))


from PlayWright import Playwright_, logger
import time
from ReadFile import ReadData
import os
import openpyxl

def get_info(url):
    try:
        Playwright_.goto(url)
        time.sleep(3)
        video_location = '//div[@aria-label="Start playback"]'
        video_count = Playwright_.get_count(video_location)
        video = 1 if video_count != 0 else video_count
        text_location = '//div[@class="text  en" or @class="content"]/p'
        content = str()
        text_count = Playwright_.get_count(text_location)
        for text_ in range(1, text_count+1):
            text = Playwright_.get_text(f'({text_location})[{text_}]')
            content += text + '\n'
        content = content.strip('\n')
        return content, str(video)
    except Exception as e:
        logger.error(f'❌ {url} 爬取失败: {e}')
        return '爬取失败', '爬取失败'


def extract_links_from_data(data_dict, sheet_name):
    """从字典数据中提取所有URL链接"""
    links = []
    for col_name, values in data_dict.items():
        for row_idx, value in enumerate(values, start=2):
            if value and isinstance(value, str) and (value.startswith('http://') or value.startswith('https://')):
                links.append({
                    'sheet': sheet_name,
                    'row': row_idx,
                    'column': col_name,
                    'url': value
                })
    return links


def append_to_original_file(file_path, sheet_name, row_num, content, video):
    """在原始文件的对应行追加爬取结果"""
    wb = openpyxl.load_workbook(file_path)

    if sheet_name not in wb.sheetnames:
        logger.error(f'警告：Sheet "{sheet_name}" 不存在')
        return False

    ws = wb[sheet_name]


    ws.cell(row=row_num, column=4, value=content)
    ws.cell(row=row_num, column=5, value=video)

    wb.save(file_path)
    return True


def main():
    excel_files = ['南方国家.xlsx']

    for excel_file in excel_files:
        if not os.path.exists(excel_file):
            logger.info(f'{excel_file} 文件不存在，跳过')
            continue

        logger.info(f'\n开始处理文件：{excel_file}')

        wb = openpyxl.load_workbook(excel_file)
        all_links = []

        for sheet_name in wb.sheetnames:
            data = ReadData.read_xlsx_col(excel_file, sheetname=sheet_name)
            if data:
                links = extract_links_from_data(data, sheet_name)
                all_links.extend(links)
                logger.info(f'  Sheet "{sheet_name}" 找到 {len(links)} 个链接')

        logger.info(f'总共找到 {len(all_links)} 个链接\n')

        for idx, link_info in enumerate(all_links, 1):
            try:
                logger.info(f'[{idx}/{len(all_links)}] 正在爬取：{link_info["url"]}')
                content, video = get_info(link_info['url'])

                append_to_original_file(
                    excel_file,
                    link_info['sheet'],
                    link_info['row'],
                    content,
                    video
                )

                logger.info(f'✓ 爬取成功，已追加到第 {link_info["row"]} 行， 视频：{video}')

            except Exception as e:
                logger.error(f'✗ 爬取失败：{e}')
                try:
                    append_to_original_file(
                        excel_file,
                        link_info['sheet'],
                        link_info['row'],
                        f'错误：{str(e)}',
                        '0'
                    )
                except Exception as e:
                    logger.error(f'✗ 爬取失败：{e}')

            time.sleep(2)

    logger.info(f'\n{"=" * 50}')
    logger.info(f'全部完成！结果已追加到原始文件')
    logger.info(f'{"=" * 50}')


if __name__ == '__main__':
    main()


