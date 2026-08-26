# coding='utf-8'
import sys

import ReadFile
import os

if getattr(sys, 'frozen', False):
    bundleDir = sys._MEIPASS
    baseDir = os.path.dirname(sys.executable)
else:
    bundleDir = os.path.dirname(os.path.abspath(__file__))
    baseDir = bundleDir

sys.path.insert(0, bundleDir)


from newPlayWright import PlayWright, logger
from openpyxl import load_workbook

host = 'https://www.xiaohongshu.com'
config_file = os.path.join(baseDir, 'config.ini')

xlsx_path = os.path.join(baseDir, '用户昵称.xlsx')
existing_names_list = ReadFile.ReadData.read_xlsx_col(xlsx_path)['昵称']
existing_names_set = set(existing_names_list)
wb = load_workbook(xlsx_path)
ws = wb.active

def write():
    ws.cell(row=1, column=4).value = '去重后昵称'
    for rowId, value in enumerate(list(dict.fromkeys(existing_names_set)), start=2):
        ws.cell(row=rowId, column=4).value = value
    wb.save(xlsx_path)

def login():
    """小红薯登录"""
    logger.info('登录小红书....')
    ele = '//li/div/a//span[text()="我"]'
    key = 'login.xiaohongshu1'
    PlayWright.login(host, ele, key, file=config_file)
    logger.info('小红书登录成功')

def get_names():
    global existing_names_set, existing_names_list
    rowsEle = '//div[@class="user-info"]/a'
    rowCount = PlayWright.get_count(rowsEle)

    flag = False
    start = max(1, rowCount - 49)
    count = 0
    for rowId in range(start, rowCount + 1):
        name = PlayWright.get_text(f'({rowsEle})[{rowId}]')
        if name not in existing_names_set:
            ws.append([len(existing_names_set) + 1, name])
            count += 1
            # existing_names_set.add(name)
            existing_names_list.append(name)
            existing_names_set = set(existing_names_list[-50:])
            flag = True
    if flag:
        logger.info(f'新增 {count} 条昵称，共计 {len(existing_names_list)} 条')
    else:
        logger.info(f'当前无可写入数据，当前共计 {len(existing_names_list)} 条')


def roll_():
    for i in range(355):
        PlayWright.mouse_wheel(800)

def main():
    login()
    PlayWright.goto('https://www.xiaohongshu.com/notification')
    PlayWright.click('//span[text()="新增关注"]')
    roll_()

    roll = 0
    while True:
        get_names()
        PlayWright.mouse_wheel(800)
        roll += 1
        if roll % 10 == 0:
            wb.save(xlsx_path)
            logger.info(f'循环{roll}次，已保存')



if __name__ == '__main__':
    main()
    write()