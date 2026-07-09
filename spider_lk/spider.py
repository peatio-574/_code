# -*- coding: utf-8 -*-
"""
抖音快客 + 麦客 数据采集 & 定时存储到 Excel

每30分钟执行一次:
  1. 登录抖音快客 → 抓取线索数据
  2. 登录麦客 → 抓取成都/济南/广州的表单数据
  3. 写入 数据.xlsx（5个sheet: 汇总/深圳/成都/济南/广州）
  4. 写入前判断手机号是否已存在于汇总表，存在则跳过
"""
import sys
from pathlib import Path


sys.path.append(str(Path(__file__).parent.parent))

from PlayWright import Playwright_, logger, get_config_value, write_config_value
import time
import os
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment
import requests

# PyInstaller 打包后 __file__ 指向只读临时目录，
# 配置文件和数据文件统一放到 exe 所在目录，确保可读写
if getattr(sys, 'frozen', False):
    _BASE_DIR = os.path.dirname(sys.executable)
else:
    _BASE_DIR = os.path.dirname(__file__)
configFile = os.path.join(_BASE_DIR, 'config.ini')
fileName = os.path.join(_BASE_DIR, '数据.xlsx')

HEADERS = ['数据来源', '姓名', '电话', '城市', '日期']
SHEETS = ['汇总', '深圳', '成都', '济南', '广州']
INTERVAL_MINUTES = 30  # 定时采集间隔（分钟）

# ============================================================================
# Excel 初始化
# ============================================================================
def init_excel():
    """创建 Excel 文件（含5个sheet + 表头），如果已存在则跳过"""
    if os.path.exists(fileName):
        wb = load_workbook(fileName)
        # 确保5个sheet都存在
        for sheet_name in SHEETS:
            if sheet_name not in wb.sheetnames:
                ws = wb.create_sheet(title=sheet_name)
                ws.append(HEADERS)
                _style_header(ws)
        wb.save(fileName)
        logger.info(f'Excel 已就绪: {fileName} (sheets={wb.sheetnames})')
        return wb
    else:
        wb = Workbook()
        # 删除默认 sheet
        wb.remove(wb.active)
        for sheet_name in SHEETS:
            ws = wb.create_sheet(title=sheet_name)
            ws.append(HEADERS)
            _style_header(ws)
        wb.save(fileName)
        logger.info(f'Excel 已创建: {fileName} (sheets={SHEETS})')
        return wb


def _style_header(ws):
    """设置表头样式"""
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')


def load_existing_phones(wb) -> set:
    """
    加载汇总表中已有的电话号码（用于去重）

    选择电话号码作为去重键的原因:
      - 每条线索的电话号码是唯一的业务标识
      - O(1) 查找（set），比逐行比对快
    """
    phones = set()
    if '汇总' in wb.sheetnames:
        ws = wb['汇总']
        for row in range(2, ws.max_row + 1):
            phone = ws.cell(row=row, column=3).value  # 电话在第3列
            if phone:
                phones.add(str(phone).strip())
    return phones


def save_data_to_sheets(wb, datas: list, existing_phones: set):
    """
    将数据写入 Excel

    去重逻辑:
      每条数据以「电话号码」为唯一标识。
      如果汇总表中已存在该号码 → 整条跳过
      如果汇总表中不存在 → 写入汇总表 + 对应城市表

    参数:
      wb:             已加载的 openpyxl Workbook
      datas:          [{数据来源, 姓名, 电话, 城市, 日期}, ...]
      existing_phones: 已有的电话号码集合
    """
    new_count = 0
    skip_count = 0

    for data in datas:
        phone = str(data['phone']).strip()
        city = data['city']
        source = data['data_source']
        name = data['username']
        date = data['date']

        # ---- 去重判断: 手机号是否已存在 ----
        if phone in existing_phones:
            skip_count += 1
            continue

        # ---- 标记为已存在 ----
        existing_phones.add(phone)

        # ---- 写入汇总表 ----
        ws_summary = wb['汇总']
        ws_summary.append([source, name, phone, city, date])
        if '深圳' not in city:
            add_info(city, name, phone)

        # ---- 写入对应城市表 ----
        if city in wb.sheetnames:
            ws_city = wb[city]
            ws_city.append([source, name, phone, city, date])

        new_count += 1

    if new_count > 0 or skip_count > 0:
        logger.info(f'写入数据: 新增{new_count}条, 跳过{skip_count}条(已存在)')


# ============================================================================
# 登录 & 数据采集
# ============================================================================
def lk_login():
    """抖音快客登录"""
    logger.info('=' * 80)
    logger.info('开始登录抖音快客....')
    url = 'https://life.douyin.com/p/liteapp/leads_life/sales/list'
    ele = '//div[contains(@class,"Profile-index-module__title--")]'
    key = 'login.lk_cookie'
    try:
        Playwright_.login(url, ele, key, file=configFile)
        logger.info('✓ 抖音快客登录成功')
        return True
    except Exception as e:
        logger.error(f'✗ 抖音快客登录失败：{e}')
        return False


def get_lk_data():
    """获取抖音快客数据"""
    source = '抖音快客'
    ele = '//tbody/tr[1]/td[3]/div/div[1]/span'
    if '**' in Playwright_.get_text(ele):
        logger.info('抖音快客号码处于隐藏状态，进行点击')
        Playwright_.click('(//tbody/tr[1]/td[3]/div/div[1]/*)[2]')
        time.sleep(2)
        Playwright_.click('//div[@class="cursor-pointer mr-1 flex items-center"]')
        time.sleep(5)
    row_ele = '//tbody/tr'
    row_count = Playwright_.get_count(row_ele)
    datas = []
    for i in range(1, row_count + 1):
        username = Playwright_.get_text(f'{row_ele}[{i}]/td[2]//a//div')
        phone = Playwright_.get_text(f'{row_ele}[{i}]/td[3]/div/div[1]/span')
        city = Playwright_.get_text(f'{row_ele}[{i}]/td[3]/div/div[2]//div')

        city = city.split('：')[1]
        # 城市映射: 山东→济南, 四川→成都, 广州→广州, 其他→深圳
        if '山东' in city:
            city = '济南'
        elif '四川' in city:
            city = '成都'
        elif '广州' in city:
            city = '广州'
        else:
            city = '深圳'

        date = Playwright_.get_text(f'{row_ele}[{i}]/td[6]')
        row_info = {
            'data_source': source,
            'username': username,
            'phone': phone,
            'city': city,
            'date': date,
        }
        datas.append(row_info)
    logger.info(f'抖音快客获取到 {len(datas)} 条数据')
    return datas


def mk_login():
    """麦客登录"""
    logger.info('=' * 80)
    logger.info('开始登录麦客....')
    url = 'https://mikeauth.com/login.php?prd=1&rg=1&d=form.php#/submit?id=200175265'
    ele = '//div[@class="h_userInfos"]'
    key = 'login.mk_cookie'
    try:
        Playwright_.login(url, ele, key, file=configFile)
        logger.info('✓ 麦客登录成功')
        close_ele = '//a[@class="pop-taCloseBtn iconfont"]'
        if Playwright_.get_count(close_ele):
            Playwright_.click(close_ele)
        return True
    except Exception as e:
        logger.error(f'✗ 麦客登录失败：{e}')
        return False


def get_mk_data():
    """获取麦客数据"""
    source = '麦客'
    ids = {'成都': 200175265, '济南': 200174417, '广州': 200174399}
    datas = []
    for city, code in ids.items():
        url = f'https://cn.mikecrm.com/form.php#/submit?id={code}'
        Playwright_.goto(url)
        time.sleep(3)
        row_ele = '//div[@class="mk_cdl_container"]/div'
        row_count = Playwright_.get_count(row_ele)
        for i in range(1, row_count + 1):
            username = Playwright_.get_text(f'{row_ele}[{i}]//a[@class="fbl_ctName"]')
            phone = Playwright_.get_text(f'{row_ele}[{i}]//div[@class="fbl_itemMobileIn"]/span')
            date = Playwright_.get_text(f'({row_ele}[{i}]//span[@class="fbl_sysInfo"])[2]')
            row_info = {
                'data_source': source,
                'username': username,
                'phone': phone,
                'city': city,
                'date': date,
            }
            datas.append(row_info)
    logger.info(f'麦客获取到 {len(datas)} 条数据')
    return datas

def bfx_login():
    """佰分象登录"""
    logger.info('=' * 80)
    logger.info('开始登录佰分象....')
    cookie = get_config_value('login', 'bfx_cookie', file=configFile)
    if cookie:
        Playwright_.add_cookie(eval(cookie))
    url = 'https://robot.jsybtx.com/#/workbench/auto-outbound'
    Playwright_.goto(url)
    time.sleep(8)
    login_ele = '//input[@placeholder="邮箱/手机号"]'
    if Playwright_.get_count(login_ele):
        Playwright_.input(login_ele, '17751137044@qq.com')
        Playwright_.input('//input[@placeholder="密码"]', 'P4XhuoUA')
        Playwright_.click('//button[@type="button"]/span[text()="登 录"]')
        time.sleep(8)
    if not Playwright_.get_count('//div[@class="menu-wrap collapsed-menu"]'):
        logger.error('佰分象登录失败')
        return False
    else:
        logger.info('佰分象登录成功')
        cookie_list = Playwright_.context.cookies()
        api_cookie = "; ".join([f"{cookie['name']}={cookie['value']}" for cookie in cookie_list])
        cookies = {'bfx_cookie': str(cookie_list), 'bfx_cookie_api': api_cookie}
        write_config_value('login', cookies, file=configFile)
        return True


def add_info(city, userName, phone):
    """添加AI呼叫"""
    tasks = {
        '广州': '8288238a-a180-4710-8320-64cddfe92e26',
        '济南': 'dacb9a20-44fe-4029-bb5b-1ded942b160e',
        '成都': '489928be-cbd8-432b-b99e-590737670c73',
        # '深圳': 'd0c0c0c0-c0c0-4c0c-8c0c-0c0c0c0c0c0c'
    }
    url = f'https://robot.jsybtx.com/saas/member/auto_dialer/task/{tasks[city]}/number'
    
    params = {
        "name": userName,
        "company": "",
        "number": phone,
        "control_select_robot": ""
    }
    headers = {
        "Content-Type": "application/json",
        'user-agent': get_config_value('login', 'user_agent', file=configFile),
        'cookie': get_config_value('login', 'bfx_cookie_api', file=configFile),
        'referer': 'https://robot.jsybtx.com/',
    }
    result = requests.post(url, headers=headers, params=params).content.decode()
    if 'Unauthenticated' in result:
        logger.error(f'❌️ {city}城市：{userName}-{phone} 添加佰分象失败，请检查')
        return False
    logger.info(f'✅️ {city}城市：{userName}-{phone} 添加佰分象成功')
    return True

# ============================================================================
# 主流程
# ============================================================================
def run_once():
    """执行一次完整的数据采集+存储流程"""
    logger.info('\n' + '█' * 60)
    logger.info(f'█  开始新一轮数据采集 {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}')
    logger.info('█' * 60 + '\n')

    # ---- 初始化 Excel ----
    wb = init_excel()
    existing_phones = load_existing_phones(wb)
    logger.info(f'汇总表中已有 {len(existing_phones)} 条记录')

    # ---- 抖音快客 ----
    if lk_login():
        dy_data = get_lk_data()
    else:
        logger.warning('抖音快客登录失败，跳过')
        dy_data = []

    # ---- 麦客 ----
    if mk_login():
        mk_data = get_mk_data()
    else:
        logger.warning('麦客登录失败，跳过')
        mk_data = []

    # ---- 佰分象 ----
    bfx_login()

    # ---- 合并 & 保存 ----
    all_data = dy_data + mk_data
    if all_data:
        save_data_to_sheets(wb, all_data, existing_phones)
        wb.save(fileName )
        logger.info(f'Excel 已保存: {fileName}')
    else:
        logger.info('本次无新数据')

    logger.info(f'本轮完成，等待 {INTERVAL_MINUTES} 分钟后执行下一轮...\n')


def main():
    """定时循环: 每30分钟执行一次"""
    logger.info(f'定时采集启动，间隔={INTERVAL_MINUTES}分钟')

    # 首次初始化 Excel
    wb = init_excel()
    wb.close()

    while True:
        try:
            run_once()
        except KeyboardInterrupt:
            logger.info('\n用户中断，程序退出')
            break
        except Exception as e:
            logger.error(f'本轮执行异常: {e}')
            import traceback
            traceback.print_exc()

        # 等待30分钟再执行下一轮
        logger.info(f'休眠 {INTERVAL_MINUTES} 分钟...')
        time.sleep(INTERVAL_MINUTES * 60)


if __name__ == '__main__':
    main()