# -*- coding: utf-8 -*-
import sys
from pathlib import Path

sys.path.append(str(Path(__file__).parent.parent))


from PlayWright import Playwright_, get_config_value, logger
from ReadFile import ReadData
import time
import openpyxl


file = 'd:/_code/spider_panjiva/数据.xlsx'

wb = openpyxl.load_workbook(file)
ws = wb['Sheet1']

def panjiva_login():
    logger.info('开始登录盘踞网....')
    url = 'https://cn.panjiva.com/shipment_search?type=us_imports'
    ele = '//span[contains(@title,"User icon")]'
    key = 'login.panjiva_cookie'
    Playwright_.login(url, ele, key)
    logger.info('盘踞网登录成功！！！')



def get_page_info():
    """获取当页数据"""
    company = []
    page = Playwright_.get_text('//ul[@class="results-by-page"]/li/div/span')
    logger.info(f'开始获取第{page}页数据')
    try:
        count_ele = '//tbody[@class="notranslate"]/tr'
        count = Playwright_.get_count(count_ele)
        for i in range(1, count+1):
            ele = f'({count_ele})[{i}]/td[5]/strong/a'
            if Playwright_.get_count(ele):
                company.append(Playwright_.get_text(ele))
    except Exception as e:
        logger.error(f'第{page}页数据获取异常：{e}')
    return company



def panjiva_search():
    """获取盘踞网数据"""
    # 限制数量
    limit_count = int(get_config_value('login', 'limit_count'))

    panjiva_login()
    logger.info('请选择自定义搜索条件，进行搜索，45秒后开始读取数据....')
    time.sleep(45)

    companys = []

    row_id = 1

    while True:
        # 页面滚动
        Playwright_.page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
        time.sleep(3)
        # 获取页面数据
        company = get_page_info()
        for company_ in company:
            if company_ not in companys:
                companys.append(company_)
                row_id += 1
                ws.cell(row=row_id, column=1, value=company_)
                wb.save(file)
                logger.info(f'已保存第{len(companys)}条数据：{company_}，行号：{row_id}')

        # 判断是否获取足够数据
        if len(companys) >= limit_count:
            logger.info(f'已获取{len(companys)}条数据，停止获取数据')
            break

        # 判断是否有下一页
        next_page_ele = '//i[@class="panjiva-icon-right"]'
        next_page_count = Playwright_.get_count(next_page_ele)
        if next_page_count > 0:
            Playwright_.click(next_page_ele)
            time.sleep(5)
        else:
            logger.info('已获取网站所有数据，停止获取数据')
            break



def qcc_login():
    logger.info('开始登录企查查....')
    url = 'https://www.qcc.com/'
    ele = '//div[contains(@class, "qcc-header-user-avatar")]'
    key = 'login.qcc_cookie'
    Playwright_.login(url, ele, key)
    logger.info('企查查登录成功！！！')



def qcc_search(company):
    """获取企查查单条数据"""
    logger.info(f'开始搜索【{company}】企查查数据')
    Playwright_.input('(//input)[1]', company, enter=True)
    time.sleep(5)

    result_count = Playwright_.get_count('//table/tr')
    result_count = min(result_count, 5)
    for i in range(1, result_count+1):
        ele = f'(//table/tr)[{i}]'

        # 验证公司存续状态
        status_ele = f'{ele}//span[contains(@class,"nstatus")]'
        status= Playwright_.get_text(status_ele)

        if status == '注销':
            continue

        company_en_ele = f'{ele}//span[@class="sf"]/span'
        company_en = Playwright_.get_text(company_en_ele) if Playwright_.get_count(company_en_ele) else ''
        if company_en == company:
            company_name_ele = f'{ele}//span[@class="copy-title"]/a/span'
            company_name = Playwright_.get_text(company_name_ele)
            logger.info(f'【{company}】已找到企查查数据：{company_name}')
            return company_name, status
    logger.info(f'【{company}】未找到企查查数据')
    return '', ''



def qcc_get_data():
    """获取企查查数据"""
    companys = ReadData.read_xlsx_col(file)['公司英文名']

    qcc_login()

    for row_id, company in enumerate(companys, start=2):
        if company == '':
            continue

        company_name, status = qcc_search(company)

        if company_name != '':
            ws.cell(row=row_id, column=2, value=company_name)
            ws.cell(row=row_id, column=3, value=status)
            wb.save(file)
            logger.info(f'已保存第{row_id}行数据：{company_name}')
    logger.info('企查查数据获取完成！！！')


if __name__ == '__main__':
    qcc_get_data()

