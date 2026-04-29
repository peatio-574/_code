import os
import pandas
from Logger import logger
from openpyxl.drawing.image import Image
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

class ReadData(object):
    @classmethod
    def read_xlsx_row(cls, file, sheetname=0, header=0, index_col=None):
        """按行读取xlsx，[{},{},{}]"""
        try:
            file = os.path.abspath(file)
            pd = pandas.read_excel(file, sheet_name=sheetname, header=header, index_col=index_col, keep_default_na=False).astype(str)  # astype将数据转为str
            col_key = [i for i in pd]
            rows = pd.shape[0]
            datalist = list()
            for row in range(rows):
                # 每行用字典保存
                row_dict = {col: pd[col][row] for col in col_key}
                datalist.append(row_dict)
            logger.info('%s文件按行读取成功' % file)
            return datalist
        except Exception as e:
            logger.error('%s文件按行读取失败：%s' % (file, e))
            return None

    @classmethod
    def read_xlsx_col(cls, file, sheetname=0, header=0, index_col=None):
        """按列读取xlsx，{col1:[], col2:[], col3:[]}"""
        try:
            file = os.path.abspath(file)
            pd = pandas.read_excel(file, sheet_name=sheetname, header=header, index_col=index_col, keep_default_na=False).astype(str)
            col_key = [i for i in pd]
            rows = pd.shape[0]
            data_dict = dict()
            for col in col_key:
                # 每列用列表保存
                col_list = [pd[col][i] for i in range(rows)]
                data_dict[col] = col_list
            logger.info('%s文件按列读取成功' % file)
            return data_dict
        except Exception as e:
            logger.error('%s文件按列读取失败：%s' % (file, e))
            return None

    @classmethod
    def write_by_row(cls, rows, filename, sheetname=0, header=0, index_col=None):
        """按行写入"""
        file = os.path.abspath(filename)
        try:
            wb = load_workbook(file)
            ws = wb.active
            ws.tile = '数据'
            for row_idx in range(len(rows)):
                row_data = rows[row_idx]
                for col_idx in range(len(row_data)):
                    value = row_data[col_idx]
                    print(value)
                    if value.endswith('.png'):
                        img = Image(value)
                        cell= get_column_letter(col_idx+1) + str(row_idx+1)
                        img.anchor = cell
                        # img.width = 100
                        # img.height = 20
                        ws.add_image(img)
                    else:
                        ws.cell(row=row_idx+1, column=col_idx+1, value=value)
            wb.save(file)
            logger.info('%s文件按行写入成功' % file)
            return True
        except Exception as e:
            logger.error('%s文件按行写入失败：%s' % (file, e))
            return False



if __name__ == '__main__':
    file = '../data/测试数据.xlsx'
    data = ReadData.read_xlsx_row(file)
    # for i in data:
    #     print(i)
    # for k, v in ReadData.read_xlsx_col(file).items():
    #     print(k, v)