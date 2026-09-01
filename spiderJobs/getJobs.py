# coding='utf-8'
import json
import requests
from Logger import logger
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import time
import os
import sys
import tkinter as tk
from tkinter import ttk
import threading


def _get_data_path(filename):
    """获取数据文件路径，兼容打包后的exe"""
    if getattr(sys, 'frozen', False):
        base = sys._MEIPASS
    else:
        base = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base, filename)


def _read_xlsx_rows(file):
    """用openpyxl按行读取xlsx，返回 [[col1,col2,...], ...]"""
    if not os.path.exists(file):
        return []
    wb = load_workbook(file, read_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(values_only=True):
        row = [str(c) if c is not None else '' for c in row]
        rows.append(row)
    wb.close()
    return rows


def parse_area_codes(file_path):
    """解析国聘省份codes"""
    with open(file_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    result = {"provinces": {}, "cities": {}}
    china = data['data'][0]
    for province in china.get('children', []):
        result["provinces"][province['value']] = province['label']
        result["cities"][province['value']] = {}
        for city in province.get('children', []):
            result["cities"][province['value']][city['name']] = city['value']
    return result


class GuoPin(object):
    # 网站名称
    web_name = '国聘'
    # 区域json
    area_codes = parse_area_codes(_get_data_path('data.json'))


    @classmethod
    def load_yesterday_data(cls, data_file):
        """读取昨天数据"""
        yesterday = time.strftime('%Y-%m-%d', time.localtime(time.time() - 86400))
        today = time.strftime('%Y-%m-%d', time.localtime(time.time()))
        all_data = _read_xlsx_rows(data_file)
        yesterday_data = []
        for row in all_data:
            if str(row[0]) not in (str(yesterday), str(today)):
                break
            yesterday_data.append(row[1:])
        return yesterday_data


    @classmethod
    def set_column_style(cls, data_file):
        """设置表头格式"""
        xlsx_headers = [
            '采集日期', '省份', '城市',
            "职位名称", "公司名称", "公司性质", "公司规模", "公司行业",
            "招聘类型", "职位性质", "职位类别", "薪资范围", "招聘人数",
            "学历要求", "经验要求", "专业要求", "工作地点", "详细地址",
            "报名截止", "职位描述"
        ]

        wb = load_workbook(filename=data_file) if os.path.exists(data_file) else Workbook()
        ws = wb.active
        ws.title = "职位数据"

        ws.freeze_panes = 'A2'
        HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")

        THIN_BORDER = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        for col, header in enumerate(xlsx_headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = THIN_BORDER
        wb.save(data_file)
        return wb, ws


    @classmethod
    def get_page_info(cls, page_id=1, province_name='北京', province_code=110000,
                      city_name='北京市', city_code=110100):
        """获取单页数据"""
        try:
            url = 'https://gp-api.iguopin.com/api/jobs/v1/recom-job'

            headers = {
                'content-type': 'application/json',
                'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/151.0.0.0 Safari/537.36'
            }

            params = {
                "search": {
                    "page": page_id,
                    "page_size": 20,
                    "district": [
                        f"000000.{province_code}.{city_code}"
                    ]
                },
                "recom": {
                    "update_time": True,
                    "company_nature": True,
                    "hot_job": True
                }
            }
            response = requests.post(url, data=json.dumps(params), headers=headers).json()
            return response['data']['list']
        except Exception as e:
            logger.error(f'{province_name}-{city_name} 第{page_id}页{cls.web_name}数据采集失败：{e}')
            return False


    @classmethod
    def parser_page_info(cls, province_name, city_name, page_info):
        """解析单页数据"""
        rows = []
        for row in page_info:
            job_name = row['job_name']
            company_name = row['company_name']
            row_info = [
                str(time.strftime('%Y-%m-%d', time.localtime(time.time() - 86400))),
                province_name,
                city_name,
                job_name,
                company_name,
                row['company_info']['nature_cn'],
                row['company_info']['scale_cn'],
                row['company_info'].get('industry_cn'),
                row['recruitment_type_cn'],
                row['nature_cn'],
                row['category_cn'],
                f"{row['min_wage']}~{row['max_wage']} {row['wage_unit_cn']}",
                row['amount'] if row['amount'] > 0 else '若干',
                row['education_cn'],
                row['experience_cn'],
                ', '.join(row.get('major_cn', [])),
                row['district_list'][0]['area_cn'] if row.get('district_list') else '',
                row['district_list'][0].get('address', '') if row.get('district_list') else '',
                row['end_time'],
                row['contents']
            ]
            rows.append(row_info)
        return rows


    @classmethod
    def get_single_city(cls, wb, ws, data_file, province_name='北京',
            province_code=110000, city_name='北京市', city_code=110100,
            yesterday_data=None, write=True):
        """采集单个城市，write=False时仅返回数据不写入"""
        all_rows = []
        page_count = 20  # 每页数量，与API page_size一致
        pages = 20  # 最大页数
        for page_id in range(1, pages + 1):
            page_info = cls.get_page_info(page_id=page_id, province_name=province_name, province_code=province_code,
                                          city_name=city_name, city_code=city_code)
            if isinstance(page_info, bool):
                continue
            elif not page_info:
                logger.info(f'【{cls.web_name}】{province_name} - {city_name} 第{page_id}页共采集0条数据')
                break

            parsed = cls.parser_page_info(province_name=province_name, city_name=city_name, page_info=page_info)
            count = 0
            for row_info in parsed:
                if row_info[1:] in yesterday_data:
                    continue
                count += 1
                all_rows.append(row_info)
                if write:
                    ws.append(row_info)
            if write:
                wb.save(data_file)

            logger.info(f'【{cls.web_name}】{province_name}-{city_name} 第{page_id}页共采集{count}条有效数据')
            if len(parsed) < page_count:
                break
        return all_rows

    @classmethod
    def run(cls, data_file='国聘数据.xlsx'):
        # 设置xlsx样式
        wb, ws = cls.set_column_style(data_file)
        yesterday_data = cls.load_yesterday_data(data_file)
        # 省份数据
        province_info = cls.area_codes.get('provinces')
        # 城市数据
        citys_info = cls.area_codes.get('cities')

        for province_code, city_info in citys_info.items():
            # 遍历省份
            province_name = province_info.get(province_code)
            for city_name, city_code in city_info.items():
                # 遍历城市
                # if province_name != '辽宁' or city_name != '大连市':
                #     continue
                args = [wb, ws, data_file, province_name, province_code, city_name, city_code, yesterday_data]
                # 单个城市采集
                logger.info(f'开始采集【{cls.web_name}{province_name}-{city_name}】[{province_code}.{city_code}]')
                cls.get_single_city(*args)


class BaseSpiderGUI:
    """爬虫可视化基类"""

    def __init__(self, title='数据采集系统', spider_name=''):
        self.root = tk.Tk()
        self.root.title(title)
        self.root.geometry('960x640')
        self.root.minsize(800, 500)

        self.spider_name = spider_name
        self.is_running = False
        self.is_paused = False
        self.pause_event = threading.Event()
        self.pause_event.set()
        self.spider_thread = None

        self.total_count = 0
        self.today_count = 0
        self.error_count = 0
        self.current_city = ''
        self.start_time = None

        self._build_ui()

    # ==================== UI 构建 ====================

    def _build_ui(self):
        self._build_header()
        self._build_control_panel()
        self._build_table()
        self._build_status_bar()

    def _build_header(self):
        header = tk.Frame(self.root, bg='#2c3e50', height=50)
        header.pack(fill='x')
        header.pack_propagate(False)
        tk.Label(header, text=f'📊 {self.spider_name}数据采集',
                 bg='#2c3e50', fg='white', font=('微软雅黑', 14, 'bold')).pack(side='left', padx=15)
        self.lbl_time = tk.Label(header, text='', bg='#2c3e50', fg='#ecf0f1', font=('微软雅黑', 9))
        self.lbl_time.pack(side='right', padx=15)
        self._update_clock()

    def _build_control_panel(self):
        frame = tk.LabelFrame(self.root, text=' 控制面板 ', font=('微软雅黑', 10, 'bold'), padx=10, pady=8)
        frame.pack(fill='x', padx=10, pady=(10, 5))

        btn_frame = tk.Frame(frame)
        btn_frame.pack(fill='x')

        self.btn_start = tk.Button(btn_frame, text='▶ 开始采集', width=12, bg='#27ae60', fg='white',
                                   font=('微软雅黑', 10, 'bold'), command=self._on_start)
        self.btn_start.pack(side='left', padx=(0, 8))

        self.btn_pause = tk.Button(btn_frame, text='⏸ 暂停', width=10, bg='#f39c12', fg='white',
                                   font=('微软雅黑', 10, 'bold'), command=self._on_pause, state='disabled')
        self.btn_pause.pack(side='left', padx=(0, 8))

        self.btn_stop = tk.Button(btn_frame, text='⏹ 停止', width=10, bg='#e74c3c', fg='white',
                                  font=('微软雅黑', 10, 'bold'), command=self._on_stop, state='disabled')
        self.btn_stop.pack(side='left')

        progress_frame = tk.Frame(frame)
        progress_frame.pack(fill='x', pady=(8, 0))

        self.progress = ttk.Progressbar(progress_frame, mode='determinate')
        self.progress.pack(side='left', fill='x', expand=True, padx=(0, 10))

        self.lbl_progress = tk.Label(progress_frame, text='0%', font=('微软雅黑', 9), width=6)
        self.lbl_progress.pack(side='right')

        status_frame = tk.Frame(frame)
        status_frame.pack(fill='x', pady=(5, 0))

        self.lbl_status = tk.Label(status_frame, text='就绪', font=('微软雅黑', 9), fg='#7f8c8d', anchor='w')
        self.lbl_status.pack(side='left', fill='x', expand=True)

        self.lbl_stats = tk.Label(status_frame, text='', font=('微软雅黑', 9), fg='#2c3e50')
        self.lbl_stats.pack(side='right')

    def _build_table(self):
        table_frame = tk.Frame(self.root)
        table_frame.pack(fill='both', expand=True, padx=10, pady=(5, 5))

        columns = ('date', 'province', 'city', 'job_name', 'company_name', 'nature', 'scale',
                   'industry', 'recruit_type', 'job_type', 'category', 'salary',
                   'amount', 'education', 'experience', 'major', 'location',
                   'address', 'deadline', 'description')
        headings = ('采集日期', '省份', '城市', '职位名称', '公司名称', '公司性质', '公司规模',
                    '公司行业', '招聘类型', '职位性质', '职位类别', '薪资范围',
                    '招聘人数', '学历要求', '经验要求', '专业要求', '工作地点',
                    '详细地址', '报名截止', '职位描述')

        self.tree = ttk.Treeview(table_frame, columns=columns, show='headings', height=18)
        self.tree.heading('#0', text='')
        self.tree.column('#0', width=0, stretch=False)

        widths = [80, 60, 70, 160, 200, 60, 80, 140, 70, 50, 100, 110,
                  50, 60, 60, 180, 90, 180, 130, 250]
        for col, heading, w in zip(columns, headings, widths):
            self.tree.heading(col, text=heading, anchor='w')
            self.tree.column(col, width=w, minwidth=40, anchor='w')

        vsb = ttk.Scrollbar(table_frame, orient='vertical', command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        self.tree.pack(side='left', fill='both', expand=True)
        vsb.pack(side='right', fill='y')

    def _build_status_bar(self):
        bar = tk.Frame(self.root, bg='#ecf0f1', height=24)
        bar.pack(fill='x', side='bottom')
        bar.pack_propagate(False)
        self.lbl_bottom = tk.Label(bar, text='就绪', bg='#ecf0f1', fg='#7f8c8d',
                                   font=('微软雅黑', 8), anchor='w')
        self.lbl_bottom.pack(side='left', padx=10)

    # ==================== 时钟 ====================

    def _update_clock(self):
        now = time.strftime('%Y-%m-%d %H:%M:%S')
        self.lbl_time.config(text=now)
        self.root.after(1000, self._update_clock)

    # ==================== 按钮事件 ====================

    def _on_start(self):
        if self.is_running and not self.is_paused:
            return
        if self.is_paused:
            self.is_paused = False
            self.pause_event.set()
            self.btn_pause.config(text='⏸ 暂停', bg='#f39c12')
            self._set_status('继续采集中...')
            return

        self.is_running = True
        self.is_paused = False
        self.total_count = 0
        self.today_count = 0
        self.error_count = 0
        self.start_time = time.time()

        self.btn_start.config(state='disabled')
        self.btn_pause.config(state='normal')
        self.btn_stop.config(state='normal')

        self.spider_thread = threading.Thread(target=self.spider_task, daemon=True)
        self.spider_thread.start()

    def _on_pause(self):
        if not self.is_running:
            return
        if self.is_paused:
            self.is_paused = False
            self.pause_event.set()
            self.btn_pause.config(text='⏸ 暂停', bg='#f39c12')
            self._set_status('继续采集中...')
        else:
            self.is_paused = True
            self.pause_event.clear()
            self.btn_pause.config(text='▶ 继续', bg='#3498db')
            self._set_status('已暂停')

    def _on_stop(self):
        if not self.is_running:
            return
        self.is_running = False
        self.is_paused = False
        self.pause_event.set()
        self.btn_start.config(state='normal')
        self.btn_pause.config(state='disabled', text='⏸ 暂停', bg='#f39c12')
        self.btn_stop.config(state='disabled')
        self._set_status('已停止')

    # ==================== 爬虫回调 ====================

    def on_progress(self, current, total, city=''):
        self.total_count = current
        pct = int(current / total * 100) if total > 0 else 0
        self.progress['value'] = pct
        self.lbl_progress.config(text=f'{pct}%')
        self.current_city = city
        elapsed = time.time() - self.start_time if self.start_time else 0
        self.lbl_stats.config(text=f'已采集: {self.today_count} | 错误: {self.error_count} | 耗时: {elapsed:.0f}s')
        self._set_status(f'正在采集: {city} ({current}/{total})')
        self.lbl_bottom.config(text=f'当前: {city} | 进度: {current}/{total}')

    def on_complete(self, msg='采集完成'):
        self.is_running = False
        self.is_paused = False
        self.root.after(0, lambda: self.btn_start.config(state='normal'))
        self.root.after(0, lambda: self.btn_pause.config(state='disabled', text='⏸ 暂停', bg='#f39c12'))
        self.root.after(0, lambda: self.btn_stop.config(state='disabled'))
        self.root.after(0, lambda: self._set_status(msg))
        self.root.after(0, lambda: self.progress.config(value=100))
        self.root.after(0, lambda: self.lbl_progress.config(text='100%'))

    def on_error(self, msg):
        self.error_count += 1
        self.root.after(0, lambda: self._set_status(f'错误: {msg}'))

    def on_data(self, row):
        self.today_count += 1
        self.root.after(0, lambda: self._insert_row(row))

    def check_pause(self):
        self.pause_event.wait()

    def _insert_row(self, row):
        values = tuple(str(v) if v is not None else '' for v in row)
        self.tree.insert('', 'end', values=values)
        self.tree.yview_moveto(1.0)

    def _set_status(self, text):
        self.lbl_status.config(text=text)

    def spider_task(self):
        raise NotImplementedError('子类必须实现 spider_task 方法')

    def run(self):
        self.root.mainloop()


class GuoPinGUI(BaseSpiderGUI):
    """国聘采集可视化界面"""

    def __init__(self):
        super().__init__(title='国聘数据采集系统', spider_name='国聘')

    def spider_task(self, data_file='国聘数据.xlsx'):
        spider = GuoPin()
        province_info = spider.area_codes.get('provinces')
        citys_info = spider.area_codes.get('cities')

        total_cities = sum(len(c) for c in citys_info.values())
        done_cities = 0

        wb, ws = spider.set_column_style(data_file)
        yesterday_data = spider.load_yesterday_data(data_file)

        for province_code, city_info in citys_info.items():
            if not self.is_running:
                break

            province_name = province_info.get(province_code, '')

            for city_name, city_code in city_info.items():
                if not self.is_running:
                    break

                self.check_pause()
                city_label = f'{province_name}-{city_name}'
                self.on_progress(done_cities, total_cities, city_label)

                try:
                    rows = spider.get_single_city(
                        wb, ws, data_file,
                        province_name, province_code,
                        city_name, city_code,
                        yesterday_data, write=True
                    )
                    for row in rows:
                        self.on_data(row)
                    done_cities += 1
                except Exception as e:
                    self.on_error(f'{city_label}: {e}')
                    done_cities += 1

        self.on_complete('国聘采集完成')


if __name__ == '__main__':
    GuoPinGUI().run()
    # GuoPin.run()
