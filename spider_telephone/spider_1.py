# -*- coding: utf-8 -*-
import sys
import time
from pathlib import Path

sys.path.append(str(Path(__file__).parent.parent))

from PlayWright import Playwright_, logger, get_config_value
from openpyxl import load_workbook
from ReadFile import ReadData

def get_info(phone, row_idx, ws, wb, file, first=False):
    try:
        if first:
            Playwright_.goto('https://shop.10086.cn/mall_280_280.html')
            time.sleep(5)
        input_location = '(//input[@iprompt="请输入手机号码"])[1]'
        Playwright_.wait_for_selector(input_location, state='visible', timeout=10000)
        Playwright_.input(input_location, str(phone), enter=True)
        time.sleep(0.75)
        count = Playwright_.get_count('//div[text()="请正确输入移动手机号码"]')
    except Exception as e:
        logger.error(f"❌ 查询 {phone} 失败: {e}")
        count = '查询失败'
    ws.cell(row=row_idx, column=2, value=count)
    wb.save(file)
    logger.info(f"行号：{row_idx}，手机号: {phone}, 结果: {count}，写入数据成功，文件：{file}")

def main(option='file'):
    file = get_config_value('login', option)
    data = ReadData.read_xlsx_col(file)['电话号码']
    logger.info(f"从{file} 中读取到 {len(data)} 个手机号")

    wb = load_workbook(file)
    ws = wb['Sheet1']

    for row_idx, phone in enumerate(data, start=2):
        get_info(phone=phone, row_idx=row_idx, ws=ws, wb=wb, file=file, first=True if row_idx == 2 else False)
    logger.info(f"✅ 写入数据完成，合计{len(data)}条数据，文件：{file}")



if __name__ == '__main__':
    main('file1')
