# coding='utf-8'
import sys

import re
from pathlib import Path

# 把项目根目录加入Python路径
sys.path.append(str(Path(__file__).parent.parent))

import os

configFile = os.path.join(os.path.dirname(__file__), 'config.ini')

from PlayWright import Playwright_, logger
from openpyxl import Workbook, load_workbook
import time

def xueQiuLogin():
    logger.info('开始登录雪球网....')
    url = 'https://xueqiu.com/S/SH688515'
    ele = '//a[text()="发帖"]'
    key = 'login.xueQiuCookie'
    Playwright_.login(url, ele, key, file=configFile)
    logger.info('✅️ 雪球登录成功')


def getXueQiuPageInfo(ws, wb, fileName):
    rowEle = '//article[@class="timeline__item"]'
    rowCount = Playwright_.get_count(rowEle)
    for rowIdx in range(1, rowCount + 1):
        transferCount = Playwright_.get_text(f'(({rowEle})[{rowIdx}]//a[contains(@class, "timeline__item__control")])[1]')
        transferCount = re.findall(r'\d+', transferCount)[0] if re.findall(r'\d+', transferCount) else '0'

        commitCount = Playwright_.get_text(f'(({rowEle})[{rowIdx}]//a[contains(@class, "timeline__item__control")])[2]')
        commitCount = re.findall(r'\d+', commitCount)[0] if re.findall(r'\d+', commitCount) else '0'

        likeCount = Playwright_.get_text(f'(({rowEle})[{rowIdx}]//a[contains(@class, "timeline__item__control")])[3]')
        likeCount = re.findall(r'\d+', likeCount)[0] if re.findall(r'\d+', likeCount) else '0'

        saveCount = Playwright_.get_text(f'(({rowEle})[{rowIdx}]//a[contains(@class, "timeline__item__control")])[4]')
        saveCount = re.findall(r'\d+', saveCount)[0] if re.findall(r'\d+', saveCount) else '0'

        publishTime = Playwright_.get_text(f'({rowEle})[{rowIdx}]//a[@class="date-and-source"]')

        titleEle = f'({rowEle})[{rowIdx}]//h3[@class="timeline__item__title"]'
        content = Playwright_.get_text(titleEle) + '\n' if Playwright_.get_count(titleEle) else ''

        logger.info(f'publishTime: {publishTime}, transferCount: {transferCount}, commitCount: {commitCount}, likeCount: {likeCount}, saveCount: {saveCount}')
        expandEle = f'({rowEle})[{rowIdx}]//a[text()="展开"]'
        if Playwright_.get_count(expandEle):  # 展开
            Playwright_.click(expandEle)
            time.sleep(3)
            aEle = f'({rowEle})[{rowIdx}]//div[@class="content content--detail"]/div/a'
            brEle = f'({rowEle})[{rowIdx}]//div[@class="content content--detail"]/div/br'
            if Playwright_.get_count(aEle) or Playwright_.get_count(brEle):
                content += Playwright_.get_text(f'({rowEle})[{rowIdx}]//div[@class="content content--detail"]/div')
            else:
                contentEle = f'({rowEle})[{rowIdx}]//div[@class="content content--detail"]/div/*'
                contentCount = Playwright_.get_count(contentEle)
                for contentId in range(1, contentCount + 1):
                    tagName = Playwright_.page.locator(f'({contentEle})[{contentId}]').evaluate("el => el.tagName")
                    if tagName == 'IMG':
                        href = Playwright_.get_attribute(f'({contentEle})[{contentId}]', 'src')
                        content += f'【链接：{href}】\n'
                    else:
                        content += Playwright_.get_text(f'({contentEle})[{contentId}]') + '\n'

            Playwright_.mouse_wheel(200)
            Playwright_.click(f'({rowEle})[{rowIdx}]//a[text()="收起"]')
            time.sleep(3)

        else:  # 无需展开
            content += Playwright_.get_text(f'(({rowEle})[{rowIdx}]//div[@class="content content--description"])[1]')
        logger.info(f'{rowIdx}, content: {content}\n')
        ws.append([publishTime, content, transferCount, commitCount, likeCount, saveCount])
    wb.save(fileName)

def getXueQiuInfo():
    fileName = os.path.join(os.path.dirname(__file__), '雪球网数据.xlsx')

    if os.path.exists(fileName):
        # 如果文件存在，加载现有工作簿
        wb = load_workbook(fileName)
        ws = wb.active
    else:
        headers = ['发布时间', '正文内容', '转发数', '评论数', '点赞数', '收藏数']
        wb = Workbook()
        ws = wb.active
        ws.title = '数据'
        ws.append(headers)
    wb.save(fileName)

    xueQiuLogin()
    time.sleep(10)
    for page in range(1, 11):
        logger.info(f'\n开始处理第{page}页数据')
        getXueQiuPageInfo(ws, wb, fileName)
        Playwright_.click('//a[text()="下一页"]')
        time.sleep(10)

if __name__ == '__main__':
    step = input('请输入操作步骤（1.爬取雪球网数据，2.爬取股吧数据）：')
    if step == '1':
        getXueQiuInfo()