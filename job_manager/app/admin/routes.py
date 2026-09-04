from flask import render_template, redirect, url_for, flash, request, jsonify, send_file, session
from flask_login import login_required, current_user
from ..models import db, User, Job, PushRecord, OperationLog, Campus, Role
from ..permissions import (
    super_admin_required, admin_required, permission_required, init_csrf,
    PERMISSION_MANAGE_JOBS, PERMISSION_MANAGE_STUDENTS, 
    PERMISSION_MANAGE_ACCOUNTS, PERMISSION_PUSH_JOBS,
    get_user_permissions, can_access_menu
)
from . import admin_bp
from datetime import datetime
import openpyxl
import io


def log_operation(action, target_type='', target_id=0, details=''):
    log = OperationLog(
        user_id=current_user.id, action=action, target_type=target_type,
        target_id=target_id, details=details, ip_address=request.remote_addr
    )
    db.session.add(log)


def _is_ajax():
    """判断是否为 AJAX 请求"""
    return (request.headers.get('X-Requested-With') == 'XMLHttpRequest'
            or request.args.get('ajax') == '1')


@admin_bp.before_request
def before_request():
    init_csrf()


def get_template_context():
    return {
        'csrf_token': session.get('csrf_token'),
        'user_permissions': get_user_permissions(current_user) if current_user.is_authenticated else [],
        'can_access_menu': lambda menu: can_access_menu(current_user, menu) if current_user.is_authenticated else False
    }


# ==================== 仪表盘 ====================
@admin_bp.route('/')
@admin_required
def dashboard():
    ctx = get_template_context()
    ctx['total_jobs'] = Job.query.filter_by(status='active', is_deleted=False).count()
    ctx['total_students'] = User.query.filter_by(user_type='student', is_active=True, is_deleted=False).count()
    ctx['total_pushes'] = PushRecord.query.filter_by(is_deleted=False).count()
    ctx['recent_pushes'] = PushRecord.query.filter_by(is_deleted=False).order_by(PushRecord.pushed_at.desc()).limit(10).all()
    return render_template('admin/dashboard.html', **ctx)


# ==================== 校区管理（超管） ====================
@admin_bp.route('/campuses/page')
@admin_required
@super_admin_required
def campuses_page():
    """校区管理页面（HTML骨架，数据由前端异步加载）"""
    ctx = get_template_context()
    return render_template('admin/campuses_list.html', **ctx)


@admin_bp.route('/campuses')
@admin_required
@super_admin_required
def campuses_list():
    """校区管理数据接口：返回JSON数据，前端负责渲染（支持搜索+分页）"""
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    keyword = request.args.get('keyword', '')
    
    if per_page not in [20, 50, 100]:
        per_page = 20
    
    query = Campus.query.filter_by(is_deleted=False)
    if keyword:
        query = query.filter(Campus.name.contains(keyword))
    
    pagination = query.order_by(Campus.updated_at.desc()).paginate(page=page, per_page=per_page, error_out=False)
    
    data = []
    for campus in pagination.items:
        admin_count = User.query.filter_by(campus_id=campus.id, user_type='admin', is_deleted=False).count()
        student_count = User.query.filter_by(campus_id=campus.id, user_type='student', is_deleted=False).count()
        data.append({
            'id': campus.id,
            'name': campus.name,
            'admin_count': admin_count,
            'student_count': student_count,
            'created_at': campus.created_at.strftime('%Y-%m-%d %H:%M') if campus.created_at else '-',
            'updated_at': campus.updated_at.strftime('%Y-%m-%d %H:%M') if campus.updated_at else '-'
        })
    
    return jsonify({
        'success': True,
        'campuses': data,
        'pagination': {
            'page': pagination.page,
            'pages': pagination.pages,
            'total': pagination.total,
            'per_page': pagination.per_page,
            'has_prev': pagination.has_prev,
            'has_next': pagination.has_next,
            'prev_num': pagination.prev_num,
            'next_num': pagination.next_num
        }
    })


@admin_bp.route('/campuses/add', methods=['POST'])
@admin_required
@super_admin_required
def campus_add():
    name = request.form.get('name', '').strip()
    if not name:
        return jsonify({'success': False, 'message': '校区名称不能为空'})
    if Campus.query.filter_by(name=name, is_deleted=False).first():
        return jsonify({'success': False, 'message': '校区名称已存在'})
    campus = Campus(name=name)
    db.session.add(campus)
    log_operation('add_campus', 'campus', 0, f'新增校区：{name}')
    db.session.commit()
    return jsonify({'success': True, 'message': '校区添加成功'})


@admin_bp.route('/campuses/delete', methods=['POST'])
@admin_required
@super_admin_required
def campus_delete():
    campus_id = request.form.get('campus_id')
    campus = Campus.query.get_or_404(campus_id)
    campus.is_deleted = True
    log_operation('delete_campus', 'campus', campus.id, f'删除校区：{campus.name}')
    db.session.commit()
    return jsonify({'success': True, 'message': '校区删除成功'})


@admin_bp.route('/campuses/edit', methods=['POST'])
@admin_required
@super_admin_required
def campus_edit():
    campus_id = request.form.get('campus_id')
    name = request.form.get('name', '').strip()
    if not name:
        return jsonify({'success': False, 'message': '校区名称不能为空'})
    
    campus = Campus.query.get_or_404(campus_id)
    existing = Campus.query.filter(Campus.name == name, Campus.id != campus_id, Campus.is_deleted==False).first()
    if existing:
        return jsonify({'success': False, 'message': '校区名称已存在'})
    
    old_name = campus.name
    campus.name = name
    log_operation('edit_campus', 'campus', campus.id, f'编辑校区：{old_name} → {name}')
    db.session.commit()
    return jsonify({'success': True, 'message': '校区更新成功'})


# ==================== 角色管理（超管） ====================
@admin_bp.route('/roles/page')
@admin_required
@super_admin_required
def roles_page():
    """角色管理页面（HTML骨架，数据由前端异步加载）"""
    ctx = get_template_context()
    return render_template('admin/roles_list.html', **ctx)


@admin_bp.route('/roles')
@admin_required
@super_admin_required
def roles_list():
    """角色管理数据接口：返回JSON数据，前端负责渲染（支持搜索+分页）"""
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    keyword = request.args.get('keyword', '')
    
    if per_page not in [20, 50, 100]:
        per_page = 20
    
    query = Role.query.filter_by(is_deleted=False)
    if keyword:
        query = query.filter(Role.name.contains(keyword))
    
    pagination = query.order_by(Role.updated_at.desc()).paginate(page=page, per_page=per_page, error_out=False)
    
    data = []
    for role in pagination.items:
        user_count = User.query.filter_by(role=role.name, is_deleted=False).count()
        data.append({
            'id': role.id,
            'name': role.name,
            'user_count': user_count,
            'is_active': role.is_active,
            'created_at': role.created_at.strftime('%Y-%m-%d %H:%M') if role.created_at else '-',
            'updated_at': role.updated_at.strftime('%Y-%m-%d %H:%M') if role.updated_at else '-'
        })
    
    return jsonify({
        'success': True,
        'roles': data,
        'pagination': {
            'page': pagination.page,
            'pages': pagination.pages,
            'total': pagination.total,
            'per_page': pagination.per_page,
            'has_prev': pagination.has_prev,
            'has_next': pagination.has_next,
            'prev_num': pagination.prev_num,
            'next_num': pagination.next_num
        }
    })


@admin_bp.route('/roles/add', methods=['POST'])
@admin_required
@super_admin_required
def role_add():
    name = request.form.get('name', '').strip()
    is_active = request.form.get('is_active') == '1'
    if not name:
        return jsonify({'success': False, 'message': '角色名称不能为空'})
    if Role.query.filter_by(name=name, is_deleted=False).first():
        return jsonify({'success': False, 'message': '角色名称已存在'})
    role = Role(name=name, is_active=is_active)
    db.session.add(role)
    log_operation('add_role', 'role', 0, f'新增角色：{name}')
    db.session.commit()
    return jsonify({'success': True, 'message': '角色添加成功'})


@admin_bp.route('/roles/delete', methods=['POST'])
@admin_required
@super_admin_required
def role_delete():
    role_id = request.form.get('role_id')
    role = Role.query.get_or_404(role_id)
    role.is_deleted = True
    log_operation('delete_role', 'role', role.id, f'删除角色：{role.name}')
    db.session.commit()
    return jsonify({'success': True, 'message': '角色删除成功'})


@admin_bp.route('/roles/toggle-status', methods=['POST'])
@admin_required
@super_admin_required
def role_toggle_status():
    role_id = request.form.get('role_id')
    role = Role.query.get_or_404(role_id)
    role.is_active = not role.is_active
    status_text = '启用' if role.is_active else '禁用'
    log_operation('toggle_role_status', 'role', role.id, f'角色{status_text}：{role.name}')
    db.session.commit()
    return jsonify({'success': True, 'message': f'角色已{status_text}'})


@admin_bp.route('/roles/edit', methods=['POST'])
@admin_required
@super_admin_required
def role_edit():
    role_id = request.form.get('role_id')
    name = request.form.get('name', '').strip()
    if not name:
        return jsonify({'success': False, 'message': '角色名称不能为空'})
    
    role = Role.query.get_or_404(role_id)
    existing = Role.query.filter(Role.name == name, Role.id != role_id, Role.is_deleted==False).first()
    if existing:
        return jsonify({'success': False, 'message': '角色名称已存在'})
    
    old_name = role.name
    role.name = name
    role.is_active = request.form.get('is_active') == '1'
    log_operation('edit_role', 'role', role.id, f'编辑角色：{old_name} → {name}')
    db.session.commit()
    return jsonify({'success': True, 'message': '角色更新成功'})


# ==================== 岗位管理（仅超管） ====================
@admin_bp.route('/jobs/page')
@admin_required
@super_admin_required
def jobs_page():
    """岗位管理页面（HTML骨架，数据由前端异步加载）"""
    ctx = get_template_context()
    return render_template('admin/jobs_list.html', **ctx)


@admin_bp.route('/jobs')
@admin_required
@super_admin_required
def jobs_list():
    """岗位管理数据接口：返回JSON数据，前端负责渲染"""
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    keyword = request.args.get('keyword', '')
    province = request.args.get('province', '')
    city = request.args.get('city', '')
    education = request.args.get('education', '')
    company_type = request.args.get('company_type', '')
    recruit_type = request.args.get('recruit_type', '')
    status_filter = request.args.get('status_filter', 'active')
    
    if per_page not in [20, 50, 100]:
        per_page = 20
    
    query = Job.query.filter_by(is_deleted=False)
    if keyword:
        query = query.filter(db.or_(Job.company_name.contains(keyword), Job.job_name.contains(keyword)))
    if province:
        query = query.filter(db.or_(Job.province == province, Job.province.like(province + '%')))
    if city:
        query = query.filter(db.or_(Job.city == city, Job.city.like(city + '%')))
    if education:
        if education == '不限':
            query = query.filter(db.or_(Job.education_req == '', Job.education_req == None, Job.education_req == '不限'))
        else:
            query = query.filter(Job.education_req == education)
    if company_type:
        query = query.filter(Job.company_type == company_type)
    if recruit_type:
        query = query.filter(Job.recruit_type.contains(recruit_type))
    if status_filter == 'active':
        query = query.filter(db.or_(Job.deadline == None, Job.deadline >= datetime.now()))
    elif status_filter == 'expired':
        query = query.filter(Job.deadline != None, Job.deadline < datetime.now())
    
    pagination = query.order_by(Job.updated_at.desc()).paginate(page=page, per_page=per_page, error_out=False)
    
    jobs = []
    for job in pagination.items:
        jobs.append({
            'id': job.id,
            'job_name': job.job_name,
            'company_name': job.company_name,
            'company_type': job.company_type,
            'recruit_type': job.recruit_type or '',
            'salary_range': job.salary_range,
            'education_req': job.education_req,
            'location': f"{job.province}-{job.city}",
            'recruit_count': job.recruit_count,
            'deadline': job.deadline.strftime('%Y-%m-%d %H:%M') if job.deadline else None,
            'is_expired': job.is_expired(),
            'is_xiaozhao': '校园' in (job.recruit_type or '')
        })
    
    return jsonify({
        'success': True,
        'jobs': jobs,
        'pagination': {
            'page': pagination.page,
            'pages': pagination.pages,
            'total': pagination.total,
            'per_page': pagination.per_page,
            'has_prev': pagination.has_prev,
            'has_next': pagination.has_next,
            'prev_num': pagination.prev_num,
            'next_num': pagination.next_num
        }
    })


@admin_bp.route('/jobs/add', methods=['GET', 'POST'])
@admin_required
@super_admin_required
def job_add():
    ctx = get_template_context()
    if request.method == 'POST':
        job = Job(
            province=request.form.get('province', ''),
            city=request.form.get('city', ''),
            job_name=request.form.get('job_name', ''),
            company_name=request.form.get('company_name', ''),
            company_type=request.form.get('company_type', ''),
            company_size=request.form.get('company_size', ''),
            company_industry=request.form.get('company_industry', ''),
            recruit_type=request.form.get('recruit_type', ''),
            job_nature=request.form.get('job_nature', ''),
            job_category=request.form.get('job_category', ''),
            salary_range=request.form.get('salary_range', ''),
            recruit_count=int(request.form.get('recruit_count', 1)),
            education_req=request.form.get('education_req', ''),
            experience_req=request.form.get('experience_req', ''),
            major_req=request.form.get('major_req', ''),
            work_location=request.form.get('work_location', ''),
            address=request.form.get('address', ''),
            deadline=datetime.strptime(request.form['deadline'], '%Y-%m-%d %H:%M:%S') if request.form.get('deadline') else None,
            job_detail=request.form.get('job_detail', ''),
            created_by=current_user.id
        )
        db.session.add(job)
        log_operation('add_job', 'job', 0, f'新增岗位：{job.job_name}')
        db.session.commit()
        message = '岗位添加成功'
        if _is_ajax():
            return jsonify({'success': True, 'message': message})
        flash(message, 'success')
        return redirect(url_for('admin.jobs_page'))
    return render_template('admin/job_form.html', **ctx)


@admin_bp.route('/jobs/edit/<int:id>', methods=['GET', 'POST'])
@admin_required
@super_admin_required
def job_edit(id):
    ctx = get_template_context()
    job = Job.query.filter_by(id=id, is_deleted=False).first_or_404()
    if request.method == 'POST':
        job.province = request.form.get('province', '')
        job.city = request.form.get('city', '')
        job.job_name = request.form.get('job_name', '')
        job.company_name = request.form.get('company_name', '')
        job.company_type = request.form.get('company_type', '')
        job.company_size = request.form.get('company_size', '')
        job.company_industry = request.form.get('company_industry', '')
        job.recruit_type = request.form.get('recruit_type', '')
        job.job_nature = request.form.get('job_nature', '')
        job.job_category = request.form.get('job_category', '')
        job.salary_range = request.form.get('salary_range', '')
        job.recruit_count = int(request.form.get('recruit_count', 1))
        job.education_req = request.form.get('education_req', '')
        job.experience_req = request.form.get('experience_req', '')
        job.major_req = request.form.get('major_req', '')
        job.work_location = request.form.get('work_location', '')
        job.address = request.form.get('address', '')
        job.deadline = datetime.strptime(request.form['deadline'], '%Y-%m-%d %H:%M:%S') if request.form.get('deadline') else None
        job.job_detail = request.form.get('job_detail', '')
        log_operation('edit_job', 'job', job.id, f'编辑岗位：{job.job_name}')
        db.session.commit()
        message = '岗位更新成功'
        if _is_ajax():
            return jsonify({'success': True, 'message': message})
        flash(message, 'success')
        return redirect(url_for('admin.jobs_page'))
    ctx['job_id'] = id
    return render_template('admin/job_form.html', **ctx)


@admin_bp.route('/jobs/<int:id>/data')
@admin_required
@super_admin_required
def job_data(id):
    """岗位编辑页数据接口：返回JSON数据，前端负责填充表单"""
    job = Job.query.filter_by(id=id, is_deleted=False).first_or_404()
    return jsonify({
        'province': job.province,
        'city': job.city,
        'job_name': job.job_name,
        'company_name': job.company_name,
        'company_type': job.company_type,
        'company_size': job.company_size,
        'company_industry': job.company_industry,
        'recruit_type': job.recruit_type or '',
        'job_nature': job.job_nature,
        'job_category': job.job_category,
        'salary_range': job.salary_range,
        'recruit_count': job.recruit_count,
        'education_req': job.education_req,
        'experience_req': job.experience_req,
        'major_req': job.major_req,
        'work_location': job.work_location,
        'address': job.address,
        'deadline': job.deadline.strftime('%Y-%m-%d %H:%M:%S') if job.deadline else '',
        'job_detail': job.job_detail
    })


@admin_bp.route('/jobs/delete', methods=['POST'])
@admin_required
@super_admin_required
def job_delete():
    ids = request.form.getlist('job_ids')
    if not ids:
        msg = '请选择要删除的岗位'
        if _is_ajax():
            return jsonify({'success': False, 'message': msg})
        flash(msg, 'warning')
        return redirect(url_for('admin.jobs_page'))
    jobs = Job.query.filter(Job.id.in_(ids), Job.is_deleted==False).all()
    if not jobs:
        msg = '未找到要删除的岗位'
        if _is_ajax():
            return jsonify({'success': False, 'message': msg})
        flash(msg, 'warning')
        return redirect(url_for('admin.jobs_page'))
    for job in jobs:
        job.is_deleted = True
        log_operation('delete_job', 'job', job.id, f'删除岗位：{job.job_name}')
    db.session.commit()
    msg = f'成功删除 {len(jobs)} 个岗位'
    if _is_ajax():
        return jsonify({'success': True, 'message': msg})
    flash(msg, 'success')
    return redirect(url_for('admin.jobs_page'))


@admin_bp.route('/jobs/toggle_status', methods=['POST'])
@admin_required
@super_admin_required
def job_toggle_status():
    ids = request.form.getlist('job_ids')
    action = request.form.get('action', 'activate')
    if not ids:
        msg = '请选择要操作的岗位'
        if _is_ajax():
            return jsonify({'success': False, 'message': msg})
        flash(msg, 'warning')
        return redirect(url_for('admin.jobs_page'))
    new_status = 'active' if action == 'activate' else 'inactive'
    jobs = Job.query.filter(Job.id.in_(ids), Job.is_deleted==False).all()
    if not jobs:
        msg = '未找到要操作的岗位'
        if _is_ajax():
            return jsonify({'success': False, 'message': msg})
        flash(msg, 'warning')
        return redirect(url_for('admin.jobs_page'))
    for job in jobs:
        job.status = new_status
    db.session.commit()
    msg = f'成功操作 {len(jobs)} 个岗位'
    if _is_ajax():
        return jsonify({'success': True, 'message': msg})
    flash(msg, 'success')
    return redirect(url_for('admin.jobs_page'))


@admin_bp.route('/jobs/import', methods=['GET', 'POST'])
@admin_required
@super_admin_required
def job_import():
    ctx = get_template_context()
    ajax = _is_ajax()
    if request.method == 'POST':
        file = request.files.get('file')
        if not file:
            msg = '请选择文件'
            if ajax:
                return jsonify({'success': False, 'message': msg})
            flash(msg, 'danger')
            return render_template('admin/job_import.html', **ctx)
        if not file.filename.endswith(('.xlsx', '.xls')):
            msg = '请上传Excel文件'
            if ajax:
                return jsonify({'success': False, 'message': msg})
            flash(msg, 'danger')
            return render_template('admin/job_import.html', **ctx)
        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            row_count = max(1, ws.max_row - 1)
            count = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[2]:
                    continue
                job = Job(
                    province=str(row[0] or ''),
                    city=str(row[1] or ''),
                    job_name=str(row[2] or ''),
                    company_name=str(row[3] or ''),
                    company_type=str(row[4] or ''),
                    company_size=str(row[5] or ''),
                    company_industry=str(row[6] or ''),
                    recruit_type=str(row[7] or '社会招聘'),
                    job_nature=str(row[8] or ''),
                    job_category=str(row[9] or ''),
                    salary_range=str(row[10] or ''),
                    recruit_count=int(row[11]) if row[11] and str(row[11]).isdigit() else 1,
                    education_req=str(row[12] or ''),
                    experience_req=str(row[13] or ''),
                    major_req=str(row[14] or ''),
                    work_location=str(row[15] or ''),
                    address=str(row[16] or ''),
                    deadline=datetime.strptime(str(row[17]), '%Y-%m-%d %H:%M:%S') if row[17] and ' ' in str(row[17]) else (datetime.strptime(str(row[17]), '%Y-%m-%d') if row[17] else None),
                    job_detail=str(row[18] or ''),
                    created_by=current_user.id
                )
                db.session.add(job)
                count += 1
            if count == 0:
                msg = '未导入任何岗位，请检查文件格式'
                if ajax:
                    return jsonify({'success': False, 'message': msg})
                flash(msg, 'warning')
                return render_template('admin/job_import.html', **ctx)
            log_operation('import_jobs', 'job', 0, f'批量导入 {count} 个岗位')
            db.session.commit()
            success_msg = f'成功导入 {count} 个岗位'
            if ajax:
                return jsonify({'success': True, 'message': success_msg, 'count': count, 'total_rows': row_count})
            flash(success_msg, 'success')
            return redirect(url_for('admin.jobs_page'))
        except Exception as e:
            db.session.rollback()
            msg = f'导入失败：{str(e)}'
            if ajax:
                return jsonify({'success': False, 'message': msg})
            flash(msg, 'danger')
    return render_template('admin/job_import.html', **ctx)


@admin_bp.route('/jobs/template')
@admin_required
@super_admin_required
def job_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '岗位导入模板'
    ws.append(['省份', '城市', '职位名称', '公司名称', '公司性质', '公司规模', '公司行业', 
               '招聘类型', '职位性质', '职位类别', '薪资范围', '招聘人数', '学历要求', 
               '经验要求', '专业要求', '工作地点', '详细地址', '报名截止(YYYY-MM-DD HH:MM:SS)', '职位描述'])
    ws.append(['新疆', '阿勒泰地区',
               '北屯 供应链组织者（应届本科，财务/统计相关专业）',
               '国药集团新疆新特药业有限公司', '国企', '1000-2000人', '批发业',
               '校园招聘', '校招', '渠道专员/助理', '5600~7000 元/月', 1,
               '本科', '应届生', '财务会计类, 统计学类', '阿勒泰', '',
               '2026-11-09 23:59:59',
               '负责资质证照的备案、盯计划、反馈缺货、协调配送、调价、退货、对账、回款核销等全链路运营操作'])
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='岗位导入模板.xlsx')


# ==================== 用户管理（仅管理员/超管） ====================
@admin_bp.route('/users/page')
@admin_required
@super_admin_required
def users_page():
    """用户管理页面（HTML骨架，数据由前端异步加载）"""
    ctx = get_template_context()
    ctx['campuses'] = Campus.query.filter_by(is_active=True, is_deleted=False).all()
    ctx['roles'] = Role.query.filter_by(is_active=True, is_deleted=False).all()
    return render_template('admin/users_list.html', **ctx)


@admin_bp.route('/users')
@admin_required
@super_admin_required
def users_list():
    """用户管理数据接口：仅返回管理员，前端负责渲染"""
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    keyword = request.args.get('keyword', '')
    campus = request.args.get('campus', '')
    role = request.args.get('role', '')
    status = request.args.get('status', '')
    
    if per_page not in [20, 50, 100]:
        per_page = 20
    
    query = User.query.filter_by(user_type='admin', is_deleted=False)
    if keyword:
        query = query.filter(db.or_(
            User.username.contains(keyword), User.real_name.contains(keyword),
            User.phone.contains(keyword)
        ))
    if campus:
        campus_obj = Campus.query.filter_by(name=campus, is_deleted=False).first()
        if campus_obj:
            query = query.filter(User.campus_id == campus_obj.id)
    if role:
        query = query.filter(User.role == role)
    if status:
        query = query.filter(User.is_active == (status == '1'))
    pagination = query.order_by(User.updated_at.desc()).paginate(page=page, per_page=per_page, error_out=False)
    
    users = []
    for user in pagination.items:
        user_data = {
            'id': user.id,
            'real_name': user.real_name,
            'id_card_last6': user.get_id_card_last6(),
            'user_type': user.user_type,
            'campus': user.campus.name if user.campus else '-',
            'role': user.role or '-',
            'phone': user.phone,
            'can_push_jobs': user.can_push_jobs,
            'can_view_jobs': user.can_view_jobs,
            'can_manage_students': user.can_manage_students,
            'creator': user.creator.real_name if user.creator else '-',
            'is_active': user.is_active,
            'created_at': user.created_at.strftime('%Y-%m-%d %H:%M') if user.created_at else '-',
            'updated_at': user.updated_at.strftime('%Y-%m-%d %H:%M') if user.updated_at else '-'
        }
        if current_user.is_super_admin():
            user_data['password'] = user.password_plain or '-'
        users.append(user_data)
    
    return jsonify({
        'success': True,
        'users': users,
        'pagination': {
            'page': pagination.page,
            'pages': pagination.pages,
            'total': pagination.total,
            'per_page': pagination.per_page,
            'has_prev': pagination.has_prev,
            'has_next': pagination.has_next,
            'prev_num': pagination.prev_num,
            'next_num': pagination.next_num
        }
    })


@admin_bp.route('/users/add', methods=['GET', 'POST'])
@admin_required
@super_admin_required
def user_add():
    ctx = get_template_context()
    ctx['campuses'] = Campus.query.filter_by(is_active=True, is_deleted=False).all()
    ctx['roles'] = Role.query.filter_by(is_active=True, is_deleted=False).all()
    ajax = _is_ajax()
    
    if request.method == 'POST':
        real_name = request.form.get('real_name', '').strip()
        phone = request.form.get('phone', '').strip()
        
        if User.query.filter_by(username=phone, is_deleted=False).first():
            msg = f'手机号"{phone}"已注册'
            if ajax:
                return jsonify({'success': False, 'message': msg})
            flash(msg, 'danger')
            return render_template('admin/user_form.html', **ctx)
        
        password = request.form.get('password', '')
        if not password:
            msg = '请输入初始密码'
            if ajax:
                return jsonify({'success': False, 'message': msg})
            flash(msg, 'danger')
            return render_template('admin/user_form.html', **ctx)
        
        user = User(
            username=phone,
            user_type='admin',
            real_name=real_name,
            phone=phone,
            campus_id=int(request.form.get('campus_id', 0)) or None,
            role=request.form.get('role', ''),
            can_push_jobs=bool(request.form.get('can_push_jobs')),
            can_view_jobs=bool(request.form.get('can_view_jobs')),
            can_manage_students=bool(request.form.get('can_manage_students')),
            avatar=request.form.get('avatar', ''),
            is_active=request.form.get('is_active') == '1',
            created_by=current_user.id
        )
        user.set_password(password)
        db.session.add(user)
        log_operation('add_user', 'user', 0, f'新增管理员：{user.real_name}')
        db.session.commit()
        msg = '管理员添加成功'
        if ajax:
            return jsonify({'success': True, 'message': msg})
        flash(msg, 'success')
        return redirect(url_for('admin.users_page'))
    
    return render_template('admin/user_form.html', **ctx)


@admin_bp.route('/users/edit/<int:id>', methods=['GET', 'POST'])
@admin_required
@super_admin_required
def user_edit(id):
    ctx = get_template_context()
    user = User.query.filter_by(id=id, is_deleted=False).first_or_404()
    if user.user_type != 'admin':
        msg = '该用户不是管理员，请到学员管理操作'
        if request.method == 'POST' and _is_ajax():
            return jsonify({'success': False, 'message': msg})
        flash(msg, 'danger')
        return redirect(url_for('admin.users_page'))
    ctx['campuses'] = Campus.query.filter_by(is_active=True, is_deleted=False).all()
    ctx['roles'] = Role.query.filter_by(is_active=True, is_deleted=False).all()
    
    if request.method == 'POST':
        user.real_name = request.form.get('real_name', '')
        user.phone = request.form.get('phone', '')
        user.campus_id = int(request.form.get('campus_id', 0)) or None
        user.role = request.form.get('role', '')
        user.can_push_jobs = bool(request.form.get('can_push_jobs'))
        user.can_view_jobs = bool(request.form.get('can_view_jobs'))
        user.can_manage_students = bool(request.form.get('can_manage_students'))
        user.avatar = request.form.get('avatar', '')
        user.is_active = request.form.get('is_active') == '1'
        
        new_password = request.form.get('password', '')
        if new_password and new_password != user.password_plain:
            user.set_password(new_password)
        
        log_operation('edit_user', 'user', user.id, f'编辑管理员：{user.real_name}')
        db.session.commit()
        msg = '管理员信息更新成功'
        if _is_ajax():
            return jsonify({'success': True, 'message': msg})
        flash(msg, 'success')
        return redirect(url_for('admin.users_page'))
    ctx['user_obj'] = user
    return render_template('admin/user_form.html', **ctx)


@admin_bp.route('/users/delete', methods=['POST'])
@admin_required
@super_admin_required
def user_delete():
    user_id = request.form.get('user_id')
    verify_name = request.form.get('verify_name', '').strip()
    verify_phone = request.form.get('verify_phone', '').strip()
    
    user = User.query.filter_by(id=user_id, is_deleted=False).first_or_404()
    
    if user.user_type != 'admin':
        msg = '该用户不是管理员，请到学员管理操作'
        if _is_ajax():
            return jsonify({'success': False, 'message': msg})
        flash(msg, 'danger')
        return redirect(url_for('admin.users_page'))
    
    if user.id == current_user.id:
        msg = '不能删除自己的账号'
        if _is_ajax():
            return jsonify({'success': False, 'message': msg})
        flash(msg, 'danger')
        return redirect(url_for('admin.users_page'))
    
    if user.real_name != verify_name or user.phone != verify_phone:
        msg = '姓名或电话验证失败'
        if _is_ajax():
            return jsonify({'success': False, 'message': msg})
        flash(msg, 'danger')
        return redirect(url_for('admin.users_page'))
    
    user.is_deleted = True
    log_operation('delete_user', 'user', user.id, f'删除管理员：{user.real_name}')
    db.session.commit()
    msg = '管理员删除成功'
    if _is_ajax():
        return jsonify({'success': True, 'message': msg})
    flash(msg, 'success')
    return redirect(url_for('admin.users_page'))


@admin_bp.route('/users/batch-delete', methods=['POST'])
@admin_required
@super_admin_required
def user_batch_delete():
    user_ids = request.form.getlist('user_ids')
    if not user_ids:
        return jsonify({'success': False, 'message': '请选择要删除的管理员'})
    
    deleted_count = 0
    skipped_self = 0
    for uid in user_ids:
        user = User.query.filter_by(id=uid, is_deleted=False).first()
        if not user:
            continue
        if user.user_type != 'admin':
            continue
        if user.id == current_user.id:
            skipped_self += 1
            continue
        user.is_deleted = True
        log_operation('delete_user', 'user', user.id, f'批量删除管理员：{user.real_name}')
        deleted_count += 1
    
    db.session.commit()
    
    msg = f'成功删除 {deleted_count} 个管理员'
    if skipped_self > 0:
        msg += f'，跳过 {skipped_self} 个（不能删除自己）'
    return jsonify({'success': True, 'message': msg})


@admin_bp.route('/users/toggle_status', methods=['POST'])
@admin_required
@super_admin_required
def user_toggle_status():
    user_id = request.form.get('user_id', type=int)
    if not user_id:
        return jsonify({'success': False, 'message': '参数错误'})
    user = User.query.filter_by(id=user_id, is_deleted=False).first()
    if not user:
        return jsonify({'success': False, 'message': '用户不存在'})
    user.is_active = not user.is_active
    log_operation('toggle_status', 'user', user.id, f'切换状态：{user.real_name} → {"启用" if user.is_active else "禁用"}')
    db.session.commit()
    return jsonify({'success': True, 'message': f'已{"启用" if user.is_active else "禁用"}', 'is_active': user.is_active})


@admin_bp.route('/users/list_json')
@admin_required
def users_list_json():
    if current_user.is_super_admin():
        students = User.query.filter_by(user_type='student', is_active=True, is_deleted=False).all()
    else:
        students = User.query.filter_by(user_type='student', is_active=True, created_by=current_user.id, is_deleted=False).all()
    return jsonify([{
        'id': s.id, 'username': s.username, 
        'real_name': s.real_name, 'id_card_last6': s.get_id_card_last6()
    } for s in students])


# ==================== 学员管理 ====================
@admin_bp.route('/students/page')
@admin_required
def students_page():
    """学员管理页面（HTML骨架，数据由前端异步加载）"""
    ctx = get_template_context()
    ctx['campuses'] = Campus.query.filter_by(is_active=True, is_deleted=False).all()
    return render_template('admin/students_list.html', **ctx)


@admin_bp.route('/students')
@admin_required
def students_list():
    """学员管理数据接口：返回JSON数据，前端负责渲染"""
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    keyword = request.args.get('keyword', '')
    campus = request.args.get('campus', '')
    education = request.args.get('education', '')
    status = request.args.get('status', '')
    
    if per_page not in [20, 50, 100]:
        per_page = 20
    
    query = User.query.filter_by(user_type='student', is_deleted=False)
    if not current_user.is_super_admin():
        query = query.filter_by(created_by=current_user.id)
    if keyword:
        query = query.filter(db.or_(
            User.username.contains(keyword), User.real_name.contains(keyword),
            User.phone.contains(keyword)
        ))
    if campus:
        campus_obj = Campus.query.filter_by(name=campus, is_deleted=False).first()
        if campus_obj:
            query = query.filter(User.campus_id == campus_obj.id)
    if education:
        query = query.filter(User.education == education)
    if status:
        if status == 'active':
            query = query.filter(User.is_active == True)
        elif status == 'disabled':
            query = query.filter(User.is_active == False)
    pagination = query.order_by(User.updated_at.desc()).paginate(page=page, per_page=per_page, error_out=False)
    
    students = []
    for stu in pagination.items:
        campus_name = stu.campus.name if stu.campus else '-'
        students.append({
            'id': stu.id,
            'real_name': stu.real_name,
            'username': stu.username,
            'campus': campus_name,
            'id_card_last6': stu.get_id_card_last6(),
            'phone': stu.phone,
            'gender': stu.gender or '-',
            'age': stu.get_age() or '-',
            'education': stu.education or '-',
            'major': stu.major or '-',
            'intention_city': stu.intention_city or '-',
            'graduation_date': stu.graduation_date.strftime('%Y-%m-%d') if stu.graduation_date else '-',
            'creator': stu.creator.username if stu.creator else '-',
            'is_active': stu.is_active
        })
    
    return jsonify({
        'success': True,
        'students': students,
        'pagination': {
            'page': pagination.page,
            'pages': pagination.pages,
            'total': pagination.total,
            'per_page': pagination.per_page,
            'has_prev': pagination.has_prev,
            'has_next': pagination.has_next,
            'prev_num': pagination.prev_num,
            'next_num': pagination.next_num
        }
    })


@admin_bp.route('/students/add', methods=['GET', 'POST'])
@admin_required
def student_add():
    ctx = get_template_context()
    ajax = _is_ajax()
    
    if request.method == 'POST':
        real_name = request.form.get('real_name', '').strip()
        phone = request.form.get('phone', '').strip()
        id_card = request.form.get('id_card', '').strip()
        
        if not id_card or len(id_card) != 18:
            msg = '请输入18位身份证号'
            if ajax:
                return jsonify({'success': False, 'message': msg})
            flash(msg, 'danger')
            return render_template('admin/student_form.html', **ctx)
        
        if User.query.filter_by(username=real_name, is_deleted=False).first():
            msg = f'姓名"{real_name}"已存在'
            if ajax:
                return jsonify({'success': False, 'message': msg})
            flash(msg, 'danger')
            return render_template('admin/student_form.html', **ctx)
        
        gender, birth_date, age = User.parse_id_card(id_card)
        password = id_card[-6:]
        
        user = User(
            username=real_name,
            user_type='student',
            real_name=real_name,
            phone=phone,
            id_card=id_card,
            gender=gender or request.form.get('gender', ''),
            birth_date=birth_date,
            education=request.form.get('education', ''),
            major=request.form.get('major', ''),
            political_status=request.form.get('political_status', ''),
            intention_city=request.form.get('intention_city', ''),
            first_intention=request.form.get('first_intention', ''),
            second_intention=request.form.get('second_intention', ''),
            third_intention=request.form.get('third_intention', ''),
            certificate=request.form.get('certificate', ''),
            remark=request.form.get('remark', ''),
            graduation_date=datetime.strptime(request.form['graduation_date'], '%Y-%m-%d').date() if request.form.get('graduation_date') else None,
            origin_place=request.form.get('origin_place', ''),
            avatar=request.form.get('avatar', ''),
            created_by=current_user.id
        )
        user.set_password(password)
        db.session.add(user)
        log_operation('add_user', 'user', 0, f'新增学员：{user.real_name}')
        db.session.commit()
        msg = '学员添加成功'
        if ajax:
            return jsonify({'success': True, 'message': msg})
        flash(msg, 'success')
        return redirect(url_for('admin.students_page'))
    
    return render_template('admin/student_form.html', **ctx)


@admin_bp.route('/students/edit/<int:id>', methods=['GET', 'POST'])
@admin_required
def student_edit(id):
    ctx = get_template_context()
    student = User.query.filter_by(id=id, user_type='student', is_deleted=False).first_or_404()
    
    if not current_user.is_super_admin() and student.created_by != current_user.id:
        msg = '无权编辑此学员'
        if request.method == 'POST' and _is_ajax():
            return jsonify({'success': False, 'message': msg})
        flash(msg, 'danger')
        return redirect(url_for('admin.students_page'))
    
    if request.method == 'POST':
        student.real_name = request.form.get('real_name', '')
        student.phone = request.form.get('phone', '')
        student.education = request.form.get('education', '')
        student.major = request.form.get('major', '')
        student.political_status = request.form.get('political_status', '')
        student.intention_city = request.form.get('intention_city', '')
        student.first_intention = request.form.get('first_intention', '')
        student.second_intention = request.form.get('second_intention', '')
        student.third_intention = request.form.get('third_intention', '')
        student.certificate = request.form.get('certificate', '')
        student.remark = request.form.get('remark', '')
        student.graduation_date = datetime.strptime(request.form['graduation_date'], '%Y-%m-%d').date() if request.form.get('graduation_date') else None
        student.origin_place = request.form.get('origin_place', '')
        student.avatar = request.form.get('avatar', '')
        
        new_password = request.form.get('password', '')
        if new_password:
            student.set_password(new_password)
        
        log_operation('edit_user', 'user', student.id, f'编辑学员：{student.real_name}')
        db.session.commit()
        msg = '学员信息更新成功'
        if _is_ajax():
            return jsonify({'success': True, 'message': msg})
        flash(msg, 'success')
        return redirect(url_for('admin.students_page'))
    ctx['student'] = student
    return render_template('admin/student_form.html', **ctx)


@admin_bp.route('/students/delete', methods=['POST'])
@admin_required
def student_delete():
    user_id = request.form.get('user_id')
    verify_name = request.form.get('verify_name', '').strip()
    verify_phone = request.form.get('verify_phone', '').strip()
    
    user = User.query.filter_by(id=user_id, user_type='student', is_deleted=False).first_or_404()
    
    if not current_user.is_super_admin() and user.created_by != current_user.id:
        msg = '无权删除此学员'
        if _is_ajax():
            return jsonify({'success': False, 'message': msg})
        flash(msg, 'danger')
        return redirect(url_for('admin.students_page'))
    
    if user.real_name != verify_name or user.phone != verify_phone:
        msg = '姓名或电话验证失败'
        if _is_ajax():
            return jsonify({'success': False, 'message': msg})
        flash(msg, 'danger')
        return redirect(url_for('admin.students_page'))
    
    user.is_deleted = True
    log_operation('delete_user', 'user', user.id, f'删除学员：{user.real_name}')
    db.session.commit()
    msg = '学员删除成功'
    if _is_ajax():
        return jsonify({'success': True, 'message': msg})
    flash(msg, 'success')
    return redirect(url_for('admin.students_page'))


@admin_bp.route('/students/toggle-status', methods=['POST'])
@admin_required
@super_admin_required
def student_toggle_status():
    student_id = request.form.get('student_id')
    student = User.query.filter_by(id=student_id, user_type='student', is_deleted=False).first_or_404()
    student.is_active = not student.is_active
    status_text = '启用' if student.is_active else '禁用'
    log_operation('toggle_student_status', 'user', student.id, f'学员{status_text}：{student.real_name}')
    db.session.commit()
    return jsonify({'success': True, 'message': f'学员已{status_text}'})


@admin_bp.route('/students/batch-delete', methods=['POST'])
@admin_required
@super_admin_required
def student_batch_delete():
    student_ids = request.form.getlist('student_ids')
    if not student_ids:
        return jsonify({'success': False, 'message': '请选择要删除的学员'})
    
    deleted_count = 0
    for sid in student_ids:
        student = User.query.filter_by(id=sid, user_type='student', is_deleted=False).first()
        if not student:
            continue
        student.is_deleted = True
        log_operation('delete_user', 'user', student.id, f'批量删除学员：{student.real_name}')
        deleted_count += 1
    
    db.session.commit()
    return jsonify({'success': True, 'message': f'成功删除 {deleted_count} 个学员'})


# ==================== 岗位推荐 ====================
@admin_bp.route('/push/page')
@admin_required
@permission_required(PERMISSION_PUSH_JOBS)
def push_page():
    """推荐记录页面（HTML骨架，数据由前端异步加载）"""
    ctx = get_template_context()
    return render_template('admin/push_list.html', **ctx)


@admin_bp.route('/push')
@admin_required
@permission_required(PERMISSION_PUSH_JOBS)
def push_list():
    """推荐记录数据接口：返回JSON数据，前端负责渲染"""
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    
    if per_page not in [20, 50, 100]:
        per_page = 20
    
    pagination = PushRecord.query.filter_by(is_deleted=False).order_by(PushRecord.updated_at.desc()).paginate(page=page, per_page=per_page, error_out=False)
    
    pushes = []
    for push in pagination.items:
        pushes.append({
            'job_name': push.job.job_name if push.job else '',
            'company_name': push.job.company_name if push.job else '',
            'student': push.student.real_name or push.student.username,
            'pusher': push.pusher.real_name or push.pusher.username,
            'pushed_at': push.pushed_at.strftime('%Y-%m-%d %H:%M'),
            'is_read': push.is_read
        })
    
    return jsonify({
        'success': True,
        'pushes': pushes,
        'pagination': {
            'page': pagination.page,
            'pages': pagination.pages,
            'total': pagination.total,
            'per_page': pagination.per_page,
            'has_prev': pagination.has_prev,
            'has_next': pagination.has_next,
            'prev_num': pagination.prev_num,
            'next_num': pagination.next_num
        }
    })


@admin_bp.route('/push/do', methods=['POST'])
@admin_required
@permission_required(PERMISSION_PUSH_JOBS)
def push_do():
    job_ids = request.form.getlist('job_ids')
    student_ids = request.form.getlist('student_ids')
    if not job_ids:
        flash('请选择岗位', 'warning')
        return redirect(url_for('admin.push_page'))
    if not student_ids:
        flash('请选择学员', 'warning')
        return redirect(url_for('admin.push_page'))
    
    count = 0
    for job_id in job_ids:
        for student_id in student_ids:
            existing = PushRecord.query.filter_by(job_id=int(job_id), student_id=int(student_id), is_deleted=False).first()
            if not existing:
                push = PushRecord(job_id=int(job_id), student_id=int(student_id), pushed_by=current_user.id)
                db.session.add(push)
                count += 1
    log_operation('push_jobs', 'push', 0, f'推送 {len(job_ids)} 个岗位给 {len(student_ids)} 个学员')
    db.session.commit()
    flash(f'成功推荐 {count} 条岗位', 'success')
    return redirect(url_for('admin.push_page'))


# ==================== 操作日志 ====================
@admin_bp.route('/logs/page')
@admin_required
@super_admin_required
def logs_page():
    """操作日志页面（HTML骨架，数据由前端异步加载）"""
    ctx = get_template_context()
    return render_template('admin/logs_list.html', **ctx)


@admin_bp.route('/logs')
@admin_required
@super_admin_required
def logs_list():
    """操作日志数据接口：返回JSON数据，前端负责渲染"""
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    action = request.args.get('action', '')
    
    if per_page not in [20, 50, 100]:
        per_page = 20
    
    query = OperationLog.query.filter_by(is_deleted=False)
    if action:
        query = query.filter_by(action=action)
    pagination = query.order_by(OperationLog.updated_at.desc()).paginate(page=page, per_page=per_page, error_out=False)
    
    logs = []
    for log in pagination.items:
        logs.append({
            'id': log.id,
            'created_at': log.created_at.strftime('%Y-%m-%d %H:%M:%S') if log.created_at else '-',
            'operator': log.user.real_name or log.user.username,
            'action': log.action,
            'details': log.details,
            'ip_address': log.ip_address
        })
    
    return jsonify({
        'success': True,
        'logs': logs,
        'pagination': {
            'page': pagination.page,
            'pages': pagination.pages,
            'total': pagination.total,
            'per_page': pagination.per_page,
            'has_prev': pagination.has_prev,
            'has_next': pagination.has_next,
            'prev_num': pagination.prev_num,
            'next_num': pagination.next_num
        }
    })


# ==================== 管理员导入（超管） ====================
@admin_bp.route('/import_admins', methods=['GET', 'POST'])
@admin_required
@super_admin_required
def import_admins():
    ctx = get_template_context()
    ajax = _is_ajax()
    if request.method == 'POST':
        file = request.files.get('file')
        if not file:
            msg = '请选择文件'
            if ajax:
                return jsonify({'success': False, 'message': msg})
            flash(msg, 'danger')
            return render_template('admin/import_admins.html', **ctx)
        if not file.filename.endswith(('.xlsx', '.xls')):
            msg = '请上传Excel文件'
            if ajax:
                return jsonify({'success': False, 'message': msg})
            flash(msg, 'danger')
            return render_template('admin/import_admins.html', **ctx)
        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            row_count = max(1, ws.max_row - 1)
            count = 0
            errors = []
            for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row[0]:
                    continue
                real_name = str(row[0] or '').strip()
                password = str(row[1] or '123456').strip()
                phone = str(row[2] or '').strip()
                campus_name = str(row[3] or '').strip()
                role = str(row[4] or '').strip()
                can_push = str(row[5] or '').lower() in ('是', 'yes', 'true', '1', '√')
                can_view = str(row[6] or '').lower() in ('是', 'yes', 'true', '1', '√')
                can_manage = str(row[7] or '').lower() in ('是', 'yes', 'true', '1', '√')
                
                if not real_name:
                    errors.append(f'第{idx}行：姓名为空')
                    continue
                if not phone:
                    errors.append(f'第{idx}行：手机号为空')
                    continue
                
                campus_id = None
                if campus_name:
                    campus = Campus.query.filter_by(name=campus_name, is_deleted=False).first()
                    if not campus:
                        campus = Campus(name=campus_name)
                        db.session.add(campus)
                        db.session.flush()
                    campus_id = campus.id
                
                username = phone
                if User.query.filter_by(username=username, is_deleted=False).first():
                    errors.append(f'第{idx}行：账号{username}已存在')
                    continue
                
                user = User(
                    username=username,
                    user_type='admin',
                    real_name=real_name,
                    phone=phone,
                    campus_id=campus_id,
                    role=role,
                    can_push_jobs=can_push,
                    can_view_jobs=can_view,
                    can_manage_students=can_manage,
                    created_by=current_user.id
                )
                user.set_password(password)
                db.session.add(user)
                count += 1
            
            if count == 0 and not errors:
                msg = '未导入任何管理员，请检查文件格式'
                if ajax:
                    return jsonify({'success': False, 'message': msg})
                flash(msg, 'warning')
                return render_template('admin/import_admins.html', **ctx)
            
            log_operation('import_admins', 'user', 0, f'批量导入 {count} 个管理员')
            db.session.commit()
            
            if errors:
                result_msg = f'成功导入 {count} 个管理员，{len(errors)} 条警告：' + '; '.join(errors[:5])
                result_type = 'warning'
            else:
                result_msg = f'成功导入 {count} 个管理员'
                result_type = 'success'
            if ajax:
                return jsonify({'success': True, 'message': result_msg, 'count': count, 'errors': errors, 'total_rows': row_count})
            flash(result_msg, result_type)
            return redirect(url_for('admin.users_page'))
        except Exception as e:
            db.session.rollback()
            msg = f'导入失败：{str(e)}'
            if ajax:
                return jsonify({'success': False, 'message': msg})
            flash(msg, 'danger')
    return render_template('admin/import_admins.html', **ctx)


@admin_bp.route('/template_admins')
@admin_required
@super_admin_required
def template_admins():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '管理员导入模板'
    ws.append(['姓名', '密码(默认123456)', '手机号', '校区名称', '角色(校长/老师等)', 
               '岗位推送权限(是/否)', '岗位查看权限(是/否)', '学员管理权限(是/否)'])
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='管理员导入模板.xlsx')


# ==================== 学员导入 ====================
@admin_bp.route('/import_students', methods=['GET', 'POST'])
@admin_required
@permission_required(PERMISSION_MANAGE_STUDENTS)
def import_students():
    ctx = get_template_context()
    if request.method == 'POST':
        file = request.files.get('file')
        ajax = _is_ajax()
        if not file:
            msg = '请选择文件'
            if ajax:
                return jsonify({'success': False, 'message': msg})
            flash(msg, 'danger')
            return render_template('admin/import_students.html', **ctx)
        if not file.filename.endswith(('.xlsx', '.xls')):
            msg = '请上传Excel文件'
            if ajax:
                return jsonify({'success': False, 'message': msg})
            flash(msg, 'danger')
            return render_template('admin/import_students.html', **ctx)
        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            count = 0
            errors = []
            for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row[0]:
                    continue
                id_card = str(row[0] or '').strip()
                real_name = str(row[1] or '').strip()
                phone = str(row[2] or '').strip()
                education = str(row[3] or '').strip()
                major = str(row[4] or '').strip()
                gender = str(row[5] or '').strip()
                political_status = str(row[6] or '').strip()
                origin_place = str(row[7] or '').strip()
                intention_city = str(row[8] or '').strip()
                first_intention = str(row[9] or '').strip()
                second_intention = str(row[10] or '').strip()
                third_intention = str(row[11] or '').strip()
                certificate = str(row[12] or '').strip()
                remark = str(row[13] or '').strip()
                graduation_date_str = str(row[14] or '').strip()
                password = str(row[15] or '123456').strip()
                
                if not id_card or len(id_card) != 18:
                    errors.append(f'第{idx}行：身份证号格式错误')
                    continue
                if not real_name:
                    errors.append(f'第{idx}行：姓名为空')
                    continue
                
                if User.query.filter_by(username=real_name, is_deleted=False).first():
                    errors.append(f'第{idx}行：姓名{real_name}已存在')
                    continue
                
                auto_gender, birth_date, age = User.parse_id_card(id_card)
                if not gender and auto_gender:
                    gender = auto_gender
                
                graduation_date = None
                if graduation_date_str:
                    try:
                        graduation_date = datetime.strptime(graduation_date_str, '%Y-%m-%d').date()
                    except:
                        pass
                
                user = User(
                    username=real_name,
                    user_type='student',
                    real_name=real_name,
                    phone=phone,
                    id_card=id_card,
                    gender=gender,
                    birth_date=birth_date,
                    education=education,
                    major=major,
                    political_status=political_status,
                    origin_place=origin_place,
                    intention_city=intention_city,
                    first_intention=first_intention,
                    second_intention=second_intention,
                    third_intention=third_intention,
                    certificate=certificate,
                    remark=remark,
                    graduation_date=graduation_date,
                    created_by=current_user.id
                )
                user.set_password(password)
                db.session.add(user)
                count += 1
            
            if count == 0 and not errors:
                msg = '未导入任何学员，请检查文件格式'
                if ajax:
                    return jsonify({'success': False, 'message': msg})
                flash(msg, 'warning')
                return render_template('admin/import_students.html', **ctx)
            
            log_operation('import_students', 'user', 0, f'批量导入 {count} 个学员')
            db.session.commit()
            
            if errors:
                result_msg = f'成功导入 {count} 个学员，{len(errors)} 条警告：' + '; '.join(errors[:5])
                result_type = 'warning'
            else:
                result_msg = f'成功导入 {count} 个学员'
                result_type = 'success'
            if ajax:
                return jsonify({'success': True, 'message': result_msg, 'count': count, 'errors': errors})
            flash(result_msg, result_type)
            return redirect(url_for('admin.students_page'))
        except Exception as e:
            db.session.rollback()
            msg = f'导入失败：{str(e)}'
            if ajax:
                return jsonify({'success': False, 'message': msg})
            flash(msg, 'danger')
    return render_template('admin/import_students.html', **ctx)


@admin_bp.route('/template_students')
@admin_required
@permission_required(PERMISSION_MANAGE_STUDENTS)
def template_students():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '学员导入模板'
    ws.append(['身份证号(18位)', '姓名', '手机号', '学历(大专/本科/硕士/博士)', '专业', 
               '性别(男/女)', '政治面貌', '生源地', '意向城市', '第一意向岗位', 
               '第二意向岗位', '第三意向岗位', '证书', '备注', '毕业时间(YYYY-MM-DD)', '密码(默认123456)'])
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='学员导入模板.xlsx')
