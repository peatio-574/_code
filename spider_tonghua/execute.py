# -*- coding: utf-8 -*-
import sys

from pathlib import Path

sys.path.append(str(Path(__file__).parent.parent))

import requests, time
from Logger import logger
import re
import os
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import load_workbook, Workbook

filename = os.path.join(os.path.dirname(__file__), '数据.xlsx')
company_file = os.path.join(os.path.dirname(__file__), 'company.txt')

copy_data = input('请输入是否备份数据（1是，0否）：')
if copy_data == '1':
    import shutil
    bak_file = os.path.join(os.path.dirname(__file__), f'{time.strftime("%Y%m%d_%H%M%S")}_数据.xlsx')
    if os.path.exists(filename):
        shutil.copyfile(filename, bak_file)
        logger.info(f'数据已备份至：{bak_file}')
    else:
        logger.info('数据文件不存在，跳过备份')

if os.path.exists(filename):
    # 如果文件存在，加载现有工作簿
    wb = load_workbook(filename)
    ws = wb.active
else:
    xlsx_headers = ['营业部名称', '详情链接', '上榜日期', '股票简称', '上榜原因',
                    '涨跌幅(%)', '买入额（万）', '卖出额（万）', '买卖净额（万）', '所属板块']
    wb = Workbook()
    ws = wb.active
    ws.title = '数据'
    ws.append(xlsx_headers)
    wb.save(filename)

existing_data = set()
for existing_row in ws.iter_rows(min_row=2, values_only=True):
    existing_data.add(existing_row)

# ---------------------------------------------------------------------------
# 多线程相关
# ---------------------------------------------------------------------------
xlsx_lock = threading.Lock()       # 保护 ws / wb / existing_data 的并发写入
MAX_WORKERS = 5                    # 并发线程数（I/O 密集型，5~8 均可）


headers = {
    "Accept": "text/html, */*; q=0.01",
    "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
    "Host": "data.10jqka.com.cn",
    "hexin-v": "Ax7qvCre-HSliywQ9clrO5NQb79l3-iudLOWIcini71clbBhMG8yaUQz5keb",
    "Referer": "https://data.10jqka.com.cn/market/longhu/",
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) Chrome 130.0.0.0 Safari/537.36",
    "X-Requested-With": "XMLHttpRequest",
}


def get_first(page=1, page_count=False):
    """上榜次数最多营业部名称，链接"""
    url = f'https://data.10jqka.com.cn/ifmarket/lhbyyb/type/1/tab/sbcs/field/sbcs/sort/desc/page/{page}/'
    try:
        resp = requests.get(url, headers=headers, timeout=30)
        resp.raise_for_status()
        info = resp.content.decode('gbk')
    except Exception as e:
        logger.error(f'获取营业部列表失败: {url} — {e}')
        return [] if not page_count else 0
    pattern = r'<a href="(.*?)" target="_blank" title="(.*?)">.*?</a>' if not page_count else '<span class="page_info">1/(.*?)</span>'
    companys = re.findall(pattern, info)
    # time.sleep(2)
    if page_count:
        return int(companys[0]) if companys else 0
    return companys


def get_second(page=1, page_count=False):
    """实力最强营业部名称，链接"""
    url = f'https://data.10jqka.com.cn/ifmarket/lhbyyb/type/1/tab/zjsl/field/zgczje/sort/desc/page/{page}/'
    try:
        resp = requests.get(url, headers=headers, timeout=30)
        resp.raise_for_status()
        info = resp.content.decode('gbk')
    except Exception as e:
        logger.error(f'获取营业部列表失败: {url} — {e}')
        return [] if not page_count else 0
    pattern = r'<a href="(.*?)" target="_blank" title="(.*?)">.*?</a>' if not page_count else '<span class="page_info">1/(.*?)</span>'
    companys = re.findall(pattern, info)
    # time.sleep(2)
    if page_count:
        return int(companys[0]) if companys else 0
    return companys


def get_companys():
    result = []
    first_page_count = get_first(page_count=True)
    logger.info(f'上榜次数最多营业部共{first_page_count}页公司名称')
    for page in range(1, first_page_count + 1):
        companys = get_first(page)
        for company in companys:
            result.append(company)
        logger.info(f'上榜次数最多营业部，已获取第{page}页数据，当前共{len(result)}家公司')

    second_page_count = get_second(page_count=True)
    logger.info(f'实力最强营业部共{second_page_count}页公司名称')
    for page in range(1, second_page_count + 1):
        companys = get_second(page)
        for company in companys:
            if company not in result:
                result.append(company)
        logger.info(f'实力最强营业部，已获取第{page}页数据，当前共{len(result)}家公司')

    result = [str(company) for company in result]
    with open(company_file, 'w', encoding='utf-8') as f:
        string = '\n'.join(result)
        f.write(string)
        logger.info('公司信息保存成功')


def get_page_detail(orgcode, page=1, total_pages=False):
    """获取营业部指定页码数据"""
    url = f'http://data.10jqka.com.cn/ifmarket/lhbhistory/orgcode/{orgcode}/field/ENDDATE/order/desc/page/{page}/'
    logger.info(f'链接:{url}')
    try:
        resp = requests.get(url, headers=headers, timeout=30)
        resp.raise_for_status()
        info = resp.content.decode('gbk')
    except Exception as e:
        logger.error(f'请求失败: {url} — {e}')
        return 0 if total_pages else []

    if not total_pages:
        result = re.findall(
            r'<tr>\s*<td>(\d{4}-\d{2}-\d{2})</td>\s*<td>\s*<a[^>]*>([^<]+)</a>\s*</td>\s*<td[^>]*>([^<]+)</td>\s*<td[^>]*>([^<]+)</td>\s*<td[^>]*>([^<]+)</td>\s*<td[^>]*>([^<]+)</td>\s*<td[^>]*>([^<]+)</td>\s*<td[^>]*>([^<]+)</td>\s*</tr>',
            info,
            re.DOTALL
        )

    else:
        page_info = re.search(r'(\d+)/(\d+)', info)
        result = int(page_info.group(2)) if page_info else 0
    # time.sleep(2)
    return result


def save_page_safe(idx, company_name, company_url, page_info, page_no):
    """
    线程安全的单页保存。
    - 去重检查、ws.append、wb.save 在锁内执行（保证线程互不干扰）
    - time.sleep 在锁外（不阻塞其他线程写文件）
    """
    try:
        new_rows = []
        dup = 0

        # ---- 锁内：去重 → 写入 → 保存 ----
        with xlsx_lock:
            for row_id, page_ in enumerate(page_info, start=1):
                row = [company_name, company_url, *page_]
                key = tuple(row)
                if key in existing_data:
                    dup += 1
                    continue
                existing_data.add(key)
                ws.append(row)
                new_rows.append(row)

            if new_rows:
                wb.save(filename)

        # ---- 锁外：日志 & 休眠 ----
        if new_rows:
            logger.info(
                f'第{idx}家：【{company_name}】第{page_no}页：'
                f'新增{len(new_rows)}条' + (f'，跳过{dup}条(已存在)' if dup else '')
            )
    except Exception as e:
        logger.error(f'第{idx}家：【{company_name}】第{page_no}页保存失败：{e}')

    time.sleep(1)  # 锁外休眠，不影响其他线程


def _process_one_company(idx, company_name, company_url):
    """
    处理单个营业部的所有页面（在子线程中执行）。
    流程：获取总页数 → 倒序逐页爬取 → 逐页保存。
    """
    try:
        orgcode = company_url.split('/')[-2]
        total_pages = get_page_detail(orgcode, page=1, total_pages=True)
        logger.info(f'第{idx}家：【{company_name}】共{total_pages}页')

        for page_no in range(total_pages, 0, -1):
            page_info = get_page_detail(orgcode, page=page_no)
            logger.info(
                f'第{idx}家：【{company_name}】第{page_no}/{total_pages}页，{len(page_info)}行'
            )
            save_page_safe(idx, company_name, company_url, page_info, page_no=page_no)

        logger.info(f'第{idx}家：【{company_name}】✓ 全部完成')
        return True

    except Exception as e:
        logger.error(f'第{idx}家：【{company_name}】异常：{e}')
        return False


def run():
    """
    多线程爬取入口。
    读取 company.txt → 线程池并发处理所有营业部 → 实时汇报进度。
    """
    with open(company_file, 'r', encoding='utf-8') as f:
        companys = f.readlines()

    # 预处理任务列表
    tasks = []
    for idx, company in enumerate(companys, start=1):
        company_url, company_name = eval(company)
        tasks.append((idx, company_name, company_url))

    logger.info(f'共{len(tasks)}家公司，启动{MAX_WORKERS}个线程并发爬取')
    logger.info('=' * 50)

    ok_count = 0
    fail_count = 0

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        future_map = {
            executor.submit(_process_one_company, idx, name, url): (idx, name)
            for idx, name, url in tasks
        }

        for future in as_completed(future_map):
            idx, name = future_map[future]
            try:
                if future.result():
                    ok_count += 1
                else:
                    fail_count += 1
            except Exception as e:
                fail_count += 1
                logger.error(f'第{idx}家：【{name}】线程崩溃：{e}')

            logger.info(f'进度：{ok_count + fail_count}/{len(tasks)}  成功={ok_count}  失败={fail_count}')

    logger.info('=' * 50)
    logger.info(f'全部完成！成功{ok_count}家，失败{fail_count}家')
    logger.info(f'数据文件：{filename}')


if __name__ == '__main__':
    while True:
        step = input('请输入操作步骤(1获取公司信息，2开始获取数据)：')
        if step == '1':
            get_companys()
        else:
            run()