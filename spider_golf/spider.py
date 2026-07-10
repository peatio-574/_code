# coding='utf-8'
"""
    @作者：彭帅
    @邮箱：acheng6@126.com
    @时间：2026/7/9 22:10
"""
import os
import time

from PlayWright import Playwright_, logger
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment
info = {
    'Alabama': 'https://golfcartresource.com/dealer-locator?keywords=&address=Alabama&directory_radius=50&center=&address_type=state&category=0',
    'Alaska': 'https://golfcartresource.com/dealer-locator?address_type=state&zoom=15&is_mile=1&directory_radius=50&keywords=&address=Alaska&directory_radius=50&center=&address_type=state&category=0',
    'Arizona': 'https://golfcartresource.com/dealer-locator?address_type=state&zoom=15&is_mile=1&directory_radius=50&keywords=&address=Arizona&directory_radius=50&center=&address_type=state&category=0',
    'Arkansas': 'https://golfcartresource.com/dealer-locator?address_type=state&zoom=15&is_mile=1&directory_radius=50&keywords=&address=Arkansas&directory_radius=50&center=&address_type=state&category=0',
    'California': 'https://golfcartresource.com/dealer-locator?address_type=state&zoom=15&is_mile=1&directory_radius=50&keywords=&address=California&directory_radius=50&center=&address_type=state&category=0',
    'Colorado': 'https://golfcartresource.com/dealer-locator?address_type=state&zoom=15&is_mile=1&directory_radius=50&keywords=&address=Colorado&directory_radius=50&center=&address_type=state&category=0',
    'Connecticut': 'https://golfcartresource.com/dealer-locator?address_type=state&zoom=15&is_mile=1&directory_radius=50&keywords=&address=Connecticut&directory_radius=50&center=&address_type=state&category=0',
    'Delaware': 'https://golfcartresource.com/dealer-locator?address_type=state&zoom=15&is_mile=1&directory_radius=50&keywords=&address=Delaware&directory_radius=50&center=&address_type=state&category=0',
    'Florida': 'https://golfcartresource.com/dealer-locator?address_type=state&zoom=15&is_mile=1&directory_radius=50&keywords=&address=Florida&directory_radius=50&center=&address_type=state&category=0',
    'Georgia': 'https://golfcartresource.com/dealer-locator?address_type=state&zoom=15&is_mile=1&directory_radius=50&keywords=&address=Georgia&directory_radius=50&center=&address_type=state&category=0',
    'Hawaii': 'https://golfcartresource.com/dealer-locator?address_type=state&zoom=15&is_mile=1&directory_radius=50&keywords=&address=Hawaii&directory_radius=50&center=&address_type=state&category=0',
    'Idaho': 'https://golfcartresource.com/dealer-locator?address_type=state&zoom=15&is_mile=1&directory_radius=50&keywords=&address=Idaho&directory_radius=50&center=&address_type=state&category=0',
    'Illinois': 'https://golfcartresource.com/dealer-locator?address_type=state&zoom=15&is_mile=1&directory_radius=50&keywords=&address=Illinois&directory_radius=50&center=&address_type=state&category=0',
    'Indiana': 'https://golfcartresource.com/dealer-locator?address_type=state&zoom=15&is_mile=1&directory_radius=50&keywords=&address=Indiana&directory_radius=50&center=&address_type=state&category=0',
    'Iowa': 'https://golfcartresource.com/dealer-locator?address_type=state&zoom=15&is_mile=1&directory_radius=50&keywords=&address=Iowa&directory_radius=50&center=&address_type=state&category=0',
    'Kansas': 'https://golfcartresource.com/dealer-locator?address_type=state&zoom=15&is_mile=1&directory_radius=50&keywords=&address=Kansas&directory_radius=50&center=&address_type=state&category=0',
    'Kentucky': 'https://golfcartresource.com/dealer-locator?address_type=state&zoom=15&is_mile=1&directory_radius=50&keywords=&address=Kentucky&directory_radius=50&center=&address_type=state&category=0',
    'Louisiana': 'https://golfcartresource.com/dealer-locator?address_type=state&zoom=15&is_mile=1&directory_radius=50&keywords=&address=Louisiana&directory_radius=50&center=&address_type=state&category=0',
    'Maine': 'https://golfcartresource.com/dealer-locator?address_type=state&zoom=15&is_mile=1&directory_radius=50&keywords=&address=Maine&directory_radius=50&center=&address_type=state&category=0',
    'Maryland': 'https://golfcartresource.com/dealer-locator?address_type=state&zoom=15&is_mile=1&directory_radius=50&keywords=&address=Maryland&directory_radius=50&center=&address_type=state&category=0',
    'Massachusetts': 'https://golfcartresource.com/dealer-locator?address_type=state&zoom=15&is_mile=1&directory_radius=50&keywords=&address=Massachusetts&directory_radius=50&center=&address_type=state&category=0',
    'Michigan': 'https://golfcartresource.com/dealer-locator?address_type=state&zoom=15&is_mile=1&directory_radius=50&keywords=&address=Michigan&directory_radius=50&center=&address_type=state&category=0',
    'Minnesota': 'https://golfcartresource.com/dealer-locator?address_type=state&zoom=15&is_mile=1&directory_radius=50&keywords=&address=Minnesota&directory_radius=50&center=&address_type=state&category=0',
    'Mississippi': 'https://golfcartresource.com/dealer-locator?address_type=state&zoom=15&is_mile=1&directory_radius=50&keywords=&address=Mississippi&directory_radius=50&center=&address_type=state&category=0',
    'Missouri': 'https://golfcartresource.com/dealer-locator?address_type=state&zoom=15&is_mile=1&directory_radius=50&keywords=&address=Missouri&directory_radius=50&center=&address_type=state&category=0',
    'Montana': 'https://golfcartresource.com/dealer-locator?address_type=state&zoom=15&is_mile=1&directory_radius=50&keywords=&address=Montana&directory_radius=50&center=&address_type=state&category=0',
    'Nebraska': 'https://golfcartresource.com/dealer-locator?address_type=state&zoom=15&is_mile=1&directory_radius=50&keywords=&address=Nebraska&directory_radius=50&center=&address_type=state&category=0',
    'Nevada': 'https://golfcartresource.com/dealer-locator?address_type=state&zoom=15&is_mile=1&directory_radius=50&keywords=&address=Nevada&directory_radius=50&center=&address_type=state&category=0',
    'New Hampshire': 'https://golfcartresource.com/dealer-locator?address_type=state&zoom=15&is_mile=1&directory_radius=50&keywords=&address=New+Hampshire&directory_radius=50&center=&address_type=state&category=0',
    'New Jersey': 'https://golfcartresource.com/dealer-locator?address_type=state&zoom=15&is_mile=1&directory_radius=50&keywords=&address=New+Jersey&directory_radius=50&center=&address_type=state&category=0',
    'New Mexico': 'https://golfcartresource.com/dealer-locator?address_type=state&zoom=15&is_mile=1&directory_radius=50&keywords=&address=New+Mexico&directory_radius=50&center=&address_type=state&category=0',
    'New York': 'https://golfcartresource.com/dealer-locator?address_type=state&zoom=15&is_mile=1&directory_radius=50&keywords=&address=New+York&directory_radius=50&center=&address_type=state&category=0',
    'North Carolina': 'https://golfcartresource.com/dealer-locator?address_type=state&zoom=15&is_mile=1&directory_radius=50&keywords=&address=North+Carolina&directory_radius=50&center=&address_type=state&category=0',
    'North Dakota': 'https://golfcartresource.com/dealer-locator?address_type=state&zoom=15&is_mile=1&directory_radius=50&keywords=&address=North+Dakota&directory_radius=50&center=&address_type=state&category=0',
    'Ohio': 'https://golfcartresource.com/dealer-locator?address_type=state&zoom=15&is_mile=1&directory_radius=50&keywords=&address=Ohio&directory_radius=50&center=&address_type=state&category=0',
    'Oklahoma': 'https://golfcartresource.com/dealer-locator?address_type=state&zoom=15&is_mile=1&directory_radius=50&keywords=&address=Oklahoma&directory_radius=50&center=&address_type=state&category=0',
    'Pennyslvania': 'https://golfcartresource.com/dealer-locator?address_type=state&zoom=15&is_mile=1&directory_radius=50&keywords=&address=Pennyslvania&directory_radius=50&center=&address_type=state&category=0',
    'South Carolina': 'https://golfcartresource.com/dealer-locator?address_type=state&zoom=15&is_mile=1&directory_radius=50&keywords=&address=South+Carolina&directory_radius=50&center=&address_type=state&category=0',
    'South Dakota': 'https://golfcartresource.com/dealer-locator?address_type=state&zoom=15&is_mile=1&directory_radius=50&keywords=&address=South+Dakota&directory_radius=50&center=&address_type=state&category=0',
    'Tennessee': 'https://golfcartresource.com/dealer-locator?address_type=state&zoom=15&is_mile=1&directory_radius=50&keywords=&address=Tennessee&directory_radius=50&center=&address_type=state&category=0',
    'Texas': 'https://golfcartresource.com/dealer-locator?address_type=state&zoom=15&is_mile=1&directory_radius=50&keywords=&address=Texas&directory_radius=50&center=&address_type=state&category=0',
    'Utah': 'https://golfcartresource.com/dealer-locator?address_type=state&zoom=15&is_mile=1&directory_radius=50&keywords=&address=Utah&directory_radius=50&center=&address_type=state&category=0',
    'Vermont': 'https://golfcartresource.com/dealer-locator?address_type=state&zoom=15&is_mile=1&directory_radius=50&keywords=&address=Vermont&directory_radius=50&center=&address_type=state&category=0',
    'Virginia': 'https://golfcartresource.com/dealer-locator?address_type=state&zoom=15&is_mile=1&directory_radius=50&keywords=&address=Virginia&directory_radius=50&center=&address_type=state&category=0',
    'Washington': 'https://golfcartresource.com/dealer-locator?address_type=state&zoom=15&is_mile=1&directory_radius=50&keywords=&address=Washington&directory_radius=50&center=&address_type=state&category=0',
    'West Virginia': 'https://golfcartresource.com/dealer-locator?address_type=state&zoom=15&is_mile=1&directory_radius=50&keywords=&address=West+Virginia&directory_radius=50&center=&address_type=state&category=0',
    'Wisconsin': 'https://golfcartresource.com/dealer-locator?address_type=state&zoom=15&is_mile=1&directory_radius=50&keywords=&address=Wisconsin&directory_radius=50&center=&address_type=state&category=0',
    'Wyoming': 'https://golfcartresource.com/dealer-locator?address_type=state&zoom=15&is_mile=1&directory_radius=50&keywords=&address=Wyoming&directory_radius=50&center=&address_type=state&category=0'
}

dataFile = os.path.join(os.path.dirname(__file__), '数据.xlsx')

if os.path.exists(dataFile):
    # 如果文件存在，加载现有工作簿
    wb = load_workbook(dataFile)
    ws = wb.active
else:
    headers = ['国家', '经销商', '品牌']
    wb = Workbook()
    ws = wb.active
    ws.title = '数据'
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    wb.save(dataFile)
exists_data = []
for row in ws.rows:
    # row 是元组，每个元素是单元格
    row_data = [cell.value for cell in row]
    exists_data.append(row_data)

def getPageinfo(city, url):
    logger.info(f'开始处理国家：{city}')
    Playwright_.goto(url)
    nextPageEle = '(//div[@class="sabai-pagination sabai-btn-group"]/a)[last()][not(contains(@class, "disabled"))]'
    while True:
        rowEle = '//div[contains(@id, "sabai-entity-content")]'
        rowCount = Playwright_.get_count(rowEle)
        for i in range(1, rowCount + 1):
            shopName = Playwright_.get_text(f'({rowEle})[{i}]//div[@class="sabai-directory-title"]/a')
            productEle = f'({rowEle})[{i}]//div[@class="sabai-directory-category"]/a'
            prodcut = []
            for product_ in range(1, Playwright_.get_count(productEle) + 1):
                prodcut.append(Playwright_.get_text(f'({productEle})[{product_}]').strip())
            prodcut = ','.join(prodcut)
            rowData = [city, shopName, prodcut]
            if rowData in exists_data:
                continue
            ws.append(rowData)
            exists_data.append(rowData)
            logger.info(str(rowData))
        wb.save(dataFile)
        nextCount = Playwright_.get_count(nextPageEle)
        if not nextCount:
            logger.info(f'{city}爬取完毕')
            break
        Playwright_.click(nextPageEle)
        time.sleep(4)


def run():
    proxy = {"server": "http://127.0.0.1:7892"}
    for city, url in info.items():
        getPageinfo(city, url)
    # url = 'https://golfcartresource.com/find-golf-cart-dealer-near/'
    url = 'https://golfcartresource.com/dealer-locator?address_type=state&zoom=15&is_mile=1&directory_radius=50&keywords=&address=Virginia&directory_radius=50&center=&address_type=state&category=0'
    Playwright_.goto(url, proxy=proxy)



if __name__ == '__main__':
    run()